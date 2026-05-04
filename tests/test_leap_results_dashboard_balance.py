from __future__ import annotations

import os
from pathlib import Path

import openpyxl
import pandas as pd
import pytest
from openpyxl.styles import Alignment

from codebase.utilities import leap_results_dashboard_balance as balance_mod
from codebase.utilities.energy_balance_template_extractor import TemplateBalanceExtractor, TemplateLayout
from codebase.utilities.leap_results_dashboard_v2.comparison_engine import build_total_component_ledger


def _write_minimal_balance_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    scenario: str,
    year: int,
    units: str,
    flow: str = "Production",
    fuel: str = "Total",
    value: float = 1.0,
) -> None:
    ws.cell(1, 1, 'LEAP Area "Test Economy"')
    ws.cell(2, 1, f"Scenario: {scenario}, Year: {year}, Units: {units}")
    ws.cell(3, 2, fuel)
    ws.cell(4, 1, flow)
    ws.cell(4, 2, value)


def _make_minimal_workbook(path: Path, *, units_2023: str = "Petajoule", value_2023: float = 1.0) -> None:
    wb = openpyxl.Workbook()
    ws_2023 = wb.active
    ws_2023.title = "EBal|2023"
    _write_minimal_balance_sheet(
        ws_2023,
        scenario="Reference",
        year=2023,
        units=units_2023,
        value=value_2023,
    )

    ws_2022 = wb.create_sheet("EBal|2022")
    _write_minimal_balance_sheet(
        ws_2022,
        scenario="Reference",
        year=2022,
        units="Petajoule",
        value=2.0,
    )
    wb.save(path)


def test_extract_balance_workbook_combines_year_sheets_and_metadata(tmp_path: Path) -> None:
    workbook_path = tmp_path / "mini_balance.xlsx"
    _make_minimal_workbook(workbook_path)

    out = balance_mod._extract_balance_workbook(
        workbook_path,
        template_sheet="EBal|2023",
        mapping_pairs_path=balance_mod.DEFAULT_MAPPING_PAIRS_PATH,
        codebook_path=balance_mod.DEFAULT_CODEBOOK_PATH,
    )

    mapped = out["mapped_long"]
    assert set(pd.to_numeric(mapped["year"], errors="coerce").dropna().astype(int).unique()) == {2022, 2023}
    assert {"EBal|2022", "EBal|2023"}.issubset(set(mapped["source_sheet"].astype(str)))
    assert out["report"]["summary"]["selected_sheet_count"] == 2


def test_extract_balance_workbook_converts_units_to_petajoule(tmp_path: Path) -> None:
    workbook_path = tmp_path / "mini_balance_units.xlsx"
    _make_minimal_workbook(workbook_path, units_2023="Gigajoule", value_2023=1000.0)

    out = balance_mod._extract_balance_workbook(
        workbook_path,
        template_sheet="EBal|2023",
        mapping_pairs_path=balance_mod.DEFAULT_MAPPING_PAIRS_PATH,
        codebook_path=balance_mod.DEFAULT_CODEBOOK_PATH,
    )

    mapped = out["mapped_long"].copy()
    row_2023 = mapped[pd.to_numeric(mapped["year"], errors="coerce").eq(2023)].iloc[0]
    assert row_2023["units_petajoule"] == "Petajoule"
    assert pytest.approx(float(row_2023["value_petajoule"]), rel=1e-9, abs=1e-12) == 0.001


def test_balance_template_extractor_preserves_duplicate_flow_label_occurrences() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(3, 2, "Fuel A")
    rows = [
        ("Total transformation sector", 0, 0.0),
        ("Transport non road", 0, 0.0),
        ("Air", 1, 1.0),
        ("Other parent", 0, 0.0),
        ("Air", 1, 2.0),
        ("Total final energy consumption", 0, 0.0),
    ]
    for row_idx, (name, indent, value) in enumerate(rows, start=4):
        cell = ws.cell(row_idx, 1, name)
        cell.alignment = Alignment(indent=indent)
        ws.cell(row_idx, 2, value)

    extractor = TemplateBalanceExtractor(
        template_sheet="unused",
        mapping_pairs_path=Path("unused.xlsx"),
        codebook_path=Path("unused.xlsx"),
    )
    result = extractor._extract_sheet_matrix(
        ws,
        template=TemplateLayout(flows=[name for name, _, _ in rows], fuels=["Fuel A"]),
    )

    air_rows = result[result["leap_sector_name_raw"].eq("Air")][["leap_sector_name_full_path", "value"]]
    assert air_rows.to_dict("records") == [
        {"leap_sector_name_full_path": "Transport non road/Air", "value": 1.0},
        {"leap_sector_name_full_path": "Other parent/Air", "value": 2.0},
    ]


def test_load_balance_leap_long_filters_full_mappings_and_is_deterministic(monkeypatch: pytest.MonkeyPatch) -> None:
    ref_rows = pd.DataFrame(
        [
            {
                "scenario": "ref",
                "year": 2022,
                "leap_sector": "flow_a",
                "leap_fuel": "fuel_a",
                "esto_flow": "esto_f1",
                "esto_product": "esto_p1",
                "leap_sector_name": "Flow A",
                "leap_fuel_name": "Fuel A",
                "value_petajoule": 1.0,
                "source_sheet": "EBal|2022",
                "source_workbook": "ref.xlsx",
            },
            {
                "scenario": "ref",
                "year": 2022,
                "leap_sector": "flow_a",
                "leap_fuel": "fuel_a",
                "esto_flow": "esto_f1",
                "esto_product": "esto_p1",
                "leap_sector_name": "Flow A",
                "leap_fuel_name": "Fuel A",
                "value_petajoule": 2.0,
                "source_sheet": "EBal|2022",
                "source_workbook": "ref.xlsx",
            },
            {
                "scenario": "ref",
                "year": 2022,
                "leap_sector": "flow_incomplete",
                "leap_fuel": "fuel_a",
                "esto_flow": "esto_f_incomplete",
                "esto_product": "",
                "leap_sector_name": "Flow Incomplete",
                "leap_fuel_name": "Fuel A",
                "value_petajoule": 3.0,
                "source_sheet": "EBal|2022",
                "source_workbook": "ref.xlsx",
            },
            {
                "scenario": "ref",
                "year": 2023,
                "leap_sector": "flow_conflict",
                "leap_fuel": "fuel_a",
                "esto_flow": "esto_conflict_1",
                "esto_product": "esto_p1",
                "leap_sector_name": "Flow Conflict",
                "leap_fuel_name": "Fuel A",
                "value_petajoule": 1.0,
                "source_sheet": "EBal|2023",
                "source_workbook": "ref.xlsx",
            },
            {
                "scenario": "ref",
                "year": 2023,
                "leap_sector": "flow_conflict",
                "leap_fuel": "fuel_a",
                "esto_flow": "esto_conflict_2",
                "esto_product": "esto_p1",
                "leap_sector_name": "Flow Conflict",
                "leap_fuel_name": "Fuel A",
                "value_petajoule": 1.0,
                "source_sheet": "EBal|2023",
                "source_workbook": "ref.xlsx",
            },
            {
                "scenario": "ref",
                "year": 2022,
                "leap_sector": "flow_unmapped",
                "leap_fuel": "fuel_a",
                "esto_flow": "esto_fu",
                "esto_product": "esto_pu",
                "leap_sector_name": "Flow Unmapped",
                "leap_fuel_name": "Fuel A",
                "value_petajoule": 4.0,
                "source_sheet": "EBal|2022",
                "source_workbook": "ref.xlsx",
            },
        ]
    )
    tgt_rows = pd.DataFrame(
        [
            {
                "scenario": "tgt",
                "year": 2022,
                "leap_sector": "flow_a",
                "leap_fuel": "fuel_a",
                "esto_flow": "esto_f1",
                "esto_product": "esto_p1",
                "leap_sector_name": "Flow A",
                "leap_fuel_name": "Fuel A",
                "value_petajoule": 5.0,
                "source_sheet": "EBal|2022",
                "source_workbook": "tgt.xlsx",
            }
        ]
    )

    calls = {"n": 0}

    def _fake_extract(*args, **kwargs):  # noqa: ANN002, ANN003
        calls["n"] += 1
        rows = ref_rows if calls["n"] == 1 else tgt_rows
        return {
            "template_sheet": "EBal|2023",
            "raw_long": rows.copy(),
            "mapped_long": rows.copy(),
            "coverage": pd.DataFrame(),
            "unit_diag": pd.DataFrame(),
            "report": {"summary": {"selected_sheet_count": 2}},
        }

    monkeypatch.setattr(balance_mod, "_extract_balance_workbook", _fake_extract)

    structure = {
        "sheet_catalog": {
            "sheet_a": {"measure": "Energy balance (PJ)"},
            "sheet_conflict": {"measure": "Energy balance (PJ)"},
        },
        "flow_to_sheet": {
            "flow_a": "sheet_a",
            "flow_conflict": "sheet_conflict",
        },
    }

    result1 = balance_mod.load_balance_leap_long(
        ref_workbook_path=balance_mod.DEFAULT_REF_WORKBOOK_PATH,
        tgt_workbook_path=balance_mod.DEFAULT_TGT_WORKBOOK_PATH,
        mapping_pairs_path=balance_mod.DEFAULT_MAPPING_PAIRS_PATH,
        codebook_path=balance_mod.DEFAULT_CODEBOOK_PATH,
        structure_config=structure,
        known_issues={"mapping_overrides": [], "label_overrides": {}, "row_filters": {}},
        projection_economy="20_USA",
    )
    calls["n"] = 0
    result2 = balance_mod.load_balance_leap_long(
        ref_workbook_path=balance_mod.DEFAULT_REF_WORKBOOK_PATH,
        tgt_workbook_path=balance_mod.DEFAULT_TGT_WORKBOOK_PATH,
        mapping_pairs_path=balance_mod.DEFAULT_MAPPING_PAIRS_PATH,
        codebook_path=balance_mod.DEFAULT_CODEBOOK_PATH,
        structure_config=structure,
        known_issues={"mapping_overrides": [], "label_overrides": {}, "row_filters": {}},
        projection_economy="20_USA",
    )

    leap_long = result1["leap_long"]
    assert set(leap_long["scenario"].unique()) == {"Reference", "Target"}
    ref_sum = leap_long[
        leap_long["scenario"].eq("Reference")
        & leap_long["sector_code_9th"].eq("flow_a")
        & leap_long["year"].eq(2022)
    ]["leap_value"].iloc[0]
    assert float(ref_sum) == 3.0

    issue_reasons = set(result1["issues"]["reason"].dropna().astype(str))
    assert "incomplete_mapping" in issue_reasons
    assert "mapping_conflict_after_aggregation" in issue_reasons
    assert "flow_not_in_structure_config" in issue_reasons

    pd.testing.assert_frame_equal(
        result1["leap_long"].reset_index(drop=True),
        result2["leap_long"].reset_index(drop=True),
    )


def test_load_balance_leap_long_applies_overrides_only_from_json(monkeypatch: pytest.MonkeyPatch) -> None:
    rows = pd.DataFrame(
        [
            {
                "scenario": "ref",
                "year": 2022,
                "leap_sector": "flow_x",
                "leap_fuel": "fuel_x",
                "esto_flow": "",
                "esto_product": "",
                "leap_sector_name": "Flow X",
                "leap_fuel_name": "Fuel X",
                "value_petajoule": 1.0,
                "source_sheet": "EBal|2022",
                "source_workbook": "ref.xlsx",
            }
        ]
    )

    def _fake_extract(*args, **kwargs):  # noqa: ANN002, ANN003
        return {
            "template_sheet": "EBal|2022",
            "raw_long": rows.copy(),
            "mapped_long": rows.copy(),
            "coverage": pd.DataFrame(),
            "unit_diag": pd.DataFrame(),
            "report": {"summary": {"selected_sheet_count": 1}},
        }

    monkeypatch.setattr(balance_mod, "_extract_balance_workbook", _fake_extract)

    structure = {
        "sheet_catalog": {"sheet_x": {"measure": "Energy balance (PJ)"}},
        "flow_to_sheet": {"flow_x": "sheet_x"},
    }

    no_override = balance_mod.load_balance_leap_long(
        ref_workbook_path=balance_mod.DEFAULT_REF_WORKBOOK_PATH,
        tgt_workbook_path=balance_mod.DEFAULT_TGT_WORKBOOK_PATH,
        mapping_pairs_path=balance_mod.DEFAULT_MAPPING_PAIRS_PATH,
        codebook_path=balance_mod.DEFAULT_CODEBOOK_PATH,
        structure_config=structure,
        known_issues={"mapping_overrides": [], "label_overrides": {}, "row_filters": {}},
        projection_economy="20_USA",
    )
    assert no_override["leap_long"].empty

    with_override = balance_mod.load_balance_leap_long(
        ref_workbook_path=balance_mod.DEFAULT_REF_WORKBOOK_PATH,
        tgt_workbook_path=balance_mod.DEFAULT_TGT_WORKBOOK_PATH,
        mapping_pairs_path=balance_mod.DEFAULT_MAPPING_PAIRS_PATH,
        codebook_path=balance_mod.DEFAULT_CODEBOOK_PATH,
        structure_config=structure,
        known_issues={
            "mapping_overrides": [
                {
                    "active": True,
                    "match": {"leap_flow": "flow_x", "leap_product": "fuel_x"},
                    "set": {"esto_flow": "esto_flow_x", "esto_product": "esto_product_x"},
                }
            ],
            "label_overrides": {},
            "row_filters": {},
        },
        projection_economy="20_USA",
    )
    assert not with_override["leap_long"].empty
    assert int(with_override["override_report"]["applied_rows"].sum()) > 0


def test_render_balance_dashboards_respects_structure_and_empty_notice(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def _fake_build_charts(comparison_long, charts_dir, backend="plotly", hide_leap_only_charts=False):  # noqa: ANN001
        charts_dir.mkdir(parents=True, exist_ok=True)
        p = charts_dir / "SheetMapped__Energy_balance__PJ__Coal.html"
        p.write_text("<html><body>chart</body></html>", encoding="utf-8")
        return [p]

    monkeypatch.setattr(balance_mod, "build_charts", _fake_build_charts)

    comparison_long = pd.DataFrame(
        [
            {
                "economy": "20_USA",
                "scenario": "Reference",
                "sheet": "SheetMapped",
                "measure": "Energy balance (PJ)",
                "fuel_label": "Coal",
                "source": "leap",
                "year": 2022,
                "value": 1.0,
            }
        ]
    )
    structure = {
        "page_tree": [
            {
                "id": "group",
                "label": "Group",
                "children": [
                    {"id": "mapped", "label": "Mapped", "children": []},
                    {"id": "empty", "label": "Empty", "children": []},
                ],
            }
        ],
        "sheet_catalog": {
            "SheetMapped": {
                "display_label": "SheetMapped",
                "path": ["Group", "Mapped"],
                "measure": "Energy balance (PJ)",
                "sort_order": 0,
            }
        },
        "flow_to_sheet": {},
        "empty_page_notice": "No mapped data on this page.",
    }

    out = balance_mod.render_balance_dashboards(
        comparison_long=comparison_long,
        mapping_status=pd.DataFrame(),
        structure_config=structure,
        output_dir=tmp_path,
        chart_backend="plotly",
        hide_leap_only_charts=False,
    )

    dashboard_index = Path(out["dashboard_index"])
    assert dashboard_index.exists()

    empty_page = Path(out["dashboards_dir"]) / balance_mod._page_filename_from_path(("Group", "Empty"))
    assert empty_page.exists()
    html = empty_page.read_text(encoding="utf-8")
    assert "No mapped data on this page." in html


def test_template_esto_axis_records_strips_code_prefix_from_fuel_label() -> None:
    """Template-declared fuel_label must match the LEAP-observed path (code prefix stripped).

    LEAP-observed rows set fuel_label via _strip_esto_code_prefix(esto_product), e.g.
    esto_product="08.01 Natural gas" → fuel_label="Natural gas".
    Template rows must produce the same label so the two paths deduplicate correctly and
    the safety guard does not fire on phantom duplicate ESTO pairs.
    """
    template = {
        "defaults": {"measure": "Energy balance (PJ)"},
        "Energy": {
            "graphs": {
                "natural_gas": {
                    "esto_flow": "08 Natural gas",
                    "products": ["08.01 Natural gas", "Total"],
                }
            }
        },
    }

    import unittest.mock as mock

    with mock.patch.object(balance_mod, "_load_dashboard_template_allowlist", return_value=template):
        result = balance_mod._dashboard_template_esto_axis_records(
            None,
            scenario_names=["Reference"],
        )

    assert not result.empty, "Expected at least one row from template"
    coded_label_rows = result[result["fuel_label"].str.contains(r"^\d{2}", regex=True)]
    assert coded_label_rows.empty, (
        f"fuel_label must not contain ESTO code prefixes; got: {coded_label_rows['fuel_label'].tolist()}"
    )
    natural_gas = result[result["esto_product"].str.contains("Natural gas", case=False)]
    assert not natural_gas.empty, "Expected a Natural gas row"
    assert (natural_gas["fuel_label"] == "Natural gas").all(), (
        f"Expected fuel_label='Natural gas', got: {natural_gas['fuel_label'].tolist()}"
    )
    total_rows = result[result["esto_product"] == "Total"]
    assert not total_rows.empty, "Expected a Total row"
    assert (total_rows["fuel_label"] == "Total").all(), (
        f"Expected fuel_label='Total' for Total product, got: {total_rows['fuel_label'].tolist()}"
    )


def test_template_esto_axis_records_emit_page_and_chart_group_fields() -> None:
    template = {
        "defaults": {"measure": "Energy balance (PJ)"},
        "Buildings": {
            "Commercial": {
                "graphs": {
                    "Electricity": {
                        "esto_flow": "16.01 Commercial and public services",
                        "products": ["17 Electricity"],
                    }
                }
            }
        },
    }

    import unittest.mock as mock

    with mock.patch.object(balance_mod, "_load_dashboard_template_allowlist", return_value=template):
        result = balance_mod._dashboard_template_esto_axis_records(
            None,
            scenario_names=["Reference"],
        )

    row = result.iloc[0]
    assert row["sheet"] == "esto__16_01__Commercial_and_public_services"
    assert row["page_label"] == "Buildings"
    assert row["page_key"] == "buildings"
    assert row["chart_group_label"] == "Commercial"
    assert row["chart_group_key"]


def test_v2_template_schema_parses_aggregate_and_by_fuel_nodes() -> None:
    template = {
        "defaults": {"measure": "Energy balance (PJ)"},
        "Industry": {
            "aggregate_graphs": {
                "fuels": "Total",
                "esto_flows": ["14.02 Construction"],
            },
            "Construction": {
                "by_fuel_graphs": {
                    "esto_flows": ["14.02 Construction"],
                    "products": "All",
                }
            },
        },
    }

    import unittest.mock as mock

    with mock.patch.object(balance_mod, "_load_dashboard_template_allowlist", return_value=template):
        structure = balance_mod.build_esto_axis_structure_from_dashboard_template(None)

    assert "14.02 Construction" in structure["esto_flow_to_sheet"]
    sheet = structure["esto_flow_to_sheet"]["14.02 Construction"]
    assert structure["sheet_catalog"][sheet]["path"] == ["Industry sector", "Construction"]


def test_template_all_products_uses_nonzero_leap_base_and_projection_rows() -> None:
    template = {
        "defaults": {"measure": "Energy balance (PJ)"},
        "Supply": {
            "Imports": {
                "by_fuel_graphs": {
                    "esto_flows": ["02 Imports"],
                    "products": "All",
                }
            }
        },
    }
    leap_working = pd.DataFrame(
        [
            {
                "scenario": "Target",
                "esto_flow": "02 Imports",
                "esto_product": "01.02 Other bituminous coal",
                "leap_value": -2.0,
                "esto_is_subtotal": False,
            }
        ]
    )
    base_df = pd.DataFrame(
        [
            {
                "economy": "20USA",
                "flows": "02 Imports",
                "products": "08.01 Natural gas",
                "2022": 3.0,
                "is_subtotal": False,
            }
        ]
    )

    import unittest.mock as mock

    def fake_projection_series(_ninth_df, *, fuel_code, projection_years, **_kwargs):  # noqa: ANN001
        value = 4.0 if fuel_code == "07_01_motor_gasoline" else 0.0
        return pd.Series({year: value for year in projection_years})

    with (
        mock.patch.object(balance_mod, "_load_dashboard_template_allowlist", return_value=template),
        mock.patch.object(balance_mod, "pull_projection_series", side_effect=fake_projection_series),
    ):
        result = balance_mod._dashboard_template_esto_axis_records(
            None,
            scenario_names=["Target"],
            leap_working=leap_working,
            base_df=base_df,
            ninth_df=pd.DataFrame({"dummy": [1]}),
            esto_to_ninth={
                ("02 Imports", "07.01 Motor gasoline"): [("sector_a", "07_01_motor_gasoline")],
                ("02 Imports", "16.01 Biogas"): [("sector_a", "16_01_biogas")],
            },
            base_year=2022,
            base_economy="20USA",
            projection_economy="20_USA",
            projection_years=[2023],
            scenario_to_projection={"target": "target"},
        )

    assert set(result["esto_product"]) == {
        "01.02 Other bituminous coal",
        "08.01 Natural gas",
        "07.01 Motor gasoline",
    }
    assert "Other bituminous coal" in set(result["fuel_label"])


def test_multi_flow_by_fuel_template_sums_component_flows_in_comparison() -> None:
    template = {
        "defaults": {"measure": "Energy balance (PJ)"},
        "Supply": {
            "Imports plus exports": {
                "by_fuel_graphs": {
                    "esto_flows": ["02 Imports", "03 Exports"],
                    "products": "All",
                }
            }
        },
    }
    leap_long = pd.DataFrame(
        [
            {
                "scenario": "Target",
                "sheet_name": "imports",
                "measure": "Energy balance (PJ)",
                "fuel_label": "Natural gas",
                "year": 2023,
                "leap_value": 10.0,
                "esto_flow": "02 Imports",
                "esto_product": "08.01 Natural gas",
            },
            {
                "scenario": "Target",
                "sheet_name": "exports",
                "measure": "Energy balance (PJ)",
                "fuel_label": "Natural gas",
                "year": 2023,
                "leap_value": -4.0,
                "esto_flow": "03 Exports",
                "esto_product": "08.01 Natural gas",
            },
        ]
    )
    mapping_status = leap_long[
        ["sheet_name", "measure", "fuel_label", "esto_flow", "esto_product"]
    ].rename(columns={"sheet_name": "sheet"})
    mapping_status["sector_code_9th"] = ""
    mapping_status["ninth_fuel_code"] = ""
    base_df = pd.DataFrame(
        [
            {"economy": "20USA", "flows": "02 Imports", "products": "08.01 Natural gas", "2022": 2.0, "is_subtotal": False},
            {"economy": "20USA", "flows": "03 Exports", "products": "08.01 Natural gas", "2022": -5.0, "is_subtotal": False},
        ]
    )

    import unittest.mock as mock

    with mock.patch.object(balance_mod, "_load_dashboard_template_allowlist", return_value=template):
        result = balance_mod.build_balance_comparison_esto_axis(
            leap_long=leap_long,
            mapping_status=mapping_status,
            base_year=2022,
            projection_years=[2023],
            base_economy="20USA",
            projection_economy="20_USA",
            scenario_map={"Target": "target"},
            base_df=base_df,
            ninth_df=pd.DataFrame(),
            chart_navigation_guide_path=None,
        )

    comparison = result["comparison_long"]
    virtual = comparison[
        comparison["sheet"].astype(str).str.startswith("template__")
        & comparison["fuel_label"].eq("Natural gas")
    ]
    assert not virtual.empty
    leap_value = virtual.loc[virtual["source"].eq("leap"), "value"].sum()
    base_value = virtual.loc[virtual["source"].eq("base"), "value"].sum()
    assert leap_value == pytest.approx(6.0)
    assert base_value == pytest.approx(-3.0)
    status = result["mapping_status"]
    virtual_status = status[
        status["sheet"].astype(str).str.startswith("template__")
        & status["fuel_label"].eq("Natural gas")
    ]
    assert set(virtual_status["esto_flow"]) == {"02 Imports", "03 Exports"}
    assert virtual_status["esto_flow_group_key"].nunique() == 1
    assert virtual_status["esto_flow_group_label"].iloc[0] == "Imports plus exports"


def test_build_balance_comparison_esto_axis_drops_parent_ninth_targets_when_children_exist() -> None:
    leap_long = pd.DataFrame(
        [
            {
                "scenario": "Target",
                "sheet_name": "steel",
                "measure": "Energy balance (PJ)",
                "fuel_label": "Anthracite",
                "year": 2023,
                "leap_value": 1.0,
                "esto_flow": "14.03.01 Iron and steel",
                "esto_product": "01.04 Anthracite",
            }
        ]
    )
    mapping_status = pd.DataFrame(
        [
            {
                "sheet": "steel",
                "measure": "Energy balance (PJ)",
                "fuel_label": "Anthracite",
                "esto_flow": "14.03.01 Iron and steel",
                "esto_product": "01.04 Anthracite",
                "sector_code_9th": "14_03_01_01_fs|14_03_01_03_ccs|14_03_01_iron_and_steel",
                "ninth_fuel_code": "01_x_thermal_coal",
            }
        ]
    )
    base_df = pd.DataFrame(
        [
            {
                "economy": "20USA",
                "flows": "14.03.01 Iron and steel",
                "products": "01.04 Anthracite",
                "2022": 0.0,
                "is_subtotal": False,
            }
        ]
    )
    ninth_df = pd.DataFrame(
        [
            {
                "economy": "20_USA",
                "scenarios": "target",
                "sectors": "14_industry_sector",
                "sub1sectors": "14_03_manufacturing",
                "sub2sectors": "14_03_01_iron_and_steel",
                "sub3sectors": "x",
                "sub4sectors": "x",
                "fuels": "01_coal",
                "subfuels": "01_x_thermal_coal",
                "subtotal_layout": True,
                "subtotal_results": True,
                "2023": 100.0,
            },
            {
                "economy": "20_USA",
                "scenarios": "target",
                "sectors": "14_industry_sector",
                "sub1sectors": "14_03_manufacturing",
                "sub2sectors": "14_03_01_iron_and_steel",
                "sub3sectors": "14_03_01_01_fs",
                "sub4sectors": "x",
                "fuels": "01_coal",
                "subfuels": "01_x_thermal_coal",
                "subtotal_layout": False,
                "subtotal_results": False,
                "2023": 7.0,
            },
            {
                "economy": "20_USA",
                "scenarios": "target",
                "sectors": "14_industry_sector",
                "sub1sectors": "14_03_manufacturing",
                "sub2sectors": "14_03_01_iron_and_steel",
                "sub3sectors": "14_03_01_03_ccs",
                "sub4sectors": "x",
                "fuels": "01_coal",
                "subfuels": "01_x_thermal_coal",
                "subtotal_layout": False,
                "subtotal_results": False,
                "2023": 3.0,
            },
        ]
    )

    result = balance_mod.build_balance_comparison_esto_axis(
        leap_long=leap_long,
        mapping_status=mapping_status,
        base_year=2022,
        projection_years=[2023],
        base_economy="20USA",
        projection_economy="20_USA",
        scenario_map={"Target": "target"},
        base_df=base_df,
        ninth_df=ninth_df,
        chart_navigation_guide_path=None,
    )

    comparison = result["comparison_long"]
    projection_value = comparison.loc[
        comparison["source"].eq("projection")
        & comparison["fuel_label"].eq("Anthracite")
        & pd.to_numeric(comparison["year"], errors="coerce").eq(2023),
        "value",
    ].sum()
    assert projection_value == pytest.approx(10.0)
    status = result["mapping_status"]
    target_text = "|".join(status["sector_code_9th"].fillna("").astype(str).unique())
    assert "14_03_01_iron_and_steel" not in target_text
    assert "14_03_01_01_fs" in target_text
    assert "14_03_01_03_ccs" in target_text


def test_build_balance_comparison_esto_axis_dedupes_shared_ninth_pairs_by_sheet() -> None:
    """Shared 9th rows should only feed the first ESTO-product chart on a sheet.

    Domestic air transport can map both gasoline-type and kerosene-type jet fuel
    ESTO products to the same 9th ``07_x_jet_fuel`` series. The pre-render
    chart_group_key can differ by input source, so the dedupe must use the
    stable ESTO-axis sheet rather than the chart-group alias.
    """
    leap_long = pd.DataFrame(
        [
            {
                "scenario": "Target",
                "sheet_name": "esto__15_01__Domestic_air_transport",
                "page_key": "transport",
                "page_label": "Transport",
                "chart_group_key": "chart_group_alias_a",
                "chart_group_label": "Domestic air transport",
                "measure": "Energy balance (PJ)",
                "fuel_label": "Gasoline type jet fuel",
                "year": 2023,
                "leap_value": 1.0,
                "esto_flow": "15.01 Domestic air transport",
                "esto_product": "07.04 Gasoline type jet fuel",
            },
            {
                "scenario": "Target",
                "sheet_name": "esto__15_01__Domestic_air_transport",
                "page_key": "transport",
                "page_label": "Transport",
                "chart_group_key": "chart_group_alias_b",
                "chart_group_label": "Domestic air transport",
                "measure": "Energy balance (PJ)",
                "fuel_label": "Kerosene type jet fuel",
                "year": 2023,
                "leap_value": 2.0,
                "esto_flow": "15.01 Domestic air transport",
                "esto_product": "07.05 Kerosene type jet fuel",
            },
        ]
    )
    mapping_status = pd.DataFrame(
        [
            {
                "sheet": "esto__15_01__Domestic_air_transport",
                "chart_group_key": "chart_group_alias_a",
                "measure": "Energy balance (PJ)",
                "fuel_label": "Gasoline type jet fuel",
                "esto_flow": "15.01 Domestic air transport",
                "esto_product": "07.04 Gasoline type jet fuel",
                "sector_code_9th": "15_01_01_passenger",
                "ninth_fuel_code": "07_x_jet_fuel",
            },
            {
                "sheet": "esto__15_01__Domestic_air_transport",
                "chart_group_key": "chart_group_alias_b",
                "measure": "Energy balance (PJ)",
                "fuel_label": "Kerosene type jet fuel",
                "esto_flow": "15.01 Domestic air transport",
                "esto_product": "07.05 Kerosene type jet fuel",
                "sector_code_9th": "15_01_01_passenger",
                "ninth_fuel_code": "07_x_jet_fuel",
            },
        ]
    )
    base_df = pd.DataFrame(
        [
            {
                "economy": "20USA",
                "flows": "15.01 Domestic air transport",
                "products": "07.04 Gasoline type jet fuel",
                "2022": 0.0,
                "is_subtotal": False,
            },
            {
                "economy": "20USA",
                "flows": "15.01 Domestic air transport",
                "products": "07.05 Kerosene type jet fuel",
                "2022": 0.0,
                "is_subtotal": False,
            },
        ]
    )
    ninth_df = pd.DataFrame(
        [
            {
                "economy": "20_USA",
                "scenarios": "target",
                "sectors": "15_transport_sector",
                "sub1sectors": "15_01_domestic_air_transport",
                "sub2sectors": "15_01_01_passenger",
                "sub3sectors": "x",
                "sub4sectors": "x",
                "fuels": "07_petroleum_products",
                "subfuels": "07_x_jet_fuel",
                "subtotal_layout": False,
                "subtotal_results": False,
                "2023": 123.0,
            },
        ]
    )

    result = balance_mod.build_balance_comparison_esto_axis(
        leap_long=leap_long,
        mapping_status=mapping_status,
        base_year=2022,
        projection_years=[2023],
        base_economy="20USA",
        projection_economy="20_USA",
        scenario_map={"Target": "target"},
        base_df=base_df,
        ninth_df=ninth_df,
        chart_navigation_guide_path=None,
    )

    projection = result["comparison_long"][
        result["comparison_long"]["source"].eq("projection")
        & pd.to_numeric(result["comparison_long"]["year"], errors="coerce").eq(2023)
    ]
    values = projection.set_index("fuel_label")["value"]

    assert values["Gasoline type jet fuel"] == pytest.approx(123.0)
    assert pd.isna(values["Kerosene type jet fuel"])
    assert projection["esto_flow_group_key"].nunique() == 1
    assert set(projection["dashboard_section_key"]) == {"chart_group_alias_a", "chart_group_alias_b"}


def test_render_balance_dashboards_exposes_multiple_chart_groups_on_one_page(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def _fake_build_charts(comparison_long, charts_dir, backend="plotly", hide_leap_only_charts=False):  # noqa: ANN001
        charts_dir.mkdir(parents=True, exist_ok=True)
        written = []
        for (sheet, measure, fuel), _ in comparison_long.groupby(["sheet", "measure", "fuel_label"]):
            file_name = f"{balance_mod._safe_token(f'{sheet}__{measure}')}__{balance_mod._safe_token(fuel)}.html"
            path = charts_dir / file_name
            path.write_text("<html><body>chart</body></html>", encoding="utf-8")
            written.append(path)
        return written

    monkeypatch.setattr(balance_mod, "build_charts", _fake_build_charts)

    template = {
        "defaults": {"measure": "Energy balance (PJ)"},
        "Buildings": {
            "Commercial": {
                "graphs": {"Electricity": {"esto_flow": "16.01 Commercial and public services", "products": ["17 Electricity"]}}
            },
            "Residential": {
                "graphs": {"Electricity": {"esto_flow": "16.02 Residential", "products": ["17 Electricity"]}}
            },
        },
    }
    structure = balance_mod.build_esto_axis_structure_from_dashboard_template(None)
    with monkeypatch.context() as m:
        m.setattr(balance_mod, "_load_dashboard_template_allowlist", lambda _path: template)
        structure = balance_mod.build_esto_axis_structure_from_dashboard_template(None)
        comparison_long = pd.DataFrame(
            [
                {
                    "economy": "20_USA",
                    "scenario": "Target",
                    "sheet": "esto__16_01__Commercial_and_public_services",
                    "measure": "Energy balance (PJ)",
                    "fuel_label": "Electricity",
                    "source": "leap",
                    "year": 2030,
                    "value": 1.0,
                },
                {
                    "economy": "20_USA",
                    "scenario": "Target",
                    "sheet": "esto__16_02__Residential",
                    "measure": "Energy balance (PJ)",
                    "fuel_label": "Electricity",
                    "source": "leap",
                    "year": 2030,
                    "value": 2.0,
                },
            ]
        )
        mapping_status = pd.DataFrame(
            [
                {
                    "sheet": "esto__16_01__Commercial_and_public_services",
                    "measure": "Energy balance (PJ)",
                    "fuel_label": "Electricity",
                    "esto_flow": "16.01 Commercial and public services",
                    "esto_product": "17 Electricity",
                },
                {
                    "sheet": "esto__16_02__Residential",
                    "measure": "Energy balance (PJ)",
                    "fuel_label": "Electricity",
                    "esto_flow": "16.02 Residential",
                    "esto_product": "17 Electricity",
                },
            ]
        )
        out = balance_mod.render_balance_dashboards(
            comparison_long=comparison_long,
            mapping_status=mapping_status,
            structure_config=structure,
            output_dir=tmp_path,
            chart_backend="plotly",
            chart_navigation_guide_path=None,
        )

    exposure = pd.read_csv(out["chart_group_exposure"])
    assert set(exposure["page_label"]) == {"Buildings"}
    assert {"Commercial", "Residential"}.issubset(set(exposure["chart_group_label"]))
    assert exposure["chart_group_key"].nunique() == 2


def test_render_balance_dashboards_allows_unmeasured_transformation_by_fuel_graphs(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def _fake_build_charts(comparison_long, charts_dir, backend="plotly", hide_leap_only_charts=False):  # noqa: ANN001
        charts_dir.mkdir(parents=True, exist_ok=True)
        written = []
        for (sheet, measure, fuel), _ in comparison_long.groupby(["sheet", "measure", "fuel_label"]):
            file_name = f"{balance_mod._safe_token(f'{sheet}__{measure}')}__{balance_mod._safe_token(fuel)}.html"
            path = charts_dir / file_name
            path.write_text("<html><body>chart</body></html>", encoding="utf-8")
            written.append(path)
        return written

    monkeypatch.setattr(balance_mod, "build_charts", _fake_build_charts)
    template = {
        "defaults": {"measure": "Energy balance (PJ)"},
        "Refining": {
            "Oil refineries": {
                "by_fuel_graphs": {
                    "esto_flows": ["09.07 Oil refineries"],
                    "products": "All",
                }
            }
        },
    }

    with monkeypatch.context() as m:
        m.setattr(balance_mod, "_load_dashboard_template_allowlist", lambda _path: template)
        structure = balance_mod.build_esto_axis_structure_from_dashboard_template(None)
        sheet = structure["esto_flow_to_sheet"]["09.07 Oil refineries"]
        comparison_long = pd.DataFrame(
            [
                {
                    "economy": "20_USA",
                    "scenario": "Target",
                    "sheet": sheet,
                    "measure": "Energy balance (PJ)",
                    "fuel_label": "Crude oil",
                    "source": "leap",
                    "year": 2030,
                    "value": -1.0,
                },
                {
                    "economy": "20_USA",
                    "scenario": "Target",
                    "sheet": sheet,
                    "measure": "Energy balance (PJ)",
                    "fuel_label": "Crude oil",
                    "source": "projection",
                    "year": 2030,
                    "value": 2.0,
                },
            ]
        )
        mapping_status = pd.DataFrame(
            [
                {
                    "sheet": sheet,
                    "measure": "Energy balance (PJ)",
                    "fuel_label": "Crude oil",
                    "esto_flow": "09.07 Oil refineries",
                    "esto_product": "06.01 Crude oil",
                }
            ]
        )
        out = balance_mod.render_balance_dashboards(
            comparison_long=comparison_long,
            mapping_status=mapping_status,
            structure_config=structure,
            output_dir=tmp_path,
            chart_backend="plotly",
            chart_navigation_guide_path=None,
        )

    exposure = pd.read_csv(out["chart_group_exposure"])
    crude_rows = exposure[
        exposure["entry_kind"].eq("direct")
        & exposure["fuel_label"].eq("Crude oil")
    ]
    assert set(crude_rows["measure"]) == {
        balance_mod.TRANSFORMATION_INPUT_MEASURE,
        balance_mod.TRANSFORMATION_OUTPUT_MEASURE,
    }


def test_render_balance_dashboards_writes_refining_aggregate_without_leap_rows(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def _fake_build_charts(comparison_long, charts_dir, backend="plotly", hide_leap_only_charts=False):  # noqa: ANN001
        charts_dir.mkdir(parents=True, exist_ok=True)
        written = []
        for (sheet, measure, fuel), _ in comparison_long.groupby(["sheet", "measure", "fuel_label"]):
            file_name = f"{balance_mod._safe_token(f'{sheet}__{measure}')}__{balance_mod._safe_token(fuel)}.html"
            path = charts_dir / file_name
            path.write_text("<html><body>chart</body></html>", encoding="utf-8")
            written.append(path)
        return written

    def _fake_make_chart(sheet, fuel, subset, output_dir, backend="plotly", display_sheet=None, file_sheet=None):  # noqa: ANN001
        output_dir.mkdir(parents=True, exist_ok=True)
        path = output_dir / f"{balance_mod._safe_token(file_sheet or sheet)}__{balance_mod._safe_token(fuel)}.html"
        path.write_text("<html><body>aggregate</body></html>", encoding="utf-8")
        return path

    template = {
        "defaults": {"measure": "Energy balance (PJ)"},
        "Refining": {
            "aggregate_graphs": {
                "fuels": "Total",
                "esto_flows": ["09.07 Oil refineries"],
                "measures": [
                    balance_mod.TRANSFORMATION_INPUT_MEASURE,
                    balance_mod.TRANSFORMATION_OUTPUT_MEASURE,
                ],
            },
            "Oil refineries": {
                "by_fuel_graphs": {
                    "esto_flows": ["09.07 Oil refineries"],
                    "products": "All",
                }
            },
        },
    }

    with monkeypatch.context() as m:
        m.setattr(balance_mod, "_load_dashboard_template_allowlist", lambda _path: template)
        m.setattr(balance_mod, "build_charts", _fake_build_charts)
        m.setattr(balance_mod, "make_chart", _fake_make_chart)
        structure = balance_mod.build_esto_axis_structure_from_dashboard_template(None)
        sheet = structure["esto_flow_to_sheet"]["09.07 Oil refineries"]
        comparison_long = pd.DataFrame(
            [
                {
                    "economy": "20_USA",
                    "scenario": "Target",
                    "sheet": sheet,
                    "measure": "Energy balance (PJ)",
                    "fuel_label": "Crude oil",
                    "source": "base",
                    "year": 2022,
                    "value": -10.0,
                },
                {
                    "economy": "20_USA",
                    "scenario": "Target",
                    "sheet": sheet,
                    "measure": "Energy balance (PJ)",
                    "fuel_label": "Crude oil",
                    "source": "projection",
                    "year": 2030,
                    "value": -11.0,
                },
                {
                    "economy": "20_USA",
                    "scenario": "Target",
                    "sheet": sheet,
                    "measure": "Energy balance (PJ)",
                    "fuel_label": "Motor gasoline",
                    "source": "base",
                    "year": 2022,
                    "value": 8.0,
                },
                {
                    "economy": "20_USA",
                    "scenario": "Target",
                    "sheet": sheet,
                    "measure": "Energy balance (PJ)",
                    "fuel_label": "Motor gasoline",
                    "source": "projection",
                    "year": 2030,
                    "value": 9.0,
                },
            ]
        )
        mapping_status = pd.DataFrame(
            [
                {
                    "sheet": sheet,
                    "measure": "Energy balance (PJ)",
                    "fuel_label": "Crude oil",
                    "esto_flow": "09.07 Oil refineries",
                    "esto_product": "06.01 Crude oil",
                },
                {
                    "sheet": sheet,
                    "measure": "Energy balance (PJ)",
                    "fuel_label": "Motor gasoline",
                    "esto_flow": "09.07 Oil refineries",
                    "esto_product": "07.01 Motor gasoline",
                },
            ]
        )
        out = balance_mod.render_balance_dashboards(
            comparison_long=comparison_long,
            mapping_status=mapping_status,
            structure_config=structure,
            output_dir=tmp_path,
            chart_backend="plotly",
            chart_navigation_guide_path=None,
        )

    exposure = pd.read_csv(out["chart_group_exposure"])
    aggregate_rows = exposure[
        exposure["entry_kind"].eq("aggregate")
        & exposure["dashboard_path"].eq("Refining")
        & exposure["fuel_label"].eq("Total")
    ]
    assert set(aggregate_rows["measure"]) == {
        balance_mod.TRANSFORMATION_INPUT_MEASURE,
        balance_mod.TRANSFORMATION_OUTPUT_MEASURE,
    }


def test_mapping_lineage_audit_attaches_only_rendered_chart_groups(tmp_path: Path) -> None:
    chart_groups = pd.DataFrame(
        [
            {
                "chart_group_id": "chart::charts/construction.html",
                "dashboard_path": "Industry > Construction",
                "chart_file": "charts/construction.html",
                "page_key": "industry",
                "page_label": "Industry",
                "chart_group_key": "industry__construction",
                "chart_group_label": "Construction",
                "section_id": "construction",
                "section_label": "Construction",
                "entry_kind": "direct",
                "sheet": "esto__14_02__Construction",
                "measure": "Energy balance (PJ)",
                "fuel_label": "Anthracite",
            }
        ]
    )
    chart_groups_path = tmp_path / "chart_group_exposure.csv"
    chart_groups.to_csv(chart_groups_path, index=False)
    all_chart_groups = chart_groups.copy()
    all_chart_groups["exposed_in_dashboard"] = True
    all_chart_groups_path = tmp_path / "all_chart_groups.csv"
    all_chart_groups.to_csv(all_chart_groups_path, index=False)

    lineage = pd.DataFrame(
        [
            {
                "dataset": "9th",
                "scenario": "Target",
                "year": 2023,
                "esto_flow": "14.02 Construction",
                "esto_product": "01.04 Anthracite",
                "source_sector": "14_02_construction",
                "source_fuel": "01_x_thermal_coal",
                "value_pj": 1.0,
            },
            {
                "dataset": "9th",
                "scenario": "Target",
                "year": 2023,
                "esto_flow": "14 Industry sector",
                "esto_product": "01.04 Anthracite",
                "source_sector": "14_industry_sector",
                "source_fuel": "01_x_thermal_coal",
                "value_pj": 2.0,
            },
        ]
    )

    attached = balance_mod.attach_chart_groups_to_mapping_lineage_audit(
        lineage,
        chart_groups_path,
        all_chart_groups_path,
    )

    assert len(attached) == 1
    row = attached.iloc[0]
    assert row["chart_group_id"] == "chart::charts/construction.html"
    assert row["esto_flow"] == "14.02 Construction"
    assert row["source_sector"] == "14_02_construction"


def test_total_component_ledger_dedup_scopes_to_chart_group_key() -> None:
    chart_rows = pd.DataFrame(
        [
            {"economy": "20_USA", "sheet": "Legacy", "page_key": "buildings", "page_label": "Buildings", "chart_group_key": "commercial", "chart_group_label": "Commercial", "measure": "Energy balance (PJ)", "fuel_label": "Electricity", "scenario": "Target", "source": "projection", "year": 2030, "value": 1.0},
            {"economy": "20_USA", "sheet": "Legacy", "page_key": "buildings", "page_label": "Buildings", "chart_group_key": "residential", "chart_group_label": "Residential", "measure": "Energy balance (PJ)", "fuel_label": "Electricity", "scenario": "Target", "source": "projection", "year": 2030, "value": 2.0},
        ]
    )
    mapping_status = pd.DataFrame(
        [
            {"sheet": "Legacy", "chart_group_key": "commercial", "measure": "Energy balance (PJ)", "fuel_label": "Electricity", "sector_code_9th": "16_01_commercial", "ninth_fuel_code": "17_electricity"},
            {"sheet": "Legacy", "chart_group_key": "residential", "measure": "Energy balance (PJ)", "fuel_label": "Electricity", "sector_code_9th": "16_02_residential", "ninth_fuel_code": "17_electricity"},
        ]
    )

    ledger = build_total_component_ledger(chart_rows, mapping_status)

    assert set(ledger["chart_group_key"]) == {"commercial", "residential"}
    assert set(ledger["duplicate_exact_comparator_key_count"]) == {1}


def test_attach_chart_groups_uses_legacy_sheet_fallback(tmp_path: Path) -> None:
    chart_groups_path = tmp_path / "chart_group_exposure.csv"
    pd.DataFrame(
        [
            {
                "chart_group_id": "chart::legacy.html",
                "dashboard_path": "Legacy Page",
                "sheet": "LegacySheet",
                "measure": "Energy balance (PJ)",
                "fuel_label": "Coal",
                "chart_file": "charts/legacy.html",
                "section_id": "sec-legacy",
                "section_label": "Legacy Page",
                "entry_kind": "direct",
            }
        ]
    ).to_csv(chart_groups_path, index=False)
    exposure = pd.DataFrame(
        [
            {
                "sheet": "LegacySheet",
                "measure": "Energy balance (PJ)",
                "fuel_label": "Coal",
                "source": "base",
                "year": 2022,
                "value": 1.0,
            }
        ]
    )

    out = balance_mod.attach_chart_groups_to_dashboard_exposure(exposure, chart_groups_path)

    assert out["chart_group_id"].iloc[0] == "chart::legacy.html"
    assert out["chart_group_key"].iloc[0]


@pytest.mark.integration
def test_balance_dashboard_workflow_integration_usa_inputs() -> None:
    if os.getenv("RUN_BALANCE_INTEGRATION", "").strip().lower() not in {"1", "true", "yes", "y"}:
        pytest.skip("Set RUN_BALANCE_INTEGRATION=1 to run full integration workflow test.")

    from codebase.leap_results_dashboard_balance_workflow import run_workflow

    result = run_workflow()
    out_dir = Path(result["comparison_long"]).resolve().parent

    required = [
        out_dir / "comparison_long.csv",
        out_dir / "comparison_wide.csv",
        out_dir / "mapping_status.xlsx",
        out_dir / "chart_line_mapping_ledger.csv",
        out_dir / "chart_total_component_ledger.csv",
        out_dir / "comparison_gap_diagnostics.csv",
        out_dir / "dashboards/index.html",
    ]
    for path in required:
        assert path.exists(), f"Missing expected output: {path}"

    comparison_long = pd.read_csv(out_dir / "comparison_long.csv")
    assert {"leap", "base", "projection"}.issubset(set(comparison_long["source"].dropna().astype(str).unique()))
    years = pd.to_numeric(comparison_long["year"], errors="coerce").dropna().astype(int)
    assert int(years.min()) <= 2022
    assert int(years.max()) >= 2060

    leap_long = pd.read_csv(out_dir / "leap_long.csv")
    assert set(leap_long["leap_units"].dropna().astype(str).unique()) == {"Petajoule"}
