#%%
"""
Load LEAP balance exports into dataframes for ESTO-axis balance comparisons.

This workflow extracts REF/TGT LEAP balance workbooks into long-form mapped
rows, then compares them against ESTO and 9th reference data using ESTO balance
rows as the chart axis. It writes the comparison tables, mapping diagnostics,
ledgers, and rendered dashboard outputs under the configured output directory.
 
Why ESTO axis:
- ESTO flow/product rows are a practical middle ground for balance-table
  comparison. They are usually less granular than the full 9th sector/fuel
  hierarchy, but more structured and comparable than raw LEAP balance labels.
- The workflow maps LEAP balance cells to ESTO flow/product pairs and maps 9th
  sector/fuel projections back to those same ESTO pairs, so LEAP, ESTO, and 9th
  can be compared on one common axis.
- Mapping lineage is the primary source of truth for charted 9th rows. The
  workflow should first use mappings carried through the LEAP mapping lineage
  audit path. Template-level use_esto_to_ninth_mapping is a fallback for
  ESTO-axis rows where no LEAP-derived mapping lineage can be created; it
  reaches directly into the ESTO-to-9th canonical mapping and should not replace
  valid LEAP-derived lineage.
- The dashboard target universe is built from the active
  ESTO -> LEAP -> 9th crosswalk in config/leap_mappings.xlsx:
  leap_combined_esto is joined to leap_combined_ninth on the LEAP sector/fuel
  path before chart rows are generated. This means mapped 9th rows can still be
  shown when the corresponding LEAP export row is zero or absent, as long as the
  9th pair is nonzero in the requested projection slice.
- LEAP-backed rows get priority when many ESTO products share the same 9th
  sector/fuel pair. Workbook-only template rows fill gaps, but should not steal
  shared 9th values from rows that have direct LEAP balance evidence.
- Transformation charts split signed balance rows at render time: negative
  values are inputs and displayed as positive magnitudes; positive values are
  outputs.
- Simple audit outputs are written for the key ESTO-axis inputs:
  LEAP-to-ESTO rows, 9th-to-ESTO rows, and the combined comparison table.
- See docs/esto_axis_balance_dashboard_system.md for the full system guide,
  including mapping expansion, subtotal rules, many-to-many behavior, and
  debugging steps.
"""

from __future__ import annotations

import json
import os
import re
import shutil
import sys
from pathlib import Path
from typing import Sequence

import openpyxl
from openpyxl import load_workbook

import pandas as pd

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from codebase.utilities.leap_results_dashboard_balance import (  # noqa: E402
    attach_chart_groups_to_dashboard_exposure,
    attach_chart_groups_to_mapping_lineage_audit,
    build_mapped_ninth_to_esto_balance_rows,
    build_mapping_lineage_audit_table,
    build_merged_esto_axis_balance_table,
    build_simple_balance_duplicate_diagnostics,
    build_ninth_balance_esto_long_table,
    build_simple_ninth_balance_table,
    build_balance_comparison_esto_axis,
    build_esto_axis_structure_from_dashboard_template,
    convert_leap_balances_to_esto_long_table,
    render_balance_dashboards,
    simplify_chart_line_mapping_ledger_output,
    simplify_chart_total_component_ledger_output,
    simplify_mapping_lineage_audit_output,
    write_balance_missing_mapping_candidates,
    write_dashboard_comparator_pair_coverage,
    write_ninth_mapping_data_coverage,
    write_runtime_missing_pair_summary,
    _split_transformation_input_output_measures,
)
from codebase.utilities.leap_results_dashboard_utils import _prepare_render_long  # noqa: E402
from codebase.utilities.leap_results_dashboard_v2.comparison_engine import (  # noqa: E402
    build_chart_line_mapping_ledger,
    build_total_component_ledger,
)
from codebase.utilities.leap_results_dashboard_v2.diagnostics import (  # noqa: E402
    write_diagnostics,
)
from codebase.utilities.leap_results_dashboard_v2.output_writer import write_core_outputs  # noqa: E402
from codebase.utilities.leap_balance_export_resolver import resolve_balance_export_workbook  # noqa: E402
from codebase.utilities.workflow_common import (  # noqa: E402
    WorkflowTimer,
    archive_config_dir_once_per_day,
)
from codebase.utilities.workflow_outputs import build_workflow_output_layout, write_output_manifest  # noqa: E402
from codebase.utilities.output_paths import BALANCE_TABLES_ROOT  # noqa: E402


#%%
def _resolve(path: Path | str) -> Path:
    raw = str(path).replace("\\", "/")
    drive_match = re.match(r"^([a-zA-Z]):/(.*)$", raw)
    if drive_match:
        drive = drive_match.group(1).lower()
        rest = drive_match.group(2)
        if os.name == "nt":
            return Path(f"{drive.upper()}:/{rest}")
        return Path(f"/mnt/{drive}/{rest}")
    candidate = Path(raw)
    return candidate if candidate.is_absolute() else (REPO_ROOT / candidate)


#%%
BALANCE_EXPORT_ECONOMY = "20_USA"
REF_BALANCE_EXPORT_DATE_ID: str | None = None
TGT_BALANCE_EXPORT_DATE_ID: str | None = None
REF_WORKBOOK_PATH = resolve_balance_export_workbook(
    economy=BALANCE_EXPORT_ECONOMY,
    scenario="REF",
    date_id=REF_BALANCE_EXPORT_DATE_ID,
)
TGT_WORKBOOK_PATH = resolve_balance_export_workbook(
    economy=BALANCE_EXPORT_ECONOMY,
    scenario="TGT",
    date_id=TGT_BALANCE_EXPORT_DATE_ID,
)

KNOWN_ISSUES_CONFIG_PATH = _resolve("config/leap_results_balance_known_issues.json")
CHART_NAVIGATION_GUIDE_PATH = _resolve("config/leap_comparison_dashboard_template_v2.json")
LEAP_ROWS_TO_REMOVE_AND_ADD_SHEET = "leap_rows_to_remove_and_add"

OUTPUT_DIR = _resolve("outputs/dashboards/leap_results_dashboard/USA")
BALANCE_TO_ESTO_LONG_OUTPUT_DIR = BALANCE_TABLES_ROOT / "leap_balance_to_esto_long" / "USA"

LEAP_TO_ESTO_MAPPING = (_resolve("config/leap_mappings.xlsx"), "leap_combined_esto")
NINTH_TO_ESTO_MAPPING = (_resolve("config/master_config.xlsx"), "ninth_pairs_to_esto_pairs")
CODEBOOK_PATH = _resolve("config/sector_fuel_codes_to_names.xlsx")
SHEET_MAP_PATH = _resolve("config/leap_results_sheet_map.csv")
BACKUP_MAPPINGS_PATH = _resolve("config/backup_leap_mappings.xlsx")
EXPLICIT_MAPPINGS_PATH = _resolve("config/leap_results_explicit_mappings.csv")
EXPLICIT_REASSIGNMENTS_PATH = _resolve("config/leap_results_explicit_reassignments.csv")
SYNTHETIC_REFERENCE_ROWS_PATH = _resolve("config/synthetic_reference_rows.csv")

BASE_TABLE_PATH = _resolve("data/00APEC_2025_low_with_subtotals.csv")
PROJECTION_TABLE_PATH = _resolve("data/merged_file_energy_ALL_20251106.csv")

BASE_YEAR = 2022
MAX_OUTPUT_YEAR = 2060
PROJECTION_YEARS: Sequence[int] = tuple(range(BASE_YEAR + 1, MAX_OUTPUT_YEAR + 1))
SCENARIO_MAP = {"Reference": "reference", "Target": "target"}
BASE_ECONOMY = "20USA"
PROJECTION_ECONOMY = "20_USA"

CHART_BACKEND = "plotly"
HIDE_LEAP_ONLY_CHARTS = False
ENABLE_WORKFLOW_TIMING = True
WRITE_WORKFLOW_TIMING_CSV = True
WORKFLOW_TIMING_FILENAME = "workflow_stage_timings.csv"

# ---------------------------------------------------------------------------
# Stage control — set False to skip a stage and reload cached outputs instead.
# Run the full workflow at least once before skipping any stage.
#   STAGE_EXTRACT           reads LEAP Excel workbooks (typically the slowest step)
#   STAGE_COMPARE           builds the ESTO-axis comparison (reads large projection CSVs)
#   STAGE_WRITE_OUTPUTS     writes comparison tables, simple balance tables, diagnostics
#   STAGE_RENDER_DASHBOARDS renders HTML dashboards (slow for large chart sets)
#   STAGE_WRITE_COVERAGE    writes coverage, runtime issues, and mapping checks
# Common skip patterns:

# Re-render dashboards only (e.g. after changing VISIBLE_COMPARISON_SERIES): set STAGE_EXTRACT=False, STAGE_COMPARE=False, STAGE_WRITE_OUTPUTS=False, STAGE_WRITE_COVERAGE=False
# Skip just the slow Excel reading: set STAGE_EXTRACT=False
# Skip everything except re-rendering: set only STAGE_RENDER_DASHBOARDS=True, all others False
# ---------------------------------------------------------------------------
STAGE_EXTRACT: bool = True
STAGE_COMPARE: bool = True
STAGE_WRITE_OUTPUTS: bool = True
STAGE_RENDER_DASHBOARDS: bool = True
STAGE_WRITE_COVERAGE: bool = True
# Controls which source/scenario series are written and rendered.
# Source labels are the normalized comparison source values:
# - "base" is ESTO base-year data
# - "projection" is 9th projection data
# - "leap" is LEAP balance-export data
VISIBLE_COMPARISON_SERIES: set[tuple[str, str]] = {
    ("base", "ESTO"),
    # ("projection", "Reference"),
    ("projection", "Target"),
    ("leap", "Target"),
}
BUNKER_SHEET_KEYS = {
    "esto__04__International_marine_bunkers",
    "esto__05__International_aviation_bunkers",
}
FAIL_ON_UNMAPPED_BALANCE_ROWS = os.getenv("FAIL_ON_UNMAPPED_BALANCE_ROWS", "1").strip().lower() in {
    "1",
    "true",
    "yes",
    "y",
}


#%%
def _load_json(path: Path) -> dict[str, object]:
    if not path.exists():
        raise FileNotFoundError(f"Required JSON config file not found: {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def _mapping_workbook(mapping_ref: tuple[Path, str]) -> Path:
    return mapping_ref[0]


def _truthy(value: object) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "on", "t"}


def _normalize_key(value: object) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _load_ignored_unmapped_issue_keys(mapping_workbook_path: Path) -> set[tuple[str, str]]:
    if not mapping_workbook_path.exists():
        return set()
    try:
        sheet = pd.read_excel(mapping_workbook_path, sheet_name=LEAP_ROWS_TO_REMOVE_AND_ADD_SHEET, dtype=str).fillna("")
    except Exception:
        return set()
    for col in ["leap_sector_name_full_path", "raw_leap_fuel_name"]:
        if col not in sheet.columns:
            sheet[col] = ""
    keys = set()
    for _, row in sheet.iterrows():
        sector = _normalize_key(row.get("leap_sector_name_full_path", ""))
        fuel = _normalize_key(row.get("raw_leap_fuel_name", ""))
        if sector and fuel:
            keys.add((sector, fuel))
    return keys


def _write_dashboard_about_supplements(
    *,
    dashboards_dir: Path,
    mapping_workbook_path: Path,
    master_config_path: Path,
    template_json_path: Path,
) -> None:
    """Copy mapping sheets and JSON template into dashboards_dir, then patch about.html."""
    xlsx_filename = "dashboard_mappings.xlsx"
    json_filename = "dashboard_template.json"

    # Build mappings xlsx with the three mapping sheets
    sheets_to_copy = [
        (mapping_workbook_path, "leap_combined_esto"),
        (mapping_workbook_path, "leap_combined_ninth"),
        (master_config_path, "ninth_pairs_to_esto_pairs"),
    ]
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)
    for src_path, sheet_name in sheets_to_copy:
        if not src_path.exists():
            continue
        try:
            src_wb = load_workbook(src_path, data_only=True, read_only=True)
            if sheet_name not in src_wb.sheetnames:
                src_wb.close()
                continue
            dst_ws = out_wb.create_sheet(title=sheet_name)
            for row in src_wb[sheet_name].iter_rows(values_only=True):
                dst_ws.append(list(row))
            src_wb.close()
        except Exception:
            pass
    out_wb.save(dashboards_dir / xlsx_filename)

    # Copy JSON template
    if template_json_path.exists():
        shutil.copy2(template_json_path, dashboards_dir / json_filename)

    # Patch about.html — inject reference files section and losses note
    about_path = dashboards_dir / "about.html"
    if not about_path.exists():
        return
    html = about_path.read_text(encoding="utf-8")
    extra = (
        '<section class="about-section"><h2>Reference files</h2>'
        "<ul>"
        f'<li><a href="{xlsx_filename}">dashboard_mappings.xlsx</a>'
        " — mapping sheets used to build this dashboard:"
        " leap_combined_esto, leap_combined_ninth, and ninth_pairs_to_esto_pairs (ESTO-to-9th canonical mapping).</li>"
        f'<li><a href="{json_filename}">dashboard_template.json</a>'
        " — the chart navigation template that defines the ESTO-axis structure and dashboard sections.</li>"
        "</ul>"
        "</section>"
        '<section class="about-section"><h2>Note on LEAP values and losses</h2>'
        "<p>LEAP does not extract losses as a separate line item."
        " Instead, losses are deducted directly from the value shown for the affected fuel,"
        " so the fuel figure already has losses subtracted from it."
        " This means transformation output values can be lower than the equivalent ESTO figure"
        " — and in some cases can switch from positive to negative —"
        " simply because the losses have been taken out of the fuel value rather than reported separately.</p>"
        "</section>"
    )
    if "</article>" in html:
        html = html.replace("</article>", extra + "</article>", 1)
        about_path.write_text(html, encoding="utf-8")


def _issue_source_key(row: pd.Series) -> tuple[str, str]:
    sector = _normalize_key(row.get("mapping_key_sector", "") or row.get("leap_sector_name_full_path", ""))
    fuel = _normalize_key(row.get("mapping_key_fuel", "") or row.get("leap_product_name", "") or row.get("leap_product", ""))
    return sector, fuel


def _filter_ignored_unmapped_issues(
    runtime_issues: pd.DataFrame,
    *,
    mapping_workbook_path: Path,
) -> pd.DataFrame:
    if runtime_issues.empty:
        return runtime_issues.copy()
    ignored_keys = _load_ignored_unmapped_issue_keys(mapping_workbook_path)
    if not ignored_keys:
        return runtime_issues.copy()

    work = runtime_issues.copy()
    reason = work["reason"].fillna("").astype(str).str.strip().str.lower() if "reason" in work.columns else pd.Series("", index=work.index)
    unmapped_mask = reason.eq("missing_esto_pair")
    if not unmapped_mask.any():
        return work

    source_keys = work.loc[unmapped_mask].apply(_issue_source_key, axis=1)
    suppress_mask = pd.Series(False, index=work.index)
    suppress_mask.loc[unmapped_mask] = source_keys.map(lambda key: key in ignored_keys).to_numpy()
    return work.loc[~suppress_mask].copy()


def _load_cached_ingestion(
    structure_config: dict,
    balance_to_esto_long_output_dir: Path,
) -> tuple[dict, dict, dict]:
    """Load ingestion sub-tables written by a previous extraction run.

    Returns (conversion, ingestion, resolved_structure) matching the shapes
    produced by convert_leap_balances_to_esto_long_table().  Fields not written
    to the shared cache (e.g. matching_diagnostics) are returned as empty
    DataFrames.
    """
    shared_dir = _resolve(balance_to_esto_long_output_dir)
    support_dir = shared_dir / "supporting_files"

    def _rcsv(path: Path) -> pd.DataFrame:
        return pd.read_csv(path, dtype=str, low_memory=False).fillna("") if path.exists() else pd.DataFrame()

    def _rjson(path: Path) -> dict:
        return json.loads(path.read_text(encoding="utf-8")) if path.exists() else {}

    ingestion: dict = {
        "leap_long": _rcsv(support_dir / "leap_balance_mapped_detail_long.csv"),
        "mapping_status": _rcsv(support_dir / "leap_balance_mapping_status.csv"),
        "issues": _rcsv(support_dir / "leap_balance_runtime_issues.csv"),
        "override_report": _rcsv(support_dir / "leap_balance_override_application_report.csv"),
        "auto_sheet_rows": _rcsv(support_dir / "auto_sheet_rows.csv"),
        "coverage": _rcsv(support_dir / "leap_balance_coverage.csv"),
        "unit_diagnostics": _rcsv(support_dir / "leap_balance_unit_diagnostics.csv"),
        "matching_diagnostics": pd.DataFrame(),
        "extraction_summary": _rjson(support_dir / "leap_balance_extraction_summary.json"),
        "resolved_structure": _rjson(support_dir / "resolved_structure_config.json"),
    }
    resolved_structure = ingestion["resolved_structure"] or structure_config
    conversion: dict = {
        "esto_long": _rcsv(shared_dir / "leap_balance_esto_long.csv"),
        "ingestion": ingestion,
        "pre_group_leap_mapped": pd.DataFrame(),
        "pre_group_incomplete_rows": pd.DataFrame(),
        "issues": ingestion["issues"],
        "override_report": ingestion["override_report"],
        "auto_sheet_rows": ingestion["auto_sheet_rows"],
        "coverage": ingestion["coverage"],
        "unit_diagnostics": ingestion["unit_diagnostics"],
        "extraction_summary": ingestion["extraction_summary"],
        "resolved_structure": resolved_structure,
    }
    return conversion, ingestion, resolved_structure


def _load_cached_comparison(
    layout,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict]:
    """Load comparison_long, mapping_status, and leap_long from a previous run.

    Returns (comparison_long, mapping_status, leap_long, comparison_stub).
    The comparison_stub contains empty DataFrames for fields not cached
    separately (base_df, ninth_df, mapping_inputs).
    """
    comparison_long_path = layout.root / "comparison_long.csv"
    mapping_status_path = layout.root / "mapping_status.xlsx"
    leap_long_path = layout.root / "leap_long.csv"

    comparison_long = (
        pd.read_csv(comparison_long_path, dtype=str, low_memory=False).fillna("")
        if comparison_long_path.exists()
        else pd.DataFrame()
    )
    mapping_status = (
        pd.read_excel(mapping_status_path, dtype=str, sheet_name="mapping_status").fillna("")
        if mapping_status_path.exists()
        else pd.DataFrame()
    )
    leap_long = (
        pd.read_csv(leap_long_path, dtype=str, low_memory=False).fillna("")
        if leap_long_path.exists()
        else pd.DataFrame()
    )
    comparison_stub: dict = {
        "base_df": pd.DataFrame(),
        "ninth_df": pd.DataFrame(),
        "mapping_inputs": {},
    }
    return comparison_long, mapping_status, leap_long, comparison_stub


def _normalize_base_scenario(comparison_long: pd.DataFrame) -> pd.DataFrame:
    """Rename all source=='base' rows to scenario='ESTO' and deduplicate.

    The base (ESTO) value is a single base-year lookup that gets duplicated once
    per LEAP scenario during comparison assembly. Renaming it to 'ESTO' and
    deduplicating ensures only one copy appears and the VISIBLE_COMPARISON_SERIES
    entry ("base", "ESTO") can match it.
    """
    if comparison_long.empty:
        return comparison_long.copy()
    out = comparison_long.copy()
    out["source"] = out["source"].fillna("").astype(str).str.strip()
    base_mask = out["source"] == "base"
    out.loc[base_mask, "scenario"] = "ESTO"
    key_cols = [c for c in out.columns if c != "value"]
    out = out.drop_duplicates(subset=key_cols, keep="first")
    return out.reset_index(drop=True)


def _filter_visible_comparison_series(
    comparison_long: pd.DataFrame,
    visible_series: set[tuple[str, str]],
) -> pd.DataFrame:
    if comparison_long.empty or not visible_series:
        return comparison_long.copy()
    out = comparison_long.copy()
    out["source"] = out["source"].fillna("").astype(str).str.strip()
    out["scenario"] = out["scenario"].fillna("").astype(str).str.strip()
    allowed = {(str(source).strip(), str(scenario).strip()) for source, scenario in visible_series}
    mask = pd.Series(list(zip(out["source"], out["scenario"])), index=out.index).isin(allowed)
    return out.loc[mask].copy().reset_index(drop=True)


def _apply_bunker_abs_values(comparison_long: pd.DataFrame) -> pd.DataFrame:
    if comparison_long.empty:
        return comparison_long.copy()
    out = comparison_long.copy()
    out["sheet"] = out["sheet"].fillna("").astype(str).str.strip()
    if "chart_group_key" not in out.columns:
        out["chart_group_key"] = ""
    out["chart_group_key"] = out["chart_group_key"].fillna("").astype(str).str.strip()
    bunker_mask = out["chart_group_key"].isin(BUNKER_SHEET_KEYS) | out["sheet"].isin(BUNKER_SHEET_KEYS)
    if bunker_mask.any():
        out.loc[bunker_mask, "value"] = pd.to_numeric(out.loc[bunker_mask, "value"], errors="coerce").abs()
    return out


def _comparison_wide_from_long(comparison_long: pd.DataFrame) -> pd.DataFrame:
    index_cols = [
        "economy",
        "scenario",
        "sheet",
        "page_key",
        "page_label",
        "chart_group_key",
        "chart_group_label",
        "measure",
        "fuel_label",
        "year",
    ]
    if comparison_long.empty:
        return pd.DataFrame(columns=index_cols)
    work = comparison_long.copy()
    for col in index_cols:
        if col not in work.columns:
            work[col] = ""
    return (
        work.pivot_table(
            index=index_cols,
            columns="source",
            values="value",
            aggfunc="sum",
        )
        .reset_index()
        .rename_axis(columns=None)
        .sort_values(["scenario", "page_key", "chart_group_key", "measure", "fuel_label", "year"], kind="mergesort")
        .reset_index(drop=True)
    )


def _raise_if_unmapped_balance_rows(runtime_issues: pd.DataFrame, runtime_issues_path: Path) -> None:
    if runtime_issues.empty:
        return
    if not FAIL_ON_UNMAPPED_BALANCE_ROWS:
        return
    reason_counts = (
        runtime_issues.groupby("reason", dropna=False)
        .size()
        .reset_index(name="row_count")
        .sort_values(["row_count", "reason"], ascending=[False, True])
    )
    counts_text = ", ".join(
        f"{row.reason}: {int(row.row_count)}" for row in reason_counts.itertuples(index=False)
    )
    raise RuntimeError(
        "Unmapped LEAP balance rows remain after writing dashboard outputs. "
        f"See {runtime_issues_path}. Counts: {counts_text}"
    )


def _write_shared_balance_to_esto_outputs(
    *,
    conversion: dict[str, object],
    mapping_status: pd.DataFrame,
    comparison_long: pd.DataFrame,
    simple_ninth_balance: pd.DataFrame,
    output_dir: Path,
) -> dict[str, str]:
    out_dir = _resolve(output_dir)
    support_dir = out_dir / "supporting_files"
    out_dir.mkdir(parents=True, exist_ok=True)
    support_dir.mkdir(parents=True, exist_ok=True)

    leap_path = out_dir / "leap_balance_esto_long.csv"
    ninth_path = out_dir / "ninth_balance_esto_long.csv"
    detail_path = support_dir / "leap_balance_mapped_detail_long.csv"
    mapping_status_path = support_dir / "leap_balance_mapping_status.csv"
    comparison_long_path = support_dir / "esto_axis_comparison_long.csv"
    ninth_semantic_path = support_dir / "ninth_balance_esto_long_semantic_columns.csv"
    runtime_issues_path = support_dir / "leap_balance_runtime_issues.csv"
    override_report_path = support_dir / "leap_balance_override_application_report.csv"
    auto_sheet_path = support_dir / "auto_sheet_rows.csv"
    coverage_path = support_dir / "leap_balance_coverage.csv"
    unit_diag_path = support_dir / "leap_balance_unit_diagnostics.csv"
    extraction_summary_path = support_dir / "leap_balance_extraction_summary.json"
    resolved_structure_path = support_dir / "resolved_structure_config.json"

    conversion["esto_long"].to_csv(leap_path, index=False)
    build_ninth_balance_esto_long_table(simple_ninth_balance).to_csv(ninth_path, index=False)
    conversion["leap_long"].to_csv(detail_path, index=False)
    mapping_status.to_csv(mapping_status_path, index=False)
    comparison_long.to_csv(comparison_long_path, index=False)
    simple_ninth_balance.to_csv(ninth_semantic_path, index=False)
    conversion["issues"].to_csv(runtime_issues_path, index=False)
    conversion["override_report"].to_csv(override_report_path, index=False)
    conversion["auto_sheet_rows"].to_csv(auto_sheet_path, index=False)
    conversion["coverage"].to_csv(coverage_path, index=False)
    conversion["unit_diagnostics"].to_csv(unit_diag_path, index=False)
    extraction_summary_path.write_text(
        json.dumps(conversion["extraction_summary"], ensure_ascii=True, indent=2),
        encoding="utf-8",
    )
    resolved_structure_path.write_text(
        json.dumps(conversion["resolved_structure"], ensure_ascii=True, indent=2),
        encoding="utf-8",
    )

    return {
        "shared_leap_balance_esto_long": str(leap_path),
        "shared_ninth_balance_esto_long": str(ninth_path),
        "shared_supporting_files_dir": str(support_dir),
    }


def run_workflow() -> dict[str, object]:
    timer = WorkflowTimer("leap_results_dashboard_balance_estoaxis", enabled=ENABLE_WORKFLOW_TIMING)
    archive_config_dir_once_per_day()
    out_dir = _resolve(OUTPUT_DIR)
    layout = build_workflow_output_layout(out_dir)
    timing_path = layout.runtime / WORKFLOW_TIMING_FILENAME

    structure_config = build_esto_axis_structure_from_dashboard_template(CHART_NAVIGATION_GUIDE_PATH)
    known_issues = _load_json(KNOWN_ISSUES_CONFIG_PATH)
    timer.lap("setup")

    # -------------------------------------------------------------------------
    # Stage: extract and map LEAP balance workbooks
    # -------------------------------------------------------------------------
    if STAGE_EXTRACT:
        conversion = convert_leap_balances_to_esto_long_table(
            ref_workbook_path=REF_WORKBOOK_PATH,
            tgt_workbook_path=TGT_WORKBOOK_PATH,
            template_sheet="EBal|2060",
            mapping_pairs_path=_mapping_workbook(LEAP_TO_ESTO_MAPPING),
            codebook_path=CODEBOOK_PATH,
            structure_config=structure_config,
            known_issues=known_issues,
            projection_economy=PROJECTION_ECONOMY,
            max_output_year=MAX_OUTPUT_YEAR,
            explicit_pair_mappings_only=True,
        )
        ingestion = conversion["ingestion"]
        resolved_structure = ingestion.get("resolved_structure", structure_config)
    else:
        print("[SKIP] Stage: extract and map LEAP balance workbooks — loading cached outputs")
        conversion, ingestion, resolved_structure = _load_cached_ingestion(
            structure_config=structure_config,
            balance_to_esto_long_output_dir=BALANCE_TO_ESTO_LONG_OUTPUT_DIR,
        )
    timer.lap("extract and map LEAP balance workbooks")

    # -------------------------------------------------------------------------
    # Stage: build ESTO-axis comparison
    # -------------------------------------------------------------------------
    if STAGE_COMPARE:
        comparison = build_balance_comparison_esto_axis(
            leap_long=ingestion["leap_long"],
            mapping_status=ingestion["mapping_status"],
            base_year=BASE_YEAR,
            projection_years=tuple(PROJECTION_YEARS),
            base_economy=BASE_ECONOMY,
            projection_economy=PROJECTION_ECONOMY,
            scenario_map=SCENARIO_MAP,
            sheet_map_path=SHEET_MAP_PATH,
            backup_mappings_path=BACKUP_MAPPINGS_PATH,
            codebook_path=CODEBOOK_PATH,
            canonical_pairs_path=NINTH_TO_ESTO_MAPPING,
            explicit_mappings_path=EXPLICIT_MAPPINGS_PATH,
            explicit_reassignments_path=EXPLICIT_REASSIGNMENTS_PATH,
            synthetic_reference_rows_path=SYNTHETIC_REFERENCE_ROWS_PATH,
            esto_table_path=BASE_TABLE_PATH,
            projection_table_path=PROJECTION_TABLE_PATH,
            chart_navigation_guide_path=CHART_NAVIGATION_GUIDE_PATH,
            balance_mapping_workbook_path=_mapping_workbook(LEAP_TO_ESTO_MAPPING),
            known_issues=known_issues,
        )
        comparison_long = comparison["comparison_long"].copy()
        mapping_status = comparison["mapping_status"].copy()
        leap_long = ingestion["leap_long"].copy()
        comparison_long = comparison_long[pd.to_numeric(comparison_long["year"], errors="coerce").le(MAX_OUTPUT_YEAR)].copy()
        comparison_long_for_audit = _normalize_base_scenario(comparison_long.copy())
        comparison_long = _normalize_base_scenario(comparison_long)
        comparison_long = _apply_bunker_abs_values(comparison_long)
        comparison_long = _filter_visible_comparison_series(comparison_long, VISIBLE_COMPARISON_SERIES)
        comparison_wide = _comparison_wide_from_long(comparison_long)
        leap_long = leap_long[pd.to_numeric(leap_long["year"], errors="coerce").le(MAX_OUTPUT_YEAR)].copy()
    else:
        print("[SKIP] Stage: build ESTO-axis comparison — loading cached outputs")
        comparison_long, mapping_status, leap_long, comparison = _load_cached_comparison(layout)
        comparison_long_for_audit = comparison_long.copy()
        comparison_wide = _comparison_wide_from_long(comparison_long)
    timer.lap("build ESTO-axis comparison")

    # Pre-declare path variables used in all downstream stages and the result dict
    # so they are always defined regardless of which stages run.
    simple_leap_balance_path = layout.root / "simple_leap_balance_mapped.csv"
    simple_ninth_balance_path = layout.root / "simple_ninth_balance_mapped.csv"
    merged_esto_axis_balance_path = layout.root / "merged_leap_ninth_esto_balance.csv"
    duplicate_summary_path = layout.diagnostics / "simple_balance_duplicate_summary.csv"
    duplicate_details_path = layout.diagnostics / "simple_balance_duplicate_details.csv"
    mapped_leap_to_esto_path = layout.mapping / "mapped_leap_to_esto_balance_rows.csv"
    mapped_ninth_to_esto_path = layout.mapping / "mapped_ninth_to_esto_balance_rows.csv"
    esto_axis_comparison_long_path = layout.root / "esto_axis_comparison_long.csv"
    mapping_lineage_audit_path = layout.mapping / "mapping_lineage_audit.csv"
    chart_line_mapping_path = layout.ledgers / "chart_line_mapping_ledger.csv"
    chart_total_component_path = layout.ledgers / "chart_total_component_ledger.csv"
    runtime_issues_path = layout.runtime / "balance_runtime_issues.csv"
    runtime_missing_pair_summary_path = layout.runtime / "balance_runtime_missing_pair_summary.xlsx"
    override_report_path = layout.runtime / "balance_override_application_report.csv"
    auto_sheet_path = layout.mapping / "auto_sheet_rows.csv"
    resolved_structure_path = layout.mapping / "resolved_structure_config.json"
    coverage_path = layout.coverage / "balance_coverage.csv"
    unit_diag_path = layout.coverage / "balance_unit_diagnostics.csv"
    matching_diag_path = layout.coverage / "balance_matching_diagnostics.csv"
    extraction_summary_path = layout.coverage / "balance_extraction_summary.json"

    core_paths: dict[str, str] = {}
    shared_conversion_paths: dict[str, str] = {}
    diagnostics_paths: dict[str, object] = {}
    comparator_pair_coverage_xlsx: object = None
    missing_mapping_candidates_path: object = None
    ninth_mapping_data_coverage_path: object = None
    filtered_runtime_issues = ingestion["issues"].copy()

    # -------------------------------------------------------------------------
    # Stage: write core outputs, simple balance tables, diagnostics
    # -------------------------------------------------------------------------
    if STAGE_WRITE_OUTPUTS:
        core_paths = write_core_outputs(
            out_dir=layout.root,
            supporting_dir=layout.supporting,
            comparison_long=comparison_long,
            comparison_wide=comparison_wide,
            mapping_status=mapping_status,
            leap_long=leap_long,
        )
        simple_leap_balance = conversion["esto_long"].copy()
        simple_ninth_balance = build_simple_ninth_balance_table(
            comparison_long=comparison_long,
            mapping_status=mapping_status,
        )
        shared_conversion_paths = _write_shared_balance_to_esto_outputs(
            conversion=conversion,
            mapping_status=mapping_status,
            comparison_long=comparison_long,
            simple_ninth_balance=simple_ninth_balance,
            output_dir=BALANCE_TO_ESTO_LONG_OUTPUT_DIR,
        )
        mapped_ninth_to_esto = comparison.get("ninth_projection_components", pd.DataFrame())
        if mapped_ninth_to_esto.empty:
            mapped_ninth_to_esto = build_mapped_ninth_to_esto_balance_rows(
                comparison_long=comparison_long,
                mapping_status=mapping_status,
            )
        merged_esto_axis_balance = build_merged_esto_axis_balance_table(
            simple_leap_balance=simple_leap_balance,
            simple_ninth_balance=simple_ninth_balance,
            comparison_long=comparison_long,
            mapping_status=mapping_status,
        )
        duplicate_summary, duplicate_details = build_simple_balance_duplicate_diagnostics(
            simple_leap_balance=simple_leap_balance,
            simple_ninth_balance=simple_ninth_balance,
        )
        simple_leap_balance.to_csv(simple_leap_balance_path, index=False)
        simple_ninth_balance.to_csv(simple_ninth_balance_path, index=False)
        merged_esto_axis_balance.to_csv(merged_esto_axis_balance_path, index=False)
        duplicate_summary.to_csv(duplicate_summary_path, index=False)
        duplicate_details.to_csv(duplicate_details_path, index=False)
        simple_leap_balance.to_csv(mapped_leap_to_esto_path, index=False)
        mapped_ninth_to_esto.to_csv(mapped_ninth_to_esto_path, index=False)
        comparison_long.to_csv(esto_axis_comparison_long_path, index=False)
        timer.lap("write core and simple ESTO-axis outputs")
        diagnostics_paths = write_diagnostics(
            comparison_long=comparison_long,
            mapping_status=mapping_status,
            out_dir=layout.root,
            base_year=BASE_YEAR,
            diagnostic_probe_year=min(2030, MAX_OUTPUT_YEAR),
            top_diagnostic_rows=40,
        )
        timer.lap("write diagnostics")
    else:
        print("[SKIP] Stage: write outputs")

    # -------------------------------------------------------------------------
    # Stage: render dashboards and write chart ledgers
    # -------------------------------------------------------------------------
    chart_line_mapping_ledger = pd.DataFrame()
    chart_total_component_ledger = pd.DataFrame()
    dashboard_paths: dict[str, object] = {}
    chart_group_exposure_df = pd.DataFrame()
    all_chart_groups_df = pd.DataFrame()

    if STAGE_RENDER_DASHBOARDS:
        split_comparison_long = _split_transformation_input_output_measures(
            comparison_long,
            resolved_structure.get("sheet_catalog", {}),
        )
        chart_input = _prepare_render_long(split_comparison_long)
        chart_line_mapping_ledger = build_chart_line_mapping_ledger(chart_input, mapping_status)
        chart_total_component_ledger = build_total_component_ledger(chart_input, mapping_status)
        dashboard_paths = render_balance_dashboards(
            comparison_long=comparison_long,
            mapping_status=mapping_status,
            structure_config=resolved_structure,
            output_dir=layout.root,
            chart_backend=CHART_BACKEND,
            hide_leap_only_charts=HIDE_LEAP_ONLY_CHARTS,
            chart_navigation_guide_path=CHART_NAVIGATION_GUIDE_PATH,
        )
        _write_dashboard_about_supplements(
            dashboards_dir=Path(str(dashboard_paths["dashboards_dir"])),
            mapping_workbook_path=_mapping_workbook(LEAP_TO_ESTO_MAPPING),
            master_config_path=NINTH_TO_ESTO_MAPPING[0],
            template_json_path=CHART_NAVIGATION_GUIDE_PATH,
        )
        timer.lap("render dashboards")
        chart_line_mapping_ledger = attach_chart_groups_to_dashboard_exposure(
            chart_line_mapping_ledger,
            dashboard_paths.get("chart_group_exposure"),
            dashboard_paths.get("all_chart_groups"),
        )
        chart_group_exposure_df = (
            pd.read_csv(dashboard_paths["chart_group_exposure"], dtype=str).fillna("")
            if dashboard_paths.get("chart_group_exposure")
            else pd.DataFrame()
        )
        all_chart_groups_df = (
            pd.read_csv(dashboard_paths["all_chart_groups"], dtype=str).fillna("")
            if dashboard_paths.get("all_chart_groups")
            else pd.DataFrame()
        )
        simplify_chart_line_mapping_ledger_output(chart_line_mapping_ledger).to_csv(chart_line_mapping_path, index=False)
        simplify_chart_total_component_ledger_output(chart_total_component_ledger).to_csv(chart_total_component_path, index=False)
        mapping_lineage_audit = build_mapping_lineage_audit_table(
            pre_group_leap_mapped=conversion.get("pre_group_leap_mapped", pd.DataFrame()),
            pre_group_incomplete_rows=conversion.get("pre_group_incomplete_rows", pd.DataFrame()),
            comparison_long_full=comparison_long,
            mapping_status=mapping_status,
            mapped_ninth_to_esto_balance_rows=comparison.get("ninth_projection_components", pd.DataFrame()),
            target_years=(BASE_YEAR, BASE_YEAR + 1, MAX_OUTPUT_YEAR),
        )
        mapping_lineage_audit = attach_chart_groups_to_mapping_lineage_audit(
            mapping_lineage_audit,
            dashboard_paths.get("chart_group_exposure"),
            dashboard_paths.get("all_chart_groups"),
        )
        simplify_mapping_lineage_audit_output(mapping_lineage_audit).to_csv(mapping_lineage_audit_path, index=False)
        timer.lap("write chart ledgers")
    else:
        print("[SKIP] Stage: render dashboards")
        if chart_line_mapping_path.exists():
            chart_line_mapping_ledger = pd.read_csv(chart_line_mapping_path, dtype=str, low_memory=False).fillna("")
        if chart_total_component_path.exists():
            chart_total_component_ledger = pd.read_csv(chart_total_component_path, dtype=str, low_memory=False).fillna("")

    # -------------------------------------------------------------------------
    # Stage: write runtime issues, mapping candidates, and coverage checks
    # -------------------------------------------------------------------------
    if STAGE_WRITE_COVERAGE:
        ingestion["issues"].to_csv(runtime_issues_path, index=False)
        filtered_runtime_issues = _filter_ignored_unmapped_issues(
            ingestion["issues"],
            mapping_workbook_path=_mapping_workbook(LEAP_TO_ESTO_MAPPING),
        )
        write_runtime_missing_pair_summary(
            runtime_issues=filtered_runtime_issues,
            output_path=runtime_missing_pair_summary_path,
        )
        comparator_pair_coverage_xlsx = write_dashboard_comparator_pair_coverage(
            mapping_status=mapping_status,
            dashboard_exposure=chart_line_mapping_ledger,
            chart_group_exposure=chart_group_exposure_df,
            all_chart_groups=all_chart_groups_df,
            base_df=comparison.get("base_df", pd.DataFrame()),
            ninth_df=comparison.get("ninth_df", pd.DataFrame()),
            output_path=layout.coverage / "dashboard_comparator_pair_coverage.xlsx",
            base_economy=BASE_ECONOMY,
            projection_economy=PROJECTION_ECONOMY,
            base_year=BASE_YEAR,
            projection_years=tuple(PROJECTION_YEARS),
            scenarios=tuple(SCENARIO_MAP.values()),
            runtime_issues=filtered_runtime_issues,
            chart_navigation_guide_path=CHART_NAVIGATION_GUIDE_PATH,
            mapping_workbook_path=_mapping_workbook(LEAP_TO_ESTO_MAPPING),
            mapping_sheet_name=LEAP_TO_ESTO_MAPPING[1],
        )
        missing_mapping_candidates_path = write_balance_missing_mapping_candidates(
            runtime_issues=filtered_runtime_issues,
            output_path=layout.mapping / "balance_missing_mapping_candidates.xlsx",
            mapping_workbook_path=_mapping_workbook(LEAP_TO_ESTO_MAPPING),
        )
        leap_combined_ninth_mapping = pd.read_excel(
            _mapping_workbook(LEAP_TO_ESTO_MAPPING),
            sheet_name="leap_combined_ninth",
            dtype=str,
        ).fillna("")
        ninth_mapping_data_coverage_path = write_ninth_mapping_data_coverage(
            ninth_df=comparison.get("ninth_df", pd.DataFrame()),
            ninth_mapping_pairs=leap_combined_ninth_mapping,
            output_path=layout.coverage / "ninth_mapping_data_coverage.xlsx",
            projection_economy=PROJECTION_ECONOMY,
            scenarios=tuple(SCENARIO_MAP.values()),
            years=tuple(PROJECTION_YEARS),
        )
        ingestion["override_report"].to_csv(override_report_path, index=False)
        ingestion.get("auto_sheet_rows", pd.DataFrame()).to_csv(auto_sheet_path, index=False)
        resolved_structure_path.write_text(json.dumps(resolved_structure, ensure_ascii=True, indent=2), encoding="utf-8")
        ingestion["coverage"].to_csv(coverage_path, index=False)
        ingestion["unit_diagnostics"].to_csv(unit_diag_path, index=False)
        ingestion.get("matching_diagnostics", pd.DataFrame()).to_csv(matching_diag_path, index=False)
        extraction_summary_path.write_text(json.dumps(ingestion["extraction_summary"], ensure_ascii=True, indent=2), encoding="utf-8")
        timer.lap("write runtime, mapping, and coverage checks")
    else:
        print("[SKIP] Stage: write coverage")

    manifest = write_output_manifest(
        out_dir=layout.root,
        primary_outputs={
            "comparison_long": str(layout.root / "comparison_long.csv"),
            "comparison_wide": str(layout.root / "comparison_wide.csv"),
            "mapping_status": str(layout.root / "mapping_status.xlsx"),
            "leap_long": str(layout.root / "leap_long.csv"),
            "simple_leap_balance_mapped": str(simple_leap_balance_path),
            "simple_ninth_balance_mapped": str(simple_ninth_balance_path),
            "merged_leap_ninth_esto_balance": str(merged_esto_axis_balance_path),
            "dashboard_index": dashboard_paths.get("dashboard_index"),
            "charts_dir": str(layout.charts),
            "dashboards_dir": str(layout.dashboards),
        },
        supporting_outputs={
            "shared_leap_balance_esto_long": shared_conversion_paths.get("shared_leap_balance_esto_long"),
            "shared_ninth_balance_esto_long": shared_conversion_paths.get("shared_ninth_balance_esto_long"),
            "shared_supporting_files_dir": shared_conversion_paths.get("shared_supporting_files_dir"),
            "mapping_lineage_audit_csv": str(mapping_lineage_audit_path),
            "gap_diagnostics": diagnostics_paths.get("gap_diagnostics"),
            "mapping_rundown_by_sheet": diagnostics_paths.get("mapping_rundown_by_sheet"),
            "mapping_rundown_details": diagnostics_paths.get("mapping_rundown_details"),
            "comparison_issue_summary": diagnostics_paths.get("comparison_issue_summary"),
            "comparison_issue_cause_summary": diagnostics_paths.get("comparison_issue_cause_summary"),
            "chart_line_mapping_ledger": str(chart_line_mapping_path),
            "chart_total_component_ledger": str(chart_total_component_path),
            "runtime_issues_csv": str(runtime_issues_path),
            "runtime_missing_pair_summary_xlsx": str(runtime_missing_pair_summary_path),
            "dashboard_comparator_pair_coverage_xlsx": comparator_pair_coverage_xlsx,
            "missing_mapping_candidates_xlsx": missing_mapping_candidates_path,
            "ninth_mapping_data_coverage_xlsx": ninth_mapping_data_coverage_path,
            "override_report_csv": str(override_report_path),
            "auto_sheet_rows_csv": str(auto_sheet_path),
            "resolved_structure_json": str(resolved_structure_path),
            "balance_coverage_csv": str(coverage_path),
            "balance_unit_diagnostics_csv": str(unit_diag_path),
            "balance_matching_diagnostics_csv": str(matching_diag_path),
            "balance_extraction_summary_json": str(extraction_summary_path),
            "workflow_stage_timings_csv": str(timing_path),
        },
        primary_output_descriptions={
            "comparison_long": "Main ESTO-axis comparison table across LEAP, ESTO, and 9th sources.",
            "comparison_wide": "Wide ESTO-axis comparison table with one column per source.",
            "mapping_status": "Mapping workbook for ESTO-axis comparison rows.",
            "leap_long": "Normalized LEAP balance rows before ESTO-axis aggregation.",
            "simple_leap_balance_mapped": "Compact LEAP balance table on the ESTO-axis structure.",
            "simple_ninth_balance_mapped": "Compact 9th balance table on the ESTO-axis structure.",
            "merged_leap_ninth_esto_balance": "Merged LEAP and 9th compact balance table on the ESTO axis.",
            "dashboard_index": "Main HTML entrypoint for the rendered ESTO-axis dashboard.",
            "charts_dir": "Chart files used by the ESTO-axis dashboards.",
            "dashboards_dir": "Rendered ESTO-axis dashboard HTML pages.",
        },
        supporting_output_descriptions={
            "shared_leap_balance_esto_long": "Reusable LEAP balance long table mapped to ESTO rows.",
            "shared_ninth_balance_esto_long": "Reusable 9th balance long table mapped to ESTO rows.",
            "shared_supporting_files_dir": "Shared supporting folder for the reusable ESTO-long conversion outputs.",
            "gap_diagnostics": "Largest gaps between LEAP and ESTO-axis comparator sources.",
            "mapping_lineage_audit_csv": "Row-level mapping lineage for LEAP/9th/ESTO at audit years (see dataset column).",
            "mapping_rundown_by_sheet": "Sheet-level summary of ESTO-axis mapping completeness.",
            "mapping_rundown_details": "Detailed mapping audit workbook for ESTO-axis rows.",
            "comparison_issue_summary": "Prioritized ESTO-axis comparison issues with gap metrics and hints.",
            "comparison_issue_cause_summary": "Frequency summary of ESTO-axis issue categories.",
            "chart_line_mapping_ledger": "Per-chart-line ledger linking visible chart rows to mapping decisions.",
            "chart_total_component_ledger": "Ledger showing how visible total lines were constructed.",
            "runtime_issues_csv": "Runtime balance rows that could not be mapped cleanly.",
            "runtime_missing_pair_summary_xlsx": "Grouped summary of missing ESTO-axis mapping pairs.",
            "dashboard_comparator_pair_coverage_xlsx": "Coverage audit for comparator pairs actually exposed in ESTO-axis dashboards.",
            "missing_mapping_candidates_xlsx": "Workbook of candidate mapping additions based on runtime misses.",
            "ninth_mapping_data_coverage_xlsx": "Coverage check for mapped 9th pairs against available 9th data.",
            "override_report_csv": "Report of which manual overrides were applied.",
            "auto_sheet_rows_csv": "Rows automatically assigned to dashboard sheets during preparation.",
            "resolved_structure_json": "Resolved dashboard structure config used for rendering.",
            "balance_coverage_csv": "Coverage summary from LEAP balance extraction.",
            "balance_unit_diagnostics_csv": "Unit normalization checks from LEAP balance extraction.",
            "balance_matching_diagnostics_csv": "Row-level detail-mode and allocation diagnostics from LEAP balance extraction.",
            "balance_extraction_summary_json": "Summary metadata from the balance extraction stage.",
            "workflow_stage_timings_csv": "Runtime duration by broad workflow stage.",
        },
        notes=[
            "Primary ESTO-axis balance outputs are written at the workflow root.",
            "Supporting diagnostics and mapping evidence are grouped under supporting_files/.",
        ],
    )
    timer.lap("write manifest")

    result = {
        **core_paths,
        **shared_conversion_paths,
        "gap_diagnostics": diagnostics_paths.get("gap_diagnostics"),
        "mapping_rundown_by_sheet": diagnostics_paths.get("mapping_rundown_by_sheet"),
        "mapping_rundown_details": diagnostics_paths.get("mapping_rundown_details"),
        "comparison_issue_summary": diagnostics_paths.get("comparison_issue_summary"),
        "comparison_issue_cause_summary": diagnostics_paths.get("comparison_issue_cause_summary"),
        "simple_leap_balance_mapped": str(simple_leap_balance_path),
        "simple_ninth_balance_mapped": str(simple_ninth_balance_path),
        "merged_leap_ninth_esto_balance": str(merged_esto_axis_balance_path),
        "simple_balance_duplicate_summary": str(duplicate_summary_path),
        "simple_balance_duplicate_details": str(duplicate_details_path),
        "mapped_leap_to_esto_balance_rows": str(mapped_leap_to_esto_path),
        "mapped_ninth_to_esto_balance_rows": str(mapped_ninth_to_esto_path),
        "esto_axis_comparison_long": str(esto_axis_comparison_long_path),
        "mapping_lineage_audit_csv": str(mapping_lineage_audit_path),
        "chart_line_mapping_ledger": str(chart_line_mapping_path),
        "chart_total_component_ledger": str(chart_total_component_path),
        "dashboard_index": dashboard_paths.get("dashboard_index"),
        "charts_written": dashboard_paths.get("charts_written"),
        "empty_pages_csv": dashboard_paths.get("empty_pages_csv"),
        "chart_navigation_hierarchy": dashboard_paths.get("chart_navigation_hierarchy"),
        "chart_navigation_hierarchy_flat": dashboard_paths.get("chart_navigation_hierarchy_flat"),
        "chart_navigation_guide": str(CHART_NAVIGATION_GUIDE_PATH),
        "chart_navigation_rendered_template": dashboard_paths.get("chart_navigation_rendered_template"),
        "runtime_issues_csv": str(runtime_issues_path),
        "runtime_missing_pair_summary_xlsx": str(runtime_missing_pair_summary_path),
        "dashboard_comparator_pair_coverage_xlsx": comparator_pair_coverage_xlsx,
        "missing_mapping_candidates_xlsx": missing_mapping_candidates_path,
        "ninth_mapping_data_coverage_xlsx": ninth_mapping_data_coverage_path,
        "override_report_csv": str(override_report_path),
        "auto_sheet_rows_csv": str(auto_sheet_path),
        "resolved_structure_json": str(resolved_structure_path),
        "balance_coverage_csv": str(coverage_path),
        "balance_unit_diagnostics_csv": str(unit_diag_path),
        "balance_matching_diagnostics_csv": str(matching_diag_path),
        "balance_extraction_summary_json": str(extraction_summary_path),
        "output_manifest": str(manifest),
        "workflow_stage_timings_csv": str(timing_path),
    }
    try:
        _raise_if_unmapped_balance_rows(filtered_runtime_issues, runtime_issues_path)
    except RuntimeError as exc:
        result["runtime_error"] = str(exc)
        result["completed_with_unmapped_rows"] = True
    timer.finish()
    if WRITE_WORKFLOW_TIMING_CSV:
        timer.write_csv(timing_path)
    return result

#%%
RUN_WORKFLOW = True
WORKFLOW_RESULT: dict[str, object] | None = None
if RUN_WORKFLOW:
    WORKFLOW_RESULT = run_workflow()
    print("[OK] Balance dashboard ESTO-axis workflow complete.")
    for key, value in WORKFLOW_RESULT.items():
        print(f"- {key}: {value}")
#%%
