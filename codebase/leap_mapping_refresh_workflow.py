#%%

"""
Refresh maintenance columns in config/leap_mappings.xlsx.

This workflow recomputes the lightweight audit columns used to maintain
`leap_combined_esto` and `leap_combined_ninth` without rerunning the full
dashboard process.
"""

from __future__ import annotations

import os
import re
import shutil
import sys
from pathlib import Path
from typing import Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))


#%%
def _resolve(path: Path | str) -> Path:
    raw = str(path).replace("\\", "/")
    drive_match = re.match(r"^([a-zA-Z]):/(.*)$", raw)
    if drive_match:
        drive = drive_match.group(1).lower()
        rest = drive_match.group(2)
        if os.name == "nt":
            return Path(f"{drive.upper()}:/{rest}")
        return Path(f"/mnt/{drive}/{rest}")
    candidate = Path(raw)
    return candidate if candidate.is_absolute() else (REPO_ROOT / candidate)


MAPPING_WORKBOOK_PATH = _resolve("config/leap_mappings.xlsx")
ESTO_TABLE_PATH = _resolve("data/00APEC_2025_low_with_subtotals.csv")
NINTH_TABLE_PATH = _resolve("data/merged_file_energy_ALL_20251106.csv")
OUTPUT_DIR = _resolve("outputs/mappings/mapping_checks")
MISSING_PAIRS_CSV_PATH = OUTPUT_DIR / "leap_mapping_missing_pairs.csv"
DUPLICATE_MAPPINGS_CSV_PATH = OUTPUT_DIR / "leap_mapping_duplicate_mappings.csv"
TRIO_PRESENCE_CSV_PATH = OUTPUT_DIR / "leap_mapping_trio_presence_check.csv"

TRIO_PRESENCE_OUTPUT_NOTE = (
    "Filter trio_presence_csv by presence_status first. Focus most on "
    "ninth_active_esto_removed, esto_active_ninth_removed, "
    "esto_removed_ninth_active, and ninth_active_esto_missing; these are mapped "
    "rows that can change expected dashboard results. Treat fuel=Total rows, "
    "same-target parent/child mappings, old incorrect fuel rows, expected "
    "losses-sector removals like 10.01.02/10.01.03, and detailed transport "
    "rows kept remove_row=True as low priority."
)

ESTO_SHEET = "leap_combined_esto"
NINTH_SHEET = "leap_combined_ninth"

BASE_YEAR = 2022
PROJECTION_YEARS: Sequence[int] = tuple(range(2023, 2061))
PROJECTION_SCENARIOS: Sequence[str] = ("reference", "target")


#%%
def _clean(value: object) -> str:
    text = str(value or "").strip()
    return "" if text.lower() in {"", "nan", "none"} else text


def _norm_text(value: object) -> str:
    return " ".join(_clean(value).lower().split())


def _truthy(value: object) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "t", "yes", "y", "on"}


def _path_key(path: object) -> str:
    parts = [part.strip() for part in str(path or "").split("/") if part.strip()]
    return "/".join(_norm_text(part) for part in parts)


def _mapping_cardinality(source_target_count: int, target_source_count: int) -> str:
    if source_target_count <= 0 or target_source_count <= 0:
        return ""
    if source_target_count == 1 and target_source_count == 1:
        return "one_to_one"
    if source_target_count > 1 and target_source_count == 1:
        return "one_to_many"
    if source_target_count == 1 and target_source_count > 1:
        return "many_to_one"
    return "many_to_many"


def _subtotal_alignment(leap_is_subtotal: bool, target_is_subtotal: bool) -> str:
    if leap_is_subtotal and target_is_subtotal:
        return "aligned_subtotal"
    if (not leap_is_subtotal) and (not target_is_subtotal):
        return "aligned_non_subtotal"
    return "mismatch"


def _active_mask(frame: pd.DataFrame) -> pd.Series:
    remove_mask = frame.get("remove_row", False)
    duplicate_mask = frame.get("duplicate_to_remove", False)
    remove_mask = pd.Series(remove_mask, index=frame.index).map(_truthy)
    duplicate_mask = pd.Series(duplicate_mask, index=frame.index).map(_truthy)
    return ~(remove_mask | duplicate_mask)


def _drop_unnamed_columns(frame: pd.DataFrame) -> pd.DataFrame:
    keep_cols = [col for col in frame.columns if not str(col).startswith("Unnamed:")]
    return frame.loc[:, keep_cols].copy()


def _drop_columns_if_present(frame: pd.DataFrame, columns: Sequence[str]) -> pd.DataFrame:
    drop_cols = [col for col in columns if col in frame.columns]
    if not drop_cols:
        return frame.copy()
    return frame.drop(columns=drop_cols).copy()


def _reorder_columns(frame: pd.DataFrame, preferred_columns: Sequence[str]) -> pd.DataFrame:
    ordered = [col for col in preferred_columns if col in frame.columns]
    trailing = [col for col in frame.columns if col not in ordered]
    return frame.loc[:, ordered + trailing].copy()


def _compute_leap_subtotals(frame: pd.DataFrame) -> pd.DataFrame:
    out = frame.copy()
    for col in ["leap_sector_name_full_path", "raw_leap_fuel_name"]:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].fillna("").astype(str).str.strip()

    active = _active_mask(out)
    active_paths = {
        _clean(value)
        for value in out.loc[active, "leap_sector_name_full_path"].tolist()
        if _clean(value)
    }

    def leap_sector_is_subtotal(path: object) -> bool:
        text = _clean(path)
        key = _path_key(text)
        if not key:
            return False
        if key.startswith("total "):
            return True
        prefix = f"{text}/"
        return any(other != text and other.startswith(prefix) for other in active_paths)

    def leap_fuel_is_subtotal(fuel: object) -> bool:
        key = _norm_text(fuel)
        return key == "total" or key.startswith("total ")

    leap_sector_is_subtotal = out["leap_sector_name_full_path"].map(leap_sector_is_subtotal)
    leap_fuel_is_subtotal = out["raw_leap_fuel_name"].map(leap_fuel_is_subtotal)
    out["leap_is_subtotal"] = leap_sector_is_subtotal.fillna(False).astype(bool) | leap_fuel_is_subtotal.fillna(False).astype(bool)
    return out


def _compute_pair_cardinality(frame: pd.DataFrame, target_sector_col: str, target_fuel_col: str) -> pd.DataFrame:
    """Compute cardinality of (leap_sector, leap_fuel) <-> (target_sector, target_fuel) pairs."""
    out = frame.copy()
    source_cols = ["leap_sector_name_full_path", "raw_leap_fuel_name"]
    target_cols = [target_sector_col, target_fuel_col]
    all_cols = source_cols + [c for c in target_cols if c not in source_cols]
    for col in all_cols:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].fillna("").astype(str).str.strip()
    active = _active_mask(out)
    valid = (
        active
        & out["leap_sector_name_full_path"].ne("")
        & out["raw_leap_fuel_name"].ne("")
        & out[target_sector_col].ne("")
        & out[target_fuel_col].ne("")
    )
    pair_frame = out.loc[valid, source_cols + [target_sector_col, target_fuel_col]].copy()
    pair_frame["_source_key"] = pair_frame["leap_sector_name_full_path"] + "|||" + pair_frame["raw_leap_fuel_name"]
    pair_frame["_target_key"] = pair_frame[target_sector_col] + "|||" + pair_frame[target_fuel_col]
    pairs = pair_frame[["_source_key", "_target_key"]].drop_duplicates()
    source_count = pairs.groupby("_source_key")["_target_key"].nunique()
    target_count = pairs.groupby("_target_key")["_source_key"].nunique()
    out["_source_key"] = out["leap_sector_name_full_path"].fillna("").astype(str).str.strip() + "|||" + out["raw_leap_fuel_name"].fillna("").astype(str).str.strip()
    out["_target_key"] = out[target_sector_col].fillna("").astype(str).str.strip() + "|||" + out[target_fuel_col].fillna("").astype(str).str.strip()
    out["pair_mapping_cardinality"] = ""
    valid_rows = out["leap_sector_name_full_path"].ne("") & out["raw_leap_fuel_name"].ne("") & out[target_sector_col].ne("") & out[target_fuel_col].ne("")
    out.loc[valid_rows, "pair_mapping_cardinality"] = out.loc[valid_rows].apply(
        lambda row: _mapping_cardinality(
            int(source_count.get(row["_source_key"], 0)),
            int(target_count.get(row["_target_key"], 0)),
        ),
        axis=1,
    )
    out = out.drop(columns=["_source_key", "_target_key"])
    return out


def _apply_auto_remove_rules(frame: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, int]]:
    """Mark obvious rows as remove_row=True and annotate the reason."""
    out = frame.copy()
    for col in [
        "leap_sector_name_full_path",
        "raw_leap_fuel_name",
        "remove_row",
        "remove_row_reason",
    ]:
        if col not in out.columns:
            out[col] = ""
    out["leap_sector_name_full_path"] = out["leap_sector_name_full_path"].fillna("").astype(str).str.strip()
    out["raw_leap_fuel_name"] = out["raw_leap_fuel_name"].fillna("").astype(str).str.strip()
    out["remove_row_reason"] = out["remove_row_reason"].fillna("").astype(str).str.strip()

    fuel_total_mask = out["raw_leap_fuel_name"].map(_norm_text).eq("total")
    def _sector_ends_with_fuel(path_value: object, fuel_value: object) -> bool:
        path_text = _clean(path_value)
        fuel_text = _clean(fuel_value)
        if not path_text or not fuel_text:
            return False
        parts = [part.strip() for part in path_text.split("/") if part.strip()]
        return len(parts) > 1 and parts[-1] == fuel_text

    suffix_mask = out.apply(
        lambda row: _sector_ends_with_fuel(row["leap_sector_name_full_path"], row["raw_leap_fuel_name"])
        and _norm_text(row["raw_leap_fuel_name"]) != "total"
        and not _clean(row["leap_sector_name_full_path"]).startswith("Electricity Generation/"),
        axis=1,
    )

    existing_remove_mask = out["remove_row"].map(_truthy)
    auto_mask = fuel_total_mask | suffix_mask
    newly_removed_mask = auto_mask & ~existing_remove_mask

    out["remove_row"] = existing_remove_mask | auto_mask

    def _append_reason(existing: str, reason: str) -> str:
        if not reason:
            return existing
        if not existing:
            return reason
        if reason in existing.split(" | "):
            return existing
        return f"{existing} | {reason}"

    def _strip_auto_reasons(existing: str) -> str:
        reasons = [part.strip() for part in existing.split(" | ") if part.strip()]
        reasons = [reason for reason in reasons if reason not in {"auto_remove_total_fuel", "auto_remove_sector_fuel_suffix"}]
        return " | ".join(reasons)

    out["remove_row_reason"] = out["remove_row_reason"].map(_strip_auto_reasons)
    out.loc[fuel_total_mask, "remove_row_reason"] = out.loc[fuel_total_mask, "remove_row_reason"].map(
        lambda reason: _append_reason(reason, "auto_remove_total_fuel")
    )
    out.loc[suffix_mask, "remove_row_reason"] = out.loc[suffix_mask, "remove_row_reason"].map(
        lambda reason: _append_reason(reason, "auto_remove_sector_fuel_suffix")
    )

    diagnostics = {
        "auto_remove_total_fuel_rows": int(fuel_total_mask.sum()),
        "auto_remove_sector_fuel_suffix_rows": int(suffix_mask.sum()),
        "auto_removed_new_rows": int(newly_removed_mask.sum()),
    }
    return out, diagnostics


def _load_esto_lookup() -> pd.DataFrame:
    base_df = pd.read_csv(ESTO_TABLE_PATH)
    work = base_df.copy()
    if "is_subtotal" not in work.columns:
        work["is_subtotal"] = False
    for col in ["economy", "flows", "products", str(BASE_YEAR), "is_subtotal"]:
        if col not in work.columns:
            work[col] = ""
    work["esto_flow"] = work["flows"].fillna("").astype(str).str.strip()
    work["esto_product"] = work["products"].fillna("").astype(str).str.strip()
    work["value"] = pd.to_numeric(work[str(BASE_YEAR)], errors="coerce").fillna(0.0)
    work["is_subtotal"] = work["is_subtotal"].fillna(False).map(_truthy)
    work = work[work["esto_flow"].ne("") & work["esto_product"].ne("")].copy()
    grouped = (
        work.groupby(["esto_flow", "esto_product"], as_index=False)
        .agg(
            pair_value_sum=("value", "sum"),
            esto_pair_is_subtotal=("is_subtotal", "max"),
        )
        .reset_index(drop=True)
    )
    grouped["esto_pair_abs_sum"] = grouped["pair_value_sum"].abs()
    return grouped


def _load_ninth_lookup() -> pd.DataFrame:
    ninth_df = pd.read_csv(NINTH_TABLE_PATH)
    work = ninth_df.copy()
    for col in [
        "economy",
        "scenarios",
        "sectors",
        "sub1sectors",
        "sub2sectors",
        "sub3sectors",
        "sub4sectors",
        "fuels",
        "subfuels",
        "subtotal_layout",
        "subtotal_results",
    ]:
        if col not in work.columns:
            work[col] = ""
    for col in ["subtotal_layout", "subtotal_results"]:
        work[col] = work[col].fillna(False).map(_truthy)
    scenario_set = {str(value).strip().lower() for value in PROJECTION_SCENARIOS}
    work = work[work["scenarios"].fillna("").astype(str).str.strip().str.lower().isin(scenario_set)].copy()
    year_cols = [str(year) for year in PROJECTION_YEARS if str(year) in work.columns]
    if not year_cols or work.empty:
        return pd.DataFrame(
            columns=[
                "ninth_sector",
                "ninth_fuel",
                "ninth_pair_is_subtotal",
                "ninth_pair_abs_sum",
            ]
        )
    values = work[year_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)
    work["ninth_sector"] = work.apply(
        lambda row: next(
            (
                _clean(row.get(col, ""))
                for col in ["sub4sectors", "sub3sectors", "sub2sectors", "sub1sectors", "sectors"]
                if _clean(row.get(col, ""))
            ),
            "",
        ),
        axis=1,
    )
    work["ninth_fuel"] = work.apply(
        lambda row: next(
            (
                _clean(row.get(col, ""))
                for col in ["subfuels", "fuels"]
                if _clean(row.get(col, ""))
            ),
            "",
        ),
        axis=1,
    )
    work["value_abs_sum_row"] = values.abs().sum(axis=1)
    work = work[work["ninth_sector"].ne("") & work["ninth_fuel"].ne("")].copy()
    grouped = (
        work.groupby(["ninth_sector", "ninth_fuel"], as_index=False)
        .agg(
            subtotal_layout=("subtotal_layout", "max"),
            subtotal_results=("subtotal_results", "max"),
            ninth_pair_abs_sum=("value_abs_sum_row", "sum"),
        )
        .reset_index(drop=True)
    )
    grouped["ninth_pair_is_subtotal"] = (
        grouped["subtotal_layout"].fillna(False).astype(bool)
        | grouped["subtotal_results"].fillna(False).astype(bool)
    )
    return grouped


def _refresh_esto_sheet(frame: pd.DataFrame, esto_lookup: pd.DataFrame) -> pd.DataFrame:
    out = _drop_unnamed_columns(frame)
    out = _drop_columns_if_present(
        out,
        [
            "esto_pair_is_subtotal",
            "esto_pair_is_subtotal_x",
            "esto_pair_is_subtotal_y",
            "esto_pair_abs_sum",
            "esto_pair_abs_sum_x",
            "esto_pair_abs_sum_y",
            "leap_sector_is_subtotal_computed",
            "leap_fuel_is_subtotal_computed",
        ],
    )
    out = _compute_leap_subtotals(out)
    out = _compute_pair_cardinality(out, "esto_flow", "esto_product")
    lookup = esto_lookup.copy()
    for col in ["esto_flow", "esto_product"]:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].fillna("").astype(str).str.strip()
    out = out.merge(
        lookup[["esto_flow", "esto_product", "esto_pair_is_subtotal", "esto_pair_abs_sum"]],
        on=["esto_flow", "esto_product"],
        how="left",
    )
    if "esto_pair_is_subtotal" not in out.columns:
        out["esto_pair_is_subtotal"] = False
    out["esto_pair_is_subtotal"] = out["esto_pair_is_subtotal"].fillna(False).astype(bool)
    if "esto_pair_abs_sum" not in out.columns:
        out["esto_pair_abs_sum"] = 0.0
    out["esto_pair_abs_sum"] = pd.to_numeric(out["esto_pair_abs_sum"], errors="coerce").fillna(0.0)
    total_mask = out["esto_product"].fillna("").astype(str).str.strip().str.lower().eq("19 total")
    out.loc[total_mask, "esto_pair_is_subtotal"] = True
    out["subtotal_alignment"] = out.apply(
        lambda row: _subtotal_alignment(bool(row.get("leap_is_subtotal", False)), bool(row.get("esto_pair_is_subtotal", False))),
        axis=1,
    )
    return _reorder_columns(
        out,
        [
            "leap_sector_name_original",
            "leap_sector_name_full_path",
            "raw_leap_fuel_name",
            "value",
            "esto_flow",
            "esto_product",
            "pair_mapping_cardinality",
            "leap_is_subtotal",
            "esto_pair_is_subtotal",
            "subtotal_mismatch_is_ok",
            "subtotal_alignment",
            "esto_pair_abs_sum",
            "many_to_many_is_ok",
            "remove_row",
            "remove_row_reason",
        ],
    )


def _refresh_ninth_sheet(frame: pd.DataFrame, ninth_lookup: pd.DataFrame) -> pd.DataFrame:
    out = _drop_unnamed_columns(frame)
    out = _drop_columns_if_present(
        out,
        [
            "ninth_pair_is_subtotal",
            "ninth_pair_is_subtotal_x",
            "ninth_pair_is_subtotal_y",
            "ninth_pair_abs_sum",
            "ninth_pair_abs_sum_x",
            "ninth_pair_abs_sum_y",
            "leap_sector_is_subtotal_computed",
            "leap_fuel_is_subtotal_computed",
        ],
    )
    out = _compute_leap_subtotals(out)
    out = _compute_pair_cardinality(out, "ninth_sector", "ninth_fuel")
    lookup = ninth_lookup.copy()
    for col in ["ninth_sector", "ninth_fuel"]:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].fillna("").astype(str).str.strip()
    out = out.merge(
        lookup[["ninth_sector", "ninth_fuel", "ninth_pair_is_subtotal", "ninth_pair_abs_sum"]],
        on=["ninth_sector", "ninth_fuel"],
        how="left",
    )
    if "ninth_pair_is_subtotal" not in out.columns:
        out["ninth_pair_is_subtotal"] = False
    out["ninth_pair_is_subtotal"] = out["ninth_pair_is_subtotal"].fillna(False).astype(bool)
    if "ninth_pair_abs_sum" not in out.columns:
        out["ninth_pair_abs_sum"] = 0.0
    out["ninth_pair_abs_sum"] = pd.to_numeric(out["ninth_pair_abs_sum"], errors="coerce").fillna(0.0)
    total_mask = out["ninth_fuel"].fillna("").astype(str).str.strip().str.lower().eq("19_total")
    out.loc[total_mask, "ninth_pair_is_subtotal"] = True
    out["subtotal_alignment"] = out.apply(
        lambda row: _subtotal_alignment(bool(row.get("leap_is_subtotal", False)), bool(row.get("ninth_pair_is_subtotal", False))),
        axis=1,
    )
    return _reorder_columns(
        out,
        [
            "leap_sector_name_original",
            "leap_sector_name_full_path",
            "raw_leap_fuel_name",
            "value",
            "ninth_sector",
            "ninth_fuel",
            "pair_mapping_cardinality",
            "leap_is_subtotal",
            "ninth_pair_is_subtotal",
            "subtotal_mismatch_is_ok",
            "subtotal_alignment",
            "ninth_pair_abs_sum",
            "many_to_many_is_ok",
            "remove_row",
            "remove_row_reason",
        ],
    )


def _active_pairs(frame: pd.DataFrame, col_a: str, col_b: str) -> set[tuple[str, str]]:
    """Return the set of (col_a, col_b) pairs in active (non-removed) rows."""
    active = frame[_active_mask(frame)].copy()
    a = active[col_a].fillna("").astype(str).str.strip() if col_a in active.columns else pd.Series("", index=active.index)
    b = active[col_b].fillna("").astype(str).str.strip() if col_b in active.columns else pd.Series("", index=active.index)
    return {(av, bv) for av, bv in zip(a, b) if av and bv}


def _build_duplicate_mappings(frame: pd.DataFrame, *, sheet_name: str, target_a: str, target_b: str) -> pd.DataFrame:
    """Return exact active duplicate source/target rows for one mapping sheet."""
    work = frame.copy().fillna("")
    source_cols = ["leap_sector_name_full_path", "raw_leap_fuel_name"]
    target_cols = [target_a, target_b]
    required_cols = [*source_cols, *target_cols, "remove_row", "duplicate_to_remove"]
    for col in required_cols:
        if col not in work.columns:
            work[col] = ""
        work[col] = work[col].fillna("").astype(str).str.strip()

    active = work[_active_mask(work)].copy()
    valid = active[source_cols + target_cols].apply(lambda col: col.map(_clean).ne("")).all(axis=1)
    active = active.loc[valid].copy()
    if active.empty:
        return pd.DataFrame(
            columns=[
                "sheet_name",
                "mapping_row_number",
                "duplicate_group_size",
                *source_cols,
                *target_cols,
            ]
        )

    duplicate_mask = active.duplicated(subset=[*source_cols, *target_cols], keep=False)
    duplicates = active.loc[duplicate_mask].copy()
    if duplicates.empty:
        return pd.DataFrame(
            columns=[
                "sheet_name",
                "mapping_row_number",
                "duplicate_group_size",
                *source_cols,
                *target_cols,
            ]
        )

    duplicates.insert(0, "mapping_row_number", duplicates.index + 2)
    duplicates.insert(0, "sheet_name", sheet_name)
    duplicates["duplicate_group_size"] = duplicates.groupby([*source_cols, *target_cols])[source_cols[0]].transform("size")
    return duplicates[[
        "sheet_name",
        "mapping_row_number",
        "duplicate_group_size",
        *source_cols,
        *target_cols,
    ]].reset_index(drop=True)


def _build_trio_presence_check(esto_sheet: pd.DataFrame, ninth_sheet: pd.DataFrame) -> pd.DataFrame:
    """Return row-level presence diagnostics for the two mapping sheets."""
    source_cols = ["leap_sector_name_original", "leap_sector_name_full_path", "raw_leap_fuel_name"]

    def _sheet_row_status(frame: pd.DataFrame, sheet_name: str, target_cols: list[str]) -> pd.DataFrame:
        work = frame.copy().fillna("")
        for col in source_cols + target_cols + ["remove_row", "duplicate_to_remove"]:
            if col not in work.columns:
                work[col] = ""
            work[col] = work[col].fillna("").astype(str).str.strip()
        valid = work[source_cols + target_cols].apply(lambda col: col.map(_clean).ne("")).all(axis=1)
        work = work.loc[valid].copy()
        if work.empty:
            return pd.DataFrame(
                columns=[
                    "sheet_name",
                    "mapping_row_number",
                    *source_cols,
                    *target_cols,
                    "this_row_status",
                    "this_row_is_removed",
                    "this_row_is_duplicate_removed",
                ]
            )
        work["sheet_name"] = sheet_name
        work["mapping_row_number"] = work.index + 2
        work["this_row_is_removed"] = work["remove_row"].map(_truthy)
        work["this_row_is_duplicate_removed"] = work["duplicate_to_remove"].map(_truthy)
        work["this_row_status"] = work.apply(
            lambda row: "removed_row_true"
            if row["this_row_is_removed"]
            else "duplicate_removed_row_true"
            if row["this_row_is_duplicate_removed"]
            else "active",
            axis=1,
        )
        return work[
            [
                "sheet_name",
                "mapping_row_number",
                *source_cols,
                *target_cols,
                "this_row_status",
                "this_row_is_removed",
                "this_row_is_duplicate_removed",
            ]
        ].reset_index(drop=True)

    def _first_non_empty(series: pd.Series) -> str:
        values = [str(value).strip() for value in series.tolist() if _clean(value)]
        unique_values = list(dict.fromkeys(values))
        return " | ".join(unique_values)

    def _sheet_source_summary(frame: pd.DataFrame, sheet_name: str, target_cols: list[str]) -> pd.DataFrame:
        work = frame.copy().fillna("")
        for col in source_cols + target_cols + ["remove_row", "duplicate_to_remove"]:
            if col not in work.columns:
                work[col] = ""
            work[col] = work[col].fillna("").astype(str).str.strip()
        valid = work[source_cols + target_cols].apply(lambda col: col.map(_clean).ne("")).all(axis=1)
        work = work.loc[valid].copy()
        if work.empty:
            return pd.DataFrame(
                columns=[
                    *source_cols,
                    *target_cols,
                    f"{sheet_name}_active_row_count",
                    f"{sheet_name}_removed_row_count",
                    f"{sheet_name}_duplicate_removed_row_count",
                    f"{sheet_name}_presence_state",
                ]
            )
        work[f"{sheet_name}_is_active"] = ~work["remove_row"].map(_truthy) & ~work["duplicate_to_remove"].map(_truthy)
        work[f"{sheet_name}_is_removed"] = work["remove_row"].map(_truthy)
        work[f"{sheet_name}_is_duplicate_removed"] = work["duplicate_to_remove"].map(_truthy)
        grouped = (
            work.groupby(source_cols, as_index=False)
            .agg(
                **{
                    f"{sheet_name}_active_row_count": (f"{sheet_name}_is_active", "sum"),
                    f"{sheet_name}_removed_row_count": (f"{sheet_name}_is_removed", "sum"),
                    f"{sheet_name}_duplicate_removed_row_count": (f"{sheet_name}_is_duplicate_removed", "sum"),
                    **{col: (col, _first_non_empty) for col in target_cols},
                }
            )
            .reset_index(drop=True)
        )
        for col in [
            f"{sheet_name}_active_row_count",
            f"{sheet_name}_removed_row_count",
            f"{sheet_name}_duplicate_removed_row_count",
        ]:
            grouped[col] = pd.to_numeric(grouped[col], errors="coerce").fillna(0).astype(int)
        grouped[f"{sheet_name}_presence_state"] = grouped.apply(
            lambda row: "active"
            if row[f"{sheet_name}_active_row_count"] > 0
            else "removed_only"
            if row[f"{sheet_name}_removed_row_count"] > 0
            else "duplicate_removed_only"
            if row[f"{sheet_name}_duplicate_removed_row_count"] > 0
            else "missing",
            axis=1,
        )
        return grouped

    def _comparison_status(sheet_name: str, this_row_status: str, counterpart_presence_state: str) -> str:
        if this_row_status == "active" and counterpart_presence_state == "active":
            return "both_active"
        if sheet_name == "esto":
            if this_row_status == "active" and counterpart_presence_state in {"removed_only", "duplicate_removed_only"}:
                return "esto_active_ninth_removed"
            if this_row_status in {"removed_row_true", "duplicate_removed_row_true"} and counterpart_presence_state == "active":
                return "esto_removed_ninth_active"
            if this_row_status == "active" and counterpart_presence_state == "missing":
                return "esto_active_ninth_missing"
            if this_row_status in {"removed_row_true", "duplicate_removed_row_true"} and counterpart_presence_state == "missing":
                return "esto_removed_ninth_missing"
        if sheet_name == "ninth":
            if this_row_status == "active" and counterpart_presence_state in {"removed_only", "duplicate_removed_only"}:
                return "ninth_active_esto_removed"
            if this_row_status in {"removed_row_true", "duplicate_removed_row_true"} and counterpart_presence_state == "active":
                return "ninth_removed_esto_active"
            if this_row_status == "active" and counterpart_presence_state == "missing":
                return "ninth_active_esto_missing"
            if this_row_status in {"removed_row_true", "duplicate_removed_row_true"} and counterpart_presence_state == "missing":
                return "ninth_removed_esto_missing"
        if this_row_status in {"removed_row_true", "duplicate_removed_row_true"} and counterpart_presence_state in {"removed_only", "duplicate_removed_only"}:
            return "both_removed"
        if counterpart_presence_state == "missing":
            return "actually_missing"
        return "mixed"

    def _issue_side(comparison_status: str) -> str:
        if comparison_status == "both_active":
            return "both_active"
        if comparison_status == "both_removed":
            return "both_removed"
        if comparison_status in {"esto_removed_ninth_active", "esto_removed_ninth_missing"}:
            return "esto_removed"
        if comparison_status in {"ninth_removed_esto_active", "ninth_removed_esto_missing"}:
            return "ninth_removed"
        if comparison_status in {"esto_active_ninth_removed"}:
            return "ninth_removed"
        if comparison_status in {"ninth_active_esto_removed"}:
            return "esto_removed"
        if comparison_status in {"esto_active_ninth_missing"}:
            return "ninth_missing"
        if comparison_status in {"ninth_active_esto_missing"}:
            return "esto_missing"
        if comparison_status == "actually_missing":
            return "missing"
        return comparison_status

    esto_rows = _sheet_row_status(esto_sheet, "esto", ["esto_flow", "esto_product"])
    ninth_rows = _sheet_row_status(ninth_sheet, "ninth", ["ninth_sector", "ninth_fuel"])
    esto_summary = _sheet_source_summary(esto_sheet, "esto", ["esto_flow", "esto_product"])
    ninth_summary = _sheet_source_summary(ninth_sheet, "ninth", ["ninth_sector", "ninth_fuel"])

    esto_rows = esto_rows.merge(
        ninth_summary[source_cols + ["ninth_sector", "ninth_fuel", "ninth_presence_state"]],
        on=source_cols,
        how="left",
    )
    ninth_rows = ninth_rows.merge(
        esto_summary[source_cols + ["esto_flow", "esto_product", "esto_presence_state"]],
        on=source_cols,
        how="left",
    )

    esto_rows["counterpart_presence_state"] = esto_rows["ninth_presence_state"].fillna("missing")
    ninth_rows["counterpart_presence_state"] = ninth_rows["esto_presence_state"].fillna("missing")

    esto_rows["presence_status"] = esto_rows.apply(
        lambda row: _comparison_status("esto", row["this_row_status"], row["counterpart_presence_state"]),
        axis=1,
    )
    ninth_rows["presence_status"] = ninth_rows.apply(
        lambda row: _comparison_status("ninth", row["this_row_status"], row["counterpart_presence_state"]),
        axis=1,
    )

    for work in [esto_rows, ninth_rows]:
        work["comparison_status"] = work["presence_status"]
        work["row_status"] = work["this_row_status"]
        work["issue_side"] = work["comparison_status"].map(_issue_side)
        work["missing_reason"] = work["comparison_status"].map(lambda value: "" if value == "both_active" else value)
        work["has_removed_row"] = work["this_row_is_removed"]
        work["has_duplicate_removed_row"] = work["this_row_is_duplicate_removed"]

    combined = pd.concat([esto_rows, ninth_rows], ignore_index=True)
    combined["issue_side"] = combined.apply(
        lambda row: _issue_side(str(row.get("comparison_status", ""))),
        axis=1,
    )
    combined["is_issue_row"] = combined["comparison_status"].ne("both_active")

    return combined.sort_values(
        ["is_issue_row", "sheet_name", *source_cols, "mapping_row_number"],
        ascending=[False, True, True, True, True, True],
    ).reset_index(drop=True)


def _build_coverage_gaps(
    esto_sheet: pd.DataFrame,
    ninth_sheet: pd.DataFrame,
    esto_lookup: pd.DataFrame,
    ninth_lookup: pd.DataFrame,
) -> pd.DataFrame:
    """
    Return a DataFrame of all coverage gaps: pairs with abs values > 0 that are
    missing from the active mapping rows.

    Columns: gap_type, sheet_name, original_dataset, original_pair_is_subtotal,
    key_col_1, key_col_2, pair_1, pair_2, abs_sum
      gap_type values:
        "esto_missing"   - esto data pair not in any active esto mapping row
        "ninth_missing"  - 9th data pair not in any active ninth mapping row
        "leap_unmapped_esto"  - LEAP pair with value > 0 but no esto target in esto sheet
        "leap_unmapped_ninth" - LEAP pair with value > 0 but no ninth target in ninth sheet
    """
    records: list[dict] = []

    def _lookup_subtotal_flag(frame: pd.DataFrame, key_a: str, key_b: str, value_a: str, value_b: str, subtotal_col: str) -> bool:
        if subtotal_col not in frame.columns:
            return False
        mask = frame[key_a].astype(str).str.strip().eq(value_a) & frame[key_b].astype(str).str.strip().eq(value_b)
        if not bool(mask.any()):
            return False
        return bool(frame.loc[mask, subtotal_col].fillna(False).astype(bool).any())

    # --- 1. ESTO data pairs missing from the esto mapping ---
    esto_data_pairs = set(
        zip(
            esto_lookup.loc[esto_lookup["esto_pair_abs_sum"] > 0, "esto_flow"].astype(str).str.strip(),
            esto_lookup.loc[esto_lookup["esto_pair_abs_sum"] > 0, "esto_product"].astype(str).str.strip(),
        )
    )
    esto_mapped_pairs = _active_pairs(esto_sheet, "esto_flow", "esto_product")
    for flow, product in sorted(esto_data_pairs - esto_mapped_pairs):
        abs_sum = float(
            esto_lookup.loc[
                esto_lookup["esto_flow"].astype(str).str.strip().eq(flow)
                & esto_lookup["esto_product"].astype(str).str.strip().eq(product),
                "esto_pair_abs_sum",
            ].sum()
        )
        records.append(
            {
                "gap_type": "esto_missing",
                "sheet_name": ESTO_SHEET,
                "original_dataset": "esto",
                "original_pair_is_subtotal": _lookup_subtotal_flag(
                    esto_lookup,
                    "esto_flow",
                    "esto_product",
                    flow,
                    product,
                    "esto_pair_is_subtotal",
                ),
                "key_col_1": "esto_flow",
                "key_col_2": "esto_product",
                "pair_1": flow,
                "pair_2": product,
                "abs_sum": abs_sum,
            }
        )

    # --- 2. 9th data pairs missing from the ninth mapping ---
    ninth_data_pairs = set(
        zip(
            ninth_lookup.loc[ninth_lookup["ninth_pair_abs_sum"] > 0, "ninth_sector"].astype(str).str.strip(),
            ninth_lookup.loc[ninth_lookup["ninth_pair_abs_sum"] > 0, "ninth_fuel"].astype(str).str.strip(),
        )
    )
    ninth_mapped_pairs = _active_pairs(ninth_sheet, "ninth_sector", "ninth_fuel")
    for sector, fuel in sorted(ninth_data_pairs - ninth_mapped_pairs):
        abs_sum = float(
            ninth_lookup.loc[
                ninth_lookup["ninth_sector"].astype(str).str.strip().eq(sector)
                & ninth_lookup["ninth_fuel"].astype(str).str.strip().eq(fuel),
                "ninth_pair_abs_sum",
            ].sum()
        )
        records.append(
            {
                "gap_type": "ninth_missing",
                "sheet_name": NINTH_SHEET,
                "original_dataset": "ninth",
                "original_pair_is_subtotal": _lookup_subtotal_flag(
                    ninth_lookup,
                    "ninth_sector",
                    "ninth_fuel",
                    sector,
                    fuel,
                    "ninth_pair_is_subtotal",
                ),
                "key_col_1": "ninth_sector",
                "key_col_2": "ninth_fuel",
                "pair_1": sector,
                "pair_2": fuel,
                "abs_sum": abs_sum,
            }
        )

    # --- 3. LEAP pairs with abs(value) > 0 that are unmapped ---
    for sheet, gap_type, sheet_name, target_a, target_b in [
        (esto_sheet, "leap_unmapped_esto", ESTO_SHEET, "esto_flow", "esto_product"),
        (ninth_sheet, "leap_unmapped_ninth", NINTH_SHEET, "ninth_sector", "ninth_fuel"),
    ]:
        active = _compute_leap_subtotals(sheet)[_active_mask(sheet)].copy()
        if "value" not in active.columns:
            continue
        values = pd.to_numeric(active["value"], errors="coerce").fillna(0.0).abs()
        leap_sector = active["leap_sector_name_full_path"].fillna("").astype(str).str.strip() if "leap_sector_name_full_path" in active.columns else pd.Series("", index=active.index)
        leap_fuel = active["raw_leap_fuel_name"].fillna("").astype(str).str.strip() if "raw_leap_fuel_name" in active.columns else pd.Series("", index=active.index)
        leap_is_subtotal = active["leap_is_subtotal"].fillna(False).astype(bool) if "leap_is_subtotal" in active.columns else pd.Series(False, index=active.index)
        ta = active[target_a].fillna("").astype(str).str.strip() if target_a in active.columns else pd.Series("", index=active.index)
        tb = active[target_b].fillna("").astype(str).str.strip() if target_b in active.columns else pd.Series("", index=active.index)
        unmapped_mask = (values > 0) & (ta.eq("") | tb.eq(""))
        for sector, fuel in sorted(set(zip(leap_sector[unmapped_mask], leap_fuel[unmapped_mask]))):
            abs_sum = float(values[unmapped_mask & leap_sector.eq(sector) & leap_fuel.eq(fuel)].sum())
            is_subtotal = bool(leap_is_subtotal[unmapped_mask & leap_sector.eq(sector) & leap_fuel.eq(fuel)].any())
            records.append(
                {
                    "gap_type": gap_type,
                    "sheet_name": sheet_name,
                    "original_dataset": "leap",
                    "original_pair_is_subtotal": is_subtotal,
                    "key_col_1": "leap_sector_name_full_path",
                    "key_col_2": "raw_leap_fuel_name",
                    "pair_1": sector,
                    "pair_2": fuel,
                    "abs_sum": abs_sum,
                }
            )

    return pd.DataFrame(
        records,
        columns=[
            "gap_type",
            "sheet_name",
            "original_dataset",
            "original_pair_is_subtotal",
            "key_col_1",
            "key_col_2",
            "pair_1",
            "pair_2",
            "abs_sum",
        ],
    )


COVERAGE_GAPS_PATH = MISSING_PAIRS_CSV_PATH


def _report_coverage_gaps(gaps: pd.DataFrame, *, error_on_gaps: bool) -> None:
    """Write gaps CSV and either raise or warn depending on *error_on_gaps*."""
    COVERAGE_GAPS_PATH.parent.mkdir(parents=True, exist_ok=True)
    gaps.to_csv(COVERAGE_GAPS_PATH, index=False)

    if gaps.empty:
        return

    summary_lines: list[str] = []
    for gap_type, group in gaps.groupby("gap_type"):
        summary_lines.append(f"  {gap_type}: {len(group)} pair(s)")
    try:
        report_path = COVERAGE_GAPS_PATH.relative_to(REPO_ROOT)
    except ValueError:
        report_path = COVERAGE_GAPS_PATH
    summary = (
        f"{len(gaps)} coverage gap(s) found in leap_mappings.xlsx "
        f"(written to {report_path}):\n" + "\n".join(summary_lines)
    )

    if error_on_gaps:
        raise ValueError(summary)
    else:
        import warnings
        warnings.warn(summary, stacklevel=3)


def _report_duplicate_mappings(duplicates: pd.DataFrame) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    duplicates.to_csv(DUPLICATE_MAPPINGS_CSV_PATH, index=False)
    if duplicates.empty:
        return
    summary = (
        f"{len(duplicates)} exact duplicate mapping row(s) found in leap_mappings.xlsx "
        f"(written to {DUPLICATE_MAPPINGS_CSV_PATH.relative_to(REPO_ROOT)})."
    )
    import warnings

    warnings.warn(summary, stacklevel=3)


def _write_trio_presence_csv(trio_presence: pd.DataFrame) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    row_exclusive = trio_presence[~trio_presence["presence_status"].isin({"both_active"})].copy()
    row_exclusive.to_csv(TRIO_PRESENCE_CSV_PATH, index=False)
    if row_exclusive.empty:
        return
    summary = row_exclusive.groupby(["sheet_name", "presence_status"], as_index=False).size().rename(columns={"size": "row_count"}).sort_values(["sheet_name", "presence_status"])
    summary_parts = [
        f"{row.sheet_name}:{row.presence_status}: {int(row.row_count)}"
        for row in summary.itertuples(index=False)
    ]
    import warnings

    warnings.warn(
        "Row presence mismatches found in leap_combined_esto and leap_combined_ninth "
        f"(written to {TRIO_PRESENCE_CSV_PATH.relative_to(REPO_ROOT)}): "
        + ", ".join(summary_parts),
        stacklevel=3,
    )


ARCHIVE_DIR = MAPPING_WORKBOOK_PATH.parent / "archive"


def _backup_workbook(path: Path) -> Path:
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    backup_path = ARCHIVE_DIR / f"{path.stem}.before_refresh_mapping_maintenance_columns_{pd.Timestamp.now():%Y%m%d_%H%M%S}{path.suffix}"
    shutil.copy2(path, backup_path)
    return backup_path


def _assert_not_open(path: Path) -> None:
    """Raise a clear error if the file is locked (open in Excel or another process)."""
    try:
        with open(path, "r+b"):
            pass
    except PermissionError:
        raise PermissionError(
            f"{path.name} is open in another application (e.g. Excel). "
            "Close it and re-run the workflow."
        ) from None


def _replace_sheet_with_dataframe(workbook_path: Path, sheet_name: str, frame: pd.DataFrame) -> None:
    """Replace one sheet in-place while preserving every other sheet in the workbook."""
    workbook = load_workbook(workbook_path)
    if sheet_name in workbook.sheetnames:
        sheet_index = workbook.sheetnames.index(sheet_name)
        del workbook[sheet_name]
        worksheet = workbook.create_sheet(title=sheet_name, index=sheet_index)
    else:
        worksheet = workbook.create_sheet(title=sheet_name)
    for row in dataframe_to_rows(frame, index=False, header=True):
        worksheet.append(row)
    workbook.save(workbook_path)


def run_workflow(*, error_on_gaps: bool = True) -> dict[str, object]:
    """
    Refresh mapping maintenance columns.

    Parameters
    ----------
    error_on_gaps:
        If True (default), raise a ValueError when coverage gaps are found.
        If False, emit a warning and continue; gaps are still written to
        outputs/mappings/mapping_checks/leap_mapping_missing_pairs.csv.
    """
    if not MAPPING_WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Missing mapping workbook: {MAPPING_WORKBOOK_PATH}")
    _assert_not_open(MAPPING_WORKBOOK_PATH)
    backup_path = _backup_workbook(MAPPING_WORKBOOK_PATH)

    esto_sheet = pd.read_excel(MAPPING_WORKBOOK_PATH, sheet_name=ESTO_SHEET, dtype=object).fillna("")
    ninth_sheet = pd.read_excel(MAPPING_WORKBOOK_PATH, sheet_name=NINTH_SHEET, dtype=object).fillna("")

    esto_lookup = _load_esto_lookup()
    ninth_lookup = _load_ninth_lookup()

    refreshed_esto = _refresh_esto_sheet(esto_sheet, esto_lookup)
    refreshed_ninth = _refresh_ninth_sheet(ninth_sheet, ninth_lookup)

    refreshed_esto, esto_auto_remove = _apply_auto_remove_rules(refreshed_esto)
    refreshed_ninth, ninth_auto_remove = _apply_auto_remove_rules(refreshed_ninth)
    refreshed_esto = _compute_pair_cardinality(refreshed_esto, "esto_flow", "esto_product")
    refreshed_ninth = _compute_pair_cardinality(refreshed_ninth, "ninth_sector", "ninth_fuel")

    auto_remove_summary = (
        "Auto-remove rules applied: "
        f"ESTO total fuels={esto_auto_remove['auto_remove_total_fuel_rows']}, "
        f"ESTO suffix matches={esto_auto_remove['auto_remove_sector_fuel_suffix_rows']}, "
        f"9th total fuels={ninth_auto_remove['auto_remove_total_fuel_rows']}, "
        f"9th suffix matches={ninth_auto_remove['auto_remove_sector_fuel_suffix_rows']}, "
        f"newly marked rows={esto_auto_remove['auto_removed_new_rows'] + ninth_auto_remove['auto_removed_new_rows']}."
    )
    import warnings

    warnings.warn(auto_remove_summary, stacklevel=3)

    gaps = _build_coverage_gaps(refreshed_esto, refreshed_ninth, esto_lookup, ninth_lookup)
    _report_coverage_gaps(gaps, error_on_gaps=error_on_gaps)

    duplicate_esto = _build_duplicate_mappings(
        refreshed_esto,
        sheet_name=ESTO_SHEET,
        target_a="esto_flow",
        target_b="esto_product",
    )
    duplicate_ninth = _build_duplicate_mappings(
        refreshed_ninth,
        sheet_name=NINTH_SHEET,
        target_a="ninth_sector",
        target_b="ninth_fuel",
    )
    duplicate_mappings = pd.concat([duplicate_esto, duplicate_ninth], ignore_index=True)
    _report_duplicate_mappings(duplicate_mappings)

    trio_presence = _build_trio_presence_check(refreshed_esto, refreshed_ninth)
    _write_trio_presence_csv(trio_presence)

    _replace_sheet_with_dataframe(MAPPING_WORKBOOK_PATH, ESTO_SHEET, refreshed_esto)
    _replace_sheet_with_dataframe(MAPPING_WORKBOOK_PATH, NINTH_SHEET, refreshed_ninth)

    return {
        "mapping_workbook": str(MAPPING_WORKBOOK_PATH),
        "backup_workbook": str(backup_path),
        "coverage_gaps_csv": str(COVERAGE_GAPS_PATH),
        "coverage_gaps_count": int(len(gaps)),
        "duplicate_mappings_csv": str(DUPLICATE_MAPPINGS_CSV_PATH),
        "duplicate_mappings_count": int(len(duplicate_mappings)),
        "trio_presence_csv": str(TRIO_PRESENCE_CSV_PATH),
        "trio_presence_count": int(len(trio_presence)),
        "auto_remove_total_fuel_rows_esto": int(esto_auto_remove["auto_remove_total_fuel_rows"]),
        "auto_remove_sector_fuel_suffix_rows_esto": int(esto_auto_remove["auto_remove_sector_fuel_suffix_rows"]),
        "auto_removed_new_rows_esto": int(esto_auto_remove["auto_removed_new_rows"]),
        "auto_remove_total_fuel_rows_ninth": int(ninth_auto_remove["auto_remove_total_fuel_rows"]),
        "auto_remove_sector_fuel_suffix_rows_ninth": int(ninth_auto_remove["auto_remove_sector_fuel_suffix_rows"]),
        "auto_removed_new_rows_ninth": int(ninth_auto_remove["auto_removed_new_rows"]),
        "leap_combined_esto_rows": int(len(refreshed_esto)),
        "leap_combined_ninth_rows": int(len(refreshed_ninth)),
    }


#%%
RUN_WORKFLOW = True
# Set to False to emit a warning instead of raising an error when coverage gaps are found.
ERROR_ON_GAPS = False

WORKFLOW_RESULT: dict[str, object] | None = None
if RUN_WORKFLOW:
    WORKFLOW_RESULT = run_workflow(error_on_gaps=ERROR_ON_GAPS)
    print("[OK] Mapping maintenance columns refreshed.")
    for key, value in WORKFLOW_RESULT.items():
        print(f"- {key}: {value}")
        if key == "trio_presence_csv":
            print(f"- trio_presence_note: {TRIO_PRESENCE_OUTPUT_NOTE}")
#%%
