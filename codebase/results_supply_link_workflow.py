#%%
"""
Link LEAP balance-demand results, transformation outputs, and supply trade branches.

This workflow uses LEAP balance exports plus transformation outputs to derive
supply imports, exports, and production targets that keep the model balanced. It
is the integrated supply path to use when demand/transformation results should
drive supply trade updates rather than running the standalone supply workflow
alone.
"""

from __future__ import annotations

from functools import lru_cache
from datetime import datetime, timezone
import json
import os
import re
import sys
import copy
import shutil
import time
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from codebase.configuration import workflow_config as workflow_cfg
from codebase.utilities.master_config import (
    MASTER_CONFIG_PATH,
    config_table_exists,
    read_config_table,
)
from codebase.configuration.all_products_and_flows import ESTO_PRODUCT_LIST, ESTO_SECTORS
from codebase.mappings.canonical_mapping import (
    DEFAULT_BACKUP_LEAP_MAPPINGS,
    DEFAULT_CODEBOOK,
    DEFAULT_NINTH_TO_ESTO,
    DEFAULT_SHEET_MAP,
    build_sector_to_esto_flow_lookup,
    load_canonical_pairs,
    load_fuel_aliases,
    load_sheet_map,
)

from codebase.functions import supply_data_pipeline, leap_api
from codebase.functions.analysis_input_write_dispatcher import (
    get_analysis_input_write_mode,
)
from codebase import leap_results_workflow, transformation_workflow, transfers_workflow
from codebase.utilities.leap_results_dashboard_balance import (
    DEFAULT_BACKUP_MAPPINGS_PATH as DEFAULT_BALANCE_BACKUP_MAPPINGS_PATH,
    DEFAULT_BASE_TABLE_PATH as DEFAULT_BALANCE_BASE_TABLE_PATH,
    DEFAULT_CODEBOOK_PATH as DEFAULT_BALANCE_CODEBOOK_PATH,
    DEFAULT_EXPLICIT_MAPPINGS_PATH as DEFAULT_BALANCE_EXPLICIT_MAPPINGS_PATH,
    DEFAULT_EXPLICIT_REASSIGNMENTS_PATH as DEFAULT_BALANCE_EXPLICIT_REASSIGNMENTS_PATH,
    DEFAULT_MAPPING_PAIRS_PATH as DEFAULT_BALANCE_MAPPING_PAIRS_PATH,
    DEFAULT_PROJECTION_TABLE_PATH as DEFAULT_BALANCE_PROJECTION_TABLE_PATH,
    DEFAULT_REF_WORKBOOK_PATH as DEFAULT_BALANCE_REF_WORKBOOK_PATH,
    DEFAULT_SHEET_MAP_PATH as DEFAULT_BALANCE_SHEET_MAP_PATH,
    DEFAULT_SYNTHETIC_REFERENCE_ROWS_PATH as DEFAULT_BALANCE_SYNTHETIC_REFERENCE_ROWS_PATH,
    DEFAULT_TGT_WORKBOOK_PATH as DEFAULT_BALANCE_TGT_WORKBOOK_PATH,
    build_balance_comparison_esto_axis,
    build_esto_axis_structure_from_dashboard_template,
    convert_leap_balances_to_esto_long_table,
)
from codebase.utilities.leap_results_dashboard_utils import (
    DEFAULT_EXPLICIT_LEAP_MAPPINGS,
    DEFAULT_EXPLICIT_LEAP_REASSIGNMENTS,
    apply_explicit_sector_reassignments,
    build_comparisons,
    load_explicit_sector_fuel_mappings,
    load_explicit_sector_reassignments,
    load_leap_workbook,
)
from codebase.scrapbook.utilities import load_augmented_reference_tables
from codebase.utilities.workflow_common import archive_config_dir_once_per_day
from codebase.utilities import workflow_common
from codebase.utilities.output_paths import BALANCE_TABLES_ROOT, INTEGRATED_LEAP_EXPORTS_ROOT


def _resolve(path: Path | str) -> Path:
    """Resolve a possibly relative path against the repo root."""
    raw = str(path).replace("\\", "/")
    candidate = Path(raw)
    return candidate if candidate.is_absolute() else (REPO_ROOT / candidate)


def _emit_completion_beep(*, success: bool = True, style: str = "simple") -> None:
    """Emit an audible completion signal (winsound, notebook audio, terminal bell)."""
    if not bool(ENABLE_COMPLETION_BEEP):
        return
    count = max(int(COMPLETION_BEEP_COUNT), 1)
    frequency = max(int(COMPLETION_BEEP_FREQUENCY_HZ), 37)
    duration = max(int(COMPLETION_BEEP_DURATION_MS), 50)
    pause_seconds = max(float(COMPLETION_BEEP_PAUSE_SECONDS), 0.0)
    if not success:
        count = max(count, 2)
        frequency = max(frequency - 180, 37)
        if style == "chime":
            style = "error"

    if style == "chime":
        tone_plan = [(659, 90), (784, 90), (988, 140)]  # E5, G5, B5
        gap_ms = 40
    elif style == "error":
        tone_plan = [(440, 140), (330, 180)]  # A4 -> E4 (descending)
        gap_ms = 60
    else:
        tone_plan = [(frequency, duration)] * count
        gap_ms = int(pause_seconds * 1000)

    try:
        import winsound  # type: ignore

        for index, (freq_hz, tone_duration_ms) in enumerate(tone_plan):
            try:
                winsound.Beep(max(int(freq_hz), 37), max(int(tone_duration_ms), 50))
            except Exception:
                winsound.MessageBeep()
            if gap_ms > 0 and index < len(tone_plan) - 1:
                time.sleep(gap_ms / 1000.0)
        return
    except Exception:
        pass

    # Jupyter kernels often ignore terminal bells; use browser audio when possible.
    try:
        from IPython import get_ipython  # type: ignore
        from IPython.display import Javascript, display  # type: ignore

        ip = get_ipython()
        shell_name = type(ip).__name__ if ip is not None else ""
        if shell_name == "ZMQInteractiveShell":
            tones_js = ", ".join(
                f"{{freq: {max(int(freq_hz), 37)}, durMs: {max(int(tone_duration_ms), 50)}}}"
                for freq_hz, tone_duration_ms in tone_plan
            )
            js = f"""
            (() => {{
              const AudioCtx = window.AudioContext || window.webkitAudioContext;
              if (!AudioCtx) return;
              const tones = [{tones_js}];
              const gapMs = {int(gap_ms)};
              const playOne = (delayMs, freq, durMs) => {{
                setTimeout(() => {{
                  const ctx = new AudioCtx();
                  const osc = ctx.createOscillator();
                  const gain = ctx.createGain();
                  osc.type = "sine";
                  osc.frequency.value = freq;
                  gain.gain.value = 0.045;
                  osc.connect(gain);
                  gain.connect(ctx.destination);
                  osc.start();
                  osc.stop(ctx.currentTime + (durMs / 1000));
                  osc.onended = () => ctx.close();
                }}, delayMs);
              }};
              let cursor = 0;
              for (const tone of tones) {{
                playOne(cursor, tone.freq, tone.durMs);
                cursor += tone.durMs + gapMs;
              }}
            }})();
            """
            display(Javascript(js))
            return
    except Exception:
        pass

    for index, _ in enumerate(tone_plan):
        print("\a", end="", flush=True)
        if gap_ms > 0 and index < len(tone_plan) - 1:
            time.sleep(gap_ms / 1000.0)
    print("", flush=True)


# -----------------------------------------------------------------------------
# Workflow configuration
# -----------------------------------------------------------------------------

# Scope settings that are applied from the bottom notebook runtime block.
EXPORT_DATASET_KEY = workflow_cfg.SUPPLY_EXPORT_DATASET_KEY  # "ninth" or "esto"

# Input/output locations.
RESULTS_DIR = REPO_ROOT / "outputs" / "leap_results_dashboard" / "USA"
COMPARISON_LONG_PATH = RESULTS_DIR / "comparison_long.csv"  # demand comparison input
MAPPING_STATUS_PATH = RESULTS_DIR / "mapping_status.xlsx"  # demand mapping input
OUTPUT_DIR = INTEGRATED_LEAP_EXPORTS_ROOT  # workflow output root
RECONCILIATION_FILENAME = "results_supply_reconciliation.csv"  # core merged output
YEARLY_BALANCE_DIR = BALANCE_TABLES_ROOT / "results_supply_link" / "yearly_balance_tables"  # simple balance tables
CONVENTIONAL_BALANCE_DIR = (
    BALANCE_TABLES_ROOT / "results_supply_link" / "conventional_balance_tables"
)  # conventional format tables
EXPORT_OUTPUT_DIR = OUTPUT_DIR / "workbooks"  # supply+transformation+transfer LEAP files
EXPORT_FILENAME_TEMPLATE = supply_data_pipeline.EXPORT_FILENAME_TEMPLATE
TRANSFORMATION_EXPORT_OUTPUT_DIR = EXPORT_OUTPUT_DIR
TRANSFORMATION_EXPORT_FILENAME_TEMPLATE = transformation_workflow.core.EXPORT_FILENAME_TEMPLATE
COMBINED_EXPORT_FILENAME_TEMPLATE = "combined_supply_transformation_leap_imports_{economy}_{scenario}.xlsx"

# LEAP results workbook discovery and refinery fallback settings.
LEAP_RESULTS_TABLES_DIR = REPO_ROOT / "data" / "leap results tables"
REFINERY_RESULTS_FILENAME_TEMPLATE = "transformation_and_supply_results_{economy}_{scenario}.xlsx"
TRANSFORMATION_RESULTS_FILENAME_TEMPLATE = "transformation_results_{economy}_{scenario}.xlsx"
REFINERY_RESULTS_SHEET_NAME = "refining output"
REFINERY_SECTOR_NAME = "Oil refineries"
REFINERY_FUEL_LABEL_ALIASES = {
    "Gas and diesel oil": "Gas/diesel oil",
}

# Demand and year controls.
DEMAND_SOURCE_PRIORITY = ("leap", "projection")
BASE_YEAR = supply_data_pipeline.EXPORT_BASE_YEAR
LEAP_IMPORT_MAX_YEAR = 2060
FINAL_YEAR = min(int(supply_data_pipeline.EXPORT_FINAL_YEAR), LEAP_IMPORT_MAX_YEAR)  # LEAP-safe horizon
BALANCE_EXPORT_YEARS = [BASE_YEAR, 2030, 2050]

# Optional external cap templates (LEAP-format workbooks). Leave empty to disable.
CONSTRAINT_TEMPLATE_PATHS: list[Path | str] = []
CONSTRAINT_TEMPLATE_SHEETS: list[str] | None = None

# Demand table shaping controls.
DROP_PARENT_DEMAND_ROWS_WHEN_CHILDREN_PRESENT = True
INCLUDE_TOP_LEVEL_DEMAND_CATEGORY_ROWS = True
DROP_DISAGGREGATED_DEMAND_SECTORS = True

# LEAP import controls.
LEAP_IMPORT_SCENARIOS: list[str] | None = None
LEAP_IMPORT_REGION = supply_data_pipeline.EXPORT_REGION
LEAP_IMPORT_CREATE_BRANCHES = True
LEAP_IMPORT_FILL_BRANCHES = True
LEAP_IMPORT_INCLUDE_CURRENT_ACCOUNTS = False
LEAP_IMPORT_SUPPLY_TO_LEAP = True
LEAP_IMPORT_TRANSFORMATION_TO_LEAP = True
LEAP_IMPORT_TRANSFERS_TO_LEAP = True
LEAP_IMPORT_LOG_LEVEL = "summary"  # detailed|summary|quiet
LEAP_IMPORT_WARNING_PRINT_LIMIT = 20

# Results packaging controls.
# - When True, workflow writes one consolidated run workbook in OUTPUT_DIR.
# - When False, legacy per-file outputs (csv/xlsx sidecars) are used.
RESULTS_SINGLE_FILE_NAME = "results_supply_link_run.xlsx"
RESULTS_SINGLE_FILE_ARCHIVE_DIR = OUTPUT_DIR / "supporting_files" / "archive"
RESULTS_CHECKS_DIR = OUTPUT_DIR / "supporting_files" / "checks"
RESULTS_RUNTIME_DIR = OUTPUT_DIR / "supporting_files" / "runtime"
RESULTS_SINGLE_FILE_ARCHIVE_EVERY_RUN = True
ENABLE_WORKFLOW_TIMING = True
WRITE_WORKFLOW_TIMING_CSV = True
WORKFLOW_TIMING_FILENAME = "workflow_stage_timings.csv"
RESULTS_UNMATCHED_ID_REPORT_FILENAME = "results_supply_link_unmatched_id_rows.csv"
RESULTS_METADATA_MISMATCH_REPORT_FILENAME = "results_supply_link_metadata_mismatches.csv"
RESULTS_CONFIG_MAPPING_MISMATCH_REPORT_FILENAME = "results_supply_link_config_mapping_mismatches.csv"
RESULTS_BALANCE_DEMAND_ISSUES_FILENAME = "results_supply_link_balance_demand_issues.csv"
RESULTS_BALANCE_MATCHING_DIAGNOSTICS_FILENAME = "results_supply_link_balance_matching_diagnostics.csv"
RESULTS_DROPPED_UNMATCHED_ZERO_SUPPLY_ROWS_FILENAME = (
    "results_supply_link_dropped_unmatched_zero_supply_rows.csv"
)
# Runtime-toggle values for archive cadence and legacy sidecars are set in the
# bottom "Notebook Runtime Variables" block.

# Optional live LEAP probe to keep fuel branch catalogs current.
LEAP_FUEL_BRANCH_PROBE_OUTPUT_PATH = (
    RESULTS_CHECKS_DIR / "transformation_supply_fuel_branch_catalog_probe.csv"
)
USE_RESULTS_VERIFICATION_EXPORT_SOURCE = True
RESULTS_VERIFICATION_EXPORT_PATH = REPO_ROOT / "data" / "full model export.xlsx"
RESULTS_VERIFICATION_EXPORT_SHEET = "Export"

# Backward-compatible aliases used by existing catalog helpers.
USE_FULL_MODEL_EXPORT_CATALOG_SOURCE = USE_RESULTS_VERIFICATION_EXPORT_SOURCE
FULL_MODEL_EXPORT_CATALOG_PATH = RESULTS_VERIFICATION_EXPORT_PATH
FULL_MODEL_EXPORT_CATALOG_SHEET = RESULTS_VERIFICATION_EXPORT_SHEET

# Transformation refresh from live LEAP Results (before reconciliation).
REFRESH_TRANSFORMATION_MEASURES_FROM_LEAP_RESULTS = False
REFRESH_TRANSFORMATION_MEASURE_SCENARIO = "Reference"
REFRESH_TRANSFORMATION_MEASURE_REGION = LEAP_IMPORT_REGION

# Trade target behavior.
# Runtime toggle is set in the bottom "Notebook Runtime Variables" block.
# Modes:
# - "legacy_split": current behaviour. Supply gets residual imports/exports and
#   transformation output fuels get import/export targets.
# - "output_share_supply_exports": trial mode. Supply imports are left at zero
#   so LEAP can autobalance them, all explicit exports stay on supply branches,
#   and transformation output-fuel import/export targets are omitted so
#   transformation activity is driven by Output Share / Process Share instead.
# - "capacity_constrained": transformation import/export targets are explicitly
#   written as zeros (to clear stale values) and process Exogenous Capacity is
#   set from projected process outputs to constrain overproduction in LEAP.
# - "capacity_unmet_iterative": workbook/manual mode only. Uses latest supply
#   results tables to proxy unmet imports, then iteratively uplifts process
#   exogenous capacity with persisted state across manual LEAP recalc passes.
# - "capacity_unmet_iterative_balanced": extends iterative mode using imports
#   deltas as unmet proxy. Positive import gaps uplift transformation and
#   primary production; negative import gaps are routed to extra exports.
DEMAND_SECTOR_PREFIXES = ("04_", "05_", "14_", "15_", "16_")

# Capacity-constrained mode knobs.
CAPACITY_CONSTRAINT_FACTOR = 1.0
CAPACITY_CONSTRAINT_UNITS = "Petajoule/Year"
CAPACITY_MAX_AVAILABILITY = 100.0
CAPACITY_CREDIT = 100.0
CAPACITY_ENDOGENOUS = 0.0
CAPACITY_CLEAR_OUTPUT_TRADE_TARGETS = True

# Capacity unmet iterative mode knobs.
# The iterative pass now prefers balance-table outputs for observed trade
# instead of legacy LEAP results workbooks.
CAPACITY_UNMET_STATE_PATH = RESULTS_RUNTIME_DIR / "capacity_unmet_iterative_state.json"
CAPACITY_UNMET_RESULTS_DIR = YEARLY_BALANCE_DIR
CAPACITY_UNMET_IMPORT_SHEETS: tuple[str, ...] = ("imports primary", "imports secondary")
CAPACITY_UNMET_EXPORT_SHEETS: tuple[str, ...] = ("exports primary", "exports secondary")
CAPACITY_UNMET_PRIORITY_BY_PRODUCT: dict[str, list[str]] = {
    "17 Electricity": [
        "Electricity generation",
        "Main activity producer CHP plants",
        "Autoproducer CHP plants",
    ],
    "18 Heat": [
        "Main activity producer CHP plants",
        "Autoproducer CHP plants",
        "Heat plants",
    ],
    "16.04 Biogas": [
        "Biogas production",
        "Biogas processing",
    ],
    "07.06 Kerosene type jet fuel": [
        "Oil refineries",
    ],
    "07.07 Gas/diesel oil": [
        "Oil refineries",
    ],
}
CAPACITY_UNMET_ALLOW_SAME_RESULTS_REUSE = False
CAPACITY_UNMET_FIRST_CLEAN_ARCHIVE_EXISTING_STATE = True
CAPACITY_UNMET_UNRESOLVED_POSITIVE_POLICY = "imports_fallback"  # fail|imports_fallback|track_only
CAPACITY_UNMET_PIN_EXPORTS_TO_9TH_PROJECTIONS = True
CAPACITY_UNMET_UNRESOLVED_POSITIVE_ALLOWLIST: set[str] = {
    "02.02 Gas coke",
}
CAPACITY_UNMET_MODULE_CAPACITY_UPPER_LIMITS: dict[str, dict[str, dict[str, float]]] = {
    "20_USA": {
        "reference": {
            # Base-year (2022) module output caps for requested sectors.
            # Values here are 2022 process-record output values (not multi-year sums).
            # Units must match CAPACITY_CONSTRAINT_UNITS.
            "Blast furnaces": 78.96,
            "Upstream liquids transfers": 0.0,
            "Refinery and blending transfers": 0.0,
            "Transfers unallocated": 0.0,
            "Liquefaction coal to oil": 0.0,
            "Charcoal processing": 0.0,
            "BKB and PB plants": 0.0,
            "Non-specified transformation": 211.367768,
            "Coke ovens": 298.348638,
            "Patent fuel plants": 0.0,
            "Natural gas blending plants": 46.94349,
            "Gas works plants": 46.943101,
        },
        "target": {
            "Blast furnaces": 78.96,
            "Upstream liquids transfers": 0.0,
            "Refinery and blending transfers": 0.0,
            "Transfers unallocated": 0.0,
            "Liquefaction coal to oil": 0.0,
            "Charcoal processing": 0.0,
            "BKB and PB plants": 0.0,
            "Non-specified transformation": 211.367768,
            "Coke ovens": 298.348638,
            "Patent fuel plants": 0.0,
            "Natural gas blending plants": 46.94349,
            "Gas works plants": 46.943101,
        },
    },
}
CAPACITY_UNMET_PRODUCTION_UPPER_LIMITS: dict[str, dict[str, dict[str, float]]] = {
    # Optional product-level production caps for balanced iterative mode.
    # Shape: economy -> scenario -> esto_product -> max production value.
    # Example:
    # "20_USA": {
    #     "reference": {
    #         "01.01 Hard coal": 123.0,
    #     },
    # }
}

_CAPACITY_UNMET_RUNTIME_CAPACITY_ADDITIONS: dict[str, float] = {}
_CAPACITY_UNMET_RUNTIME_PRIMARY_ADDITIONS: dict[str, float] = {}
_CAPACITY_UNMET_RUNTIME_EXPORT_ADJUSTMENTS: dict[str, float] = {}
_CAPACITY_UNMET_RUNTIME_PASS_SUMMARY: dict[str, object] | None = None

# Optional hard reset of supply/transformation import-export values to zero.
# Runtime toggle is set in the bottom "Notebook Runtime Variables" block.
# Optional scope filters for reset helper. Use None for category defaults.
RESET_SCOPE_ECONOMIES: list[str] | None = None
RESET_SCOPE_SCENARIOS: list[str] | None = None
RESET_SCOPE_SECTOR_TITLES: list[str] | None = None
RESET_SCOPE_ESTO_PRODUCTS: list[str] | None = None
RESET_SCOPE_YEARS: list[int] | None = None


def _format_scope_preview(
    values: Iterable[object] | None,
    *,
    default_label: str,
    limit: int = 6,
) -> str:
    """Return a compact preview string for scope filters."""
    if not values:
        return default_label
    normalized = [str(item).strip() for item in values if str(item or "").strip()]
    if not normalized:
        return default_label
    if len(normalized) <= limit:
        return ", ".join(normalized)
    head = ", ".join(normalized[:limit])
    return f"{head}, ... (+{len(normalized) - limit} more)"


def _print_reset_mode_reminder(
    *,
    run_economies: Iterable[str],
    run_scenarios: Iterable[str],
) -> None:
    """Warn users when reset mode is off; summarize active scope when on."""
    run_economy_preview = _format_scope_preview(
        run_economies,
        default_label="all run economies",
    )
    run_scenario_preview = _format_scope_preview(
        run_scenarios,
        default_label="all run scenarios",
    )
    if RUN_RESET_SUPPLY_AND_TRANSFORMATION_IMPORT_EXPORT:
        economy_preview = _format_scope_preview(
            RESET_SCOPE_ECONOMIES,
            default_label=run_economy_preview,
        )
        scenario_preview = _format_scope_preview(
            RESET_SCOPE_SCENARIOS,
            default_label=run_scenario_preview,
        )
        year_preview = _format_scope_preview(
            RESET_SCOPE_YEARS,
            default_label=f"{BASE_YEAR}-{FINAL_YEAR}",
        )
        print(
            "[INFO] Reset reminder: supply/transformation import-export reset is ENABLED "
            f"for economies [{economy_preview}], scenarios [{scenario_preview}], years [{year_preview}]."
        )
        return
    print(
        "[WARN] Reset reminder: supply/transformation import-export reset is DISABLED. "
        "Stale LEAP Imports/Exports/targets may persist across runs. "
        "Set RUN_RESET_SUPPLY_AND_TRANSFORMATION_IMPORT_EXPORT=True "
        "to force zero reset before filling."
    )

# Reset scope configuration.
# Prefer deriving reset module/fuel scope from the canonical LEAP export workbook.
RESET_SCOPE_USE_FULL_MODEL_EXPORT = True
RESET_SCOPE_REQUIRE_FULL_MODEL_EXPORT = False
# Optional manual additions on top of derived (or fallback) scope.
TRANSFORMATION_RESET_MODULES_MANUAL_OVERRIDES: dict[str, list[str]] = {}
TRANSFORMATION_RESET_FUELS_MANUAL_OVERRIDES: dict[str, list[str]] = {}

# Legacy fallback reset catalogs (used when workbook-derived scope is unavailable).
TRANSFORMATION_RESET_MODULES: dict[str, list[str]] = {
    # Legacy fallback list for transformation modules.
    "all": [
        "Upstream liquids transfers",
        "Refinery and blending transfers",
        "NG Liquefaction",
        "LNG gasification",
        "Gas works plants",
        "Natural gas blending plants",
        "Coke ovens",
        "Blast furnaces",
        "Patent fuel plants",
        "BKB and PB plants",
        "Liquefaction coal to oil",
        # "Electric boilers",
        # "Chemical heat for electricity production",
        # "Petrochemical industry",
        # "Biofuels processing",
        # "Coal mines",
        "Charcoal processing",
        "Non specified transformation",
        "Hydrogen transformation",
        "Transfers unallocated",
        
    ],
}

TRANSFORMATION_RESET_FUELS = {'all': ['Coal',
        'Coking coal',
        'Other bituminous coal',
        'Sub bituminous coal',
        'Anthracite',
        'Lignite',
        'Coal nonspecified',
        'Coal products',
        'Coke oven coke',
        'Gas coke',
        'Coke oven gas',
        'Blast furnace gas',
        'Other recovered gases',
        'Patent fuel',
        'Coal tar',
        'BKB and PB',
        'Peat',
        'Peat products',
        'Oil shale and oil sands',
        'Crude oil and NGL',
        'Crude oil',
        'Natural gas liquids',
        'Refinery feedstocks',
        'Additives and oxygenates',
        'Other hydrocarbons',
        'Petroleum products',
        'Motor gasoline',
        'Aviation gasoline',
        'Naphtha',
        'Gasoline type jet fuel',
        'Kerosene type jet fuel',
        'Kerosene',
        'Gas and diesel oil',
        'Fuel oil',
        'LPG',
        'Refinery gas not liquefied',
        'Ethane',
        'White spirit SBP',
        'Lubricants',
        'Bitumen',
        'Paraffin waxes',
        'Petroleum coke',
        'Other products',
        'PetProd nonspecified',
        'Gas',
        'Natural gas',
        'LNG',
        'Gas works gas',
        'Gas nonspecified',
        'Nuclear',
        'Hydro',
        'Geothermal',
        'Solar',
        'of which Photovoltaics',
        'Solar nonspecified',
        'Tide wave ocean',
        'Wind',
        'Solid biomass',
        'Fuelwood and woodwaste',
        'Bagasse',
        'Charcoal',
        'Black liqour',
        'Other biomass',
        'Others',
        'Biogas',
        'Industrial waste',
        'Municipal solid waste renewable',
        'Municipal solid waste non renewable',
        'Biogasoline',
        'Biodiesel',
        'Bio jet kerosene',
        'Other liquid biofuels',
        'Other sources',
        'Electricity',]}

_RESET_SCOPE_FROM_EXPORT_CACHE: dict[str, object] | None = None

# Demand source strategy.
# Runtime toggles are set in the bottom "Notebook Runtime Variables" block.

# Demand mapping/reference inputs.
DIRECT_DEMAND_SHEET_MAP_PATH = DEFAULT_SHEET_MAP
DIRECT_DEMAND_MAPPING_WORKBOOK = REPO_ROOT / "config" / "leap_mappings.xlsx"
DIRECT_DEMAND_ESTO_MAPPING_SHEET = "leap_combined_esto"
DIRECT_DEMAND_NINTH_MAPPING_SHEET = "leap_combined_ninth"
DIRECT_DEMAND_BASE_TABLE_PATH = REPO_ROOT / "data/00APEC_2025_low_with_subtotals.csv"
DIRECT_DEMAND_PROJECTION_TABLE_PATH = REPO_ROOT / "data/merged_file_energy_ALL_20251106.csv"
DIRECT_DEMAND_REFERENCE_CACHE_DIR = REPO_ROOT / "data/.cache/results_supply_link_reference_tables"
DIRECT_DEMAND_BASE_YEAR = 2022
DIRECT_DEMAND_PROJECTION_YEARS: tuple[int, ...] = tuple(range(2023, 2071))
DIRECT_DEMAND_BASE_ECONOMY = "20USA"
DIRECT_DEMAND_PROJECTION_ECONOMY = "20_USA"
DIRECT_DEMAND_SCENARIO_MAP = {"reference": "reference", "target": "target"}
DIRECT_DEMAND_USE_ESTO_AGG_ONLY = False
DIRECT_DEMAND_SIBLING_COMPARATOR_MODE = "aggregate_to_parent"
DIRECT_DEMAND_INCLUDE_SIBLING_PARENT_TOTALS = True

BALANCE_DEMAND_REF_WORKBOOK_PATH = DEFAULT_BALANCE_REF_WORKBOOK_PATH
BALANCE_DEMAND_TGT_WORKBOOK_PATH = DEFAULT_BALANCE_TGT_WORKBOOK_PATH
BALANCE_DEMAND_LEAP_TO_ESTO_MAPPING_WORKBOOK = DIRECT_DEMAND_MAPPING_WORKBOOK
BALANCE_DEMAND_NINTH_TO_ESTO_MAPPING: tuple[Path, str] = (
    _resolve(DEFAULT_BALANCE_MAPPING_PAIRS_PATH),
    "ninth_pairs_to_esto_pairs",
)
BALANCE_DEMAND_CODEBOOK_PATH = _resolve(DEFAULT_BALANCE_CODEBOOK_PATH)
BALANCE_DEMAND_SHEET_MAP_PATH = _resolve(DEFAULT_BALANCE_SHEET_MAP_PATH)
BALANCE_DEMAND_BACKUP_MAPPINGS_PATH = _resolve(DEFAULT_BALANCE_BACKUP_MAPPINGS_PATH)
BALANCE_DEMAND_EXPLICIT_MAPPINGS_PATH = _resolve(DEFAULT_BALANCE_EXPLICIT_MAPPINGS_PATH)
BALANCE_DEMAND_EXPLICIT_REASSIGNMENTS_PATH = _resolve(DEFAULT_BALANCE_EXPLICIT_REASSIGNMENTS_PATH)
BALANCE_DEMAND_SYNTHETIC_REFERENCE_ROWS_PATH = _resolve(DEFAULT_BALANCE_SYNTHETIC_REFERENCE_ROWS_PATH)
BALANCE_DEMAND_BASE_TABLE_PATH = _resolve(DEFAULT_BALANCE_BASE_TABLE_PATH)
BALANCE_DEMAND_PROJECTION_TABLE_PATH = _resolve(DEFAULT_BALANCE_PROJECTION_TABLE_PATH)
BALANCE_DEMAND_CHART_NAVIGATION_GUIDE_PATH = REPO_ROOT / "config" / "leap_comparison_dashboard_template.json"
BALANCE_DEMAND_KNOWN_ISSUES_CONFIG_PATH = REPO_ROOT / "config" / "leap_results_balance_known_issues.json"
BALANCE_DEMAND_TEMPLATE_SHEET = "EBal|2060"


def _use_legacy_trade_split_mode() -> bool:
    """Return True when exports should use the legacy split-target behavior."""
    return str(TRADE_TARGET_EXPORT_MODE).strip().lower() == "legacy_split"


def _use_output_share_supply_exports_mode() -> bool:
    """Return True when supply exports should carry explicit trade values and imports stay zero."""
    return str(TRADE_TARGET_EXPORT_MODE).strip().lower() == "output_share_supply_exports"


def _use_capacity_unmet_iterative_mode() -> bool:
    """Return True when capacity is manually uplifted using iterative unmet-import passes."""
    return str(TRADE_TARGET_EXPORT_MODE).strip().lower() == "capacity_unmet_iterative"


def _use_capacity_unmet_iterative_balanced_mode() -> bool:
    """Return True when iterative mode handles both positive and negative net-trade residuals."""
    return str(TRADE_TARGET_EXPORT_MODE).strip().lower() == "capacity_unmet_iterative_balanced"


def _use_capacity_unmet_iterative_any_mode() -> bool:
    """Return True for any iterative unmet-capacity mode."""
    return _use_capacity_unmet_iterative_mode() or _use_capacity_unmet_iterative_balanced_mode()


def _use_capacity_constrained_mode() -> bool:
    """Return True when exports should set process capacities and clear trade targets."""
    return str(TRADE_TARGET_EXPORT_MODE).strip().lower() == "capacity_constrained"


def _use_capacity_like_mode() -> bool:
    """Return True when transformation exports should write capacity variables."""
    return _use_capacity_constrained_mode() or _use_capacity_unmet_iterative_any_mode()


def _flatten_reset_scope_values(values_by_group: dict[str, list[str]] | None) -> list[str]:
    """Flatten grouped reset-scope values to a de-duplicated ordered list."""
    ordered: list[str] = []
    seen: set[str] = set()
    for values in (values_by_group or {}).values():
        for item in (values or []):
            token = str(item or "").strip()
            if not token:
                continue
            key = token.lower()
            if key in seen:
                continue
            seen.add(key)
            ordered.append(token)
    return ordered


def _load_reset_scope_from_full_model_export() -> tuple[list[str], list[str]]:
    """Return transformation module/fuel reset scope derived from full-model export."""
    global _RESET_SCOPE_FROM_EXPORT_CACHE
    if isinstance(_RESET_SCOPE_FROM_EXPORT_CACHE, dict):
        return (
            list(_RESET_SCOPE_FROM_EXPORT_CACHE.get("modules") or []),
            list(_RESET_SCOPE_FROM_EXPORT_CACHE.get("fuels") or []),
        )

    if not RESET_SCOPE_USE_FULL_MODEL_EXPORT:
        _RESET_SCOPE_FROM_EXPORT_CACHE = {
            "modules": [],
            "fuels": [],
            "module_output_fuels": {},
        }
        return [], []

    try:
        rows = _extract_catalog_rows_from_full_model_export(
            source_path=RESULTS_VERIFICATION_EXPORT_PATH,
            sheet_name=RESULTS_VERIFICATION_EXPORT_SHEET,
        )
    except Exception as exc:
        print(
            "[WARN] Failed deriving reset scope from full model export: "
            f"{exc}"
        )
        rows = []

    modules: list[str] = []
    fuels: list[str] = []
    module_output_fuels: dict[str, list[str]] = {}
    seen_modules: set[str] = set()
    seen_fuels: set[str] = set()
    seen_module_output_fuels: dict[str, set[str]] = {}
    for row in rows:
        if str(row.get("catalog_type") or "").strip().lower() != "transformation":
            continue
        module = str(row.get("module_or_root") or "").strip()
        fuel = str(row.get("fuel_name") or "").strip()
        fuel_group = str(row.get("fuel_group") or "").strip().lower()
        if module:
            key = module.lower()
            if key not in seen_modules:
                seen_modules.add(key)
                modules.append(module)
        if fuel:
            key = fuel.lower()
            if key not in seen_fuels:
                seen_fuels.add(key)
                fuels.append(fuel)
        if module and fuel and fuel_group == "output fuels":
            canonical_fuel = _canonical_transformation_fuel_label(fuel)
            if canonical_fuel:
                module_key = module.lower()
                module_bucket = module_output_fuels.setdefault(module_key, [])
                seen_bucket = seen_module_output_fuels.setdefault(module_key, set())
                canonical_key = canonical_fuel.lower()
                if canonical_key not in seen_bucket:
                    seen_bucket.add(canonical_key)
                    module_bucket.append(canonical_fuel)

    if modules or fuels:
        module_scoped_count = sum(
            1 for labels in module_output_fuels.values() if labels
        )
        print(
            "[INFO] Derived reset scope from full model export: "
            f"modules={len(modules)}, fuels={len(fuels)}, "
            f"module_output_scopes={module_scoped_count} "
            f"(source={_resolve(RESULTS_VERIFICATION_EXPORT_PATH)})"
        )
    else:
        print(
            "[WARN] No transformation reset scope derived from full model export "
            f"(source={_resolve(RESULTS_VERIFICATION_EXPORT_PATH)})."
        )

    _RESET_SCOPE_FROM_EXPORT_CACHE = {
        "modules": modules,
        "fuels": fuels,
        "module_output_fuels": module_output_fuels,
    }
    return modules, fuels


def _configured_reset_module_names() -> set[str]:
    """Return normalized module names configured for reset operations."""
    legacy_modules = _flatten_reset_scope_values(TRANSFORMATION_RESET_MODULES)
    manual_modules = _flatten_reset_scope_values(TRANSFORMATION_RESET_MODULES_MANUAL_OVERRIDES)
    derived_modules, _ = _load_reset_scope_from_full_model_export()

    if RESET_SCOPE_USE_FULL_MODEL_EXPORT and derived_modules:
        base_modules = derived_modules
    elif RESET_SCOPE_USE_FULL_MODEL_EXPORT and RESET_SCOPE_REQUIRE_FULL_MODEL_EXPORT:
        raise ValueError(
            "Reset scope requires full model export derivation, but no module scope "
            f"was derived from {RESULTS_VERIFICATION_EXPORT_PATH} "
            f"(sheet={RESULTS_VERIFICATION_EXPORT_SHEET})."
        )
    else:
        base_modules = legacy_modules

    tokens = [
        str(item).strip()
        for item in [*base_modules, *manual_modules]
        if str(item or "").strip()
    ]
    return {token.lower() for token in tokens}


def _configured_reset_fuel_labels() -> list[str]:
    """Return unique configured reset fuel labels (preserve first-seen order)."""
    legacy_fuels = _flatten_reset_scope_values(TRANSFORMATION_RESET_FUELS)
    manual_fuels = _flatten_reset_scope_values(TRANSFORMATION_RESET_FUELS_MANUAL_OVERRIDES)
    _, derived_fuels = _load_reset_scope_from_full_model_export()

    if RESET_SCOPE_USE_FULL_MODEL_EXPORT and derived_fuels:
        base_fuels = derived_fuels
    elif RESET_SCOPE_USE_FULL_MODEL_EXPORT and RESET_SCOPE_REQUIRE_FULL_MODEL_EXPORT:
        raise ValueError(
            "Reset scope requires full model export derivation, but no fuel scope "
            f"was derived from {RESULTS_VERIFICATION_EXPORT_PATH} "
            f"(sheet={RESULTS_VERIFICATION_EXPORT_SHEET})."
        )
    else:
        base_fuels = legacy_fuels

    labels: list[str] = []
    seen: set[str] = set()
    for item in [*base_fuels, *manual_fuels]:
        token = str(item or "").strip()
        if not token:
            continue
        key = token.lower()
        if key in seen:
            continue
        seen.add(key)
        labels.append(token)
    return labels


def _configured_reset_output_fuel_labels_by_module(
    module_names: Iterable[str] | None = None,
) -> dict[str, list[str]]:
    """
    Return module-specific Output Fuels reset labels.

    Keys are lower-cased transformation module names.
    Values are canonicalized fuel labels in first-seen order.
    """
    requested_modules = {
        str(item or "").strip().lower()
        for item in (module_names or [])
        if str(item or "").strip()
    }
    mapping: dict[str, list[str]] = {}
    seen: dict[str, set[str]] = {}

    def _append(module_key: str, fuel_label: object) -> None:
        module_token = str(module_key or "").strip().lower()
        if not module_token:
            return
        canonical = _canonical_transformation_fuel_label(str(fuel_label or ""))
        if not canonical:
            return
        module_seen = seen.setdefault(module_token, set())
        canonical_key = canonical.lower()
        if canonical_key in module_seen:
            return
        module_seen.add(canonical_key)
        mapping.setdefault(module_token, []).append(canonical)

    _load_reset_scope_from_full_model_export()
    cached = _RESET_SCOPE_FROM_EXPORT_CACHE if isinstance(_RESET_SCOPE_FROM_EXPORT_CACHE, dict) else {}
    raw_module_map = cached.get("module_output_fuels") if isinstance(cached, dict) else {}
    if isinstance(raw_module_map, dict):
        for module_key, labels in raw_module_map.items():
            module_token = str(module_key or "").strip().lower()
            if not module_token:
                continue
            if requested_modules and module_token not in requested_modules:
                continue
            for label in (labels or []):
                _append(module_token, label)

    manual_overrides = TRANSFORMATION_RESET_FUELS_MANUAL_OVERRIDES or {}
    if not isinstance(manual_overrides, dict):
        return mapping
    for scope_key, labels in manual_overrides.items():
        scope_token = str(scope_key or "").strip().lower()
        if not scope_token:
            continue
        if scope_token == "all":
            target_modules = (
                set(requested_modules)
                if requested_modules
                else (
                    set(mapping.keys())
                    or set(_configured_reset_module_names())
                )
            )
        else:
            target_modules = {scope_token}
            if requested_modules and scope_token not in requested_modules:
                continue
        for module_token in sorted(target_modules):
            for label in (labels or []):
                _append(module_token, label)
    return mapping


def _canonical_transformation_fuel_label(label: str) -> str:
    """Return a stable display fuel label used by transformation LEAP branch paths."""
    token = str(label or "").strip()
    if not token:
        return ""
    try:
        mapped = transformation_workflow.core.map_code_label(
            token,
            transformation_workflow.core.code_to_name_mapping,
        )
    except Exception:
        mapped = token
    normalized = str(mapped or "").strip()
    if not normalized:
        return token
    try:
        # Match LEAP branch-name sanitization so aliases like "Gas/diesel oil" and
        # "Gas and diesel oil" collapse to one canonical key before export rows.
        sanitized = transformation_workflow.core.build_branch_path([normalized])
        sanitized_token = str(sanitized or "").strip()
        if sanitized_token:
            return sanitized_token
    except Exception:
        pass
    return normalized


def _load_code_to_name_table() -> pd.DataFrame:
    """Load the code-to-name workbook used across supply/transformation workflows."""
    for workbook_path in supply_data_pipeline.CODE_TO_NAME_PATHS:
        if not config_table_exists(workbook_path, sheet_name="code_to_name"):
            continue
        return read_config_table(workbook_path, sheet_name="code_to_name", dtype=str).fillna("")
    raise FileNotFoundError(
        "No code-to-name table found via the configured legacy workbook paths or "
        f"{MASTER_CONFIG_PATH} for {supply_data_pipeline.CODE_TO_NAME_PATHS}"
    )


def _normalize_label_for_lookup(value: object) -> str:
    """Normalize fuel/sector labels for tolerant crosswalk matching."""
    text = str(value or "").strip().lower()
    if not text:
        return ""
    text = text.replace("&", " and ")
    text = text.replace("/", " and ")
    text = text.replace("-", " ")
    text = text.replace("(", " ")
    text = text.replace(")", " ")
    text = text.replace(":", " ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def _normalize_esto_product_for_match(value: object) -> str:
    """Normalize esto_product strings, stripping leading numeric code prefixes."""
    text = str(value or "").strip()
    if not text:
        return ""
    # Convert "07.07 Gas/diesel oil" -> "Gas/diesel oil" before fuzzy normalization.
    text = re.sub(r"^\d+(?:\.\d+)*\s+", "", text)
    return _normalize_label_for_lookup(text)


def _build_label_to_esto_product_lookup() -> dict[str, str]:
    """Map human-readable labels and known codes back to ESTO products."""
    table = _load_code_to_name_table()
    lookup: dict[str, str] = {}
    for _, row in table.iterrows():
        esto_product = str(row.get("esto_label") or "").strip()
        if not esto_product:
            continue
        keys = [
            row.get("name"),
            row.get("esto_label"),
            row.get("9th_label"),
            row.get("code"),
        ]
        for key in keys:
            normalized = str(key or "").strip()
            if normalized:
                lookup.setdefault(normalized, esto_product)
                lookup.setdefault(normalized.lower(), esto_product)
                fuzzy_key = _normalize_label_for_lookup(normalized)
                if fuzzy_key:
                    lookup.setdefault(fuzzy_key, esto_product)
    return lookup


def _state_token(value: object) -> str:
    """Normalize a state key token for case-insensitive comparisons."""
    return str(value or "").strip().lower()


def _capacity_addition_state_key(
    economy: str,
    scenario: str,
    module: str,
    process: str,
    instance: int,
    year: int,
) -> str:
    """Build state key for cumulative process-level capacity additions."""
    return "|".join(
        [
            _state_token(economy),
            _state_token(scenario),
            _state_token(module),
            _state_token(process),
            str(int(instance)),
            str(int(year)),
        ]
    )


def _output_addition_state_key(
    economy: str,
    scenario: str,
    esto_product: str,
    year: int,
) -> str:
    """Build state key for cumulative output additions by product/year."""
    return "|".join(
        [
            _state_token(economy),
            _state_token(scenario),
            _state_token(esto_product),
            str(int(year)),
        ]
    )


def _results_signature_state_key(economy: str, scenario: str) -> str:
    """Build state key for last processed results signatures."""
    return "|".join([_state_token(economy), _state_token(scenario)])


def _capacity_unmet_default_state() -> dict[str, object]:
    """Return empty state payload for iterative unmet-capacity runs."""
    return {
        "version": 1,
        "cumulative_capacity_additions": {},
        "cumulative_output_additions": {},
        "cumulative_primary_additions": {},
        "cumulative_export_adjustments": {},
        "last_results_signatures": {},
        "passes": [],
    }


def _resolve_capacity_unmet_iteration_run_mode() -> str:
    """Return validated run mode for iterative unmet-capacity passes."""
    mode = str(CAPACITY_UNMET_ITERATION_RUN_MODE or "").strip().lower() or "consecutive"
    valid = {"consecutive", "first_clean"}
    if mode not in valid:
        raise ValueError(
            "Invalid CAPACITY_UNMET_ITERATION_RUN_MODE="
            f"{CAPACITY_UNMET_ITERATION_RUN_MODE!r}. Valid values: {sorted(valid)}"
        )
    return mode


def _is_capacity_unmet_first_clean_run_mode() -> bool:
    """Return True when iterative unmet workflow should run baseline-only first pass."""
    return _resolve_capacity_unmet_iteration_run_mode() == "first_clean"


def _read_capacity_unmet_state(
    state_path: Path | str = CAPACITY_UNMET_STATE_PATH,
    *,
    run_mode: str | None = None,
) -> dict[str, object]:
    """Load iterative capacity state JSON from disk (or reset for first_clean mode)."""
    path = _resolve(state_path)
    mode = str(run_mode or _resolve_capacity_unmet_iteration_run_mode()).strip().lower()
    default_state = _capacity_unmet_default_state()
    if mode == "first_clean":
        if path.exists() and bool(CAPACITY_UNMET_FIRST_CLEAN_ARCHIVE_EXISTING_STATE):
            archive_dir = _resolve(RESULTS_SINGLE_FILE_ARCHIVE_DIR)
            archive_dir.mkdir(parents=True, exist_ok=True)
            stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            archive_path = archive_dir / f"{path.stem}_{stamp}{path.suffix}"
            try:
                shutil.copy2(path, archive_path)
                print(
                    "[CAPACITY_UNMET] first_clean mode: archived existing state to "
                    f"{archive_path}"
                )
            except Exception as exc:
                print(
                    "[WARN] Failed archiving existing capacity unmet state in first_clean mode: "
                    f"{exc}"
                )
        print(
            "[CAPACITY_UNMET] first_clean mode: ignoring persisted iterative state and "
            "starting from empty cumulative additions."
        )
        return default_state
    if not path.exists():
        return default_state
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise ValueError(
            f"Failed reading capacity unmet iterative state file '{path}': {exc}"
        ) from exc
    if not isinstance(payload, dict):
        raise ValueError(f"Invalid capacity unmet iterative state payload in '{path}'.")
    for key, default_value in default_state.items():
        value = payload.get(key)
        if isinstance(default_value, dict):
            payload[key] = value if isinstance(value, dict) else {}
        elif isinstance(default_value, list):
            payload[key] = value if isinstance(value, list) else []
        else:
            payload.setdefault(key, default_value)
    return payload


def _write_capacity_unmet_state(
    state: dict[str, object],
    state_path: Path | str = CAPACITY_UNMET_STATE_PATH,
) -> Path:
    """Persist iterative capacity state JSON to disk."""
    path = _resolve(state_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(state, indent=2), encoding="utf-8")
    return path


def _parse_year_column_token(value: object) -> int | None:
    """Parse an integer year from sheet column headers like 2030 or 2030.0."""
    text = str(value or "").strip()
    if not text:
        return None
    if text.endswith(".0"):
        text = text[:-2]
    if not text.isdigit():
        return None
    year = int(text)
    if year < BASE_YEAR or year > FINAL_YEAR:
        return None
    return year


def _find_supply_results_header_row(raw: pd.DataFrame) -> int:
    """Return the header row index for LEAP results-table sheets."""
    for idx in range(len(raw.index)):
        row_values = [str(item or "").strip().lower() for item in raw.iloc[idx].tolist()]
        if "fuel" in row_values:
            return int(idx)
    raise ValueError("Could not locate 'Fuel' header row in supply results table.")


def _read_supply_results_trade_sheet(
    workbook_path: Path,
    sheet_name: str,
    economy: str,
    scenario: str,
    label_to_product: dict[str, str],
    value_field: str,
) -> tuple[pd.DataFrame, list[str]]:
    """Read one supply results trade sheet into economy/scenario/product/year rows."""
    raw = pd.read_excel(workbook_path, sheet_name=sheet_name, header=None)
    header_row = _find_supply_results_header_row(raw)
    data = raw.iloc[header_row + 1 :].copy()
    data.columns = raw.iloc[header_row].tolist()
    if "Fuel" not in data.columns:
        raise ValueError(
            f"Workbook '{workbook_path.name}' sheet '{sheet_name}' missing 'Fuel' column."
        )

    year_columns: list[tuple[object, int]] = []
    for column in data.columns:
        year = _parse_year_column_token(column)
        if year is not None:
            year_columns.append((column, year))
    if not year_columns:
        raise ValueError(
            f"Workbook '{workbook_path.name}' sheet '{sheet_name}' has no {BASE_YEAR}-{FINAL_YEAR} year columns."
        )

    rows: list[dict[str, object]] = []
    unmapped_fuels: set[str] = set()
    for _, row in data.iterrows():
        fuel_label = str(row.get("Fuel") or "").strip()
        if not fuel_label:
            continue
        fuel_lookup = (
            label_to_product.get(fuel_label)
            or label_to_product.get(fuel_label.lower())
            or label_to_product.get(_normalize_label_for_lookup(fuel_label))
        )
        if not fuel_lookup:
            unmapped_fuels.add(fuel_label)
            continue
        for col, year in year_columns:
            numeric = pd.to_numeric(row.get(col), errors="coerce")
            if pd.isna(numeric):
                continue
            rows.append(
                {
                    "economy": str(economy),
                    "scenario": str(scenario),
                    "esto_product": str(fuel_lookup),
                    "year": int(year),
                    value_field: max(float(numeric), 0.0),
                }
            )
    if not rows:
        return pd.DataFrame(
            columns=["economy", "scenario", "esto_product", "year", value_field]
        ), sorted(unmapped_fuels)
    out = (
        pd.DataFrame(rows)
        .groupby(
            ["economy", "scenario", "esto_product", "year"],
            as_index=False,
            dropna=False,
        )[value_field]
        .sum(min_count=1)
    )
    return out, sorted(unmapped_fuels)


def _read_supply_results_import_sheet(
    workbook_path: Path,
    sheet_name: str,
    economy: str,
    scenario: str,
    label_to_product: dict[str, str],
) -> tuple[pd.DataFrame, list[str]]:
    """Read one supply results imports sheet into economy/scenario/product/year rows."""
    return _read_supply_results_trade_sheet(
        workbook_path=workbook_path,
        sheet_name=sheet_name,
        economy=economy,
        scenario=scenario,
        label_to_product=label_to_product,
        value_field="observed_imports",
    )


def _read_supply_results_export_sheet(
    workbook_path: Path,
    sheet_name: str,
    economy: str,
    scenario: str,
    label_to_product: dict[str, str],
) -> tuple[pd.DataFrame, list[str]]:
    """Read one supply results exports sheet into economy/scenario/product/year rows."""
    return _read_supply_results_trade_sheet(
        workbook_path=workbook_path,
        sheet_name=sheet_name,
        economy=economy,
        scenario=scenario,
        label_to_product=label_to_product,
        value_field="observed_exports",
    )


def _balance_table_csv_candidates(results_source: Path | str | Iterable[Path | str]) -> list[Path]:
    """Return explicit balance-table CSV candidates from a directory or path list."""
    if isinstance(results_source, (str, Path)):
        root = _resolve(results_source)
        if root.is_dir():
            return sorted(root.glob("balance_table_*.csv"))
        return [root] if root.suffix.lower() == ".csv" else []
    candidates: list[Path] = []
    for value in results_source:
        path = _resolve(value)
        if path.suffix.lower() == ".csv":
            candidates.append(path)
    return sorted(candidates)


def _collect_observed_trade_from_balance_tables(
    *,
    scenario_pairs: list[tuple[str, str]],
    results_dir: Path | str | Iterable[Path | str],
    include_exports: bool,
) -> tuple[pd.DataFrame, dict[str, object], list[dict[str, object]]]:
    """Collect observed imports/exports from yearly balance-table CSVs."""
    candidates = _balance_table_csv_candidates(results_dir)
    if not candidates:
        raise FileNotFoundError(
            f"No yearly balance-table CSV files were found in '{results_dir}'."
        )

    required_columns = {
        "economy",
        "scenario",
        "year",
        "esto_product",
        "balance_component",
        "value",
    }
    frames: list[pd.DataFrame] = []
    for path in candidates:
        table = pd.read_csv(path)
        missing = [column for column in required_columns if column not in table.columns]
        if missing:
            raise ValueError(
                f"Balance table '{path}' is missing required columns: {missing}"
            )
        frame = table[
            [
                "economy",
                "scenario",
                "year",
                "esto_product",
                "balance_component",
                "value",
            ]
        ].copy()
        frame["economy"] = frame["economy"].astype(str).str.strip()
        frame["scenario"] = frame["scenario"].astype(str).str.strip()
        frame["economy_key"] = frame["economy"].map(_state_token)
        frame["scenario_key"] = frame["scenario"].map(_state_token)
        frame["year"] = pd.to_numeric(frame["year"], errors="coerce").astype("Int64")
        frame["balance_component"] = frame["balance_component"].astype(str).str.strip()
        frame["value"] = pd.to_numeric(frame["value"], errors="coerce")
        frames.append(frame)

    combined = pd.concat(frames, ignore_index=True, sort=False)
    combined = combined[
        combined["balance_component"].isin({"adjusted_imports", "adjusted_exports"})
    ].copy()
    if combined.empty:
        raise FileNotFoundError(
            f"Balance tables in '{results_dir}' did not contain adjusted import/export rows."
        )

    observed_rows: list[pd.DataFrame] = []
    missing_pairs: list[tuple[str, str]] = []
    for economy, scenario_key in scenario_pairs:
        economy_key = _state_token(economy)
        scenario_key = _state_token(scenario_key)
        subset = combined[
            (combined["economy_key"] == economy_key)
            & (combined["scenario_key"] == scenario_key)
        ].copy()
        if subset.empty:
            missing_pairs.append((str(economy), str(scenario_key)))
            continue

        import_rows = subset[subset["balance_component"] == "adjusted_imports"][
            ["economy", "scenario", "esto_product", "year", "value"]
        ].copy()
        import_rows["scenario"] = scenario_key
        import_rows["value"] = pd.to_numeric(import_rows["value"], errors="coerce").abs()
        import_rows = import_rows.rename(columns={"value": "observed_imports"})
        observed_rows.append(import_rows)

        if include_exports:
            export_rows = subset[subset["balance_component"] == "adjusted_exports"][
                ["economy", "scenario", "esto_product", "year", "value"]
            ].copy()
            export_rows["scenario"] = scenario_key
            export_rows["value"] = pd.to_numeric(export_rows["value"], errors="coerce").abs()
            export_rows = export_rows.rename(columns={"value": "observed_exports"})
            observed_rows.append(export_rows)

    if missing_pairs:
        preview = ", ".join(f"{economy}/{scenario}" for economy, scenario in missing_pairs[:6])
        raise FileNotFoundError(
            "Could not locate balance-table rows for economy/scenario: "
            f"{preview}. source='{results_dir}'."
        )

    import_rows = (
        pd.concat(
            [frame for frame in observed_rows if "observed_imports" in frame.columns],
            ignore_index=True,
            sort=False,
        )
        if observed_rows
        else pd.DataFrame(
            columns=["economy", "scenario", "esto_product", "year", "observed_imports"]
        )
    )
    if not import_rows.empty:
        import_rows = (
            import_rows.groupby(
                ["economy", "scenario", "esto_product", "year"],
                as_index=False,
                dropna=False,
            )["observed_imports"]
            .sum(min_count=1)
        )

    if include_exports:
        export_rows = (
            pd.concat(
                [frame for frame in observed_rows if "observed_exports" in frame.columns],
                ignore_index=True,
                sort=False,
            )
            if observed_rows
            else pd.DataFrame(
                columns=["economy", "scenario", "esto_product", "year", "observed_exports"]
            )
        )
        if not export_rows.empty:
            export_rows = (
                export_rows.groupby(
                    ["economy", "scenario", "esto_product", "year"],
                    as_index=False,
                    dropna=False,
                )["observed_exports"]
                .sum(min_count=1)
            )
    else:
        export_rows = pd.DataFrame(
            columns=["economy", "scenario", "esto_product", "year", "observed_exports"]
        )

    observed = import_rows
    if include_exports:
        observed = observed.merge(
            export_rows,
            on=["economy", "scenario", "esto_product", "year"],
            how="outer",
        )

    signature_map: dict[str, object] = {}
    signature_payload = {
        "source": "balance_tables",
        "files": [_build_results_signature(path) for path in candidates],
    }
    for economy, scenario_key in scenario_pairs:
        signature_map[_results_signature_state_key(economy, scenario_key)] = signature_payload

    return observed, signature_map, []


def _select_supply_results_workbook(
    *,
    economy: str,
    scenario: str,
    results_dir: Path | str = CAPACITY_UNMET_RESULTS_DIR,
) -> Path:
    """Select the best matching supply results workbook for economy/scenario."""
    root = _resolve(results_dir)
    candidates = sorted(root.glob("supply_results_*.xlsx"))
    if not candidates:
        raise FileNotFoundError(f"No supply results workbooks found in '{root}'.")

    economy_tokens = {
        _state_token(economy),
        _state_token(str(economy).replace("_", "")),
    }
    scenario_tokens = {
        _state_token(scenario),
        _state_token(str(scenario).replace(" ", "")),
        _state_token(str(scenario).replace("_", "")),
    }
    scenario_tokens.update(_state_token(item) for item in _scenario_filename_candidates(scenario))
    economy_tokens = {token for token in economy_tokens if token}
    scenario_tokens = {token for token in scenario_tokens if token}

    scored: list[tuple[int, float, Path]] = []
    for path in candidates:
        name_token = _state_token(path.stem.replace("_", ""))
        econ_score = max((1 if token and token in name_token else 0) for token in economy_tokens) if economy_tokens else 0
        scen_score = max((1 if token and token in name_token else 0) for token in scenario_tokens) if scenario_tokens else 0
        if econ_score == 0 or scen_score == 0:
            continue
        try:
            stat = path.stat()
            mtime = float(stat.st_mtime)
        except Exception:
            mtime = 0.0
        scored.append((econ_score + scen_score, mtime, path))
    if not scored:
        raise FileNotFoundError(
            "Could not locate supply results workbook for economy/scenario: "
            f"economy='{economy}', scenario='{scenario}', dir='{root}'."
        )
    scored.sort(key=lambda item: (item[0], item[1]))
    return scored[-1][2]


def _build_results_signature(path: Path) -> dict[str, object]:
    """Return file signature payload used for same-results reuse guard."""
    stat = path.stat()
    return {
        "path": str(path.resolve()),
        "name": path.name,
        "size_bytes": int(stat.st_size),
        "mtime_utc": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
    }


def _parse_runtime_capacity_additions_from_state(
    additions: dict[str, object] | None,
) -> dict[str, float]:
    """Normalize state capacity-addition payload into key->float map."""
    out: dict[str, float] = {}
    if not isinstance(additions, dict):
        return out
    for key, value in additions.items():
        numeric = pd.to_numeric(value, errors="coerce")
        if pd.isna(numeric):
            continue
        value_float = float(numeric)
        if abs(value_float) <= 0.0:
            continue
        out[str(key)] = value_float
    return out


def _lookup_runtime_capacity_additions_for_record(
    *,
    economy: str,
    scenario: str,
    module: str,
    process: str,
    instance: int,
) -> dict[int, float]:
    """Return per-year cumulative exogenous-capacity additions for one process record."""
    additions_by_year: dict[int, float] = {}
    scenario_token = _state_token(scenario)
    aliases = {scenario_token}
    if scenario_token in {"current accounts", "current account"}:
        aliases.add("reference")
    for scenario_alias in aliases:
        for year in range(BASE_YEAR, FINAL_YEAR + 1):
            key = _capacity_addition_state_key(
                economy=economy,
                scenario=scenario_alias,
                module=module,
                process=process,
                instance=instance,
                year=year,
            )
            value = _CAPACITY_UNMET_RUNTIME_CAPACITY_ADDITIONS.get(key, 0.0)
            if value <= 0.0:
                continue
            additions_by_year[year] = additions_by_year.get(year, 0.0) + float(value)
    return additions_by_year


def _lookup_runtime_primary_addition(
    *,
    economy: str,
    scenario: str,
    esto_product: str,
    year: int,
) -> float:
    """Return cumulative primary-production addition for one product-year."""
    scenario_token = _state_token(scenario)
    aliases = {scenario_token}
    if scenario_token in {"current accounts", "current account"}:
        aliases.add("reference")
    value = 0.0
    for scenario_alias in aliases:
        key = _output_addition_state_key(
            economy=economy,
            scenario=scenario_alias,
            esto_product=esto_product,
            year=year,
        )
        value += float(_CAPACITY_UNMET_RUNTIME_PRIMARY_ADDITIONS.get(key, 0.0))
    return max(value, 0.0)


def _lookup_runtime_export_adjustment(
    *,
    economy: str,
    scenario: str,
    esto_product: str,
    year: int,
) -> float:
    """Return cumulative extra exports adjustment for one product-year."""
    scenario_token = _state_token(scenario)
    aliases = {scenario_token}
    if scenario_token in {"current accounts", "current account"}:
        aliases.add("reference")
    value = 0.0
    for scenario_alias in aliases:
        key = _output_addition_state_key(
            economy=economy,
            scenario=scenario_alias,
            esto_product=esto_product,
            year=year,
        )
        value += float(_CAPACITY_UNMET_RUNTIME_EXPORT_ADJUSTMENTS.get(key, 0.0))
    return max(value, 0.0)


def _is_primary_esto_product(esto_product: str) -> bool:
    """Return True when ESTO product is classified as primary supply."""
    token = str(esto_product or "").strip()
    classification = supply_data_pipeline.ESTO_PRODUCT_CLASSIFICATION.get(token)
    if classification in {"primary", "secondary"}:
        return classification == "primary"
    return True


def _lookup_module_capacity_upper_limit(
    *,
    economy: str,
    scenario: str,
    module: str,
) -> float | None:
    """Return optional module-level output cap for iterative modes."""
    root = CAPACITY_UNMET_MODULE_CAPACITY_UPPER_LIMITS
    if not isinstance(root, dict):
        return None
    economy_payload = root.get(str(economy))
    if not isinstance(economy_payload, dict):
        economy_payload = root.get(_state_token(economy))
    if not isinstance(economy_payload, dict):
        lower_lookup = {
            _state_token(key): value
            for key, value in root.items()
            if isinstance(value, dict)
        }
        economy_payload = lower_lookup.get(_state_token(economy))
    if not isinstance(economy_payload, dict):
        return None

    scenario_payload = economy_payload.get(str(scenario))
    if not isinstance(scenario_payload, dict):
        scenario_payload = economy_payload.get(_state_token(scenario))
    if not isinstance(scenario_payload, dict):
        lower_lookup = {
            _state_token(key): value
            for key, value in economy_payload.items()
            if isinstance(value, dict)
        }
        scenario_payload = lower_lookup.get(_state_token(scenario))
    if not isinstance(scenario_payload, dict):
        return None

    value = scenario_payload.get(str(module))
    if value is None:
        lower_lookup = {
            _state_token(key): val
            for key, val in scenario_payload.items()
            if val is not None
        }
        value = lower_lookup.get(_state_token(module))
    numeric = pd.to_numeric(value, errors="coerce")
    if pd.isna(numeric):
        return None
    return max(float(numeric), 0.0)


def _lookup_production_upper_limit(
    *,
    economy: str,
    scenario: str,
    esto_product: str,
) -> float | None:
    """Return optional product-level production cap for balanced iterative mode."""
    root = CAPACITY_UNMET_PRODUCTION_UPPER_LIMITS
    if not isinstance(root, dict):
        return None
    economy_payload = root.get(str(economy))
    if not isinstance(economy_payload, dict):
        economy_payload = root.get(_state_token(economy))
    if not isinstance(economy_payload, dict):
        lower_lookup = {
            _state_token(key): value
            for key, value in root.items()
            if isinstance(value, dict)
        }
        economy_payload = lower_lookup.get(_state_token(economy))
    if not isinstance(economy_payload, dict):
        return None

    scenario_payload = economy_payload.get(str(scenario))
    if not isinstance(scenario_payload, dict):
        scenario_payload = economy_payload.get(_state_token(scenario))
    if not isinstance(scenario_payload, dict):
        lower_lookup = {
            _state_token(key): value
            for key, value in economy_payload.items()
            if isinstance(value, dict)
        }
        scenario_payload = lower_lookup.get(_state_token(scenario))
    if not isinstance(scenario_payload, dict):
        return None

    value = scenario_payload.get(str(esto_product))
    if value is None:
        lower_lookup = {
            _normalize_esto_product_for_match(key): val
            for key, val in scenario_payload.items()
            if val is not None
        }
        value = lower_lookup.get(_normalize_esto_product_for_match(esto_product))
    numeric = pd.to_numeric(value, errors="coerce")
    if pd.isna(numeric):
        return None
    return max(float(numeric), 0.0)


def _build_module_baseline_output_lookup(process_catalog: pd.DataFrame) -> dict[tuple[str, str, int], float]:
    """Return normalized lookup of baseline module output totals by economy/module/year."""
    if process_catalog is None or process_catalog.empty:
        return {}
    grouped = (
        process_catalog.groupby(["economy", "module", "year"], as_index=False)["module_total_output"]
        .sum(min_count=1)
    )
    out: dict[tuple[str, str, int], float] = {}
    for _, row in grouped.iterrows():
        year_value = pd.to_numeric(row.get("year"), errors="coerce")
        if pd.isna(year_value):
            continue
        output_value = pd.to_numeric(row.get("module_total_output"), errors="coerce")
        output_float = 0.0 if pd.isna(output_value) else max(float(output_value), 0.0)
        key = (
            _state_token(row.get("economy")),
            _state_token(row.get("module")),
            int(year_value),
        )
        out[key] = output_float
    return out


def _build_module_added_output_lookup(
    additions: dict[str, float],
) -> dict[tuple[str, str, str, int], float]:
    """Aggregate process-level capacity additions to module-year totals."""
    out: dict[tuple[str, str, str, int], float] = {}
    if not isinstance(additions, dict):
        return out
    for key, value in additions.items():
        parts = str(key or "").split("|")
        if len(parts) != 6:
            continue
        economy, scenario, module, _process, _instance, year_text = parts
        year_value = pd.to_numeric(year_text, errors="coerce")
        if pd.isna(year_value):
            continue
        amount = pd.to_numeric(value, errors="coerce")
        if pd.isna(amount):
            continue
        out_key = (_state_token(economy), _state_token(scenario), _state_token(module), int(year_value))
        out[out_key] = out.get(out_key, 0.0) + max(float(amount), 0.0)
    return out


def _normalize_esto_product_token(value: object) -> str:
    """Normalize ESTO product token for case-insensitive matching."""
    return str(value or "").strip().lower()


def _resolve_unresolved_positive_policy() -> str:
    """Return unresolved-positive policy token with validation."""
    policy = str(CAPACITY_UNMET_UNRESOLVED_POSITIVE_POLICY or "").strip().lower() or "fail"
    valid = {"fail", "imports_fallback", "track_only"}
    if policy not in valid:
        raise ValueError(
            "Invalid CAPACITY_UNMET_UNRESOLVED_POSITIVE_POLICY="
            f"{CAPACITY_UNMET_UNRESOLVED_POSITIVE_POLICY!r}. Valid values: {sorted(valid)}"
        )
    return policy


def _is_unresolved_allowlisted(esto_product: object) -> bool:
    """Return True when unresolved fuel is allowlisted for non-fatal handling."""
    allowlist = globals().get("CAPACITY_UNMET_UNRESOLVED_POSITIVE_ALLOWLIST", set())
    if not isinstance(allowlist, (set, list, tuple)):
        return False
    normalized = {_normalize_esto_product_token(item) for item in allowlist}
    return _normalize_esto_product_token(esto_product) in normalized


def _save_unresolved_positive_report(
    *,
    mode: str,
    unresolved_rows: list[dict[str, object]],
) -> tuple[Path, Path]:
    """Persist unresolved-positive diagnostics to CSV and JSON artifacts."""
    output_root = _resolve(RESULTS_CHECKS_DIR)
    output_root.mkdir(parents=True, exist_ok=True)
    csv_path = output_root / f"{mode}_unresolved_positive_residuals.csv"
    json_path = output_root / f"{mode}_unresolved_positive_residuals.json"
    frame = pd.DataFrame(unresolved_rows)
    _sort_output_frame_for_csv(frame).to_csv(csv_path, index=False)
    payload = {
        "mode": mode,
        "count": int(len(unresolved_rows)),
        "rows": unresolved_rows,
    }
    json_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return csv_path, json_path


def _sort_output_frame_for_csv(
    frame: pd.DataFrame,
    *,
    exclude_sort_columns: Iterable[str] = (),
    defer_sort_columns: Iterable[str] = ("year", "source", "source_sheet"),
) -> pd.DataFrame:
    """Return a stably sorted copy for human-facing CSV outputs."""
    if frame is None:
        return pd.DataFrame()
    if frame.empty:
        return frame.copy()

    exclude = {
        str(column).strip()
        for column in exclude_sort_columns
        if str(column).strip() and str(column).strip() in frame.columns
    }
    deferred = [
        str(column).strip()
        for column in defer_sort_columns
        if str(column).strip()
        and str(column).strip() in frame.columns
        and str(column).strip() not in exclude
    ]
    primary = [
        column
        for column in frame.columns
        if column not in exclude and column not in deferred
    ]
    sort_columns = primary + deferred
    if not sort_columns:
        return frame.copy().reset_index(drop=True)

    out = frame.copy()
    try:
        return out.sort_values(by=sort_columns, kind="mergesort", na_position="last").reset_index(
            drop=True
        )
    except Exception:
        normalized = pd.DataFrame(index=out.index)
        for column in sort_columns:
            normalized[column] = out[column].map(
                lambda value: "" if pd.isna(value) else str(value)
            )
        sorted_index = normalized.sort_values(
            by=sort_columns,
            kind="mergesort",
            na_position="last",
        ).index
        return out.loc[sorted_index].reset_index(drop=True)


def _is_no_eligible_transformation_producer_case(row: dict[str, object]) -> bool:
    """Return True when unresolved row reflects lack of transformation producer mapping."""
    reason = str(row.get("reason") or "").strip().lower()
    return "no eligible transformation process outputs this fuel in this year" in reason


def _split_unresolved_rows_by_policy(
    unresolved_rows: list[dict[str, object]],
    *,
    mode: str = "",
) -> tuple[list[dict[str, object]], list[dict[str, object]], str]:
    """Split unresolved rows into fatal vs handled rows according to policy/allowlist."""
    policy = _resolve_unresolved_positive_policy()
    if not unresolved_rows:
        return [], [], policy
    fatal_rows: list[dict[str, object]] = []
    handled_rows: list[dict[str, object]] = []
    for row in unresolved_rows:
        entry = dict(row)
        allowlisted = bool(_is_unresolved_allowlisted(entry.get("esto_product")))
        no_producer_case = _is_no_eligible_transformation_producer_case(entry)
        entry["allowlisted"] = allowlisted
        # Only in balanced mode, always allow imports-fallback behavior for
        # "no eligible producer" rows so LEAP can satisfy via imports.
        if no_producer_case and str(mode).strip().lower() == "capacity_unmet_iterative_balanced":
            entry["policy_applied"] = "imports_fallback_no_producer"
            handled_rows.append(entry)
            continue
        if policy == "fail" and not allowlisted:
            entry["policy_applied"] = "fail"
            fatal_rows.append(entry)
            continue
        entry["policy_applied"] = (
            "imports_fallback_allowlist"
            if policy == "fail" and allowlisted
            else policy
        )
        handled_rows.append(entry)
    return fatal_rows, handled_rows, policy


def _scenario_filename_candidates(scenario: str) -> list[str]:
    """Return scenario tokens to try in refinery-results filenames."""
    raw = str(scenario or "").strip()
    if not raw:
        return []
    compact = raw.replace(" ", "")
    title = raw.title()
    return list(dict.fromkeys([raw, compact, title]))


def _resolve_refinery_results_workbook(economy: str, scenario: str) -> Path | None:
    """Resolve scenario-specific transformation+supply workbook for refinery fallback."""
    for token in _scenario_filename_candidates(scenario):
        filename = REFINERY_RESULTS_FILENAME_TEMPLATE.format(economy=economy, scenario=token)
        candidate = LEAP_RESULTS_TABLES_DIR / filename
        if candidate.exists():
            return candidate
    return None


def _resolve_transformation_results_workbook(economy: str, scenario: str) -> Path | None:
    """Resolve scenario-specific transformation results template workbook."""
    for token in _scenario_filename_candidates(scenario):
        filename = TRANSFORMATION_RESULTS_FILENAME_TEMPLATE.format(economy=economy, scenario=token)
        candidate = LEAP_RESULTS_TABLES_DIR / filename
        if candidate.exists():
            return candidate
    return None


def _normalize_sector_match_key(value: object) -> str:
    """Return a forgiving sector key for cross-source name matching."""
    text = str(value or "").strip().lower()
    return "".join(ch for ch in text if ch.isalnum())


def _sector_match_keys(value: object) -> list[str]:
    """Return candidate normalized keys for matching coded and display sector names."""
    raw = str(value or "").strip().lower()
    if not raw:
        return []
    keys: list[str] = []
    direct = _normalize_sector_match_key(raw)
    if direct:
        keys.append(direct)
    # Handle coded sector names like `09_13_hydrogen_transformation`.
    stripped = re.sub(r"^\d+(?:[_.]\d+)*(?:[ _.-]+)?", "", raw).strip()
    stripped_key = _normalize_sector_match_key(stripped)
    if stripped_key and stripped_key not in keys:
        keys.append(stripped_key)
    return keys


@lru_cache(maxsize=32)
def _load_transformation_template_variable_sets(
    economy: str,
    scenario: str,
) -> tuple[dict[str, set[str]], dict[str, set[str]], dict[str, str]]:
    """
    Load transformation template variables by sector from results workbook.

    Returns:
    - dict keyed by sector title (second token in Transformation\\<sector>\\...)
    - each value is a set of requested variables for that sector in the template.
    """
    workbook = _resolve_transformation_results_workbook(economy, scenario)
    if workbook is None:
        raise FileNotFoundError(
            "Transformation results template workbook not found for "
            f"economy={economy}, scenario={scenario} in {LEAP_RESULTS_TABLES_DIR}"
        )
    wb = load_workbook(workbook, data_only=False)
    variables_by_sector: dict[str, set[str]] = {}
    variables_by_sector_norm: dict[str, set[str]] = {}
    sector_name_by_norm: dict[str, str] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        meta = leap_results_workflow.parse_template_worksheet(ws)
        branch = str(meta.get("branch") or "").strip()
        variable = str(meta.get("variable") or "").strip()
        if not branch or not variable:
            continue
        bits = [part for part in branch.split("\\") if part]
        if len(bits) < 2 or bits[0] != "Transformation":
            continue
        sector = bits[1].strip()
        if not sector:
            continue
        variables_by_sector.setdefault(sector, set()).add(variable)
        norm_key = _normalize_sector_match_key(sector)
        if norm_key:
            variables_by_sector_norm.setdefault(norm_key, set()).add(variable)
            sector_name_by_norm.setdefault(norm_key, sector)
    return variables_by_sector, variables_by_sector_norm, sector_name_by_norm


def _find_refinery_sheet_header_row(raw: pd.DataFrame) -> int | None:
    """Find the header row containing Fuel + year columns in refinery output sheet."""
    for idx in range(len(raw.index)):
        values = [_normalize_template_header_value(item) for item in raw.iloc[idx].tolist()]
        lowered = {item.strip().lower() for item in values if str(item).strip()}
        has_fuel = "fuel" in lowered
        has_year = any(str(item).isdigit() for item in values)
        if has_fuel and has_year:
            return int(idx)
    return None


@lru_cache(maxsize=32)
def _load_refinery_fallback_table(economy: str, scenario: str) -> pd.DataFrame:
    """Load refinery output rows from LEAP results workbook into long format."""
    workbook = _resolve_refinery_results_workbook(economy, scenario)
    if workbook is None:
        return pd.DataFrame(columns=["economy", "scenario", "year", "sector", "esto_product", "value"])
    try:
        raw = pd.read_excel(workbook, sheet_name=REFINERY_RESULTS_SHEET_NAME, header=None)
    except Exception:
        return pd.DataFrame(columns=["economy", "scenario", "year", "sector", "esto_product", "value"])

    header_row = _find_refinery_sheet_header_row(raw)
    if header_row is None:
        return pd.DataFrame(columns=["economy", "scenario", "year", "sector", "esto_product", "value"])

    header_values = [_normalize_template_header_value(item) for item in raw.iloc[header_row].tolist()]
    data = raw.iloc[header_row + 1 :].copy()
    data.columns = header_values
    data = data.dropna(how="all").reset_index(drop=True)
    if "Fuel" not in data.columns:
        return pd.DataFrame(columns=["economy", "scenario", "year", "sector", "esto_product", "value"])

    label_to_product = _build_label_to_esto_product_lookup()
    year_columns = [str(col) for col in data.columns if str(col).isdigit()]
    rows: list[dict[str, object]] = []
    for _, record in data.iterrows():
        fuel_label = str(record.get("Fuel") or "").strip()
        if not fuel_label or fuel_label.lower() == "total":
            continue
        esto_product = label_to_product.get(fuel_label) or label_to_product.get(fuel_label.lower())
        if not esto_product:
            alias = REFINERY_FUEL_LABEL_ALIASES.get(fuel_label)
            if alias:
                esto_product = label_to_product.get(alias) or label_to_product.get(alias.lower())
        if not esto_product:
            continue
        for year_col in year_columns:
            value = pd.to_numeric(record.get(year_col), errors="coerce")
            if pd.isna(value):
                continue
            rows.append(
                {
                    "economy": str(economy),
                    "scenario": str(scenario),
                    "year": int(year_col),
                    "sector": REFINERY_SECTOR_NAME,
                    "esto_product": str(esto_product),
                    "value": float(value),
                }
            )
    if not rows:
        return pd.DataFrame(columns=["economy", "scenario", "year", "sector", "esto_product", "value"])
    return pd.DataFrame(rows)


def _get_refinery_fallback_rows_for_balance(
    *,
    economy: str,
    scenario: str,
    year: int,
) -> pd.DataFrame:
    """Return refinery fallback rows for one economy/scenario/year."""
    table = _load_refinery_fallback_table(str(economy), str(scenario))
    if table.empty:
        return pd.DataFrame(columns=["sector", "esto_product", "value"])
    year_value = int(year)
    filtered = table[
        (table["economy"].astype(str) == str(economy))
        & (table["scenario"].astype(str) == str(scenario))
        & (pd.to_numeric(table["year"], errors="coerce").astype("Int64") == year_value)
    ].copy()
    if filtered.empty:
        return pd.DataFrame(columns=["sector", "esto_product", "value"])
    return (
        filtered.groupby(["sector", "esto_product"], dropna=False, as_index=False)["value"]
        .sum(min_count=1)
    )


def _pick_preferred_source(
    row: pd.Series,
    source_priority: tuple[str, ...],
) -> tuple[float | None, str | None]:
    """Return the first non-null source value using the configured precedence."""
    for source in source_priority:
        if source not in row.index:
            continue
        value = pd.to_numeric(row[source], errors="coerce")
        if pd.notna(value):
            return float(value), source
    return None, None


def _split_sector_codes(raw_value: object) -> list[str]:
    """Split one-or-many mapped sector codes using the same separators as the dashboard."""
    text = str(raw_value or "").strip()
    if not text or text.lower() == "nan":
        return []
    parts = re.split(r"\s*(?:,|;|\||\band\b)\s*", text, flags=re.IGNORECASE)
    out: list[str] = []
    seen: set[str] = set()
    for part in parts:
        token = str(part or "").strip()
        if not token:
            continue
        key = token.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(token)
    return out


def _is_demand_sector_mapping(sector_code_text: object) -> bool:
    """Return True when any mapped 9th sector code belongs to demand/bunkers groups."""
    for code in _split_sector_codes(sector_code_text):
        token = str(code or "").strip().lower()
        if any(token.startswith(prefix) for prefix in DEMAND_SECTOR_PREFIXES):
            return True
    return False


def _sector_code_sequence(value: object) -> tuple[int, ...]:
    """Return the numeric hierarchy sequence from a 9th sector code."""
    token = str(value or "").strip()
    if not token:
        return ()
    parts = [part for part in token.split("_") if part]
    seq: list[int] = []
    for part in parts:
        if not part.isdigit():
            break
        seq.append(int(part))
    return tuple(seq)


def _select_primary_sector_code(raw_value: object) -> str:
    """Pick the deepest mapped sector code for hierarchy comparisons."""
    codes = _split_sector_codes(raw_value)
    if not codes:
        return ""
    ranked = sorted(
        codes,
        key=lambda item: (len(_sector_code_sequence(item)), len(str(item))),
        reverse=True,
    )
    return str(ranked[0])


def _build_esto_parent_product_lookup() -> dict[str, str]:
    """Map each ESTO product label to its top-level parent label when available."""
    top_level_by_code: dict[str, str] = {}
    for item in ESTO_PRODUCT_LIST:
        text = str(item or "").strip()
        if not text:
            continue
        code = text.split(" ", 1)[0]
        if "." in code:
            continue
        top_level_by_code[code] = text

    lookup: dict[str, str] = {}
    for item in ESTO_PRODUCT_LIST:
        text = str(item or "").strip()
        if not text:
            continue
        code = text.split(" ", 1)[0]
        top_code = code.split(".", 1)[0]
        lookup[text] = top_level_by_code.get(top_code, text)
    return lookup


def _get_sector_to_esto_flow_lookup() -> dict[str, str]:
    """Load the shared 9th-sector -> ESTO flow lookup used by the dashboard mapping."""
    try:
        return build_sector_to_esto_flow_lookup()
    except Exception:
        return {}


SECTOR_TO_ESTO_FLOW_LOOKUP = _get_sector_to_esto_flow_lookup()
ESTO_PARENT_PRODUCT_LOOKUP = _build_esto_parent_product_lookup()


def _run_leap_results_template_scrape() -> dict[str, object]:
    """Refresh LEAP result templates in-place using the LEAP API workflow."""
    if not leap_api.is_available():
        raise RuntimeError(
            "LEAP API is unavailable; cannot scrape LEAP results templates. "
            "Set SCRAPE_LEAP_RESULTS = False or enable LEAP API."
        )
    try:
        return leap_results_workflow.run_template_fill()
    except Exception as exc:
        message = str(exc or "")
        normalized_message = message.lower()
        load_shape_error = (
            "all fuels produced by optimized modules must have load shapes"
            in normalized_message
        )
        if not load_shape_error:
            raise
        workbook_dir = _resolve(LEAP_RESULTS_TABLES_DIR)
        fallback_workbooks = (
            sorted(workbook_dir.glob("*.xls*")) if workbook_dir.exists() else []
        )
        if fallback_workbooks:
            print(
                "[WARN] LEAP template scrape skipped because LEAP calculation failed "
                "with missing optimized-module load-shape requirements."
            )
            print(
                "[WARN] Continuing with existing LEAP results-table files in "
                f"{workbook_dir} ({len(fallback_workbooks)} workbook(s))."
            )
            print(
                "[WARN] Results may be stale until LEAP load-shape settings are fixed "
                "and templates are scraped again."
            )
            return {
                "status": "skipped_due_load_shape_calculation_error",
                "error": message,
                "fallback_workbooks": [str(path) for path in fallback_workbooks],
            }
        raise RuntimeError(
            "LEAP template scrape failed due missing load-shape requirements and no "
            "fallback workbooks were found in "
            f"{workbook_dir}. Fix LEAP load shapes or set "
            "SCRAPE_LEAP_RESULTS=False and provide pre-scraped workbooks."
        ) from exc


def _economy_tokens_for_workbook_match(economy: str) -> set[str]:
    """Build filename match tokens from an economy label such as 20_USA."""
    text = str(economy or "").strip()
    if not text:
        return set()
    tokens = {text.lower(), text.replace("_", "").lower()}
    match = re.match(r"^\s*\d{2}_([A-Za-z]{3})\s*$", text)
    if match:
        tokens.add(match.group(1).lower())
    return {token for token in tokens if token}


def _discover_direct_demand_workbooks(
    workbook_dir: Path | str,
    economies: Iterable[str],
    scenarios: Iterable[str],
) -> list[Path]:
    """Find LEAP results-table workbooks for the requested economy/scenario set."""
    root = _resolve(workbook_dir)
    if not root.exists():
        raise FileNotFoundError(f"LEAP results tables directory not found: {root}")

    economy_tokens: set[str] = set()
    for economy in economies:
        economy_tokens.update(_economy_tokens_for_workbook_match(str(economy)))
    scenario_tokens = {str(scenario or "").strip().lower() for scenario in scenarios if str(scenario or "").strip()}

    candidates = sorted(root.glob("*.xls*"))
    matched: list[Path] = []
    for path in candidates:
        name = path.name.lower()
        if economy_tokens and not any(token in name for token in economy_tokens):
            continue
        if scenario_tokens and not any(token in name for token in scenario_tokens):
            continue
        matched.append(path)
    if not matched:
        raise FileNotFoundError(
            f"No LEAP workbooks found in {root} for economies {sorted(economy_tokens)} "
            f"and scenarios {sorted(scenario_tokens)}."
        )
    return matched


def _infer_economy_from_workbook_name(path: Path) -> str:
    """Infer economy code from workbook filename tokens."""
    stem = str(path.stem)
    match = re.search(r"_(\d{2}_[A-Z]{3})_", stem, flags=re.IGNORECASE)
    if match:
        token = match.group(1).upper()
        return token[:2] + "_" + token[3:]
    match = re.search(r"_(\d{2}[A-Z]{3})_", stem, flags=re.IGNORECASE)
    if match:
        token = match.group(1).upper()
        return token[:2] + "_" + token[2:]
    return ""


def _truthy_flag(value: object) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "t", "yes", "y", "on"}


def _load_active_direct_demand_mapping_sheet(sheet_name: str) -> pd.DataFrame:
    frame = read_config_table(DIRECT_DEMAND_MAPPING_WORKBOOK, sheet_name=sheet_name).fillna("").copy()
    if "remove_row" not in frame.columns:
        frame["remove_row"] = False
    if "duplicate_to_remove" not in frame.columns:
        frame["duplicate_to_remove"] = False
    active_mask = ~frame["remove_row"].map(_truthy_flag) & ~frame["duplicate_to_remove"].map(_truthy_flag)
    return frame.loc[active_mask].copy()


def _annotate_balance_demand_issue_scope(balance_demand_issues: pd.DataFrame) -> pd.DataFrame:
    """Mark which balance-demand mapping issues can affect demand-side inputs."""
    if balance_demand_issues is None or balance_demand_issues.empty:
        return balance_demand_issues.copy()

    issues = balance_demand_issues.copy()
    issues["mapping_key_sector"] = issues.get("mapping_key_sector", "").fillna("").astype(str).str.strip()
    issues["mapping_key_fuel"] = issues.get("mapping_key_fuel", "").fillna("").astype(str).str.strip()
    issues["leap_sector_name_full_path"] = (
        issues.get("leap_sector_name_full_path", "").fillna("").astype(str).str.strip()
    )
    issues["leap_product_name"] = issues.get("leap_product_name", "").fillna("").astype(str).str.strip()
    issues["issue_sector_key"] = issues["mapping_key_sector"].where(
        issues["mapping_key_sector"].ne(""),
        issues["leap_sector_name_full_path"],
    )
    issues["issue_fuel_key"] = issues["mapping_key_fuel"].where(
        issues["mapping_key_fuel"].ne(""),
        issues["leap_product_name"],
    )

    try:
        active_ninth = _load_active_direct_demand_mapping_sheet(DIRECT_DEMAND_NINTH_MAPPING_SHEET)
    except Exception as exc:
        issues["demand_relevant"] = True
        issues["demand_relevance_basis"] = f"fallback_keep_all:{type(exc).__name__}"
        return issues

    required_cols = ["leap_sector_name_full_path", "raw_leap_fuel_name", "ninth_sector"]
    missing_cols = [col for col in required_cols if col not in active_ninth.columns]
    if missing_cols:
        issues["demand_relevant"] = True
        issues["demand_relevance_basis"] = "fallback_keep_all:missing_ninth_columns"
        return issues

    ninth_scope = active_ninth[required_cols].copy()
    for col in required_cols:
        ninth_scope[col] = ninth_scope[col].fillna("").astype(str).str.strip()
    ninth_scope["ninth_sector_is_demand"] = ninth_scope["ninth_sector"].map(_is_demand_sector_mapping)

    pair_scope = (
        ninth_scope.groupby(["leap_sector_name_full_path", "raw_leap_fuel_name"], dropna=False, as_index=False)[
            "ninth_sector_is_demand"
        ]
        .max()
        .rename(
            columns={
                "leap_sector_name_full_path": "issue_sector_key",
                "raw_leap_fuel_name": "issue_fuel_key",
                "ninth_sector_is_demand": "pair_is_demand",
            }
        )
    )
    sector_scope = (
        ninth_scope.groupby("leap_sector_name_full_path", dropna=False, as_index=False)["ninth_sector_is_demand"]
        .max()
        .rename(
            columns={
                "leap_sector_name_full_path": "issue_sector_key",
                "ninth_sector_is_demand": "sector_is_demand",
            }
        )
    )

    issues = issues.merge(pair_scope, on=["issue_sector_key", "issue_fuel_key"], how="left")
    issues = issues.merge(sector_scope, on="issue_sector_key", how="left")
    issues["pair_scope_matched"] = issues["pair_is_demand"].notna()
    issues["sector_scope_matched"] = issues["sector_is_demand"].notna()
    issues["pair_is_demand"] = issues["pair_is_demand"].fillna(False).astype(bool)
    issues["sector_is_demand"] = issues["sector_is_demand"].fillna(False).astype(bool)

    issues["demand_relevant"] = False
    issues.loc[issues["pair_scope_matched"], "demand_relevant"] = issues.loc[
        issues["pair_scope_matched"], "pair_is_demand"
    ]
    sector_only_mask = ~issues["pair_scope_matched"] & issues["sector_scope_matched"]
    issues.loc[sector_only_mask, "demand_relevant"] = issues.loc[sector_only_mask, "sector_is_demand"]

    issues["demand_relevance_basis"] = "unclassified_non_demand"
    issues.loc[issues["pair_scope_matched"] & issues["pair_is_demand"], "demand_relevance_basis"] = (
        "pair_match_demand_sector"
    )
    issues.loc[issues["pair_scope_matched"] & ~issues["pair_is_demand"], "demand_relevance_basis"] = (
        "pair_match_non_demand_sector"
    )
    issues.loc[sector_only_mask & issues["sector_is_demand"], "demand_relevance_basis"] = (
        "sector_match_demand_sector"
    )
    issues.loc[sector_only_mask & ~issues["sector_is_demand"], "demand_relevance_basis"] = (
        "sector_match_non_demand_sector"
    )
    return issues


def _mapping_priority_rank(full_path: object) -> tuple[int, int, str]:
    text = str(full_path or "").strip()
    return (text.count("/"), len(text), text.lower())


def _pick_single_mapping_value(values: pd.Series, *, preferred: object = "") -> str:
    unique_values = sorted({str(value or "").strip() for value in values if str(value or "").strip()})
    if not unique_values:
        return ""
    preferred_text = str(preferred or "").strip()
    if preferred_text and preferred_text in unique_values:
        return preferred_text
    return unique_values[0]


def _build_codebook_name_to_esto_flow_lookup(codebook_path: Path | str) -> dict[str, str]:
    try:
        codebook = read_config_table(codebook_path, sheet_name="code_to_name").fillna("")
    except Exception:
        return {}
    lookup: dict[str, str] = {}
    for _, row in codebook.iterrows():
        esto_column = str(row.get("esto_column", "")).strip().lower()
        if esto_column != "flows":
            continue
        esto_label = str(row.get("esto_label", "")).strip()
        name = str(row.get("name", "")).strip()
        if name and esto_label:
            lookup[name.lower()] = esto_label
    return lookup


def _build_direct_demand_mapping_status(
    *,
    sheet_map: pd.DataFrame,
    leap_long: pd.DataFrame,
) -> pd.DataFrame:
    """Build a minimal mapping-status table from leap_combined_ninth/esto."""
    active_esto = _load_active_direct_demand_mapping_sheet(DIRECT_DEMAND_ESTO_MAPPING_SHEET)
    active_ninth = _load_active_direct_demand_mapping_sheet(DIRECT_DEMAND_NINTH_MAPPING_SHEET)

    required_esto = ["leap_sector_name_full_path", "raw_leap_fuel_name", "esto_flow", "esto_product"]
    required_ninth = ["leap_sector_name_full_path", "raw_leap_fuel_name", "ninth_sector", "ninth_fuel"]
    missing_esto = [col for col in required_esto if col not in active_esto.columns]
    missing_ninth = [col for col in required_ninth if col not in active_ninth.columns]
    if missing_esto:
        raise KeyError(
            f"{DIRECT_DEMAND_ESTO_MAPPING_SHEET} is missing required columns for results_supply_link: {missing_esto}"
        )
    if missing_ninth:
        raise KeyError(
            f"{DIRECT_DEMAND_NINTH_MAPPING_SHEET} is missing required columns for results_supply_link: {missing_ninth}"
        )

    join_cols = ["leap_sector_name_full_path", "raw_leap_fuel_name"]
    merged = active_ninth[required_ninth].merge(
        active_esto[required_esto],
        on=join_cols,
        how="inner",
    ).drop_duplicates()
    if merged.empty:
        raise RuntimeError(
            "Direct demand mapping join between leap_combined_ninth and leap_combined_esto returned no active rows."
        )

    leap_sheet_fuels = leap_long[["sheet_name", "fuel_label"]].drop_duplicates().copy()
    leap_sheet_fuels["sheet_name"] = leap_sheet_fuels["sheet_name"].astype(str).str.strip()
    leap_sheet_fuels["fuel_label"] = leap_sheet_fuels["fuel_label"].astype(str).str.strip()

    demand_sheet_map = sheet_map.copy()
    demand_sheet_map["sheet_name"] = demand_sheet_map["sheet_name"].astype(str).str.strip()
    demand_sheet_map["sector_code_9th"] = demand_sheet_map["sector_code_9th"].astype(str).str.strip()
    if "sector_name" not in demand_sheet_map.columns:
        demand_sheet_map["sector_name"] = ""
    demand_sheet_map["sector_name"] = demand_sheet_map["sector_name"].astype(str).str.strip()
    demand_sheet_map = demand_sheet_map[
        demand_sheet_map["sector_code_9th"].map(_is_demand_sector_mapping)
    ][["sheet_name", "sector_code_9th", "sector_name"]].drop_duplicates()

    leap_sheet_fuels = leap_sheet_fuels.merge(demand_sheet_map, on="sheet_name", how="inner")
    if leap_sheet_fuels.empty:
        return pd.DataFrame(
            columns=[
                "sheet",
                "fuel_label",
                "sector_code_9th",
                "ninth_fuel_code",
                "esto_flow",
                "esto_product",
                "mapping_source",
                "mapping_note",
            ]
        )

    fuel_aliases = load_fuel_aliases(
        _resolve(DEFAULT_BACKUP_LEAP_MAPPINGS) if DEFAULT_BACKUP_LEAP_MAPPINGS else None,
        _resolve(DEFAULT_CODEBOOK),
    )
    sector_flow_lookup = build_sector_to_esto_flow_lookup(_resolve(DEFAULT_CODEBOOK))
    name_to_flow_lookup = _build_codebook_name_to_esto_flow_lookup(_resolve(DEFAULT_CODEBOOK))

    merged["raw_leap_fuel_name"] = merged["raw_leap_fuel_name"].astype(str).str.strip()
    merged["ninth_sector"] = merged["ninth_sector"].astype(str).str.strip()
    merged["ninth_fuel"] = merged["ninth_fuel"].astype(str).str.strip()
    merged["esto_flow"] = merged["esto_flow"].astype(str).str.strip()
    merged["esto_product"] = merged["esto_product"].astype(str).str.strip()
    merged["leap_sector_name_full_path"] = merged["leap_sector_name_full_path"].astype(str).str.strip()
    active_ninth["raw_leap_fuel_name"] = active_ninth["raw_leap_fuel_name"].astype(str).str.strip()
    active_ninth["ninth_sector"] = active_ninth["ninth_sector"].astype(str).str.strip()
    active_ninth["ninth_fuel"] = active_ninth["ninth_fuel"].astype(str).str.strip()
    active_esto["raw_leap_fuel_name"] = active_esto["raw_leap_fuel_name"].astype(str).str.strip()
    active_esto["esto_flow"] = active_esto["esto_flow"].astype(str).str.strip()
    active_esto["esto_product"] = active_esto["esto_product"].astype(str).str.strip()

    sector_flow_fallbacks = (
        merged[["ninth_sector", "esto_flow"]]
        .drop_duplicates()
        .groupby("ninth_sector", dropna=False)["esto_flow"]
        .apply(list)
        .to_dict()
    )
    fuel_product_fallbacks = (
        active_esto[["raw_leap_fuel_name", "esto_product"]]
        .drop_duplicates()
        .groupby("raw_leap_fuel_name", dropna=False)["esto_product"]
        .apply(list)
        .to_dict()
    )

    rows: list[dict[str, object]] = []
    for row in leap_sheet_fuels.itertuples(index=False):
        sector_codes = _split_sector_codes(row.sector_code_9th)
        if not sector_codes:
            sector_codes = [str(row.sector_code_9th)]

        exact_ninth = active_ninth[
            active_ninth["ninth_sector"].isin(sector_codes)
            & active_ninth["raw_leap_fuel_name"].eq(str(row.fuel_label))
        ].copy()
        matched = merged[
            merged["ninth_sector"].isin(sector_codes)
            & merged["raw_leap_fuel_name"].eq(str(row.fuel_label))
        ].copy()
        if exact_ninth.empty and matched.empty:
            rows.append(
                {
                    "sheet": str(row.sheet_name),
                    "fuel_label": str(row.fuel_label),
                    "sector_code_9th": str(row.sector_code_9th),
                    "ninth_fuel_code": "",
                    "esto_flow": "",
                    "esto_product": "",
                    "mapping_source": "",
                    "mapping_note": "no active leap_combined_ninth/leap_combined_esto match for sheet fuel",
                }
            )
            continue

        if not matched.empty:
            matched = matched.sort_values(
                by="leap_sector_name_full_path",
                key=lambda series: series.map(_mapping_priority_rank),
            )
        preferred_flow = ""
        if not matched.empty and matched["esto_flow"].nunique(dropna=True) > 1:
            preferred_flow = next(
                (
                    str(sector_flow_lookup.get(str(code).strip().lower(), "")).strip()
                    for code in sector_codes
                    if str(sector_flow_lookup.get(str(code).strip().lower(), "")).strip()
                ),
                "",
            )

        chosen_ninth_fuel = _pick_single_mapping_value(
            exact_ninth["ninth_fuel"] if not exact_ninth.empty else matched["ninth_fuel"]
        )
        if not chosen_ninth_fuel and str(row.fuel_label).strip().lower() == "total":
            chosen_ninth_fuel = "19_total"
        chosen_esto_product = _pick_single_mapping_value(matched["esto_product"])
        if not chosen_esto_product:
            chosen_esto_product = _pick_single_mapping_value(
                pd.Series(fuel_product_fallbacks.get(str(row.fuel_label), []), dtype="object")
            )
        if not chosen_esto_product:
            chosen_esto_product = str(
                map_fuel_label(str(row.fuel_label), fuel_aliases).get("esto_product", "")
            ).strip()
        if not chosen_esto_product and str(row.fuel_label).strip().lower() == "total":
            chosen_esto_product = "19 Total"

        chosen_esto_flow = _pick_single_mapping_value(
            matched["esto_flow"] if "esto_flow" in matched.columns else pd.Series(dtype="object"),
            preferred=preferred_flow,
        )
        if not chosen_esto_flow:
            fallback_flow_candidates: list[str] = []
            for code in sector_codes:
                fallback_flow_candidates.extend(
                    [str(item).strip() for item in sector_flow_fallbacks.get(str(code), []) if str(item).strip()]
                )
                codebook_flow = str(sector_flow_lookup.get(str(code).strip().lower(), "")).strip()
                if codebook_flow:
                    fallback_flow_candidates.append(codebook_flow)
            for name_candidate in [str(row.sheet_name).strip(), str(getattr(row, "sector_name", "")).strip()]:
                if name_candidate:
                    named_flow = str(name_to_flow_lookup.get(name_candidate.lower(), "")).strip()
                    if named_flow:
                        fallback_flow_candidates.append(named_flow)
            chosen_esto_flow = _pick_single_mapping_value(pd.Series(fallback_flow_candidates, dtype="object"))

        note_parts: list[str] = []
        if not matched.empty and matched["leap_sector_name_full_path"].nunique(dropna=True) > 1:
            note_parts.append(
                f"{int(matched['leap_sector_name_full_path'].nunique())} active LEAP paths share this demand sector/fuel mapping"
            )
        if not exact_ninth.empty and exact_ninth["ninth_fuel"].nunique(dropna=True) > 1:
            note_parts.append(
                "multiple ninth_fuel targets present; first stable active target selected"
            )
        if not matched.empty and matched["esto_flow"].nunique(dropna=True) > 1:
            note_parts.append(
                "multiple esto_flow targets present; first stable active target selected"
            )
        if not matched.empty and matched["esto_product"].nunique(dropna=True) > 1:
            note_parts.append(
                "multiple esto_product targets present; first stable active target selected"
            )
        if not exact_ninth.empty and matched.empty:
            note_parts.append("esto side fell back beyond direct leap_combined overlap")

        rows.append(
            {
                "sheet": str(row.sheet_name),
                "fuel_label": str(row.fuel_label),
                "sector_code_9th": str(row.sector_code_9th),
                "ninth_fuel_code": chosen_ninth_fuel,
                "esto_flow": chosen_esto_flow,
                "esto_product": chosen_esto_product,
                "mapping_source": "leap_combined_join",
                "mapping_note": "; ".join(note_parts),
            }
        )

    out = pd.DataFrame(rows).drop_duplicates(subset=["sheet", "fuel_label"], keep="first")
    return out.reset_index(drop=True)


def _load_direct_demand_reference_tables() -> tuple[pd.DataFrame, pd.DataFrame]:
    """Load ESTO and 9th reference tables without reusing old direct-demand mappings."""
    base_df, ninth_df = load_augmented_reference_tables(
        esto_path=_resolve(DIRECT_DEMAND_BASE_TABLE_PATH),
        ninth_path=_resolve(DIRECT_DEMAND_PROJECTION_TABLE_PATH),
        cache_dir=DIRECT_DEMAND_REFERENCE_CACHE_DIR,
        apply_esto_subtotal_map=False,
        filter_esto_subtotals_flag=False,
        filter_ninth_subtotals_flag=False,
    )
    return base_df, ninth_df


def _build_projection_rows_from_ninth(
    mapping_status: pd.DataFrame,
    *,
    ninth_df: pd.DataFrame,
    scenarios: Iterable[str],
) -> pd.DataFrame:
    if mapping_status.empty or ninth_df.empty:
        return pd.DataFrame(columns=["economy", "scenario", "sheet", "fuel_label", "year", "value", "source"])

    scenario_map = {str(k).strip().lower(): str(v).strip() for k, v in DIRECT_DEMAND_SCENARIO_MAP.items()}
    scenario_tokens = {
        str(item).strip().lower(): str(item).strip()
        for item in scenarios
        if str(item).strip()
    }

    ninth = ninth_df.copy()
    ninth["economy"] = ninth["economy"].astype(str).str.strip()
    ninth["scenarios"] = ninth["scenarios"].astype(str).str.strip().str.lower()
    sector_cols = ["sectors", "sub1sectors", "sub2sectors", "sub3sectors", "sub4sectors"]
    fuel_cols = ["fuels", "subfuels"]
    for col in [*sector_cols, *fuel_cols]:
        ninth[col] = ninth[col].fillna("").astype(str).str.strip()

    def _resolve_deepest(tokens: pd.Series) -> str:
        values = [str(value).strip() for value in tokens.tolist() if str(value).strip() and str(value).strip().lower() != "x"]
        return values[-1] if values else ""

    ninth["ninth_sector"] = ninth[sector_cols].apply(_resolve_deepest, axis=1)
    ninth["ninth_fuel"] = ninth[fuel_cols].apply(_resolve_deepest, axis=1)
    if "subtotal_results" in ninth.columns:
        subtotal_mask = ninth["subtotal_results"].astype(str).str.strip().str.lower().isin({"true", "1", "yes"})
        ninth = ninth.loc[~subtotal_mask].copy()

    year_cols = [str(year) for year in DIRECT_DEMAND_PROJECTION_YEARS if str(year) in ninth.columns]
    if not year_cols:
        return pd.DataFrame(columns=["economy", "scenario", "sheet", "fuel_label", "year", "value", "source"])

    ninth_long = ninth[
        ["economy", "scenarios", "ninth_sector", "ninth_fuel", *year_cols]
    ].melt(
        id_vars=["economy", "scenarios", "ninth_sector", "ninth_fuel"],
        value_vars=year_cols,
        var_name="year",
        value_name="value",
    )
    ninth_long["year"] = pd.to_numeric(ninth_long["year"], errors="coerce").astype("Int64")
    ninth_long["value"] = pd.to_numeric(ninth_long["value"], errors="coerce")
    ninth_long = ninth_long[
        (ninth_long["economy"] == DIRECT_DEMAND_PROJECTION_ECONOMY)
        & ninth_long["scenarios"].isin(scenario_map.keys())
        & ninth_long["year"].notna()
    ].copy()
    if ninth_long.empty:
        return pd.DataFrame(columns=["economy", "scenario", "sheet", "fuel_label", "year", "value", "source"])
    ninth_long["scenario"] = ninth_long["scenarios"].map(scenario_map)

    mapping_subset = mapping_status[
        ["sheet", "fuel_label", "sector_code_9th", "ninth_fuel_code"]
    ].copy()
    mapping_subset["sheet"] = mapping_subset["sheet"].astype(str).str.strip()
    mapping_subset["fuel_label"] = mapping_subset["fuel_label"].astype(str).str.strip()
    mapping_subset["sector_code_9th"] = mapping_subset["sector_code_9th"].astype(str).str.strip()
    mapping_subset["ninth_fuel_code"] = mapping_subset["ninth_fuel_code"].astype(str).str.strip()
    mapping_subset = mapping_subset[
        mapping_subset["sector_code_9th"].ne("")
        & mapping_subset["ninth_fuel_code"].ne("")
    ].drop_duplicates()
    if mapping_subset.empty:
        return pd.DataFrame(columns=["economy", "scenario", "sheet", "fuel_label", "year", "value", "source"])

    projection_rows = mapping_subset.merge(
        ninth_long,
        left_on=["sector_code_9th", "ninth_fuel_code"],
        right_on=["ninth_sector", "ninth_fuel"],
        how="inner",
    )
    if projection_rows.empty:
        return pd.DataFrame(columns=["economy", "scenario", "sheet", "fuel_label", "year", "value", "source"])

    projection_rows = projection_rows.rename(columns={"value": "value"})
    projection_rows["source"] = "projection"
    projection_rows = projection_rows[
        ["economy", "scenario", "sheet", "fuel_label", "year", "value", "source"]
    ].copy()
    if scenario_tokens:
        projection_rows = projection_rows[
            projection_rows["scenario"].astype(str).isin(set(scenario_tokens.values()))
        ].copy()
    return projection_rows.reset_index(drop=True)


def _collect_direct_demand_mapping_gaps(mapping_status: pd.DataFrame) -> pd.DataFrame:
    """Return unresolved direct-demand mapping rows that should fail after outputs are written."""
    base_columns = [
        "sheet",
        "fuel_label",
        "sector_code_9th",
        "ninth_fuel_code",
        "esto_flow",
        "esto_product",
        "mapping_source",
        "mapping_note",
        "gap_reason",
    ]
    if mapping_status is None or mapping_status.empty:
        return pd.DataFrame(columns=base_columns)

    work = mapping_status.copy()
    for col in base_columns[:-1]:
        if col not in work.columns:
            work[col] = ""
        work[col] = work[col].fillna("").astype(str).str.strip()

    work["mapping_note_lower"] = work["mapping_note"].str.lower()
    reasons: list[pd.Series] = []
    reasons.append(pd.Series("", index=work.index, dtype="object"))
    reasons[-1] = reasons[-1].mask(work["ninth_fuel_code"].eq(""), "missing_ninth_mapping")
    reasons[-1] = reasons[-1].mask(
        work["esto_flow"].eq(""),
        reasons[-1].where(reasons[-1].eq(""), reasons[-1] + "; ") + "missing_esto_flow_mapping",
    )
    reasons[-1] = reasons[-1].mask(
        work["esto_product"].eq(""),
        reasons[-1].where(reasons[-1].eq(""), reasons[-1] + "; ") + "missing_esto_product_mapping",
    )
    fallback_mask = work["mapping_note_lower"].str.contains(
        "fell back beyond direct leap_combined overlap",
        na=False,
    )
    reasons[-1] = reasons[-1].mask(
        fallback_mask,
        reasons[-1].where(reasons[-1].eq(""), reasons[-1] + "; ")
        + "exact_child_path_missing_in_leap_combined_esto",
    )
    work["gap_reason"] = reasons[-1].fillna("").astype(str).str.strip("; ").str.strip()
    gaps = work[work["gap_reason"].ne("")].copy()
    if gaps.empty:
        return pd.DataFrame(columns=base_columns)
    gaps = gaps[base_columns].drop_duplicates()
    gaps = gaps.sort_values(
        ["sheet", "fuel_label", "sector_code_9th", "gap_reason"],
        ascending=[True, True, True, True],
    ).reset_index(drop=True)
    return gaps


def _load_optional_json_dict(path: Path | str) -> dict[str, object]:
    """Load an optional JSON object config file, returning {} when absent."""
    resolved = _resolve(path)
    if not resolved.exists():
        return {}
    payload = json.loads(resolved.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError(f"Expected JSON object in {resolved}, found {type(payload).__name__}.")
    return payload


def _build_balance_demand_scenario_map(scenarios: Iterable[str]) -> dict[str, str]:
    """Map workflow scenario labels to the lowercase balance projection labels."""
    scenario_map: dict[str, str] = {}
    for value in scenarios:
        label = str(value or "").strip()
        if not label:
            continue
        lowered = label.lower()
        if lowered in {"reference", "target"}:
            scenario_map[label] = lowered
    return scenario_map


def load_balance_demand_inputs(
    *,
    economies: Iterable[str],
    scenarios: Iterable[str],
    workbook_dir: Path | str = LEAP_RESULTS_TABLES_DIR,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Build comparison_long + mapping_status in-memory from LEAP balance exports."""
    economy_list = workflow_common.normalize_economies(economies or ECONOMIES)
    unsupported = [
        value for value in economy_list if str(value).strip() != DIRECT_DEMAND_PROJECTION_ECONOMY
    ]
    if unsupported:
        raise ValueError(
            "Balance-export demand loading is currently configured only for "
            f"{DIRECT_DEMAND_PROJECTION_ECONOMY}. Unsupported economies: {unsupported}"
        )

    balance_scenarios = _filter_balance_scenarios(scenarios)
    scenario_map = _build_balance_demand_scenario_map(balance_scenarios)
    if not scenario_map:
        raise RuntimeError(
            "No balance-export demand scenarios remain after filtering non-balance "
            f"entries from {list(scenarios)}."
        )

    structure_config = build_esto_axis_structure_from_dashboard_template(BALANCE_DEMAND_CHART_NAVIGATION_GUIDE_PATH)
    known_issues = _load_optional_json_dict(BALANCE_DEMAND_KNOWN_ISSUES_CONFIG_PATH)

    conversion = convert_leap_balances_to_esto_long_table(
        ref_workbook_path=BALANCE_DEMAND_REF_WORKBOOK_PATH,
        tgt_workbook_path=BALANCE_DEMAND_TGT_WORKBOOK_PATH,
        template_sheet=BALANCE_DEMAND_TEMPLATE_SHEET,
        mapping_pairs_path=BALANCE_DEMAND_LEAP_TO_ESTO_MAPPING_WORKBOOK,
        codebook_path=BALANCE_DEMAND_CODEBOOK_PATH,
        structure_config=structure_config,
        known_issues=known_issues,
        projection_economy=DIRECT_DEMAND_PROJECTION_ECONOMY,
        max_output_year=FINAL_YEAR,
        explicit_pair_mappings_only=True,
    )
    comparison = build_balance_comparison_esto_axis(
        leap_long=conversion["leap_long"],
        mapping_status=conversion["mapping_status"],
        base_year=DIRECT_DEMAND_BASE_YEAR,
        projection_years=tuple(year for year in DIRECT_DEMAND_PROJECTION_YEARS if year <= FINAL_YEAR),
        base_economy=DIRECT_DEMAND_BASE_ECONOMY,
        projection_economy=DIRECT_DEMAND_PROJECTION_ECONOMY,
        scenario_map=scenario_map,
        sheet_map_path=BALANCE_DEMAND_SHEET_MAP_PATH,
        backup_mappings_path=BALANCE_DEMAND_BACKUP_MAPPINGS_PATH,
        codebook_path=BALANCE_DEMAND_CODEBOOK_PATH,
        canonical_pairs_path=BALANCE_DEMAND_NINTH_TO_ESTO_MAPPING,
        explicit_mappings_path=BALANCE_DEMAND_EXPLICIT_MAPPINGS_PATH,
        explicit_reassignments_path=BALANCE_DEMAND_EXPLICIT_REASSIGNMENTS_PATH,
        synthetic_reference_rows_path=BALANCE_DEMAND_SYNTHETIC_REFERENCE_ROWS_PATH,
        esto_table_path=BALANCE_DEMAND_BASE_TABLE_PATH,
        projection_table_path=BALANCE_DEMAND_PROJECTION_TABLE_PATH,
        chart_navigation_guide_path=None,
        known_issues=known_issues,
    )

    issues = conversion["issues"].copy()
    matching_diagnostics = conversion.get("matching_diagnostics", pd.DataFrame()).copy()
    scenario_set = {str(item).strip().lower() for item in balance_scenarios if str(item).strip()}
    comparison_long = comparison["comparison_long"].copy()
    mapping_status = comparison["mapping_status"].copy()
    if scenario_set:
        comparison_long = comparison_long[
            comparison_long["scenario"].astype(str).str.strip().str.lower().isin(scenario_set)
        ].copy()
        if "scenario" in mapping_status.columns:
            mapping_status = mapping_status[
                mapping_status["scenario"].astype(str).str.strip().str.lower().isin(scenario_set)
            ].copy()
        if "scenario" in issues.columns:
            issues = issues[
                issues["scenario"].astype(str).str.strip().str.lower().isin(scenario_set)
            ].copy()
        if "scenario" in matching_diagnostics.columns:
            matching_diagnostics = matching_diagnostics[
                matching_diagnostics["scenario"].astype(str).str.strip().str.lower().isin(scenario_set)
            ].copy()

    comparison_long["year"] = pd.to_numeric(comparison_long["year"], errors="coerce").astype("Int64")
    comparison_long["value"] = pd.to_numeric(comparison_long["value"], errors="coerce")
    return (
        comparison_long.reset_index(drop=True),
        mapping_status.reset_index(drop=True),
        issues.reset_index(drop=True),
        matching_diagnostics.reset_index(drop=True),
    )


def load_direct_leap_demand_inputs(
    *,
    economies: Iterable[str],
    scenarios: Iterable[str],
    workbook_dir: Path | str = LEAP_RESULTS_TABLES_DIR,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Backward-compatible alias for the balance-export demand loader."""
    comparison_long, mapping_status, _, _ = load_balance_demand_inputs(
        economies=economies,
        scenarios=scenarios,
        workbook_dir=workbook_dir,
    )
    return comparison_long, mapping_status


def _collect_transformation_and_transfer_rows(
    economies: Iterable[str] | None = None,
) -> list[dict]:
    """Return combined process records from transformation and transfers workflows."""
    economy_list = workflow_common.normalize_economies(economies or ECONOMIES)
    transformation_rows = transformation_workflow.collect_transformation_rows(economies=economy_list)
    if REFRESH_TRANSFORMATION_MEASURES_FROM_LEAP_RESULTS:
        transformation_rows = _refresh_transformation_measures_from_leap_results(
            transformation_rows,
            scenario=REFRESH_TRANSFORMATION_MEASURE_SCENARIO,
            region=REFRESH_TRANSFORMATION_MEASURE_REGION,
            base_year=BASE_YEAR,
            final_year=FINAL_YEAR,
        )
    transfer_rows: list[dict] = []
    for economy in economy_list:
        try:
            transfer_rows.extend(
                transfers_workflow.build_transfer_process_records(
                    economy=economy,
                    use_output_targets=False,
                )
            )
        except Exception as exc:
            print(f"[WARN] Failed to build transfer process records for {economy}: {exc}")
    return list(transformation_rows) + transfer_rows


def _query_leap_value_series_for_fuels(
    app,
    *,
    branch_candidates: list[str],
    variable_name: str,
    scenario: str,
    region: str,
    years: list[int],
    fuel_labels: list[str],
    filter_dimensions: tuple[str, ...],
    required: bool = False,
) -> dict[str, dict[int, float]]:
    """Query LEAP ValueRS for specific fuel labels across candidate branches."""
    series_by_label: dict[str, dict[int, float]] = {}
    variable_obj = None
    resolved_branch = ""

    for branch_path in branch_candidates:
        try:
            resolved = leap_results_workflow._resolve_existing_branch_path(app, branch_path)
            candidate_var = leap_results_workflow._resolve_branch_variable(
                app,
                resolved,
                variable_name,
                allow_substitution=False,
            )
            leap_results_workflow.set_axes(app, x_axis="Years", legend="Fuel")
            leap_results_workflow.set_context(
                app,
                scenario=scenario,
                region=region,
                branch_path=resolved,
            )
            try:
                app.ShowResultsViewTable()
            except Exception:
                pass
            variable_obj = candidate_var
            resolved_branch = resolved
            break
        except Exception:
            continue

    if variable_obj is None:
        if required:
            raise RuntimeError(
                "Failed to resolve required LEAP variable on candidate branches: "
                f"variable={variable_name}, branches={branch_candidates}, scenario={scenario}, region={region}"
            )
        return {}

    for label in fuel_labels:
        label_text = str(label or "").strip()
        if not label_text:
            continue
        values: dict[int, float] = {}
        for year in years:
            value = None
            for dim_label in filter_dimensions:
                filter_str = f"{dim_label}={label_text}"
                try:
                    queried = variable_obj.ValueRS(region, scenario, int(year), "", filter_str)
                except Exception:
                    continue
                numeric = pd.to_numeric(queried, errors="coerce")
                if pd.isna(numeric):
                    continue
                value = float(numeric)
                break
            if value is None:
                continue
            values[int(year)] = float(value)
        if values:
            series_by_label[label_text] = values
    if not series_by_label:
        if required:
            raise RuntimeError(
                "Required LEAP Results query returned no values: "
                f"branch={resolved_branch}, variable={variable_name}, scenario={scenario}, "
                f"region={region}, fuels={fuel_labels[:8]}"
            )
        print(
            f"[INFO] LEAP Results refresh found no values for {resolved_branch} / {variable_name} "
            f"with {len(fuel_labels)} candidate fuel(s)."
        )
    return series_by_label


def _refresh_transformation_measures_from_leap_results(
    rows: list[dict],
    *,
    scenario: str,
    region: str,
    base_year: int,
    final_year: int,
) -> list[dict]:
    """
    Refresh transformation output/feedstock series from LEAP Results by fuel filters.

    Method:
    - Query parent/process branches (not fuel-child branches)
    - Use explicit fuel-label filters against Results variables
    - Keep original record values when LEAP queries do not return data
    """
    if not rows:
        return rows
    if not leap_api.is_available():
        print("[INFO] LEAP API unavailable; skipping transformation Results refresh.")
        return rows
    try:
        app = leap_results_workflow.connect_leap()
    except Exception as exc:
        print(f"[WARN] Failed to connect LEAP for transformation Results refresh: {exc}")
        return rows

    years = list(range(int(base_year), int(final_year) + 1))
    refreshed: list[dict] = []
    refreshed_output_count = 0
    refreshed_feedstock_count = 0
    refreshed_feedstock_variable_counts = {"Inputs": 0, "Outputs by Feedstock Fuel": 0}

    for record in rows:
        out = copy.deepcopy(record)
        sector_name = str(out.get("sector_title") or "").strip()
        process_name = str(out.get("process_name") or "").strip()
        economy = str(out.get("economy") or "").strip()
        if not sector_name:
            refreshed.append(out)
            continue
        if not economy:
            raise RuntimeError(
                "Transformation Results refresh requires 'economy' on each record "
                f"(sector={sector_name}, process={process_name})."
            )

        (
            template_variables_by_sector,
            template_variables_by_sector_norm,
            template_sector_name_by_norm,
        ) = _load_transformation_template_variable_sets(economy, scenario)
        sector_template_variables = template_variables_by_sector.get(sector_name, set())
        sector_branch_name = sector_name
        if not sector_template_variables:
            for norm_key in _sector_match_keys(sector_name):
                sector_template_variables = template_variables_by_sector_norm.get(norm_key, set())
                if sector_template_variables:
                    sector_branch_name = template_sector_name_by_norm.get(norm_key, sector_name)
                    break
        if not sector_template_variables:
            raise RuntimeError(
                "No transformation template variables found for sector in results template workbook: "
                f"economy={economy}, scenario={scenario}, sector={sector_name}"
            )

        sector_branch = f"Transformation\\{sector_branch_name}"
        process_collection_branch = f"{sector_branch}\\Processes"
        process_branch = (
            f"{process_collection_branch}\\{process_name}"
            if process_name
            else process_collection_branch
        )
        branch_candidates = [process_branch, process_collection_branch, sector_branch]

        output_labels = [
            str(label).strip()
            for label in (out.get("output_values") or {}).keys()
            if str(label).strip()
        ]
        feedstock_labels = [
            str(label).strip()
            for label in (out.get("feedstock_values") or {}).keys()
            if str(label).strip()
        ]

        if output_labels:
            if "Outputs by Output Fuel" not in sector_template_variables:
                raise RuntimeError(
                    "Required template measure missing for output extraction: "
                    f"economy={economy}, scenario={scenario}, sector={sector_name}, "
                    "required='Outputs by Output Fuel'"
                )
            refreshed_output = _query_leap_value_series_for_fuels(
                app,
                branch_candidates=branch_candidates,
                variable_name="Outputs by Output Fuel",
                scenario=scenario,
                region=region,
                years=years,
                fuel_labels=output_labels,
                filter_dimensions=("Output Fuel", "Fuel"),
                required=True,
            )
            out["output_values"] = refreshed_output
            refreshed_output_count += 1

        if feedstock_labels:
            feedstock_variable_candidates = [
                name
                for name in ("Inputs", "Outputs by Feedstock Fuel")
                if name in sector_template_variables
            ]
            if not feedstock_variable_candidates:
                raise RuntimeError(
                    "No required feedstock measure found in transformation template for sector: "
                    f"economy={economy}, scenario={scenario}, sector={sector_name}, "
                    "required_one_of=['Inputs', 'Outputs by Feedstock Fuel']"
                )

            refreshed_feedstock: dict[str, dict[int, float]] = {}
            feedstock_variable_used = ""
            last_exc: Exception | None = None
            for feedstock_variable in feedstock_variable_candidates:
                try:
                    refreshed_feedstock = _query_leap_value_series_for_fuels(
                        app,
                        branch_candidates=branch_candidates,
                        variable_name=feedstock_variable,
                        scenario=scenario,
                        region=region,
                        years=years,
                        fuel_labels=feedstock_labels,
                        filter_dimensions=("Feedstock Fuel", "Fuel"),
                        required=True,
                    )
                except Exception as exc:
                    last_exc = exc
                    continue
                if refreshed_feedstock:
                    feedstock_variable_used = feedstock_variable
                    break
            if not refreshed_feedstock:
                if last_exc:
                    raise last_exc
                raise RuntimeError(
                    "Required feedstock extraction failed for all candidate variables: "
                    f"economy={economy}, scenario={scenario}, sector={sector_name}, process={process_name}"
                )
            out["feedstock_values"] = refreshed_feedstock
            refreshed_feedstock_count += 1
            if feedstock_variable_used:
                refreshed_feedstock_variable_counts[feedstock_variable_used] += 1

        refreshed.append(out)

    print(
        "[INFO] Transformation Results refresh summary: "
        f"records={len(rows)}, output_refreshed={refreshed_output_count}, "
        f"feedstock_refreshed={refreshed_feedstock_count}, "
        f"feedstock_inputs_used={refreshed_feedstock_variable_counts['Inputs']}, "
        "feedstock_outputs_by_feedstock_used="
        f"{refreshed_feedstock_variable_counts['Outputs by Feedstock Fuel']}, "
        f"scenario={scenario}, region={region}"
    )
    return refreshed


def _normalize_template_header_value(value: object) -> str:
    """Normalize LEAP import header cells into stable string column names."""
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_leap_template_sheet(path: Path | str, sheet_name: str) -> pd.DataFrame:
    """Read a LEAP-style import sheet by locating the Branch Path header row."""
    raw = pd.read_excel(_resolve(path), sheet_name=sheet_name, header=None)
    header_row: int | None = None
    for idx in range(len(raw.index)):
        values = {
            _normalize_template_header_value(value).lower()
            for value in raw.iloc[idx].tolist()
        }
        if "branch path" in values or "branch_path" in values:
            header_row = int(idx)
            break
    if header_row is None:
        raise ValueError(
            f"Could not find a LEAP-style 'Branch Path' header row in {path} ({sheet_name})."
        )
    columns = [_normalize_template_header_value(value) for value in raw.iloc[header_row].tolist()]
    data = raw.iloc[header_row + 1 :].copy()
    data.columns = columns
    data = data.dropna(how="all").reset_index(drop=True)
    return data


def _parse_data_expression(expression: object) -> dict[int, float]:
    """Parse a LEAP Data(...) expression into a year->value mapping."""
    text = str(expression or "").strip()
    if not text:
        return {}
    match = re.match(r"^\s*Data\s*\((.*)\)\s*$", text, flags=re.IGNORECASE)
    if not match:
        return {}
    body = match.group(1).strip()
    if not body:
        return {}
    parts = [part.strip() for part in body.split(",") if str(part).strip()]
    if len(parts) < 2:
        return {}
    values: dict[int, float] = {}
    for idx in range(0, len(parts) - 1, 2):
        year = pd.to_numeric(parts[idx], errors="coerce")
        value = pd.to_numeric(parts[idx + 1], errors="coerce")
        if pd.isna(year) or pd.isna(value):
            continue
        values[int(year)] = float(value)
    return values


def _infer_constraint_economies(
    template_path: Path | str,
    economies: Iterable[str] | None,
) -> list[str]:
    """Infer which economy/economies a constraint workbook should apply to."""
    economy_list = workflow_common.normalize_economies(economies or ECONOMIES)
    if not economy_list:
        return []
    template_token = re.sub(r"[^a-z0-9]+", "", _resolve(template_path).stem.lower())
    exact_matches = []
    for economy in economy_list:
        economy_token = re.sub(r"[^a-z0-9]+", "", str(economy).lower())
        if economy_token and economy_token in template_token:
            exact_matches.append(str(economy))
    if exact_matches:
        return exact_matches
    if len(economy_list) == 1:
        return [str(economy_list[0])]
    print(
        "[WARN] Skipping constraint workbook because its filename does not identify a single target economy: "
        f"{_resolve(template_path).name}"
    )
    return []


def _load_constraint_value_table(
    template_paths: Iterable[Path | str] | None = None,
    sheet_names: Iterable[str] | None = None,
    economies: Iterable[str] | None = None,
) -> pd.DataFrame:
    """Load LEAP-style template values into a long branch/variable/year table."""
    if not template_paths:
        return pd.DataFrame(
            columns=["economy", "scenario", "branch_path", "variable", "year", "value"]
        )

    rows: list[dict[str, object]] = []
    for template_path in template_paths:
        resolved_path = _resolve(template_path)
        if not resolved_path.exists():
            print(f"[WARN] Constraint workbook not found and will be skipped: {resolved_path}")
            continue
        target_economies = _infer_constraint_economies(resolved_path, economies)
        if not target_economies:
            continue
        try:
            workbook = pd.ExcelFile(resolved_path)
        except Exception as exc:
            print(f"[WARN] Failed to open constraint workbook {resolved_path}: {exc}")
            continue
        target_sheets = list(sheet_names) if sheet_names else list(workbook.sheet_names)
        for sheet_name in target_sheets:
            if str(sheet_name).strip().lower() in {"instructions", "for_viewing"}:
                continue
            if sheet_name not in workbook.sheet_names:
                continue
            try:
                sheet = _read_leap_template_sheet(resolved_path, sheet_name)
            except ValueError:
                continue
            except Exception as exc:
                print(
                    f"[WARN] Failed to read constraint sheet {resolved_path.name}::{sheet_name}: {exc}"
                )
                continue

            branch_column = (
                "Branch Path"
                if "Branch Path" in sheet.columns
                else ("Branch_Path" if "Branch_Path" in sheet.columns else None)
            )
            variable_column = "Variable" if "Variable" in sheet.columns else None
            scenario_column = "Scenario" if "Scenario" in sheet.columns else None
            if not branch_column or not variable_column or not scenario_column:
                continue

            year_columns = [
                str(column)
                for column in sheet.columns
                if str(column).isdigit()
            ]
            for _, row in sheet.iterrows():
                branch_path = str(row.get(branch_column) or "").strip()
                variable = str(row.get(variable_column) or "").strip()
                scenario = str(row.get(scenario_column) or "").strip()
                if not branch_path or not variable or not scenario:
                    continue

                year_values: dict[int, float] = {}
                if year_columns:
                    for column in year_columns:
                        numeric = pd.to_numeric(row.get(column), errors="coerce")
                        if pd.isna(numeric):
                            continue
                        year_values[int(column)] = float(numeric)
                elif "Expression" in sheet.columns:
                    year_values = _parse_data_expression(row.get("Expression"))

                if not year_values:
                    continue

                for economy in target_economies:
                    for year, value in year_values.items():
                        if year < BASE_YEAR or year > FINAL_YEAR:
                            continue
                        rows.append(
                            {
                                "economy": str(economy),
                                "scenario": scenario,
                                "branch_path": branch_path,
                                "variable": variable,
                                "year": int(year),
                                "value": float(value),
                            }
                        )

    if not rows:
        return pd.DataFrame(
            columns=["economy", "scenario", "branch_path", "variable", "year", "value"]
        )
    return pd.DataFrame(rows)


def _classify_supply_constraint_variable(variable: object) -> str | None:
    """Map a LEAP supply variable label to a recognized cap field."""
    text = str(variable or "").strip().lower()
    if not text or "unmet" in text:
        return None
    if "import" in text:
        return "max_imports"
    if "export" in text:
        return "max_exports"
    if any(token in text for token in ("production", "availability")):
        return "max_production"
    return None


def _classify_transformation_constraint_variable(
    branch_path: object,
    variable: object,
) -> str | None:
    """Map a LEAP transformation variable label to a recognized cap field."""
    branch_text = str(branch_path or "").strip().lower()
    variable_text = str(variable or "").strip().lower()
    if "\\output fuels\\" not in branch_text:
        return None
    if "import target" in variable_text or "export target" in variable_text:
        return None
    if "output" in variable_text:
        return "max_transformation_output"
    if "max" in variable_text and any(token in variable_text for token in ("production", "availability")):
        return "max_transformation_output"
    return None


def load_leap_constraint_tables(
    template_paths: Iterable[Path | str] | None = None,
    sheet_names: Iterable[str] | None = None,
    economies: Iterable[str] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Load optional supply and transformation caps from LEAP-style template workbooks."""
    value_table = _load_constraint_value_table(
        template_paths=template_paths,
        sheet_names=sheet_names,
        economies=economies,
    )
    empty_supply = pd.DataFrame(
        columns=[
            "economy",
            "scenario",
            "esto_product",
            "year",
            "max_imports",
            "max_exports",
            "max_production",
        ]
    )
    empty_transformation = pd.DataFrame(
        columns=[
            "economy",
            "scenario",
            "esto_product",
            "year",
            "max_transformation_output",
        ]
    )
    if value_table.empty:
        return empty_supply, empty_transformation

    label_to_product = _build_label_to_esto_product_lookup()

    def _lookup_product(label: object) -> str:
        token = str(label or "").strip()
        if not token:
            return ""
        return str(label_to_product.get(token) or label_to_product.get(token.lower()) or "")

    supply_rows: list[dict[str, object]] = []
    transformation_rows: list[dict[str, object]] = []
    for _, row in value_table.iterrows():
        branch_path = str(row.get("branch_path") or "").strip()
        variable = str(row.get("variable") or "").strip()
        branch_bits = [part.strip() for part in branch_path.split("\\") if str(part).strip()]
        if not branch_bits:
            continue
        branch_head = branch_bits[0].lower()
        fuel_label = branch_bits[-1]
        esto_product = _lookup_product(fuel_label)
        if not esto_product:
            continue

        if branch_head == "resources":
            constraint_field = _classify_supply_constraint_variable(variable)
            if constraint_field:
                supply_rows.append(
                    {
                        "economy": str(row["economy"]),
                        "scenario": str(row["scenario"]),
                        "esto_product": esto_product,
                        "year": int(row["year"]),
                        "constraint_field": constraint_field,
                        "value": max(float(row["value"]), 0.0),
                    }
                )
        elif branch_head == "transformation":
            constraint_field = _classify_transformation_constraint_variable(branch_path, variable)
            if constraint_field:
                transformation_rows.append(
                    {
                        "economy": str(row["economy"]),
                        "scenario": str(row["scenario"]),
                        "esto_product": esto_product,
                        "year": int(row["year"]),
                        "constraint_field": constraint_field,
                        "value": max(float(row["value"]), 0.0),
                    }
                )

    if not supply_rows:
        supply_constraints = empty_supply
    else:
        supply_constraints = (
            pd.DataFrame(supply_rows)
            .pivot_table(
                index=["economy", "scenario", "esto_product", "year"],
                columns="constraint_field",
                values="value",
                aggfunc="max",
            )
            .reset_index()
        )
        supply_constraints.columns.name = None
        for column in ["max_imports", "max_exports", "max_production"]:
            if column not in supply_constraints.columns:
                supply_constraints[column] = pd.NA
        supply_constraints = supply_constraints[
            ["economy", "scenario", "esto_product", "year", "max_imports", "max_exports", "max_production"]
        ]

    if not transformation_rows:
        transformation_constraints = empty_transformation
    else:
        transformation_constraints = (
            pd.DataFrame(transformation_rows)
            .pivot_table(
                index=["economy", "scenario", "esto_product", "year"],
                columns="constraint_field",
                values="value",
                aggfunc="max",
            )
            .reset_index()
        )
        transformation_constraints.columns.name = None
        if "max_transformation_output" not in transformation_constraints.columns:
            transformation_constraints["max_transformation_output"] = pd.NA
        transformation_constraints = transformation_constraints[
            ["economy", "scenario", "esto_product", "year", "max_transformation_output"]
        ]

    return supply_constraints, transformation_constraints


def load_results_demand_table(
    comparison_long_path: Path | str = COMPARISON_LONG_PATH,
    mapping_status_path: Path | str = MAPPING_STATUS_PATH,
    source_priority: tuple[str, ...] = DEMAND_SOURCE_PRIORITY,
    comparison_long_df: pd.DataFrame | None = None,
    mapping_status_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Aggregate mapped LEAP demand to ESTO products, with projection fallback."""
    sector_table = load_results_sector_demand_table(
        comparison_long_path=comparison_long_path,
        mapping_status_path=mapping_status_path,
        source_priority=source_priority,
        comparison_long_df=comparison_long_df,
        mapping_status_df=mapping_status_df,
    )
    if sector_table.empty:
        return pd.DataFrame(
            columns=["economy", "scenario", "esto_product", "year", "demand_value", "demand_source"]
        )
    source_counts = (
        sector_table.groupby(
            ["economy", "scenario", "esto_product", "year"],
            dropna=False,
            as_index=False,
        )["demand_source"]
        .nunique()
        .rename(columns={"demand_source": "source_count"})
    )
    grouped = (
        sector_table.groupby(
            ["economy", "scenario", "esto_product", "year"],
            dropna=False,
            as_index=False,
        )["demand_value"]
        .sum(min_count=1)
    )
    grouped = grouped.merge(
        source_counts,
        on=["economy", "scenario", "esto_product", "year"],
        how="left",
    )
    grouped["demand_source"] = grouped["source_count"].map(
        lambda count: "mixed" if pd.notna(count) and int(count) > 1 else "leap_or_projection"
    )
    return grouped[
        ["economy", "scenario", "esto_product", "year", "demand_value", "demand_source"]
    ].reset_index(drop=True)


def load_results_sector_demand_table(
    comparison_long_path: Path | str = COMPARISON_LONG_PATH,
    mapping_status_path: Path | str = MAPPING_STATUS_PATH,
    source_priority: tuple[str, ...] = DEMAND_SOURCE_PRIORITY,
    comparison_long_df: pd.DataFrame | None = None,
    mapping_status_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Return LEAP demand rows by sheet/sector and mapped ESTO product."""
    if comparison_long_df is None:
        comparison_path = _resolve(comparison_long_path)
        comparison_long = pd.read_csv(comparison_path)
    else:
        comparison_long = comparison_long_df.copy()

    if mapping_status_df is None:
        mapping_path = _resolve(mapping_status_path)
        mapping_status = pd.read_excel(mapping_path, sheet_name="mapping_status")
    else:
        mapping_status = mapping_status_df.copy()

    required_mapping_cols = ["sheet", "fuel_label", "esto_product", "sector_code_9th", "esto_flow"]
    missing_cols = [col for col in required_mapping_cols if col not in mapping_status.columns]
    if missing_cols:
        raise KeyError(f"mapping_status is missing required columns: {missing_cols}")

    merge_cols = ["sheet", "fuel_label"]
    if "measure" in comparison_long.columns and "measure" in mapping_status.columns:
        merge_cols.append("measure")
        required_mapping_cols = ["measure", *required_mapping_cols]

    mapping_subset = mapping_status[required_mapping_cols].copy()
    if "measure" in mapping_subset.columns:
        mapping_subset["measure"] = mapping_subset["measure"].fillna("").astype(str).str.strip()
    mapping_subset["sheet"] = mapping_subset["sheet"].astype(str)
    mapping_subset["fuel_label"] = mapping_subset["fuel_label"].astype(str)
    mapping_subset["esto_product"] = mapping_subset["esto_product"].fillna("").astype(str).str.strip()
    mapping_subset["sector_code_9th"] = mapping_subset["sector_code_9th"].fillna("").astype(str).str.strip()
    mapping_subset["esto_flow"] = mapping_subset["esto_flow"].fillna("").astype(str).str.strip()
    mapping_subset = mapping_subset[
        mapping_subset["sector_code_9th"].map(_is_demand_sector_mapping)
    ].copy()
    if mapping_subset.empty:
        return pd.DataFrame(
            columns=[
                "economy",
                "scenario",
                "sheet",
                "esto_product",
                "sector_code_9th",
                "esto_flow",
                "year",
                "demand_value",
                "demand_source",
            ]
        )
    mapping_subset = mapping_subset.drop_duplicates(
        subset=merge_cols,
        keep="first",
    )

    merged = comparison_long.merge(
        mapping_subset,
        on=merge_cols,
        how="left",
    )
    merged["source"] = merged["source"].astype(str).str.strip().str.lower()
    merged["esto_product"] = merged["esto_product"].fillna("").astype(str).str.strip()
    merged["value"] = pd.to_numeric(merged["value"], errors="coerce")
    merged["year"] = pd.to_numeric(merged["year"], errors="coerce").astype("Int64")
    merged = merged[
        merged["esto_product"].ne("")
        & merged["source"].isin(source_priority)
        & merged["year"].notna()
    ].copy()
    merged = merged[
        (merged["year"] >= BASE_YEAR)
        & (merged["year"] <= FINAL_YEAR)
    ].copy()

    grouped = (
        merged.groupby(
            ["economy", "scenario", "sheet", "esto_product", "sector_code_9th", "esto_flow", "year", "source"],
            dropna=False,
            as_index=False,
        )["value"]
        .sum(min_count=1)
    )
    if grouped.empty:
        return pd.DataFrame(
            columns=[
                "economy",
                "scenario",
                "sheet",
                "esto_product",
                "sector_code_9th",
                "esto_flow",
                "year",
                "demand_value",
                "demand_source",
            ]
        )

    wide = (
        grouped.pivot_table(
            index=["economy", "scenario", "sheet", "esto_product", "sector_code_9th", "esto_flow", "year"],
            columns="source",
            values="value",
            aggfunc="first",
        )
        .reset_index()
    )

    selections = wide.apply(
        lambda row: _pick_preferred_source(row, source_priority),
        axis=1,
        result_type="expand",
    )
    wide["demand_value"] = selections[0]
    wide["demand_source"] = selections[1]
    wide = wide[wide["demand_value"].notna()].copy()
    return wide[
        [
            "economy",
            "scenario",
            "sheet",
            "esto_product",
            "sector_code_9th",
            "esto_flow",
            "year",
            "demand_value",
            "demand_source",
        ]
    ].reset_index(drop=True)


def _iter_year_value_items(
    labeled_values: dict | None,
    base_year: int,
    final_year: int,
):
    """Yield (label, year, value) triples from a transformation record payload."""
    if not isinstance(labeled_values, dict):
        return
    for label, raw_value in labeled_values.items():
        year_map = supply_data_pipeline.coerce_value_by_year(raw_value, base_year, final_year)
        for year, value in year_map.items():
            year_int = int(year)
            if year_int < base_year or year_int > final_year:
                continue
            yield str(label), year_int, float(value)


def build_transformation_balance_table(
    economies: Iterable[str] | None = None,
    base_year: int = BASE_YEAR,
    final_year: int = FINAL_YEAR,
) -> pd.DataFrame:
    """Aggregate transformation+transfer output/input/loss values by ESTO product."""
    label_to_product = _build_label_to_esto_product_lookup()
    rows = _collect_transformation_and_transfer_rows(economies=economies)
    buckets: dict[tuple[str, str, int], dict[str, float]] = {}
    unmapped_labels: set[str] = set()

    def _accumulate(economy: str, product: str, year: int, field: str, value: float) -> None:
        key = (economy, product, year)
        bucket = buckets.setdefault(
            key,
            {
                "economy": economy,
                "esto_product": product,
                "year": year,
                "transformation_output": 0.0,
                "transformation_input": 0.0,
                "transformation_losses": 0.0,
            },
        )
        bucket[field] += float(value)

    for record in rows:
        economy = str(record.get("economy") or "").strip()
        if not economy:
            continue
        for label, year, value in _iter_year_value_items(record.get("output_values"), base_year, final_year):
            product = label_to_product.get(label) or label_to_product.get(label.lower())
            if not product:
                unmapped_labels.add(label)
                continue
            _accumulate(economy, product, year, "transformation_output", value)
        for label, year, value in _iter_year_value_items(record.get("feedstock_values"), base_year, final_year):
            product = label_to_product.get(label) or label_to_product.get(label.lower())
            if not product:
                unmapped_labels.add(label)
                continue
            _accumulate(economy, product, year, "transformation_input", abs(value))
        for label, year, value in _iter_year_value_items(record.get("loss_values"), base_year, final_year):
            product = label_to_product.get(label) or label_to_product.get(label.lower())
            if not product:
                unmapped_labels.add(label)
                continue
            _accumulate(economy, product, year, "transformation_losses", abs(value))

    if unmapped_labels:
        preview = ", ".join(sorted(unmapped_labels)[:10])
        print(
            "[WARN] Some transformation labels could not be mapped back to ESTO products "
            f"and were skipped: {preview}"
        )
    if not buckets:
        return pd.DataFrame(
            columns=[
                "economy",
                "esto_product",
                "year",
                "transformation_output",
                "transformation_input",
                "transformation_losses",
            ]
        )
    return pd.DataFrame(buckets.values()).sort_values(
        ["economy", "esto_product", "year"]
    ).reset_index(drop=True)


def build_transformation_sector_table(
    economies: Iterable[str] | None = None,
    base_year: int = BASE_YEAR,
    final_year: int = FINAL_YEAR,
) -> pd.DataFrame:
    """Aggregate transformation+transfer process rows into balance-style sector lines."""
    label_to_product = _build_label_to_esto_product_lookup()
    rows = _collect_transformation_and_transfer_rows(economies=economies)
    buckets: dict[tuple[str, str, int, str], float] = {}

    def _add(economy: str, scenario: str, year: int, product: str, value: float) -> None:
        key = (economy, scenario, year, product)
        buckets[key] = buckets.get(key, 0.0) + float(value)

    for record in rows:
        economy = str(record.get("economy") or "").strip()
        sector_name = _normalize_conventional_sector_name(
            record.get("sector_title") or record.get("process_name") or ""
        )
        if not economy or not sector_name:
            continue
        for label, year, value in _iter_year_value_items(record.get("output_values"), base_year, final_year):
            product = label_to_product.get(label) or label_to_product.get(label.lower())
            if product:
                _add(economy, sector_name, year, product, abs(value))
        for label, year, value in _iter_year_value_items(record.get("feedstock_values"), base_year, final_year):
            product = label_to_product.get(label) or label_to_product.get(label.lower())
            if product:
                _add(economy, sector_name, year, product, -abs(value))
        for label, year, value in _iter_year_value_items(record.get("loss_values"), base_year, final_year):
            product = label_to_product.get(label) or label_to_product.get(label.lower())
            if product:
                _add(economy, sector_name, year, product, -abs(value))

    if not buckets:
        return pd.DataFrame(
            columns=["economy", "sector", "year", "esto_product", "value"]
        )
    output_rows = [
        {
            "economy": economy,
            "sector": sector,
            "year": year,
            "esto_product": product,
            "value": value,
        }
        for (economy, sector, year, product), value in buckets.items()
    ]
    return pd.DataFrame(output_rows).sort_values(
        ["economy", "sector", "year", "esto_product"]
    ).reset_index(drop=True)


def prepare_projected_supply_table(
    economies: Iterable[str] | None = None,
    dataset_key: str = EXPORT_DATASET_KEY,
) -> tuple[pd.DataFrame, tuple]:
    """Build the existing supply projection table by ESTO product/year."""
    output_columns = [
        "economy",
        "esto_product",
        "year",
        "projected_imports",
        "projected_exports",
        "projected_net_imports",
    ]
    assets = supply_data_pipeline.prepare_supply_assets(economies=economies)
    dataset_map, sector_config, code_to_name_mapping, _, _ = assets
    data, year_cols = supply_data_pipeline.resolve_dataset(dataset_map, dataset_key)
    flow_codes = supply_data_pipeline.FLOW_CODES_BY_DATASET.get(dataset_key)
    if not flow_codes:
        raise KeyError(f"Unknown supply dataset key: {dataset_key}")

    economy_list = workflow_common.normalize_economies(
        economies or supply_data_pipeline.ECONOMIES_TO_ANALYZE
    )
    if not economy_list:
        economy_list = supply_data_pipeline.get_economy_list(data, None)

    rows: list[dict[str, object]] = []
    for economy in economy_list:
        for fuel_key, entry in sorted(sector_config.items()):
            imports_by_year = supply_data_pipeline.build_supply_value_by_year(
                data,
                year_cols,
                economy,
                entry,
                "imports",
                flow_codes.get("imports"),
                BASE_YEAR,
                FINAL_YEAR,
                projection_lookup=supply_data_pipeline.SUPPLY_PROJECTION_LOOKUP,
                projection_years=supply_data_pipeline.PROJECTION_YEAR_RANGE,
                code_to_name_mapping=code_to_name_mapping,
            )
            exports_by_year = supply_data_pipeline.build_supply_value_by_year(
                data,
                year_cols,
                economy,
                entry,
                "exports",
                flow_codes.get("exports"),
                BASE_YEAR,
                FINAL_YEAR,
                projection_lookup=supply_data_pipeline.SUPPLY_PROJECTION_LOOKUP,
                projection_years=supply_data_pipeline.PROJECTION_YEAR_RANGE,
                code_to_name_mapping=code_to_name_mapping,
            )
            for year in range(BASE_YEAR, FINAL_YEAR + 1):
                imports_value = float(imports_by_year.get(year, 0.0))
                exports_value = float(exports_by_year.get(year, 0.0))
                rows.append(
                    {
                        "economy": economy,
                        "esto_product": fuel_key,
                        "year": year,
                        "projected_imports": imports_value,
                        "projected_exports": exports_value,
                        "projected_net_imports": imports_value - exports_value,
                    }
                )
    supply_projection = pd.DataFrame(rows, columns=output_columns)
    return supply_projection, assets


def prepare_supply_primary_table(
    assets: tuple,
    economies: Iterable[str] | None = None,
    dataset_key: str = EXPORT_DATASET_KEY,
) -> pd.DataFrame:
    """Build production and stock-change rows by fuel/year from the supply dataset."""
    output_columns = [
        "economy",
        "year",
        "esto_product",
        "production",
        "stock_changes",
    ]
    dataset_map, sector_config, code_to_name_mapping, _, _ = assets
    data, year_cols = supply_data_pipeline.resolve_dataset(dataset_map, dataset_key)
    flow_codes = supply_data_pipeline.FLOW_CODES_BY_DATASET.get(dataset_key)
    if not flow_codes:
        raise KeyError(f"Unknown supply dataset key: {dataset_key}")

    economy_list = workflow_common.normalize_economies(
        economies or supply_data_pipeline.ECONOMIES_TO_ANALYZE
    )
    if not economy_list:
        economy_list = supply_data_pipeline.get_economy_list(data, None)

    rows: list[dict[str, object]] = []
    for economy in economy_list:
        for fuel_key, entry in sorted(sector_config.items()):
            production_by_year = supply_data_pipeline.build_supply_value_by_year(
                data,
                year_cols,
                economy,
                entry,
                "production",
                flow_codes.get("production"),
                BASE_YEAR,
                FINAL_YEAR,
                projection_lookup=supply_data_pipeline.SUPPLY_PROJECTION_LOOKUP,
                projection_years=supply_data_pipeline.PROJECTION_YEAR_RANGE,
                code_to_name_mapping=code_to_name_mapping,
            )
            stock_changes_by_year = supply_data_pipeline.build_supply_value_by_year(
                data,
                year_cols,
                economy,
                entry,
                "stock_changes",
                flow_codes.get("stock_changes"),
                BASE_YEAR,
                FINAL_YEAR,
                projection_lookup=supply_data_pipeline.SUPPLY_PROJECTION_LOOKUP,
                projection_years=supply_data_pipeline.PROJECTION_YEAR_RANGE,
                code_to_name_mapping=code_to_name_mapping,
            )
            for year in range(BASE_YEAR, FINAL_YEAR + 1):
                rows.append(
                    {
                        "economy": economy,
                        "year": year,
                        "esto_product": fuel_key,
                        "production": float(production_by_year.get(year, 0.0)),
                        "stock_changes": float(stock_changes_by_year.get(year, 0.0)),
                    }
                )
    return pd.DataFrame(rows, columns=output_columns)


def build_reconciliation_table(
    demand_table: pd.DataFrame,
    transformation_table: pd.DataFrame,
    supply_projection_table: pd.DataFrame,
    supply_primary_table: pd.DataFrame | None = None,
    supply_constraints: pd.DataFrame | None = None,
    transformation_constraints: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Combine demand, transformation, and supply into a trade-adjustment table."""
    key_columns = ["economy", "scenario", "esto_product", "year"]
    scenario_values: list[str] = []
    if isinstance(demand_table, pd.DataFrame) and not demand_table.empty and "scenario" in demand_table.columns:
        scenario_values.extend(
            str(value).strip()
            for value in demand_table["scenario"].dropna().astype(str).tolist()
            if str(value).strip()
        )
    for constraint_df in (supply_constraints, transformation_constraints):
        if isinstance(constraint_df, pd.DataFrame) and not constraint_df.empty and "scenario" in constraint_df.columns:
            scenario_values.extend(
                str(value).strip()
                for value in constraint_df["scenario"].dropna().astype(str).tolist()
                if str(value).strip()
            )
    scenario_values = sorted(dict.fromkeys(scenario_values))
    if not scenario_values:
        scenario_values = ["Reference"]

    key_frames: list[pd.DataFrame] = []
    if isinstance(demand_table, pd.DataFrame) and not demand_table.empty:
        key_frames.append(
            demand_table[["economy", "scenario", "esto_product", "year"]].copy()
        )

    def _expand_non_scenario_keys(table: pd.DataFrame | None, table_name: str) -> None:
        if not isinstance(table, pd.DataFrame) or table.empty:
            return
        required = ["economy", "esto_product", "year"]
        missing = [column for column in required if column not in table.columns]
        if missing:
            raise KeyError(
                f"{table_name} is missing required reconciliation columns: {missing}"
            )
        base = table[["economy", "esto_product", "year"]].drop_duplicates().copy()
        if base.empty:
            return
        scenario_df = pd.DataFrame({"scenario": scenario_values})
        base["__tmp_key"] = 1
        scenario_df["__tmp_key"] = 1
        expanded = (
            base.merge(scenario_df, on="__tmp_key", how="inner")
            .drop(columns=["__tmp_key"])
            .loc[:, key_columns]
        )
        key_frames.append(expanded)

    _expand_non_scenario_keys(transformation_table, "transformation_table")
    _expand_non_scenario_keys(supply_projection_table, "supply_projection_table")
    _expand_non_scenario_keys(supply_primary_table, "supply_primary_table")

    if not key_frames:
        return pd.DataFrame(columns=key_columns)

    merged = (
        pd.concat(key_frames, ignore_index=True)
        .drop_duplicates(subset=key_columns, keep="first")
        .reset_index(drop=True)
    )
    if isinstance(demand_table, pd.DataFrame) and not demand_table.empty:
        demand_cols = ["economy", "scenario", "esto_product", "year", "demand_value", "demand_source"]
        demand_merge = demand_table.reindex(columns=demand_cols).copy()
    else:
        demand_merge = pd.DataFrame(columns=["economy", "scenario", "esto_product", "year", "demand_value", "demand_source"])
    merged = merged.merge(
        demand_merge,
        on=["economy", "scenario", "esto_product", "year"],
        how="left",
    )
    merged = merged.merge(
        transformation_table,
        on=["economy", "esto_product", "year"],
        how="left",
    ).merge(
        supply_projection_table,
        on=["economy", "esto_product", "year"],
        how="left",
    )
    if isinstance(supply_primary_table, pd.DataFrame) and not supply_primary_table.empty:
        merged = merged.merge(
            supply_primary_table,
            on=["economy", "esto_product", "year"],
            how="left",
        )
    if isinstance(supply_constraints, pd.DataFrame) and not supply_constraints.empty:
        merged = merged.merge(
            supply_constraints,
            on=["economy", "scenario", "esto_product", "year"],
            how="left",
        )
    if isinstance(transformation_constraints, pd.DataFrame) and not transformation_constraints.empty:
        merged = merged.merge(
            transformation_constraints,
            on=["economy", "scenario", "esto_product", "year"],
            how="left",
        )
    if "demand_source" not in merged.columns:
        merged["demand_source"] = "none"
    merged["demand_source"] = merged["demand_source"].fillna("none").astype(str)

    for column in [
        "demand_value",
        "transformation_output",
        "transformation_input",
        "transformation_losses",
        "projected_imports",
        "projected_exports",
        "projected_net_imports",
        "production",
        "stock_changes",
    ]:
        if column not in merged.columns:
            merged[column] = 0.0
        merged[column] = pd.to_numeric(merged[column], errors="coerce").fillna(0.0)
    for column in [
        "max_imports",
        "max_exports",
        "max_production",
        "max_transformation_output",
    ]:
        if column not in merged.columns:
            merged[column] = pd.NA
        merged[column] = pd.to_numeric(merged[column], errors="coerce")

    max_transformation_output = merged["max_transformation_output"].where(
        merged["max_transformation_output"].notna(),
        float("inf"),
    )
    max_production = merged["max_production"].where(
        merged["max_production"].notna(),
        float("inf"),
    )
    merged["constrained_transformation_output"] = merged["transformation_output"].clip(lower=0.0).where(
        merged["transformation_output"] <= max_transformation_output,
        max_transformation_output,
    )
    merged["constrained_production"] = merged["production"].clip(lower=0.0).where(
        merged["production"] <= max_production,
        max_production,
    )

    merged["required_net_imports"] = (
        pd.to_numeric(merged["demand_value"], errors="coerce").fillna(0.0)
        + merged["transformation_input"]
        + merged["transformation_losses"]
        - merged["constrained_transformation_output"]
        - merged["constrained_production"]
        - merged["stock_changes"]
    )
    merged["trade_adjustment"] = (
        merged["required_net_imports"] - merged["projected_net_imports"]
    )
    merged["uncapped_adjusted_imports"] = (
        merged["projected_imports"] + merged["trade_adjustment"].clip(lower=0.0)
    )
    merged["uncapped_adjusted_exports"] = (
        merged["projected_exports"] + (-merged["trade_adjustment"]).clip(lower=0.0)
    )
    max_imports = merged["max_imports"].where(merged["max_imports"].notna(), float("inf"))
    max_exports = merged["max_exports"].where(merged["max_exports"].notna(), float("inf"))
    merged["adjusted_imports"] = merged["uncapped_adjusted_imports"].clip(lower=0.0).where(
        merged["uncapped_adjusted_imports"] <= max_imports,
        max_imports,
    )
    merged["adjusted_exports"] = merged["uncapped_adjusted_exports"].clip(lower=0.0).where(
        merged["uncapped_adjusted_exports"] <= max_exports,
        max_exports,
    )
    merged["imports_cap_binding"] = (
        merged["uncapped_adjusted_imports"] - merged["adjusted_imports"]
    ).clip(lower=0.0)
    merged["exports_cap_binding"] = (
        merged["uncapped_adjusted_exports"] - merged["adjusted_exports"]
    ).clip(lower=0.0)
    merged["adjusted_net_imports"] = (
        merged["adjusted_imports"] - merged["adjusted_exports"]
    )
    merged["adjusted_balance"] = (
        merged["adjusted_net_imports"]
        + merged["constrained_transformation_output"]
        + merged["constrained_production"]
        + merged["stock_changes"]
        - merged["transformation_input"]
        - merged["transformation_losses"]
        - pd.to_numeric(merged["demand_value"], errors="coerce").fillna(0.0)
    )
    return merged.sort_values(
        ["economy", "scenario", "esto_product", "year"]
    ).reset_index(drop=True)


def build_transformation_trade_target_rows(
    economies: Iterable[str] | None = None,
    base_year: int = BASE_YEAR,
    final_year: int = FINAL_YEAR,
    process_records: list[dict] | None = None,
) -> tuple[pd.DataFrame, list[dict]]:
    """Return process-level transformation import/export target rows."""
    label_to_product = _build_label_to_esto_product_lookup()
    records = process_records if process_records is not None else transformation_workflow.collect_transformation_rows(economies=economies)
    rows: list[dict[str, object]] = []
    for record_index, record in enumerate(records):
        economy = str(record.get("economy") or "").strip()
        sector_title = str(record.get("sector_title") or "").strip()
        process_name = str(record.get("process_name") or "").strip()
        if not economy:
            continue
        for label, year, value in _iter_year_value_items(
            record.get("output_import_targets"),
            base_year,
            final_year,
        ):
            product = label_to_product.get(label) or label_to_product.get(label.lower())
            if not product:
                continue
            rows.append(
                {
                    "record_index": int(record_index),
                    "economy": economy,
                    "sector_title": sector_title,
                    "process_name": process_name,
                    "direction": "import",
                    "label": str(label),
                    "esto_product": str(product),
                    "year": int(year),
                    "value": max(float(value), 0.0),
                }
            )
        for label, year, value in _iter_year_value_items(
            record.get("output_export_targets"),
            base_year,
            final_year,
        ):
            product = label_to_product.get(label) or label_to_product.get(label.lower())
            if not product:
                continue
            rows.append(
                {
                    "record_index": int(record_index),
                    "economy": economy,
                    "sector_title": sector_title,
                    "process_name": process_name,
                    "direction": "export",
                    "label": str(label),
                    "esto_product": str(product),
                    "year": int(year),
                    "value": max(float(value), 0.0),
                }
            )
    if not rows:
        return pd.DataFrame(
            columns=[
                "record_index",
                "economy",
                "sector_title",
                "process_name",
                "direction",
                "label",
                "esto_product",
                "year",
                "value",
            ]
        ), records
    return pd.DataFrame(rows), records


def apply_trade_split_between_transformation_and_supply(
    reconciliation_table: pd.DataFrame,
    transformation_target_rows: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Split gross imports/exports into transformation targets and supply residuals."""
    if reconciliation_table.empty:
        return reconciliation_table.copy()
    merged = reconciliation_table.copy()
    for required in ["projected_imports", "projected_exports", "adjusted_imports", "adjusted_exports"]:
        if required not in merged.columns:
            merged[required] = 0.0
        merged[required] = pd.to_numeric(merged[required], errors="coerce").fillna(0.0)

    if isinstance(transformation_target_rows, pd.DataFrame) and not transformation_target_rows.empty:
        totals = (
            transformation_target_rows.groupby(
                ["economy", "esto_product", "year", "direction"],
                dropna=False,
                as_index=False,
            )["value"]
            .sum(min_count=1)
        )
        import_totals = totals[totals["direction"] == "import"].rename(
            columns={"value": "baseline_transformation_import_target"}
        )[["economy", "esto_product", "year", "baseline_transformation_import_target"]]
        export_totals = totals[totals["direction"] == "export"].rename(
            columns={"value": "baseline_transformation_export_target"}
        )[["economy", "esto_product", "year", "baseline_transformation_export_target"]]
        merged = merged.merge(
            import_totals,
            on=["economy", "esto_product", "year"],
            how="left",
        ).merge(
            export_totals,
            on=["economy", "esto_product", "year"],
            how="left",
        )
    else:
        merged["baseline_transformation_import_target"] = 0.0
        merged["baseline_transformation_export_target"] = 0.0

    for column in [
        "baseline_transformation_import_target",
        "baseline_transformation_export_target",
    ]:
        merged[column] = pd.to_numeric(merged.get(column), errors="coerce").fillna(0.0)

    projected_imports = merged["projected_imports"].clip(lower=0.0)
    projected_exports = merged["projected_exports"].clip(lower=0.0)
    import_share = (
        merged["baseline_transformation_import_target"] / projected_imports.where(projected_imports > 0.0, pd.NA)
    ).fillna(0.0).clip(lower=0.0, upper=1.0)
    export_share = (
        merged["baseline_transformation_export_target"] / projected_exports.where(projected_exports > 0.0, pd.NA)
    ).fillna(0.0).clip(lower=0.0, upper=1.0)

    merged["transformation_import_share"] = import_share
    merged["transformation_export_share"] = export_share
    merged["transformation_import_target"] = (
        merged["adjusted_imports"].clip(lower=0.0) * merged["transformation_import_share"]
    ).clip(lower=0.0)
    merged["transformation_export_target"] = (
        merged["adjusted_exports"].clip(lower=0.0) * merged["transformation_export_share"]
    ).clip(lower=0.0)
    merged["supply_imports_residual"] = (
        merged["adjusted_imports"] - merged["transformation_import_target"]
    ).clip(lower=0.0)
    merged["supply_exports_residual"] = (
        merged["adjusted_exports"] - merged["transformation_export_target"]
    ).clip(lower=0.0)
    merged["combined_net_imports_after_split"] = (
        merged["supply_imports_residual"]
        + merged["transformation_import_target"]
        - merged["supply_exports_residual"]
        - merged["transformation_export_target"]
    )
    return merged


def build_supply_overrides(reconciliation_table: pd.DataFrame) -> dict[str, dict[str, dict[str, dict[str, dict[int, float]]]]]:
    """Convert the reconciliation table into supply override payloads."""
    overrides: dict[str, dict[str, dict[str, dict[str, dict[int, float]]]]] = {}
    if reconciliation_table.empty:
        return overrides
    use_legacy_split = _use_legacy_trade_split_mode()
    use_output_share_supply_exports = _use_output_share_supply_exports_mode()
    use_capacity_unmet_iterative = _use_capacity_unmet_iterative_mode()
    use_capacity_unmet_balanced = _use_capacity_unmet_iterative_balanced_mode()
    balanced_first_clean_mode = (
        use_capacity_unmet_balanced and _is_capacity_unmet_first_clean_run_mode()
    )
    for _, row in reconciliation_table.iterrows():
        economy = str(row["economy"])
        scenario = str(row["scenario"])
        product = str(row["esto_product"])
        year = int(row["year"])
        product_bucket = (
            overrides.setdefault(economy, {})
            .setdefault(scenario, {})
            .setdefault(product, {"imports": {}, "exports": {}})
        )
        if use_legacy_split:
            imports_value = row.get("supply_imports_residual", row.get("adjusted_imports", 0.0))
            exports_value = row.get("supply_exports_residual", row.get("adjusted_exports", 0.0))
        elif use_output_share_supply_exports or use_capacity_unmet_iterative:
            # Keep explicit exports on supply branches to align with trade projections,
            # while leaving imports at zero so LEAP can auto-balance imports.
            imports_value = 0.0
            exports_value = row.get("adjusted_exports", row.get("projected_exports", 0.0))
        elif use_capacity_unmet_balanced:
            # Always keep imports at zero in iterative-balanced mode.
            # first_clean/consecutive differences only affect export adjustments and
            # whether runtime residual allocations are applied.
            imports_value = 0.0
            if CAPACITY_UNMET_PIN_EXPORTS_TO_9TH_PROJECTIONS:
                # Keep exports anchored to 9th trade projections in iterative-balanced mode.
                exports_value = row.get("projected_exports", 0.0)
            else:
                exports_value = row.get("adjusted_exports", row.get("projected_exports", 0.0))
                if not balanced_first_clean_mode:
                    exports_value = float(exports_value) + _lookup_runtime_export_adjustment(
                        economy=economy,
                        scenario=scenario,
                        esto_product=product,
                        year=year,
                    )
        else:
            # Capacity-constrained mode (and other non-legacy modes) writes zeros
            # so stale LEAP trade values are explicitly cleared during import.
            imports_value = 0.0
            exports_value = 0.0
        product_bucket["imports"][year] = max(float(imports_value), 0.0)
        product_bucket["exports"][year] = max(float(exports_value), 0.0)
        if use_capacity_unmet_balanced:
            primary_add = 0.0
            if not balanced_first_clean_mode:
                primary_add = _lookup_runtime_primary_addition(
                    economy=economy,
                    scenario=scenario,
                    esto_product=product,
                    year=year,
                )
            base_production = pd.to_numeric(row.get("constrained_production"), errors="coerce")
            base_production_value = 0.0 if pd.isna(base_production) else max(float(base_production), 0.0)
            production_target = max(base_production_value + float(primary_add), 0.0)
            max_production_value = pd.to_numeric(row.get("max_production"), errors="coerce")
            if pd.isna(max_production_value):
                max_production_target = production_target
            else:
                max_production_target = max(float(max_production_value), production_target)
            product_bucket.setdefault("max_production", {})
            product_bucket["max_production"][year] = float(max_production_target)
    return overrides


def _build_capacity_process_catalog(
    process_records: list[dict],
) -> tuple[pd.DataFrame, list[str]]:
    """Build per-process output/yield rows keyed by economy/product/year."""
    if not process_records:
        return pd.DataFrame(), []
    label_to_product = _build_label_to_esto_product_lookup()
    rows: list[dict[str, object]] = []
    unmapped_labels: set[str] = set()
    instance_counter: dict[tuple[str, str, str], int] = {}
    for record_index, record in enumerate(process_records):
        economy = str(record.get("economy") or "").strip()
        module = str(record.get("sector_title") or "").strip() or "__unknown_module__"
        process = str(record.get("process_name") or "").strip() or "__unknown_process__"
        if not economy:
            continue
        counter_key = (_state_token(economy), _state_token(module), _state_token(process))
        instance_counter[counter_key] = int(instance_counter.get(counter_key, 0)) + 1
        instance = int(instance_counter[counter_key])

        product_output_by_year: dict[tuple[str, int], float] = {}
        total_output_by_year: dict[int, float] = {}
        for label, year, value in _iter_year_value_items(
            record.get("output_values"),
            BASE_YEAR,
            FINAL_YEAR,
        ):
            numeric = max(float(value), 0.0)
            if numeric <= 0.0:
                continue
            product = (
                label_to_product.get(label)
                or label_to_product.get(label.lower())
                or label_to_product.get(_normalize_label_for_lookup(label))
            )
            if not product:
                unmapped_labels.add(str(label))
                continue
            product_key = str(product)
            product_output_by_year[(product_key, int(year))] = (
                product_output_by_year.get((product_key, int(year)), 0.0) + numeric
            )
            total_output_by_year[int(year)] = total_output_by_year.get(int(year), 0.0) + numeric

        for (product_key, year), product_output in product_output_by_year.items():
            total_output = total_output_by_year.get(int(year), 0.0)
            if total_output <= 0.0:
                continue
            output_yield = float(product_output) / float(total_output)
            if output_yield <= 0.0:
                continue
            rows.append(
                {
                    "record_index": int(record_index),
                    "economy": economy,
                    "module": module,
                    "process": process,
                    "instance": int(instance),
                    "esto_product": product_key,
                    "year": int(year),
                    "product_output": float(product_output),
                    "module_total_output": float(total_output),
                    "yield": float(output_yield),
                }
            )

    if not rows:
        return pd.DataFrame(), sorted(unmapped_labels)
    catalog = pd.DataFrame(rows).sort_values(
        ["economy", "module", "process", "instance", "esto_product", "year"]
    ).reset_index(drop=True)
    return catalog, sorted(unmapped_labels)


def _resolve_capacity_priority_modules(esto_product: str) -> list[str]:
    """Return ordered priority module names configured for one ESTO product."""
    candidates = [
        CAPACITY_UNMET_PRIORITY_BY_PRODUCT.get(str(esto_product)),
        CAPACITY_UNMET_PRIORITY_BY_PRODUCT.get(str(esto_product).lower()),
        CAPACITY_UNMET_PRIORITY_BY_PRODUCT.get(_normalize_esto_product_for_match(esto_product)),
    ]
    for item in candidates:
        if isinstance(item, list) and item:
            return [str(value).strip() for value in item if str(value or "").strip()]
    return []


def _rank_capacity_candidates(
    candidate_rows: pd.DataFrame,
    esto_product: str,
) -> list[dict[str, object]]:
    """Return ranked candidate process rows for one fuel/year."""
    if candidate_rows.empty:
        return []
    ordered: list[dict[str, object]] = []
    remaining = candidate_rows.copy()
    remaining["module_key"] = remaining["module"].astype(str).str.strip().str.lower()
    priority_modules = _resolve_capacity_priority_modules(esto_product)
    for module_name in priority_modules:
        module_key = str(module_name).strip().lower()
        if not module_key:
            continue
        matched = remaining[remaining["module_key"] == module_key].copy()
        if matched.empty:
            continue
        matched = matched.sort_values(
            ["product_output", "module_total_output"],
            ascending=False,
        )
        ordered.extend(matched.to_dict("records"))
        remaining = remaining[remaining["module_key"] != module_key].copy()
    if not remaining.empty:
        remaining = remaining.sort_values(
            ["product_output", "module_total_output"],
            ascending=False,
        )
        ordered.extend(remaining.to_dict("records"))
    return ordered


def _collect_observed_trade_from_supply_results(
    *,
    scenario_pairs: list[tuple[str, str]],
    label_to_product: dict[str, str],
    results_dir: Path | str | Iterable[Path | str],
    include_exports: bool,
) -> tuple[pd.DataFrame, dict[str, object], list[dict[str, object]]]:
    """Collect observed imports/exports from the current run's balance-table CSVs."""
    return _collect_observed_trade_from_balance_tables(
        scenario_pairs=scenario_pairs,
        results_dir=results_dir,
        include_exports=include_exports,
    )


def _run_capacity_unmet_iterative_pass(
    *,
    reconciliation_table: pd.DataFrame,
    process_records: list[dict],
    economies: Iterable[str],
    scenarios: Iterable[str],
    results_dir: Path | str | Iterable[Path | str] = CAPACITY_UNMET_RESULTS_DIR,
    state_path: Path | str = CAPACITY_UNMET_STATE_PATH,
    allow_same_results_reuse: bool = CAPACITY_UNMET_ALLOW_SAME_RESULTS_REUSE,
) -> dict[str, object]:
    """Compute one manual unmet-capacity pass and persist cumulative state."""
    global _CAPACITY_UNMET_RUNTIME_CAPACITY_ADDITIONS
    if reconciliation_table.empty:
        raise ValueError("Cannot run capacity_unmet_iterative with empty reconciliation table.")

    process_catalog, unmapped_process_labels = _build_capacity_process_catalog(process_records)
    if process_catalog.empty:
        raise ValueError(
            "capacity_unmet_iterative mode requires transformation process output rows "
            "to infer fuel yields; none were found."
        )
    if unmapped_process_labels:
        preview = ", ".join(unmapped_process_labels[:12])
        print(
            "[WARN] Some transformation output labels could not be mapped to ESTO products "
            f"for capacity_unmet_iterative: {preview}"
        )

    run_mode = _resolve_capacity_unmet_iteration_run_mode()
    state = _read_capacity_unmet_state(state_path=state_path, run_mode=run_mode)
    cumulative_capacity_map = _parse_runtime_capacity_additions_from_state(
        state.get("cumulative_capacity_additions")
    )
    cumulative_output_map = _parse_runtime_capacity_additions_from_state(
        state.get("cumulative_output_additions")
    )
    module_baseline_output_lookup = _build_module_baseline_output_lookup(process_catalog)
    module_added_output_lookup = _build_module_added_output_lookup(cumulative_capacity_map)
    last_signatures = state.get("last_results_signatures")
    if not isinstance(last_signatures, dict):
        last_signatures = {}

    reconciliation = reconciliation_table.copy()
    reconciliation["scenario_key"] = (
        reconciliation["scenario"].astype(str).str.strip().str.lower()
    )
    reconciliation["adjusted_imports"] = pd.to_numeric(
        reconciliation.get("adjusted_imports"), errors="coerce"
    ).fillna(0.0)
    reconciliation["max_transformation_output"] = pd.to_numeric(
        reconciliation.get("max_transformation_output"), errors="coerce"
    )
    reconciliation["constrained_transformation_output"] = pd.to_numeric(
        reconciliation.get("constrained_transformation_output"), errors="coerce"
    ).fillna(0.0)

    scenario_pairs: list[tuple[str, str]] = []
    seen_pairs: set[tuple[str, str]] = set()
    for economy in [str(item).strip() for item in economies if str(item).strip()]:
        for scenario in [str(item).strip() for item in scenarios if str(item).strip()]:
            scenario_key = _state_token(
                _resolve_reconciliation_scenario_key(reconciliation_table, scenario)
            )
            pair = (str(economy), scenario_key)
            if pair in seen_pairs:
                continue
            seen_pairs.add(pair)
            scenario_pairs.append(pair)
    if not scenario_pairs:
        raise ValueError("capacity_unmet_iterative mode needs at least one economy/scenario pair.")

    label_to_product = _build_label_to_esto_product_lookup()
    observed_trade, signature_map, unmatched_result_fuels = _collect_observed_trade_from_supply_results(
        scenario_pairs=scenario_pairs,
        label_to_product=label_to_product,
        results_dir=results_dir,
        include_exports=False,
    )

    if (
        not allow_same_results_reuse
        and signature_map
        and last_signatures
        and signature_map == last_signatures
    ):
        print(
            "[WARN] capacity_unmet_iterative mode detected no new LEAP results artifacts since the "
            "previous pass. Continuing with reused results artifacts. "
            "Import workbook into LEAP, recalculate, refresh results tables, then rerun "
            "to use fresh results."
        )

    observed_imports = observed_trade
    if observed_imports.empty:
        raise ValueError(
            "capacity_unmet_iterative mode could not parse any observed imports from supply "
            f"results sheets {CAPACITY_UNMET_IMPORT_SHEETS} in '{_resolve(results_dir)}'."
        )

    requested_scenarios = {scenario for _, scenario in scenario_pairs}
    requested_economies = {economy for economy, _ in scenario_pairs}
    baseline_imports = reconciliation[
        reconciliation["economy"].astype(str).isin(requested_economies)
        & reconciliation["scenario_key"].astype(str).isin(requested_scenarios)
    ][
        [
            "economy",
            "scenario_key",
            "esto_product",
            "year",
            "adjusted_imports",
            "max_transformation_output",
            "constrained_transformation_output",
        ]
    ].copy()
    baseline_imports = baseline_imports.rename(columns={"scenario_key": "scenario"})
    if baseline_imports.empty:
        raise ValueError(
            "capacity_unmet_iterative mode found no reconciliation rows for run economy/scenario scope."
        )

    unmet_table = baseline_imports.merge(
        observed_imports,
        on=["economy", "scenario", "esto_product", "year"],
        how="outer",
    )
    unmet_table["adjusted_imports"] = pd.to_numeric(
        unmet_table.get("adjusted_imports"), errors="coerce"
    ).fillna(0.0)
    unmet_table["observed_imports"] = pd.to_numeric(
        unmet_table.get("observed_imports"), errors="coerce"
    ).fillna(0.0)
    unmet_table["max_transformation_output"] = pd.to_numeric(
        unmet_table.get("max_transformation_output"), errors="coerce"
    )
    unmet_table["constrained_transformation_output"] = pd.to_numeric(
        unmet_table.get("constrained_transformation_output"), errors="coerce"
    ).fillna(0.0)
    unmet_table["unmet_proxy"] = (
        unmet_table["observed_imports"] - unmet_table["adjusted_imports"]
    ).clip(lower=0.0)

    allocation_rows: list[dict[str, object]] = []
    clipping_rows: list[dict[str, object]] = []
    unresolved_rows: list[dict[str, object]] = []
    pass_capacity_additions: dict[str, float] = {}
    pass_output_additions: dict[str, float] = {}

    unmet_candidates = unmet_table[unmet_table["unmet_proxy"] > 0.0].copy()
    unmet_candidates = unmet_candidates.sort_values(
        ["economy", "scenario", "esto_product", "year"]
    )
    for _, row in unmet_candidates.iterrows():
        economy = str(row.get("economy") or "").strip()
        scenario_key = str(row.get("scenario") or "").strip().lower()
        esto_product = str(row.get("esto_product") or "").strip()
        year = int(pd.to_numeric(row.get("year"), errors="coerce"))
        unmet_value = max(float(row.get("unmet_proxy", 0.0)), 0.0)
        if not economy or not scenario_key or not esto_product or unmet_value <= 0.0:
            continue

        output_state_key = _output_addition_state_key(
            economy=economy,
            scenario=scenario_key,
            esto_product=esto_product,
            year=year,
        )
        prior_added_output = float(cumulative_output_map.get(output_state_key, 0.0))
        cap_value = pd.to_numeric(row.get("max_transformation_output"), errors="coerce")
        constrained_value = max(float(row.get("constrained_transformation_output", 0.0)), 0.0)
        if pd.isna(cap_value):
            headroom = float("inf")
        else:
            headroom = max(float(cap_value) - constrained_value - prior_added_output, 0.0)

        if headroom <= 0.0:
            clipping_rows.append(
                {
                    "economy": economy,
                    "scenario": scenario_key,
                    "esto_product": esto_product,
                    "year": int(year),
                    "requested_output_uplift": float(unmet_value),
                    "allocated_output_uplift": 0.0,
                    "clipped_output_uplift": float(unmet_value),
                    "reason": "No remaining cap headroom after constrained output + prior additions.",
                }
            )
            continue

        requested_output = float(unmet_value)
        allocatable_output = min(requested_output, headroom)
        if allocatable_output < requested_output:
            clipping_rows.append(
                {
                    "economy": economy,
                    "scenario": scenario_key,
                    "esto_product": esto_product,
                    "year": int(year),
                    "requested_output_uplift": float(requested_output),
                    "allocated_output_uplift": float(allocatable_output),
                    "clipped_output_uplift": float(requested_output - allocatable_output),
                    "reason": "Requested uplift exceeded max_transformation_output headroom.",
                }
            )
        if allocatable_output <= 0.0:
            continue

        candidates = process_catalog[
            (process_catalog["economy"].astype(str) == economy)
            & (process_catalog["esto_product"].astype(str) == esto_product)
            & (process_catalog["year"].astype(int) == int(year))
        ].copy()
        if candidates.empty:
            unresolved_rows.append(
                {
                    "economy": economy,
                    "scenario": scenario_key,
                    "esto_product": esto_product,
                    "year": int(year),
                    "unresolved_output_uplift": float(allocatable_output),
                    "reason": "No eligible transformation process outputs this fuel in this year.",
                }
            )
            continue

        ranked = _rank_capacity_candidates(candidates, esto_product)
        remaining_output = float(allocatable_output)
        for candidate in ranked:
            if remaining_output <= 0.0:
                break
            module_name = str(candidate.get("module") or "")
            module_upper_limit = _lookup_module_capacity_upper_limit(
                economy=economy,
                scenario=scenario_key,
                module=module_name,
            )
            module_headroom = float("inf")
            if module_upper_limit is not None:
                baseline_module_output = module_baseline_output_lookup.get(
                    (_state_token(economy), _state_token(module_name), int(year)),
                    0.0,
                )
                prior_module_added = module_added_output_lookup.get(
                    (_state_token(economy), _state_token(scenario_key), _state_token(module_name), int(year)),
                    0.0,
                )
                module_headroom = max(
                    float(module_upper_limit) - float(baseline_module_output) - float(prior_module_added),
                    0.0,
                )
            if module_headroom <= 0.0:
                clipping_rows.append(
                    {
                        "economy": economy,
                        "scenario": scenario_key,
                        "esto_product": esto_product,
                        "year": int(year),
                        "requested_output_uplift": float(remaining_output),
                        "allocated_output_uplift": 0.0,
                        "clipped_output_uplift": float(remaining_output),
                        "reason": (
                            f"Module upper limit reached for '{module_name}'. "
                            "Set CAPACITY_UNMET_MODULE_CAPACITY_UPPER_LIMITS to adjust."
                        ),
                    }
                )
                continue
            output_yield = pd.to_numeric(candidate.get("yield"), errors="coerce")
            if pd.isna(output_yield) or float(output_yield) <= 0.0:
                continue
            allocated_output = min(float(remaining_output), float(module_headroom))
            if allocated_output <= 0.0:
                continue
            capacity_increment = float(allocated_output) / float(output_yield)
            cap_key = _capacity_addition_state_key(
                economy=economy,
                scenario=scenario_key,
                module=module_name,
                process=str(candidate.get("process") or ""),
                instance=int(candidate.get("instance") or 1),
                year=year,
            )
            pass_capacity_additions[cap_key] = pass_capacity_additions.get(cap_key, 0.0) + capacity_increment
            module_added_key = (_state_token(economy), _state_token(scenario_key), _state_token(module_name), int(year))
            module_added_output_lookup[module_added_key] = (
                module_added_output_lookup.get(module_added_key, 0.0) + float(capacity_increment)
            )
            pass_output_additions[output_state_key] = pass_output_additions.get(output_state_key, 0.0) + allocated_output
            allocation_rows.append(
                {
                    "economy": economy,
                    "scenario": scenario_key,
                    "esto_product": esto_product,
                    "year": int(year),
                    "module": str(candidate.get("module") or ""),
                    "process": str(candidate.get("process") or ""),
                    "instance": int(candidate.get("instance") or 1),
                    "allocated_output_uplift": float(allocated_output),
                    "yield": float(output_yield),
                    "capacity_increment": float(capacity_increment),
                    "priority_modules": _resolve_capacity_priority_modules(esto_product),
                }
            )
            remaining_output -= allocated_output
        if remaining_output > 1e-9:
            unresolved_rows.append(
                {
                    "economy": economy,
                    "scenario": scenario_key,
                    "esto_product": esto_product,
                    "year": int(year),
                    "unresolved_output_uplift": float(remaining_output),
                    "reason": "Eligible processes found but no positive yield available.",
                }
            )

    fatal_unresolved_rows, handled_unresolved_rows, unresolved_policy = _split_unresolved_rows_by_policy(
        unresolved_rows,
        mode="capacity_unmet_iterative",
    )
    unresolved_csv_path: Path | None = None
    unresolved_json_path: Path | None = None
    if handled_unresolved_rows:
        unresolved_csv_path, unresolved_json_path = _save_unresolved_positive_report(
            mode="capacity_unmet_iterative",
            unresolved_rows=handled_unresolved_rows,
        )
        print(
            "[CAPACITY_UNMET_ITERATIVE][WARN] Unresolved positive residuals handled by policy "
            f"'{unresolved_policy}': {len(handled_unresolved_rows)} "
            f"(csv={unresolved_csv_path}, json={unresolved_json_path})"
        )
    if fatal_unresolved_rows:
        preview = fatal_unresolved_rows[:12]
        raise RuntimeError(
            "capacity_unmet_iterative could not allocate unmet imports to eligible transformation "
            f"processes. Examples: {preview}"
        )

    for key, value in pass_capacity_additions.items():
        cumulative_capacity_map[key] = cumulative_capacity_map.get(key, 0.0) + float(value)
    for key, value in pass_output_additions.items():
        cumulative_output_map[key] = cumulative_output_map.get(key, 0.0) + float(value)

    _CAPACITY_UNMET_RUNTIME_CAPACITY_ADDITIONS = dict(cumulative_capacity_map)
    state["cumulative_capacity_additions"] = cumulative_capacity_map
    state["cumulative_output_additions"] = cumulative_output_map
    state["last_results_signatures"] = signature_map

    unmet_total = float(unmet_candidates["unmet_proxy"].sum()) if not unmet_candidates.empty else 0.0
    allocated_total = float(sum(pass_output_additions.values()))
    clipped_total = float(
        sum(float(item.get("clipped_output_uplift", 0.0)) for item in clipping_rows)
    )
    pass_summary = {
        "timestamp_utc": datetime.now(timezone.utc).isoformat(),
        "mode": "capacity_unmet_iterative",
        "iteration_run_mode": run_mode,
        "state_path": str(_resolve(state_path)),
        "results_signature_used": signature_map,
        "baseline_import_total": float(unmet_table["adjusted_imports"].sum()),
        "observed_import_total": float(unmet_table["observed_imports"].sum()),
        "unmet_proxy_total": unmet_total,
        "allocated_output_total": allocated_total,
        "clipped_output_total": clipped_total,
        "allocation_rows": allocation_rows,
        "clipping_rows": clipping_rows,
        "unresolved_positive_rows": handled_unresolved_rows,
        "unresolved_positive_policy": unresolved_policy,
        "unresolved_positive_csv": str(unresolved_csv_path) if unresolved_csv_path else "",
        "unresolved_positive_json": str(unresolved_json_path) if unresolved_json_path else "",
        "unmatched_results_fuels": unmatched_result_fuels,
        "next_manual_step": (
            "Import generated workbook into LEAP, recalculate, refresh results tables, then rerun."
        ),
    }
    pass_history = state.get("passes")
    if not isinstance(pass_history, list):
        pass_history = []
    pass_history.append(pass_summary)
    state["passes"] = pass_history[-50:]
    persisted_path = _write_capacity_unmet_state(state, state_path=state_path)

    print("\n" + "=" * 96)
    print("[CAPACITY_UNMET_ITERATIVE] Pass summary")
    print(f"[CAPACITY_UNMET_ITERATIVE] State file: {persisted_path}")
    print(
        "[CAPACITY_UNMET_ITERATIVE] Baseline imports="
        f"{pass_summary['baseline_import_total']:.3f}, observed imports={pass_summary['observed_import_total']:.3f}, "
        f"unmet proxy={pass_summary['unmet_proxy_total']:.3f}"
    )
    print(
        "[CAPACITY_UNMET_ITERATIVE] Allocated output uplift="
        f"{allocated_total:.3f}, clipped={clipped_total:.3f}, allocations={len(allocation_rows)}"
    )
    if clipping_rows:
        print(f"[CAPACITY_UNMET_ITERATIVE][WARN] Clipped rows: {len(clipping_rows)}")
        for item in clipping_rows[:20]:
            print(
                "  - economy={economy} scenario={scenario} fuel={fuel} year={year} "
                "requested={requested:.3f} clipped={clipped:.3f} reason={reason}".format(
                    economy=item.get("economy"),
                    scenario=item.get("scenario"),
                    fuel=item.get("esto_product"),
                    year=item.get("year"),
                    requested=float(item.get("requested_output_uplift", 0.0)),
                    clipped=float(item.get("clipped_output_uplift", 0.0)),
                    reason=item.get("reason"),
                )
            )
        if len(clipping_rows) > 20:
            print(f"  ... plus {len(clipping_rows) - 20} more clipping rows")
    if unmatched_result_fuels:
        print(
            "[CAPACITY_UNMET_ITERATIVE][WARN] Unmapped Fuel labels in results sheets: "
            f"{len(unmatched_result_fuels)}"
        )
    print(
        "[CAPACITY_UNMET_ITERATIVE] Next step: "
        "Import workbook into LEAP, recalc, refresh results tables, rerun this workflow."
    )
    print("=" * 96 + "\n")
    return pass_summary


def _run_capacity_unmet_iterative_balanced_pass(
    *,
    reconciliation_table: pd.DataFrame,
    process_records: list[dict],
    economies: Iterable[str],
    scenarios: Iterable[str],
    results_dir: Path | str | Iterable[Path | str] = CAPACITY_UNMET_RESULTS_DIR,
    state_path: Path | str = CAPACITY_UNMET_STATE_PATH,
    allow_same_results_reuse: bool = CAPACITY_UNMET_ALLOW_SAME_RESULTS_REUSE,
) -> dict[str, object]:
    """Compute one iterative pass using observed imports gaps as unmet proxy."""
    global _CAPACITY_UNMET_RUNTIME_CAPACITY_ADDITIONS
    global _CAPACITY_UNMET_RUNTIME_PRIMARY_ADDITIONS
    global _CAPACITY_UNMET_RUNTIME_EXPORT_ADJUSTMENTS
    if reconciliation_table.empty:
        raise ValueError("Cannot run capacity_unmet_iterative_balanced with empty reconciliation table.")

    process_catalog, unmapped_process_labels = _build_capacity_process_catalog(process_records)
    if process_catalog.empty:
        raise ValueError(
            "capacity_unmet_iterative_balanced mode requires transformation process output rows "
            "to infer fuel yields; none were found."
        )
    if unmapped_process_labels:
        preview = ", ".join(unmapped_process_labels[:12])
        print(
            "[WARN] Some transformation output labels could not be mapped to ESTO products "
            f"for capacity_unmet_iterative_balanced: {preview}"
        )

    run_mode = _resolve_capacity_unmet_iteration_run_mode()
    state = _read_capacity_unmet_state(state_path=state_path, run_mode=run_mode)
    cumulative_capacity_map = _parse_runtime_capacity_additions_from_state(
        state.get("cumulative_capacity_additions")
    )
    cumulative_output_map = _parse_runtime_capacity_additions_from_state(
        state.get("cumulative_output_additions")
    )
    cumulative_primary_map = _parse_runtime_capacity_additions_from_state(
        state.get("cumulative_primary_additions")
    )
    cumulative_export_map = _parse_runtime_capacity_additions_from_state(
        state.get("cumulative_export_adjustments")
    )
    module_baseline_output_lookup = _build_module_baseline_output_lookup(process_catalog)
    module_added_output_lookup = _build_module_added_output_lookup(cumulative_capacity_map)
    last_signatures = state.get("last_results_signatures")
    if not isinstance(last_signatures, dict):
        last_signatures = {}

    reconciliation = reconciliation_table.copy()
    reconciliation["scenario_key"] = (
        reconciliation["scenario"].astype(str).str.strip().str.lower()
    )
    reconciliation["adjusted_imports"] = pd.to_numeric(
        reconciliation.get("adjusted_imports"), errors="coerce"
    ).fillna(0.0)
    reconciliation["adjusted_exports"] = pd.to_numeric(
        reconciliation.get("adjusted_exports"), errors="coerce"
    ).fillna(0.0)
    reconciliation["max_transformation_output"] = pd.to_numeric(
        reconciliation.get("max_transformation_output"), errors="coerce"
    )
    reconciliation["constrained_transformation_output"] = pd.to_numeric(
        reconciliation.get("constrained_transformation_output"), errors="coerce"
    ).fillna(0.0)
    reconciliation["max_production"] = pd.to_numeric(
        reconciliation.get("max_production"), errors="coerce"
    )
    reconciliation["constrained_production"] = pd.to_numeric(
        reconciliation.get("constrained_production"), errors="coerce"
    ).fillna(0.0)

    scenario_pairs: list[tuple[str, str]] = []
    seen_pairs: set[tuple[str, str]] = set()
    for economy in [str(item).strip() for item in economies if str(item).strip()]:
        for scenario in [str(item).strip() for item in scenarios if str(item).strip()]:
            scenario_key = _state_token(
                _resolve_reconciliation_scenario_key(reconciliation_table, scenario)
            )
            pair = (str(economy), scenario_key)
            if pair in seen_pairs:
                continue
            seen_pairs.add(pair)
            scenario_pairs.append(pair)
    if not scenario_pairs:
        raise ValueError("capacity_unmet_iterative_balanced needs at least one economy/scenario pair.")

    label_to_product = _build_label_to_esto_product_lookup()
    observed_trade, signature_map, unmatched_result_fuels = _collect_observed_trade_from_supply_results(
        scenario_pairs=scenario_pairs,
        label_to_product=label_to_product,
        results_dir=results_dir,
        include_exports=True,
    )
    if (
        not allow_same_results_reuse
        and signature_map
        and last_signatures
        and signature_map == last_signatures
    ):
        print(
            "[WARN] capacity_unmet_iterative_balanced detected no new LEAP results artifacts since the "
            "previous pass. Continuing with reused results artifacts. "
            "Import workbook into LEAP, recalculate, refresh results tables, then rerun "
            "to use fresh results."
        )
    if observed_trade.empty:
        raise ValueError(
            "capacity_unmet_iterative_balanced could not parse imports/exports from supply "
            f"results tables in '{_resolve(results_dir)}'."
        )
    observed_trade["observed_imports"] = pd.to_numeric(
        observed_trade.get("observed_imports"), errors="coerce"
    ).fillna(0.0)
    observed_trade["observed_exports"] = pd.to_numeric(
        observed_trade.get("observed_exports"), errors="coerce"
    ).fillna(0.0)
    observed_trade["observed_net_imports"] = (
        observed_trade["observed_imports"] - observed_trade["observed_exports"]
    )

    requested_scenarios = {scenario for _, scenario in scenario_pairs}
    requested_economies = {economy for economy, _ in scenario_pairs}
    baseline = reconciliation[
        reconciliation["economy"].astype(str).isin(requested_economies)
        & reconciliation["scenario_key"].astype(str).isin(requested_scenarios)
    ][
        [
            "economy",
            "scenario_key",
            "esto_product",
            "year",
            "adjusted_imports",
            "adjusted_exports",
            "max_transformation_output",
            "constrained_transformation_output",
            "max_production",
            "constrained_production",
        ]
    ].copy()
    baseline = baseline.rename(columns={"scenario_key": "scenario"})
    if baseline.empty:
        raise ValueError(
            "capacity_unmet_iterative_balanced found no reconciliation rows for run economy/scenario scope."
        )
    baseline["baseline_net_imports"] = (
        baseline["adjusted_imports"] - baseline["adjusted_exports"]
    )

    delta = baseline.merge(
        observed_trade[
            [
                "economy",
                "scenario",
                "esto_product",
                "year",
                "observed_imports",
                "observed_exports",
                "observed_net_imports",
            ]
        ],
        on=["economy", "scenario", "esto_product", "year"],
        how="left",
    )
    for column in [
        "adjusted_imports",
        "adjusted_exports",
        "baseline_net_imports",
        "observed_imports",
        "observed_exports",
        "observed_net_imports",
        "constrained_transformation_output",
        "constrained_production",
    ]:
        delta[column] = pd.to_numeric(delta.get(column), errors="coerce").fillna(0.0)
    delta["max_transformation_output"] = pd.to_numeric(
        delta.get("max_transformation_output"), errors="coerce"
    )
    delta["max_production"] = pd.to_numeric(delta.get("max_production"), errors="coerce")
    # Imports gap is the unmet proxy:
    # +ve: LEAP needed more imports than expected baseline -> uplift output/capacity.
    # -ve: LEAP needed fewer imports than expected baseline -> route to extra exports.
    delta["imports_gap"] = delta["observed_imports"] - delta["adjusted_imports"]

    positive_rows = delta[delta["imports_gap"] > 0.0].copy().sort_values(
        ["economy", "scenario", "esto_product", "year"]
    )
    negative_rows = delta[delta["imports_gap"] < 0.0].copy().sort_values(
        ["economy", "scenario", "esto_product", "year"]
    )

    allocation_rows: list[dict[str, object]] = []
    clipping_rows: list[dict[str, object]] = []
    unresolved_rows: list[dict[str, object]] = []
    export_rows: list[dict[str, object]] = []
    pass_capacity_additions: dict[str, float] = {}
    pass_output_additions: dict[str, float] = {}
    pass_primary_additions: dict[str, float] = {}
    pass_export_adjustments: dict[str, float] = {}

    for _, row in positive_rows.iterrows():
        economy = str(row.get("economy") or "").strip()
        scenario_key = str(row.get("scenario") or "").strip().lower()
        esto_product = str(row.get("esto_product") or "").strip()
        year = int(pd.to_numeric(row.get("year"), errors="coerce"))
        remaining_output = max(float(row.get("imports_gap", 0.0)), 0.0)
        if not economy or not scenario_key or not esto_product or remaining_output <= 0.0:
            continue

        if _is_primary_esto_product(esto_product):
            primary_key = _output_addition_state_key(
                economy=economy,
                scenario=scenario_key,
                esto_product=esto_product,
                year=year,
            )
            prior_primary = float(cumulative_primary_map.get(primary_key, 0.0))
            max_prod = pd.to_numeric(row.get("max_production"), errors="coerce")
            configured_max_prod = _lookup_production_upper_limit(
                economy=economy,
                scenario=scenario_key,
                esto_product=esto_product,
            )
            if configured_max_prod is not None:
                if pd.isna(max_prod):
                    max_prod = float(configured_max_prod)
                else:
                    max_prod = min(float(max_prod), float(configured_max_prod))
            constrained_prod = max(float(row.get("constrained_production", 0.0)), 0.0)
            if pd.isna(max_prod):
                primary_headroom = float("inf")
            else:
                primary_headroom = max(float(max_prod) - constrained_prod - prior_primary, 0.0)
            primary_alloc = min(remaining_output, primary_headroom)
            if primary_alloc > 0.0:
                pass_primary_additions[primary_key] = pass_primary_additions.get(primary_key, 0.0) + primary_alloc
                allocation_rows.append(
                    {
                        "economy": economy,
                        "scenario": scenario_key,
                        "esto_product": esto_product,
                        "year": int(year),
                        "module": "Resources\\Primary",
                        "process": "Indigenous Production",
                        "instance": 1,
                        "allocated_output_uplift": float(primary_alloc),
                        "yield": 1.0,
                        "capacity_increment": float(primary_alloc),
                        "allocation_type": "primary_production",
                    }
                )
                remaining_output -= primary_alloc
            if primary_headroom < float(row.get("imports_gap", 0.0)):
                clipped = max(float(row.get("imports_gap", 0.0)) - primary_alloc, 0.0)
                if clipped > 0.0:
                    clipping_rows.append(
                        {
                            "economy": economy,
                            "scenario": scenario_key,
                            "esto_product": esto_product,
                            "year": int(year),
                            "requested_output_uplift": float(row.get("imports_gap", 0.0)),
                            "allocated_output_uplift": float(primary_alloc),
                            "clipped_output_uplift": float(clipped),
                            "reason": "Primary production capped by max_production headroom.",
                        }
                    )

        if remaining_output <= 0.0:
            continue

        output_state_key = _output_addition_state_key(
            economy=economy,
            scenario=scenario_key,
            esto_product=esto_product,
            year=year,
        )
        prior_added_output = float(cumulative_output_map.get(output_state_key, 0.0))
        cap_value = pd.to_numeric(row.get("max_transformation_output"), errors="coerce")
        constrained_value = max(float(row.get("constrained_transformation_output", 0.0)), 0.0)
        if pd.isna(cap_value):
            headroom = float("inf")
        else:
            headroom = max(float(cap_value) - constrained_value - prior_added_output, 0.0)
        allocatable_output = min(remaining_output, headroom)
        if allocatable_output < remaining_output:
            clipping_rows.append(
                {
                    "economy": economy,
                    "scenario": scenario_key,
                    "esto_product": esto_product,
                    "year": int(year),
                    "requested_output_uplift": float(remaining_output),
                    "allocated_output_uplift": float(allocatable_output),
                    "clipped_output_uplift": float(remaining_output - allocatable_output),
                    "reason": "Transformation output capped by max_transformation_output headroom.",
                }
            )
        if allocatable_output <= 0.0:
            if remaining_output > 0.0:
                unresolved_rows.append(
                    {
                        "economy": economy,
                        "scenario": scenario_key,
                        "esto_product": esto_product,
                        "year": int(year),
                        "unresolved_output_uplift": float(remaining_output),
                        "reason": "No remaining transformation headroom after caps.",
                    }
                )
            continue

        candidates = process_catalog[
            (process_catalog["economy"].astype(str) == economy)
            & (process_catalog["esto_product"].astype(str) == esto_product)
            & (process_catalog["year"].astype(int) == int(year))
        ].copy()
        if candidates.empty:
            unresolved_rows.append(
                {
                    "economy": economy,
                    "scenario": scenario_key,
                    "esto_product": esto_product,
                    "year": int(year),
                    "unresolved_output_uplift": float(allocatable_output),
                    "reason": "No eligible transformation process outputs this fuel in this year.",
                }
            )
            continue
        ranked = _rank_capacity_candidates(candidates, esto_product)
        remaining_transform = float(allocatable_output)
        for candidate in ranked:
            if remaining_transform <= 0.0:
                break
            module_name = str(candidate.get("module") or "")
            output_yield = pd.to_numeric(candidate.get("yield"), errors="coerce")
            if pd.isna(output_yield) or float(output_yield) <= 0.0:
                continue
            output_yield_value = float(output_yield)
            cap_key = _capacity_addition_state_key(
                economy=economy,
                scenario=scenario_key,
                module=module_name,
                process=str(candidate.get("process") or ""),
                instance=int(candidate.get("instance") or 1),
                year=year,
            )
            # Use max-style capacity across co-products for a process-year:
            # if this process already got capacity this pass for another fuel, reuse it first.
            existing_pass_capacity = float(pass_capacity_additions.get(cap_key, 0.0))
            reusable_output = min(float(remaining_transform), existing_pass_capacity * output_yield_value)
            if reusable_output > 0.0:
                pass_output_additions[output_state_key] = (
                    pass_output_additions.get(output_state_key, 0.0) + float(reusable_output)
                )
                allocation_rows.append(
                    {
                        "economy": economy,
                        "scenario": scenario_key,
                        "esto_product": esto_product,
                        "year": int(year),
                        "module": str(candidate.get("module") or ""),
                        "process": str(candidate.get("process") or ""),
                        "instance": int(candidate.get("instance") or 1),
                        "allocated_output_uplift": float(reusable_output),
                        "yield": float(output_yield_value),
                        "capacity_increment": 0.0,
                        "priority_modules": _resolve_capacity_priority_modules(esto_product),
                        "allocation_type": "transformation",
                    }
                )
                remaining_transform -= float(reusable_output)
                if remaining_transform <= 0.0:
                    break
            module_upper_limit = _lookup_module_capacity_upper_limit(
                economy=economy,
                scenario=scenario_key,
                module=module_name,
            )
            module_headroom = float("inf")
            if module_upper_limit is not None:
                baseline_module_output = module_baseline_output_lookup.get(
                    (_state_token(economy), _state_token(module_name), int(year)),
                    0.0,
                )
                prior_module_added = module_added_output_lookup.get(
                    (_state_token(economy), _state_token(scenario_key), _state_token(module_name), int(year)),
                    0.0,
                )
                module_headroom = max(
                    float(module_upper_limit) - float(baseline_module_output) - float(prior_module_added),
                    0.0,
                )
            if module_headroom <= 0.0:
                clipping_rows.append(
                    {
                        "economy": economy,
                        "scenario": scenario_key,
                        "esto_product": esto_product,
                        "year": int(year),
                        "requested_output_uplift": float(remaining_transform),
                        "allocated_output_uplift": 0.0,
                        "clipped_output_uplift": float(remaining_transform),
                        "reason": (
                            f"Module upper limit reached for '{module_name}'. "
                            "Set CAPACITY_UNMET_MODULE_CAPACITY_UPPER_LIMITS to adjust."
                        ),
                    }
                )
                continue
            required_capacity_increment = float(remaining_transform) / float(output_yield_value)
            capacity_increment = min(required_capacity_increment, float(module_headroom))
            if capacity_increment <= 0.0:
                continue
            allocated_output = float(capacity_increment) * float(output_yield_value)
            if allocated_output <= 0.0:
                continue
            pass_capacity_additions[cap_key] = existing_pass_capacity + float(capacity_increment)
            module_added_key = (_state_token(economy), _state_token(scenario_key), _state_token(module_name), int(year))
            module_added_output_lookup[module_added_key] = (
                module_added_output_lookup.get(module_added_key, 0.0) + float(capacity_increment)
            )
            pass_output_additions[output_state_key] = pass_output_additions.get(output_state_key, 0.0) + allocated_output
            allocation_rows.append(
                {
                    "economy": economy,
                    "scenario": scenario_key,
                    "esto_product": esto_product,
                    "year": int(year),
                    "module": str(candidate.get("module") or ""),
                    "process": str(candidate.get("process") or ""),
                    "instance": int(candidate.get("instance") or 1),
                    "allocated_output_uplift": float(allocated_output),
                    "yield": float(output_yield_value),
                    "capacity_increment": float(capacity_increment),
                    "priority_modules": _resolve_capacity_priority_modules(esto_product),
                    "allocation_type": "transformation",
                }
            )
            remaining_transform -= allocated_output
        if remaining_transform > 1e-9:
            unresolved_rows.append(
                {
                    "economy": economy,
                    "scenario": scenario_key,
                    "esto_product": esto_product,
                    "year": int(year),
                    "unresolved_output_uplift": float(remaining_transform),
                    "reason": "Eligible processes found but no positive yield available.",
                }
            )

    for _, row in negative_rows.iterrows():
        economy = str(row.get("economy") or "").strip()
        scenario_key = str(row.get("scenario") or "").strip().lower()
        esto_product = str(row.get("esto_product") or "").strip()
        year = int(pd.to_numeric(row.get("year"), errors="coerce"))
        residual = float(row.get("imports_gap", 0.0))
        if not economy or not scenario_key or not esto_product or residual >= 0.0:
            continue
        if CAPACITY_UNMET_PIN_EXPORTS_TO_9TH_PROJECTIONS:
            # In pinned-export mode, do not convert negative import gaps into extra exports.
            # This prevents iterative state from drifting exports away from 9th projections.
            continue
        extra_exports = abs(residual)
        export_key = _output_addition_state_key(
            economy=economy,
            scenario=scenario_key,
            esto_product=esto_product,
            year=year,
        )
        pass_export_adjustments[export_key] = pass_export_adjustments.get(export_key, 0.0) + extra_exports
        export_rows.append(
            {
                "economy": economy,
                "scenario": scenario_key,
                "esto_product": esto_product,
                "year": int(year),
                "extra_exports": float(extra_exports),
                "reason": "Observed imports below baseline; route residual to explicit exports.",
            }
        )

    fatal_unresolved_rows, handled_unresolved_rows, unresolved_policy = _split_unresolved_rows_by_policy(
        unresolved_rows,
        mode="capacity_unmet_iterative_balanced",
    )
    unresolved_csv_path: Path | None = None
    unresolved_json_path: Path | None = None
    if handled_unresolved_rows:
        unresolved_csv_path, unresolved_json_path = _save_unresolved_positive_report(
            mode="capacity_unmet_iterative_balanced",
            unresolved_rows=handled_unresolved_rows,
        )
        print(
            "[CAPACITY_UNMET_ITERATIVE_BALANCED][WARN] Unresolved positive residuals handled by policy "
            f"'{unresolved_policy}': {len(handled_unresolved_rows)} "
            f"(csv={unresolved_csv_path}, json={unresolved_json_path})"
        )
    if fatal_unresolved_rows:
        preview = fatal_unresolved_rows[:12]
        raise RuntimeError(
            "capacity_unmet_iterative_balanced could not allocate positive residuals to "
            f"eligible production/transformation. Examples: {preview}"
        )

    for key, value in pass_capacity_additions.items():
        cumulative_capacity_map[key] = cumulative_capacity_map.get(key, 0.0) + float(value)
    for key, value in pass_output_additions.items():
        cumulative_output_map[key] = cumulative_output_map.get(key, 0.0) + float(value)
    for key, value in pass_primary_additions.items():
        cumulative_primary_map[key] = cumulative_primary_map.get(key, 0.0) + float(value)
    for key, value in pass_export_adjustments.items():
        cumulative_export_map[key] = cumulative_export_map.get(key, 0.0) + float(value)

    _CAPACITY_UNMET_RUNTIME_CAPACITY_ADDITIONS = dict(cumulative_capacity_map)
    _CAPACITY_UNMET_RUNTIME_PRIMARY_ADDITIONS = dict(cumulative_primary_map)
    _CAPACITY_UNMET_RUNTIME_EXPORT_ADJUSTMENTS = dict(cumulative_export_map)
    state["cumulative_capacity_additions"] = cumulative_capacity_map
    state["cumulative_output_additions"] = cumulative_output_map
    state["cumulative_primary_additions"] = cumulative_primary_map
    state["cumulative_export_adjustments"] = cumulative_export_map
    state["last_results_signatures"] = signature_map

    positive_total = float(positive_rows["imports_gap"].sum()) if not positive_rows.empty else 0.0
    negative_total = float((-negative_rows["imports_gap"]).sum()) if not negative_rows.empty else 0.0
    allocated_transform_total = float(sum(pass_output_additions.values()))
    allocated_primary_total = float(sum(pass_primary_additions.values()))
    extra_export_total = float(sum(pass_export_adjustments.values()))
    clipped_total = float(
        sum(float(item.get("clipped_output_uplift", 0.0)) for item in clipping_rows)
    )
    pass_summary = {
        "timestamp_utc": datetime.now(timezone.utc).isoformat(),
        "mode": "capacity_unmet_iterative_balanced",
        "iteration_run_mode": run_mode,
        "state_path": str(_resolve(state_path)),
        "results_signature_used": signature_map,
        "baseline_import_total": float(delta["adjusted_imports"].sum()),
        "observed_import_total": float(delta["observed_imports"].sum()),
        "positive_import_gap_total": positive_total,
        "negative_import_gap_total": negative_total,
        "baseline_net_import_total": float(delta["baseline_net_imports"].sum()),
        "observed_net_import_total": float(delta["observed_net_imports"].sum()),
        "positive_residual_total": positive_total,
        "negative_residual_total": negative_total,
        "allocated_transformation_output_total": allocated_transform_total,
        "allocated_primary_output_total": allocated_primary_total,
        "extra_export_total": extra_export_total,
        "clipped_output_total": clipped_total,
        "allocation_rows": allocation_rows,
        "export_rows": export_rows,
        "clipping_rows": clipping_rows,
        "unresolved_positive_rows": handled_unresolved_rows,
        "unresolved_positive_policy": unresolved_policy,
        "unresolved_positive_csv": str(unresolved_csv_path) if unresolved_csv_path else "",
        "unresolved_positive_json": str(unresolved_json_path) if unresolved_json_path else "",
        "unmatched_results_fuels": unmatched_result_fuels,
        "next_manual_step": (
            "Import generated workbook into LEAP, recalculate, refresh results tables, then rerun."
        ),
    }
    pass_history = state.get("passes")
    if not isinstance(pass_history, list):
        pass_history = []
    pass_history.append(pass_summary)
    state["passes"] = pass_history[-50:]
    persisted_path = _write_capacity_unmet_state(state, state_path=state_path)

    print("\n" + "=" * 96)
    print("[CAPACITY_UNMET_ITERATIVE_BALANCED] Pass summary")
    print(f"[CAPACITY_UNMET_ITERATIVE_BALANCED] State file: {persisted_path}")
    print(
        "[CAPACITY_UNMET_ITERATIVE_BALANCED] Baseline imports="
        f"{pass_summary['baseline_import_total']:.3f}, observed imports={pass_summary['observed_import_total']:.3f}"
    )
    print(
        "[CAPACITY_UNMET_ITERATIVE_BALANCED] Positive imports gap="
        f"{positive_total:.3f}, negative imports gap={negative_total:.3f}"
    )
    print(
        "[CAPACITY_UNMET_ITERATIVE_BALANCED] Allocated transformation="
        f"{allocated_transform_total:.3f}, primary={allocated_primary_total:.3f}, "
        f"extra exports={extra_export_total:.3f}, clipped={clipped_total:.3f}"
    )
    if clipping_rows:
        print(f"[CAPACITY_UNMET_ITERATIVE_BALANCED][WARN] Clipped rows: {len(clipping_rows)}")
    if unmatched_result_fuels:
        print(
            "[CAPACITY_UNMET_ITERATIVE_BALANCED][WARN] Unmapped Fuel labels in results sheets: "
            f"{len(unmatched_result_fuels)}"
        )
    print(
        "[CAPACITY_UNMET_ITERATIVE_BALANCED] Next step: "
        "Import workbook into LEAP, recalc, refresh results tables, rerun this workflow."
    )
    print("=" * 96 + "\n")
    return pass_summary


def reset_supply_and_transformation_import_export_to_zero(
    reconciliation_table: pd.DataFrame,
    transformation_process_records: list[dict] | None = None,
    *,
    economies: Iterable[str] | None = None,
    scenarios: Iterable[str] | None = None,
    sector_titles: Iterable[str] | None = None,
    esto_products: Iterable[str] | None = None,
    years: Iterable[int] | None = None,
) -> tuple[pd.DataFrame, list[dict] | None]:
    """
    Zero supply/transformation import-export values for selected scopes.

    Filters are optional and combined with logical AND when provided.
    When a filter is omitted, it is expanded to the full available set from the
    provided reconciliation/process data.
    - `reconciliation_table` columns are zeroed for matched rows.
    - `transformation_process_records` output import/export targets are zeroed for
      matched process records and target fuels.
    """
    if not isinstance(reconciliation_table, pd.DataFrame):
        raise TypeError("reconciliation_table must be a pandas DataFrame")

    def _norm_set(values: Iterable[str] | None) -> set[str]:
        if not values:
            return set()
        return {
            str(item).strip().lower()
            for item in values
            if str(item or "").strip()
        }

    def _resolve_product_filter_set(values: Iterable[str] | None) -> set[str]:
        """Resolve mixed fuel labels/codes into reconciliation esto_product tokens."""
        raw_values = [
            str(item).strip()
            for item in (values or [])
            if str(item or "").strip()
        ]
        if not raw_values:
            return set()
        lookup = _build_label_to_esto_product_lookup()
        resolved: set[str] = set()
        for token in raw_values:
            mapped = (
                lookup.get(token)
                or lookup.get(token.lower())
                or lookup.get(_normalize_label_for_lookup(token))
            )
            if mapped:
                resolved.add(str(mapped).strip())
            else:
                # Keep caller-provided raw token to support direct esto_product inputs.
                resolved.add(token)
        return resolved

    economy_set = _norm_set(economies)
    scenario_set = _norm_set(scenarios)
    configured_modules = sorted(_configured_reset_module_names())
    configured_fuels = _configured_reset_fuel_labels()
    sector_set = _norm_set(sector_titles or configured_modules)
    product_set = _resolve_product_filter_set(esto_products or configured_fuels)
    year_set = {
        int(item)
        for item in (years or [])
    }

    # Expand omitted filters to explicit "all available" sets.
    if not economy_set and "economy" in reconciliation_table.columns:
        economy_set = {
            str(item).strip().lower()
            for item in reconciliation_table["economy"].dropna().astype(str).tolist()
            if str(item).strip()
        }
    if not scenario_set and "scenario" in reconciliation_table.columns:
        scenario_set = {
            str(item).strip().lower()
            for item in reconciliation_table["scenario"].dropna().astype(str).tolist()
            if str(item).strip()
        }
    if not product_set and "esto_product" in reconciliation_table.columns:
        product_set = {
            str(item).strip()
            for item in reconciliation_table["esto_product"].dropna().astype(str).tolist()
            if str(item).strip()
        }
    if not year_set and "year" in reconciliation_table.columns:
        year_values = pd.to_numeric(reconciliation_table["year"], errors="coerce").dropna()
        year_set = {int(item) for item in year_values.tolist()}

    updated_reconciliation = reconciliation_table.copy()
    mask = pd.Series(True, index=updated_reconciliation.index)
    if economy_set and "economy" in updated_reconciliation.columns:
        mask &= updated_reconciliation["economy"].astype(str).str.strip().str.lower().isin(economy_set)
    if scenario_set and "scenario" in updated_reconciliation.columns:
        mask &= updated_reconciliation["scenario"].astype(str).str.strip().str.lower().isin(scenario_set)
    if product_set and "esto_product" in updated_reconciliation.columns:
        product_values = updated_reconciliation["esto_product"].astype(str).str.strip()
        product_mask = product_values.isin(product_set)
        normalized_product_set = {
            _normalize_esto_product_for_match(item)
            for item in product_set
            if _normalize_esto_product_for_match(item)
        }
        if normalized_product_set:
            product_mask |= product_values.map(_normalize_esto_product_for_match).isin(
                normalized_product_set
            )
        mask &= product_mask
    if year_set and "year" in updated_reconciliation.columns:
        year_values = pd.to_numeric(updated_reconciliation["year"], errors="coerce").astype("Int64")
        mask &= year_values.isin(year_set)

    reconciliation_zero_columns = [
        "projected_imports",
        "projected_exports",
        "projected_net_imports",
        "trade_adjustment",
        "required_net_imports",
        "uncapped_adjusted_imports",
        "uncapped_adjusted_exports",
        "adjusted_imports",
        "adjusted_exports",
        "adjusted_net_imports",
        "imports_cap_binding",
        "exports_cap_binding",
        "baseline_transformation_import_target",
        "baseline_transformation_export_target",
        "transformation_import_target",
        "transformation_export_target",
        "supply_imports_residual",
        "supply_exports_residual",
        "combined_net_imports_after_split",
    ]
    for column in reconciliation_zero_columns:
        if column in updated_reconciliation.columns:
            updated_reconciliation.loc[mask, column] = 0.0

    if transformation_process_records is None:
        return updated_reconciliation, None

    label_to_product = _build_label_to_esto_product_lookup()
    updated_records = copy.deepcopy(transformation_process_records)

    if not economy_set:
        economy_set = {
            str(record.get("economy") or "").strip().lower()
            for record in updated_records
            if str(record.get("economy") or "").strip()
        }
    if not sector_set:
        sector_set = {
            str(record.get("sector_title") or "").strip().lower()
            for record in updated_records
            if str(record.get("sector_title") or "").strip()
        }
    if not product_set:
        derived_products: set[str] = set()
        for record in updated_records:
            for payload_key in (
                "output_values",
                "feedstock_values",
                "loss_values",
                "output_import_targets",
                "output_export_targets",
            ):
                payload = record.get(payload_key)
                if not isinstance(payload, dict):
                    continue
                for label in payload.keys():
                    token = str(label or "").strip()
                    if not token:
                        continue
                    mapped = label_to_product.get(token) or label_to_product.get(token.lower())
                    if mapped:
                        derived_products.add(str(mapped).strip())
        product_set = derived_products

    target_years = tuple(sorted(year_set)) if year_set else tuple(range(BASE_YEAR, FINAL_YEAR + 1))

    def _record_matches(record: dict) -> bool:
        if economy_set:
            economy_value = str(record.get("economy") or "").strip().lower()
            if economy_value not in economy_set:
                return False
        if sector_set:
            sector_value = str(record.get("sector_title") or "").strip().lower()
            if sector_value not in sector_set:
                return False
        return True

    def _labels_for_product_filter(target_map: dict) -> list[str]:
        labels: list[str] = []
        for label in target_map.keys():
            token = str(label or "").strip()
            if not token:
                continue
            if not product_set:
                labels.append(token)
                continue
            mapped = label_to_product.get(token) or label_to_product.get(token.lower())
            if mapped in product_set:
                labels.append(token)
        return labels

    for record in updated_records:
        if not _record_matches(record):
            continue
        for key in ("output_import_targets", "output_export_targets"):
            target_map = record.get(key)
            if not isinstance(target_map, dict):
                continue
            for label in _labels_for_product_filter(target_map):
                year_values = target_map.get(label)
                if isinstance(year_values, dict):
                    for year in target_years:
                        year_int = int(year)
                        year_values[year_int] = 0.0
                        year_values[str(year_int)] = 0.0
                else:
                    target_map[label] = {int(year): 0.0 for year in target_years}

    return updated_reconciliation, updated_records


def _build_supply_measures_for_trade_mode() -> list[dict[str, object]]:
    """Return supply export measure definitions for the active trade mode."""
    measures = [dict(item) for item in supply_data_pipeline.SUPPLY_MEASURES]
    if not _use_capacity_unmet_iterative_any_mode():
        return measures
    # In iterative consecutive runs, leave Imports unchanged in LEAP by omitting
    # import rows from workbook exports entirely.
    if _resolve_capacity_unmet_iteration_run_mode() == "consecutive":
        measures = [
            measure
            for measure in measures
            if str(measure.get("name") or "").strip().lower() != "imports"
        ]
    measures.extend(
        [
            {
                "name": "Maximum Production",
                "flow_key": "max_production",
                "units": "Petajoule",
                "per": "",
                "branch_root": "primary",
            },
        ]
    )
    return measures


def _build_transformation_target_multiplier_table(
    reconciliation_table: pd.DataFrame,
    process_target_rows: pd.DataFrame,
    scenario: str,
) -> pd.DataFrame:
    """Return per-product/year multipliers to scale process target rows."""
    if reconciliation_table.empty or process_target_rows.empty:
        return pd.DataFrame(
            columns=["economy", "esto_product", "year", "direction", "multiplier"]
        )
    scenario_key = _resolve_reconciliation_scenario_key(reconciliation_table, scenario)
    scenario_table = reconciliation_table[
        reconciliation_table["scenario"].astype(str).str.strip().str.lower() == scenario_key
    ].copy()
    if scenario_table.empty:
        return pd.DataFrame(
            columns=["economy", "esto_product", "year", "direction", "multiplier"]
        )
    desired_imports = scenario_table[
        ["economy", "esto_product", "year", "transformation_import_target"]
    ].rename(columns={"transformation_import_target": "desired_value"})
    desired_imports["direction"] = "import"
    desired_exports = scenario_table[
        ["economy", "esto_product", "year", "transformation_export_target"]
    ].rename(columns={"transformation_export_target": "desired_value"})
    desired_exports["direction"] = "export"
    desired = pd.concat([desired_imports, desired_exports], ignore_index=True)

    baseline = (
        process_target_rows.groupby(
            ["economy", "esto_product", "year", "direction"],
            dropna=False,
            as_index=False,
        )["value"]
        .sum(min_count=1)
        .rename(columns={"value": "baseline_value"})
    )
    merged = desired.merge(
        baseline,
        on=["economy", "esto_product", "year", "direction"],
        how="left",
    )
    merged["desired_value"] = pd.to_numeric(merged["desired_value"], errors="coerce").fillna(0.0).clip(lower=0.0)
    merged["baseline_value"] = pd.to_numeric(merged["baseline_value"], errors="coerce").fillna(0.0).clip(lower=0.0)
    merged["multiplier"] = (
        merged["desired_value"] / merged["baseline_value"].where(merged["baseline_value"] > 0.0, pd.NA)
    ).fillna(0.0)
    return merged[["economy", "esto_product", "year", "direction", "multiplier"]]


def _resolve_reconciliation_scenario_key(
    reconciliation_table: pd.DataFrame,
    scenario: str,
) -> str:
    """Return the best scenario key available in reconciliation rows."""
    requested_key = str(scenario or "").strip().lower()
    if reconciliation_table.empty or "scenario" not in reconciliation_table.columns:
        return requested_key
    available_keys = {
        str(value).strip().lower()
        for value in reconciliation_table["scenario"].dropna().astype(str).tolist()
        if str(value).strip()
    }
    if requested_key in available_keys:
        return requested_key
    if requested_key in {"current accounts", "current account"}:
        for fallback_key in ("reference", "target"):
            if fallback_key in available_keys:
                return fallback_key
        if len(available_keys) == 1:
            return next(iter(available_keys))
    if "reference" in available_keys:
        return "reference"
    return requested_key


def apply_transformation_target_overrides_for_scenario(
    process_records: list[dict],
    process_target_rows: pd.DataFrame,
    reconciliation_table: pd.DataFrame,
    scenario: str,
) -> list[dict]:
    """Scale process-level transformation import/export targets for one scenario."""
    if not process_records:
        return []
    records = copy.deepcopy(process_records)
    label_to_product = _build_label_to_esto_product_lookup()
    use_legacy_split = _use_legacy_trade_split_mode()
    scaled = pd.DataFrame(
        columns=[
            "record_index",
            "economy",
            "sector_title",
            "process_name",
            "direction",
            "label",
            "esto_product",
            "year",
            "value",
            "multiplier",
            "scaled_value",
        ]
    )
    if use_legacy_split and isinstance(process_target_rows, pd.DataFrame) and not process_target_rows.empty:
        multipliers = _build_transformation_target_multiplier_table(
            reconciliation_table,
            process_target_rows,
            scenario,
        )
        if not multipliers.empty:
            scaled = process_target_rows.merge(
                multipliers,
                on=["economy", "esto_product", "year", "direction"],
                how="left",
            )
            scaled["multiplier"] = pd.to_numeric(scaled["multiplier"], errors="coerce").fillna(0.0)
            scaled["scaled_value"] = (
                pd.to_numeric(scaled["value"], errors="coerce").fillna(0.0).clip(lower=0.0) * scaled["multiplier"]
            )

    for record in records:
        record["output_import_targets"] = {}
        record["output_export_targets"] = {}
        record["process_share_by_year"] = {}
        record.pop("exogenous_capacity_by_year", None)
        record.pop("endogenous_capacity_by_year", None)
        record.pop("maximum_availability_by_year", None)
        record.pop("capacity_credit_by_year", None)
        record.pop("historical_production_by_year", None)

    grouped = (
        scaled.groupby(
            ["record_index", "direction", "label", "year"],
            dropna=False,
            as_index=False,
        )["scaled_value"]
        .sum(min_count=1)
    )
    for _, row in grouped.iterrows():
        index = int(row["record_index"])
        if index < 0 or index >= len(records):
            continue
        direction = str(row["direction"])
        label = str(row["label"])
        year = int(row["year"])
        value = max(float(row["scaled_value"]), 0.0)
        key = "output_import_targets" if direction == "import" else "output_export_targets"
        target_map = records[index].setdefault(key, {})
        label_map = target_map.setdefault(label, {})
        label_map[year] = value

    # Process Share policy:
    # - single-process module: always 100%
    # - multi-process module: split by per-year activity
    #   (scaled trade targets first when available, otherwise output values)
    target_activity_by_record: dict[int, dict[int, float]] = {}
    if use_legacy_split and not scaled.empty:
        target_activity = (
            scaled.groupby(
                ["record_index", "year"],
                dropna=False,
                as_index=False,
            )["scaled_value"]
            .sum(min_count=1)
        )
        for _, row in target_activity.iterrows():
            index = int(row["record_index"])
            if index < 0 or index >= len(records):
                continue
            year = int(row["year"])
            value = max(float(row["scaled_value"]), 0.0)
            target_activity_by_record.setdefault(index, {})[year] = value

    output_activity_by_record: dict[int, dict[int, float]] = {}
    for index, record in enumerate(records):
        output_values = record.get("output_values")
        if not isinstance(output_values, dict) or not output_values:
            continue
        for label, raw_value in output_values.items():
            product = label_to_product.get(str(label)) or label_to_product.get(str(label).lower())
            if not product:
                continue
            year_map = supply_data_pipeline.coerce_value_by_year(raw_value, BASE_YEAR, FINAL_YEAR)
            for year, value in year_map.items():
                year_int = int(year)
                output_value = max(float(value), 0.0)
                if output_value <= 0.0:
                    continue
                output_activity_by_record.setdefault(index, {})
                output_activity_by_record[index][year_int] = (
                    output_activity_by_record[index].get(year_int, 0.0) + output_value
                )

    module_to_indices: dict[tuple[str, str], list[int]] = {}
    for index, record in enumerate(records):
        economy = str(record.get("economy") or "").strip()
        sector_title = str(record.get("sector_title") or "").strip()
        if not economy or not sector_title:
            module_key = (f"__record_{index}", "")
        else:
            module_key = (economy, sector_title)
        module_to_indices.setdefault(module_key, []).append(index)

    all_years = tuple(range(BASE_YEAR, FINAL_YEAR + 1))
    for _, indices in module_to_indices.items():
        if not indices:
            continue
        if len(indices) == 1:
            records[indices[0]]["process_share_by_year"] = {year: 100.0 for year in all_years}
            continue

        for year in all_years:
            has_target_activity = any(
                target_activity_by_record.get(index, {}).get(year, 0.0) > 0.0 for index in indices
            )
            activity_lookup = target_activity_by_record if has_target_activity else output_activity_by_record
            total_activity = sum(activity_lookup.get(index, {}).get(year, 0.0) for index in indices)
            if total_activity > 0.0:
                for index in indices:
                    share_value = (activity_lookup.get(index, {}).get(year, 0.0) / total_activity) * 100.0
                    records[index].setdefault("process_share_by_year", {})[year] = max(0.0, min(share_value, 100.0))
            else:
                equal_share = 100.0 / float(len(indices))
                for index in indices:
                    records[index].setdefault("process_share_by_year", {})[year] = equal_share

    if _use_capacity_like_mode():
        reset_modules = _configured_reset_module_names()
        reset_output_fuels_by_module = _configured_reset_output_fuel_labels_by_module(
            reset_modules
        )
        scenario_key_for_capacity = _state_token(
            _resolve_reconciliation_scenario_key(reconciliation_table, scenario)
        )
        instance_counter: dict[tuple[str, str, str], int] = {}
        missing_output_scope_modules: set[str] = set()
        for record in records:
            economy_name = str(record.get("economy") or "").strip()
            module_name = str(record.get("sector_title") or "").strip() or "__unknown_module__"
            process_name = str(record.get("process_name") or "").strip() or "__unknown_process__"
            counter_key = (
                _state_token(economy_name),
                _state_token(module_name),
                _state_token(process_name),
            )
            instance_counter[counter_key] = int(instance_counter.get(counter_key, 0)) + 1
            instance = int(instance_counter[counter_key])
            output_total_by_year: dict[int, float] = {int(year): 0.0 for year in all_years}
            output_values = record.get("output_values") or {}
            output_labels: set[str] = set()
            for label in output_values.keys():
                canonical_label = _canonical_transformation_fuel_label(label)
                if canonical_label:
                    output_labels.add(canonical_label)
            for label, raw_value in output_values.items():
                if not str(label or "").strip():
                    continue
                year_map = supply_data_pipeline.coerce_value_by_year(raw_value, BASE_YEAR, FINAL_YEAR)
                for year, value in year_map.items():
                    year_int = int(year)
                    if year_int < BASE_YEAR or year_int > FINAL_YEAR:
                        continue
                    output_total_by_year[year_int] = output_total_by_year.get(year_int, 0.0) + max(float(value), 0.0)
            capacity_additions_by_year = _lookup_runtime_capacity_additions_for_record(
                economy=economy_name,
                scenario=scenario_key_for_capacity,
                module=module_name,
                process=process_name,
                instance=instance,
            )
            for year, add_value in capacity_additions_by_year.items():
                if year < BASE_YEAR or year > FINAL_YEAR:
                    continue
                output_total_by_year[int(year)] = output_total_by_year.get(int(year), 0.0) + max(float(add_value), 0.0)

            if CAPACITY_CLEAR_OUTPUT_TRADE_TARGETS:
                sector_name = str(record.get("sector_title") or "").strip().lower()
                zero_map = {int(year): 0.0 for year in all_years}
                target_labels = set(output_labels)
                if reset_modules and sector_name in reset_modules:
                    module_reset_fuels = reset_output_fuels_by_module.get(
                        sector_name, []
                    )
                    if not module_reset_fuels and RESET_SCOPE_USE_FULL_MODEL_EXPORT:
                        missing_output_scope_modules.add(
                            str(record.get("sector_title") or "").strip()
                        )
                    for label in module_reset_fuels:
                        canonical_label = _canonical_transformation_fuel_label(label)
                        if canonical_label:
                            target_labels.add(canonical_label)
                record["output_import_targets"] = {label: dict(zero_map) for label in sorted(target_labels)}
                record["output_export_targets"] = {label: dict(zero_map) for label in sorted(target_labels)}

            record["exogenous_capacity_by_year"] = {
                int(year): max(float(value), 0.0) * float(CAPACITY_CONSTRAINT_FACTOR)
                for year, value in output_total_by_year.items()
            }
            record["capacity_units"] = str(CAPACITY_CONSTRAINT_UNITS)
            record["historical_production_by_year"] = {
                int(year): max(float(value), 0.0)
                for year, value in output_total_by_year.items()
            }
        if missing_output_scope_modules:
            missing_preview = ", ".join(sorted({item for item in missing_output_scope_modules if item}))
            print(
                "[WARN] Missing module-specific 'Output Fuels' scope from full model export "
                "for transformation module(s): "
                f"{missing_preview}. "
                "Only observed output labels from process records were reset for these modules."
            )
    return records


def save_transformation_exports_with_split_targets(
    reconciliation_table: pd.DataFrame,
    process_target_rows: pd.DataFrame,
    process_records: list[dict],
    scenarios: Iterable[str],
    output_dir: Path | str = TRANSFORMATION_EXPORT_OUTPUT_DIR,
    filename_template: str = TRANSFORMATION_EXPORT_FILENAME_TEMPLATE,
    full_branch_catalog_df: pd.DataFrame | None = None,
) -> list[Path]:
    """Save scenario-specific transformation LEAP workbooks with split import/export targets."""
    if reconciliation_table.empty or not process_records:
        return []
    output_path = _resolve(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    scenario_list = [str(item) for item in scenarios if str(item).strip()]
    saved_paths: list[Path] = []

    def _projection_scenario_for_export(scenario_name: str) -> str:
        text = str(scenario_name or "").strip().lower()
        if text == "target":
            return "target"
        return "reference"

    base_economies = sorted(
        {
            str(record.get("economy")).strip()
            for record in process_records
            if str(record.get("economy") or "").strip()
        }
    )
    for scenario in scenario_list:
        scenario_process_records = process_records
        scenario_process_targets = process_target_rows
        projection_scenario = _projection_scenario_for_export(scenario)
        try:
            scenario_process_records = transformation_workflow.collect_transformation_rows(
                economies=base_economies or None,
                projection_scenario=projection_scenario,
            )
            scenario_process_targets, scenario_process_records = build_transformation_trade_target_rows(
                economies=base_economies or None,
                process_records=scenario_process_records,
            )
        except Exception as exc:
            print(
                f"[WARN] Failed to build scenario-specific transformation baseline for "
                f"{scenario} (projection={projection_scenario}); falling back to default baseline: {exc}"
            )
        scenario_records = apply_transformation_target_overrides_for_scenario(
            scenario_process_records,
            scenario_process_targets,
            reconciliation_table,
            scenario,
        )
        transformation_workflow.core.consolidate_transformation_output_rows(
            scenario_records,
            include_output_series=transformation_workflow.core.INCLUDE_OUTPUT_SERIES_IN_LEAP_EXPORT,
            use_output_targets=bool(
                transformation_workflow.core.TRANSFORMATION_OUTPUT_VARIABLES.get("output_import_target")
                or transformation_workflow.core.TRANSFORMATION_OUTPUT_VARIABLES.get("output_export_target")
            ),
        )
        economy_label = transformation_workflow._infer_primary_economy(scenario_records)
        export_filename = transformation_workflow.format_export_filename(
            economy_label,
            [scenario],
            filename_template,
        )
        export_path = transformation_workflow.core.save_transformation_export(
            scenario_records,
            transformation_workflow.core.EXPORT_REGION,
            transformation_workflow.core.EXPORT_BASE_YEAR,
            transformation_workflow.core.EXPORT_FINAL_YEAR,
            transformation_workflow.core.code_to_name_mapping,
            str(output_path),
            export_filename,
            transformation_workflow.core.EXPORT_MODEL_NAME,
            [scenario],
            full_branch_catalog_df=full_branch_catalog_df,
        )
        if export_path:
            export_file = Path(export_path)
            saved_paths.append(export_file)
    return saved_paths


def save_transfer_exports_with_supply_overrides(
    reconciliation_table: pd.DataFrame,
    economies: Iterable[str],
    scenarios: Iterable[str],
    output_dir: Path | str = TRANSFORMATION_EXPORT_OUTPUT_DIR,
    filename_template: str = transfers_workflow.EXPORT_FILENAME_TEMPLATE,
) -> list[Path]:
    """Save scenario-specific transfer workbooks with supply-linked Process Share overrides."""
    output_path = _resolve(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    scenario_list = [str(item) for item in scenarios if str(item).strip()]
    economy_list = [str(item) for item in economies if str(item).strip()]
    if not scenario_list or not economy_list:
        return []

    base_transfer_records_by_economy: dict[str, list[dict]] = {}
    for economy in economy_list:
        economy_records = transfers_workflow.build_transfer_rows(
            economy=economy,
            use_output_targets=False,
        )
        if economy_records:
            base_transfer_records_by_economy[str(economy)] = transfers_workflow.merge_transfer_rows(
                economy_records
            )
    if not base_transfer_records_by_economy:
        return []

    empty_target_rows = pd.DataFrame(
        columns=[
            "record_index",
            "economy",
            "sector_title",
            "process_name",
            "direction",
            "label",
            "esto_product",
            "year",
            "value",
        ]
    )
    saved_paths: list[Path] = []
    for scenario in scenario_list:
        for economy, economy_records in base_transfer_records_by_economy.items():
            scenario_records = apply_transformation_target_overrides_for_scenario(
                economy_records,
                empty_target_rows,
                reconciliation_table,
                scenario,
            )
            economy_label = str(economy).strip() or transfers_workflow._infer_primary_economy(scenario_records)
            export_filename = transfers_workflow.format_export_filename(
                economy_label,
                [scenario],
                filename_template,
            )
            export_path = transformation_workflow.core.save_transformation_export(
                scenario_records,
                transformation_workflow.core.EXPORT_REGION,
                transformation_workflow.core.EXPORT_BASE_YEAR,
                transformation_workflow.core.EXPORT_FINAL_YEAR,
                transformation_workflow.core.code_to_name_mapping,
                str(output_path),
                export_filename,
                transformation_workflow.core.EXPORT_MODEL_NAME,
                [scenario],
            )
            if export_path:
                export_file = Path(export_path)
                legacy_paths = _find_legacy_transfer_branch_paths(export_file)
                if legacy_paths:
                    sample = "; ".join(legacy_paths[:3])
                    raise ValueError(
                        "Transfer export still contains legacy generic transfer branches "
                        f"in {export_file.name}: {sample}"
                    )
                saved_paths.append(export_file)
    return saved_paths


def _read_workbook_sheet_with_header_detection(
    workbook_path: Path | str,
    sheet_name: str,
) -> tuple[pd.DataFrame, pd.DataFrame, list]:
    """Return (preamble_rows, data_rows, header_values) for a LEAP-style sheet."""
    path = _resolve(workbook_path)
    raw = pd.read_excel(path, sheet_name=sheet_name, header=None)
    header_row = None
    for idx in range(len(raw.index)):
        values = {_normalize_template_header_value(item).lower() for item in raw.iloc[idx].tolist()}
        if "branch path" in values and "variable" in values:
            header_row = int(idx)
            break
    if header_row is None:
        raise ValueError(f"Could not locate LEAP sheet header in {path.name}::{sheet_name}")
    header_values = raw.iloc[header_row].tolist()
    preamble = raw.iloc[:header_row].copy()
    data = raw.iloc[header_row + 1 :].copy()
    data.columns = header_values
    data = data.dropna(how="all").reset_index(drop=True)
    return preamble, data, header_values


def _merge_workbook_sheets(
    workbook_paths: Iterable[Path | str],
    sheet_name: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Merge multiple LEAP-style sheets into one standardized table."""
    source_paths = [Path(item) for item in workbook_paths if item and Path(item).exists()]
    if not source_paths:
        return pd.DataFrame(), pd.DataFrame()
    preamble, first_data, first_header = _read_workbook_sheet_with_header_detection(
        source_paths[0],
        sheet_name=sheet_name,
    )
    ordered_columns = list(first_data.columns)
    merged = [first_data]
    for path in source_paths[1:]:
        _, data, _ = _read_workbook_sheet_with_header_detection(path, sheet_name=sheet_name)
        for col in data.columns:
            if col not in ordered_columns:
                ordered_columns.append(col)
        merged.append(data)
    normalized = [frame.reindex(columns=ordered_columns) for frame in merged]
    merged_data = pd.concat(normalized, ignore_index=True, sort=False)
    dedupe_cols = [col for col in ["Branch Path", "Variable", "Scenario", "Region", "Expression"] if col in merged_data.columns]
    if dedupe_cols:
        merged_data = merged_data.drop_duplicates(subset=dedupe_cols, keep="last")
    else:
        merged_data = merged_data.drop_duplicates(keep="last")
    if "Branch Path" in merged_data.columns and "Variable" in merged_data.columns:
        merged_data = merged_data.sort_values(["Branch Path", "Variable"]).reset_index(drop=True)
    return preamble, merged_data


def _drop_wide_year_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Return a copy without wide year columns such as 2022/2022.0."""
    if df.empty:
        return df.copy()
    year_like = re.compile(r"^\d{4}(?:\.0)?$")
    keep_columns = [col for col in df.columns if not year_like.match(str(col).strip())]
    return df.loc[:, keep_columns].copy()


def _find_legacy_transfer_branch_paths(workbook_path: Path | str) -> list[str]:
    """Return any branch paths that still use the legacy generic Transfers root."""
    path = _resolve(workbook_path)
    if not path.exists():
        return []
    try:
        _, leap_data, _ = _read_workbook_sheet_with_header_detection(path, "LEAP")
    except Exception:
        return []
    if leap_data.empty or "Branch Path" not in leap_data.columns:
        return []
    branch_paths = leap_data["Branch Path"].dropna().astype(str).map(str.strip)
    return sorted(
        {
            value
            for value in branch_paths
            if value.startswith("Transformation\\Transfers\\")
        }
    )


def save_combined_supply_transformation_export(
    *,
    supply_export_paths: Iterable[Path],
    transformation_export_paths: Iterable[Path],
    transfer_export_paths: Iterable[Path],
    output_dir: Path | str = EXPORT_OUTPUT_DIR,
    filename_template: str = COMBINED_EXPORT_FILENAME_TEMPLATE,
    economy_label: str = "economy",
    scenarios: Iterable[str] | None = None,
) -> Path | None:
    """Save a single workbook that combines supply + transformation + transfers rows."""
    paths = [Path(item) for item in [*supply_export_paths, *transformation_export_paths, *transfer_export_paths] if Path(item).exists()]
    if not paths:
        return None
    leap_preamble, leap_data = _merge_workbook_sheets(paths, "LEAP")
    if leap_data.empty:
        return None
    leap_data = _drop_wide_year_columns(leap_data)
    viewing_preamble, viewing_data = _merge_workbook_sheets(paths, "FOR_VIEWING")
    if viewing_data.empty:
        viewing_preamble = leap_preamble.copy()
        viewing_data = leap_data.copy()

    scenario_list = [str(item) for item in (scenarios or []) if str(item).strip()]
    scenario_token = workflow_common.format_filename_segment("_".join(scenario_list)) or "scenario"
    economy_token = workflow_common.format_filename_segment(economy_label) or "economy"
    filename = filename_template.format(economy=economy_token, scenario=scenario_token)
    output_path = _resolve(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    combined_path = output_path / filename

    with pd.ExcelWriter(combined_path, engine="openpyxl", mode="w") as writer:
        leap_preamble.to_excel(writer, sheet_name="LEAP", index=False, header=False)
        pd.DataFrame([list(leap_data.columns)]).to_excel(
            writer,
            sheet_name="LEAP",
            index=False,
            header=False,
            startrow=len(leap_preamble),
        )
        leap_data.to_excel(
            writer,
            sheet_name="LEAP",
            index=False,
            header=False,
            startrow=len(leap_preamble) + 1,
        )
        viewing_preamble.to_excel(writer, sheet_name="FOR_VIEWING", index=False, header=False)
        pd.DataFrame([list(viewing_data.columns)]).to_excel(
            writer,
            sheet_name="FOR_VIEWING",
            index=False,
            header=False,
            startrow=len(viewing_preamble),
        )
        viewing_data.to_excel(
            writer,
            sheet_name="FOR_VIEWING",
            index=False,
            header=False,
            startrow=len(viewing_preamble) + 1,
        )
    print(f"Saved combined supply+transformation workbook to {combined_path}")
    return combined_path


def run_results_linked_leap_import(
    supply_export_paths: Iterable[Path],
    transformation_export_paths: Iterable[Path],
    scenarios: Iterable[str],
    transfer_export_paths: Iterable[Path] | None = None,
    import_scenarios: Iterable[str] | str | None = None,
    region: str = LEAP_IMPORT_REGION,
    create_branches: bool = LEAP_IMPORT_CREATE_BRANCHES,
    fill_branches: bool = LEAP_IMPORT_FILL_BRANCHES,
    include_current_accounts: bool = LEAP_IMPORT_INCLUDE_CURRENT_ACCOUNTS,
    import_supply_to_leap: bool = LEAP_IMPORT_SUPPLY_TO_LEAP,
    import_transformation_to_leap: bool = LEAP_IMPORT_TRANSFORMATION_TO_LEAP,
    import_transfers_to_leap: bool = LEAP_IMPORT_TRANSFERS_TO_LEAP,
) -> dict[str, list[Path]]:
    """Import the generated supply + transformation workbooks into LEAP via API."""
    if RUN_RESET_SUPPLY_AND_TRANSFORMATION_IMPORT_EXPORT and not include_current_accounts:
        print(
            "[INFO] Enabling Current Accounts fill pass for LEAP import because "
            "RUN_RESET_SUPPLY_AND_TRANSFORMATION_IMPORT_EXPORT=True."
        )
        include_current_accounts = True
    print(
        "[INFO] LEAP import toggles: "
        f"supply={import_supply_to_leap}, "
        f"transformation={import_transformation_to_leap}, "
        f"transfers={import_transfers_to_leap}, "
        f"include_current_accounts={include_current_accounts}"
    )
    if get_analysis_input_write_mode() == "api" and not leap_api.is_available():
        print("[INFO] LEAP API unavailable in this environment; skipping LEAP import.")
        return {"supply_imported": [], "transformation_imported": [], "transfer_imported": []}

    scenario_choices = workflow_common.resolve_import_scenarios(
        [str(item) for item in scenarios if str(item).strip()],
        import_scenarios,
    )
    if not scenario_choices:
        return {"supply_imported": [], "transformation_imported": [], "transfer_imported": []}

    supply_imported: list[Path] = []
    transformation_imported: list[Path] = []
    transfer_imported: list[Path] = []

    if import_supply_to_leap:
        for export_path in [Path(item) for item in supply_export_paths]:
            if not export_path.exists():
                continue
            for index, scenario in enumerate(scenario_choices):
                try:
                    supply_data_pipeline.run_supply_leap_import(
                        export_directory=export_path.parent,
                        filename=export_path.name,
                        scenario_to_run=scenario,
                        region=region,
                        handle_current_accounts=include_current_accounts and index == 0,
                        fill_branches=fill_branches,
                    )
                    supply_imported.append(export_path)
                except Exception as exc:
                    print(
                        f"[WARN] Supply LEAP import failed for {export_path.name} ({scenario}): {exc}"
                    )
    elif supply_export_paths:
        print(
            "[INFO] Skipping supply LEAP import; workbook(s) were still generated "
            "for manual LEAP import."
        )

    if import_transformation_to_leap:
        for export_path in [Path(item) for item in transformation_export_paths]:
            if not export_path.exists():
                continue
            try:
                available = transformation_workflow.list_export_scenarios(export_path)
            except Exception:
                available = []
            target_scenarios = [item for item in scenario_choices if item in available] or available
            for index, scenario in enumerate(target_scenarios):
                try:
                    transformation_workflow.import_transformation_workbook_to_leap(
                        export_directory=export_path.parent,
                        filename=export_path.name,
                        scenario_to_run=scenario,
                        region=region,
                        include_current_accounts=include_current_accounts and index == 0,
                        create_branches=create_branches and index == 0,
                        fill_branches=fill_branches,
                        raise_on_missing_branch=False,
                    )
                    transformation_imported.append(export_path)
                except Exception as exc:
                    print(
                        f"[WARN] Transformation LEAP import failed for {export_path.name} ({scenario}): {exc}"
                    )
    elif transformation_export_paths:
        print(
            "[INFO] Skipping transformation LEAP import; workbook(s) were still generated "
            "for manual LEAP import."
        )

    if import_transfers_to_leap:
        for export_path in [Path(item) for item in (transfer_export_paths or [])]:
            if not export_path.exists():
                continue
            legacy_paths = _find_legacy_transfer_branch_paths(export_path)
            if legacy_paths:
                sample = "; ".join(legacy_paths[:3])
                print(
                    "[WARN] Skipping transfer LEAP import for "
                    f"{export_path.name}: legacy generic transfer branches detected ({sample})."
                )
                continue
            try:
                available = transfers_workflow.list_export_scenarios(export_path)
            except Exception:
                available = []
            target_scenarios = [item for item in scenario_choices if item in available] or available
            for index, scenario in enumerate(target_scenarios):
                try:
                    transfers_workflow.import_transfer_workbook_to_leap(
                        export_directory=export_path.parent,
                        filename=export_path.name,
                        scenario_to_run=scenario,
                        region=region,
                        include_current_accounts=include_current_accounts and index == 0,
                        create_branches=create_branches and index == 0,
                        fill_branches=fill_branches,
                        raise_on_missing_branch=False,
                    )
                    transfer_imported.append(export_path)
                except Exception as exc:
                    print(
                        f"[WARN] Transfer LEAP import failed for {export_path.name} ({scenario}): {exc}"
                    )
    elif transfer_export_paths:
        print(
            "[INFO] Skipping transfer LEAP import; workbook(s) were still generated "
            "for manual LEAP import."
        )

    return {
        "supply_imported": supply_imported,
        "transformation_imported": transformation_imported,
        "transfer_imported": transfer_imported,
    }


def _resolve_existing_results_supply_export_paths(
    *,
    economies: Iterable[str],
    scenarios: Iterable[str],
    export_dir: Path | str = EXPORT_OUTPUT_DIR,
) -> tuple[list[Path], list[Path], list[Path]]:
    """Resolve expected supply/transformation/transfer export workbooks from disk."""
    economy_list = workflow_common.normalize_economies(economies or ECONOMIES)
    scenario_list = workflow_common.normalize_workflow_scenarios(scenarios, SCENARIOS)
    scenario_filename = supply_data_pipeline.format_scenario_label_for_filename(scenario_list)
    root = _resolve(export_dir)

    supply_paths: list[Path] = []
    transformation_paths: list[Path] = []
    transfer_paths: list[Path] = []
    missing: list[str] = []

    def _norm_token(text: str) -> str:
        return "".join(ch.lower() for ch in str(text or "") if ch.isalnum())

    def _pick_existing_workbook(
        *,
        prefix: str,
        economy: str,
        scenario_tokens: list[str],
    ) -> Path | None:
        econ_key = _norm_token(economy)
        token_keys = [_norm_token(token) for token in scenario_tokens if _norm_token(token)]
        candidates = sorted(root.glob(f"{prefix}_*.xlsx"))
        scored: list[tuple[int, Path]] = []
        for path in candidates:
            stem_key = _norm_token(path.stem)
            if econ_key and econ_key not in stem_key:
                continue
            token_hits = sum(1 for token in token_keys if token in stem_key)
            # Require at least one scenario token hit when scenarios were requested.
            if token_keys and token_hits == 0:
                continue
            scored.append((token_hits, path))
        if not scored:
            return None
        scored.sort(key=lambda item: (item[0], str(item[1]).lower()))
        return scored[-1][1]

    for economy in economy_list:
        supply_name = EXPORT_FILENAME_TEMPLATE.format(
            economy=str(economy),
            scenarios=scenario_filename,
        )
        transformation_name = transformation_workflow.format_export_filename(
            str(economy),
            scenario_list,
            TRANSFORMATION_EXPORT_FILENAME_TEMPLATE,
        )
        transfer_name = transfers_workflow.format_export_filename(
            str(economy),
            scenario_list,
            transfers_workflow.EXPORT_FILENAME_TEMPLATE,
        )

        supply_path = root / supply_name
        transformation_path = root / transformation_name
        transfer_path = root / transfer_name

        resolved_supply = supply_path if supply_path.exists() else _pick_existing_workbook(
            prefix="supply_leap_imports",
            economy=str(economy),
            scenario_tokens=scenario_list,
        )
        resolved_transformation = (
            transformation_path
            if transformation_path.exists()
            else _pick_existing_workbook(
                prefix="transformation_leap_imports",
                economy=str(economy),
                scenario_tokens=scenario_list,
            )
        )
        resolved_transfer = transfer_path if transfer_path.exists() else _pick_existing_workbook(
            prefix="transfer_leap_imports",
            economy=str(economy),
            scenario_tokens=scenario_list,
        )

        if resolved_supply is not None:
            supply_paths.append(resolved_supply)
        else:
            missing.append(str(supply_path))
        if resolved_transformation is not None:
            transformation_paths.append(resolved_transformation)
        else:
            missing.append(str(transformation_path))
        if resolved_transfer is not None:
            transfer_paths.append(resolved_transfer)
        else:
            missing.append(str(transfer_path))

    if missing:
        preview = "\n".join(missing[:12])
        raise FileNotFoundError(
            "Resume import could not find required export workbook(s). "
            f"First missing paths:\n{preview}"
        )
    return supply_paths, transformation_paths, transfer_paths


def resume_results_linked_leap_import_from_existing_exports(
    *,
    economies: Iterable[str] | None = None,
    scenarios: Iterable[str] | None = None,
    import_scenarios: Iterable[str] | str | None = LEAP_IMPORT_SCENARIOS,
    export_dir: Path | str = EXPORT_OUTPUT_DIR,
    region: str = LEAP_IMPORT_REGION,
    create_branches: bool = LEAP_IMPORT_CREATE_BRANCHES,
    fill_branches: bool = LEAP_IMPORT_FILL_BRANCHES,
    include_current_accounts: bool = LEAP_IMPORT_INCLUDE_CURRENT_ACCOUNTS,
    import_supply_to_leap: bool = LEAP_IMPORT_SUPPLY_TO_LEAP,
    import_transformation_to_leap: bool = LEAP_IMPORT_TRANSFORMATION_TO_LEAP,
    import_transfers_to_leap: bool = LEAP_IMPORT_TRANSFERS_TO_LEAP,
) -> dict[str, object]:
    """
    Resume only the LEAP import step using already-generated export workbooks.

    Use this after a prior workflow run reached export generation but failed or
    was interrupted during LEAP import.
    """
    os.environ["LEAP_IMPORT_LOG_LEVEL"] = str(LEAP_IMPORT_LOG_LEVEL).strip()
    os.environ["LEAP_IMPORT_WARNING_PRINT_LIMIT"] = str(LEAP_IMPORT_WARNING_PRINT_LIMIT)
    if RUN_LEAP_FUEL_BRANCH_PROBE_AT_START:
        refresh_fuel_branch_catalog_from_leap(output_path=LEAP_FUEL_BRANCH_PROBE_OUTPUT_PATH)

    economy_list = workflow_common.normalize_economies(economies or ECONOMIES)
    scenario_list = workflow_common.normalize_workflow_scenarios(scenarios, SCENARIOS)
    supply_paths, transformation_paths, transfer_paths = _resolve_existing_results_supply_export_paths(
        economies=economy_list,
        scenarios=scenario_list,
        export_dir=export_dir,
    )
    print(
        "[INFO] Resuming LEAP import from existing exports: "
        f"supply={len(supply_paths)}, transformation={len(transformation_paths)}, transfers={len(transfer_paths)}"
    )
    leap_import_result = run_results_linked_leap_import(
        supply_paths,
        transformation_paths,
        transfer_export_paths=transfer_paths,
        scenarios=scenario_list,
        import_scenarios=import_scenarios,
        region=region,
        create_branches=create_branches,
        fill_branches=fill_branches,
        include_current_accounts=include_current_accounts,
        import_supply_to_leap=import_supply_to_leap,
        import_transformation_to_leap=import_transformation_to_leap,
        import_transfers_to_leap=import_transfers_to_leap,
    )
    return {
        "supply_export_paths": supply_paths,
        "transformation_export_paths": transformation_paths,
        "transfer_export_paths": transfer_paths,
        "leap_import_result": leap_import_result,
    }


def _filter_transformation_workbook_to_trade_targets(
    workbook_path: Path | str,
    allowed_variables: tuple[str, ...] = ("Import Target", "Export Target"),
) -> None:
    """Keep only trade-target rows in transformation LEAP export sheets."""
    path = _resolve(workbook_path)
    if not path.exists():
        return
    xl = pd.ExcelFile(path)
    allowed = {str(item).strip().lower() for item in allowed_variables if str(item).strip()}
    output_sheets: dict[str, pd.DataFrame] = {}

    def _find_header_row(raw: pd.DataFrame) -> int | None:
        for idx in range(len(raw.index)):
            values = {_normalize_template_header_value(item).lower() for item in raw.iloc[idx].tolist()}
            if "branch path" in values and "variable" in values:
                return int(idx)
        return None

    for sheet_name in xl.sheet_names:
        raw = pd.read_excel(path, sheet_name=sheet_name, header=None)
        if sheet_name not in {"LEAP", "FOR_VIEWING"}:
            output_sheets[sheet_name] = raw
            continue
        header_row = _find_header_row(raw)
        if header_row is None:
            output_sheets[sheet_name] = raw
            continue
        header_values = raw.iloc[header_row].tolist()
        preamble = raw.iloc[: header_row + 1].copy()
        data = raw.iloc[header_row + 1 :].copy()
        data.columns = header_values

        variable_col = None
        for col in data.columns:
            if _normalize_template_header_value(col).lower() == "variable":
                variable_col = col
                break
        if variable_col is None:
            output_sheets[sheet_name] = raw
            continue
        keep_mask = data[variable_col].astype(str).str.strip().str.lower().isin(allowed)
        filtered_data = data.loc[keep_mask].copy()
        if filtered_data.empty:
            output_sheets[sheet_name] = preamble.reset_index(drop=True)
        else:
            filtered_data = filtered_data.reindex(columns=header_values)
            filtered_data.columns = list(range(len(filtered_data.columns)))
            preamble.columns = list(range(len(preamble.columns)))
            output_sheets[sheet_name] = pd.concat([preamble, filtered_data], ignore_index=True)

    with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
        for sheet_name in xl.sheet_names:
            output_sheets[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False, header=False)


def _read_leap_sheet_data_rows(workbook_path: Path | str, sheet_name: str = "LEAP") -> pd.DataFrame:
    """Read data rows from a LEAP-format export workbook sheet."""
    path = _resolve(workbook_path)
    if not path.exists():
        return pd.DataFrame()
    raw = pd.read_excel(path, sheet_name=sheet_name, header=None)
    header_row = None
    for idx in range(len(raw.index)):
        values = {_normalize_template_header_value(item).lower() for item in raw.iloc[idx].tolist()}
        if "branch path" in values and "variable" in values:
            header_row = int(idx)
            break
    if header_row is None:
        return pd.DataFrame()
    data = raw.iloc[header_row + 1 :].copy()
    data.columns = raw.iloc[header_row].tolist()
    if "Branch Path" not in data.columns:
        return pd.DataFrame()
    data = data[data["Branch Path"].notna()].copy()
    return data


def _read_branch_variable_rows(
    source_path: Path | str,
    sheet_name: str = "Export",
) -> pd.DataFrame:
    """Read a generic branch-variable table (xlsx/csv) with a discoverable header row."""
    path = _resolve(source_path)
    if not path.exists():
        return pd.DataFrame()

    suffix = path.suffix.lower()
    if suffix == ".csv":
        df = pd.read_csv(path)
        if {"Branch Path", "Variable"}.issubset(df.columns):
            return df.copy()
        return pd.DataFrame()

    raw = pd.read_excel(path, sheet_name=sheet_name, header=None)
    header_row = None
    for idx in range(len(raw.index)):
        values = {_normalize_template_header_value(item).lower() for item in raw.iloc[idx].tolist()}
        if "branch path" in values and "variable" in values:
            header_row = int(idx)
            break
    if header_row is None:
        return pd.DataFrame()
    data = raw.iloc[header_row + 1 :].copy()
    data.columns = raw.iloc[header_row].tolist()
    if "Branch Path" not in data.columns:
        return pd.DataFrame()
    data = data[data["Branch Path"].notna()].copy()
    return data


def _extract_catalog_rows_from_full_model_export(
    source_path: Path | str = FULL_MODEL_EXPORT_CATALOG_PATH,
    sheet_name: str = FULL_MODEL_EXPORT_CATALOG_SHEET,
) -> list[dict[str, object]]:
    """Parse full-model export into transformation/supply fuel catalog rows."""
    path = _resolve(source_path)
    if not path.exists():
        return []
    try:
        data = _read_branch_variable_rows(path, sheet_name=sheet_name)
    except Exception as exc:
        print(f"[WARN] Failed reading full model export catalog source {path}: {exc}")
        return []
    if data.empty:
        return []

    rows: list[dict[str, object]] = []

    def _parts(path_value: str) -> list[str]:
        return [part.strip() for part in str(path_value or "").split("\\") if str(part or "").strip()]

    for _, row in data.iterrows():
        branch_path = str(row.get("Branch Path") or "").strip()
        if not branch_path:
            continue
        variable = str(row.get("Variable") or "")
        scenario = str(row.get("Scenario") or "")
        parts = _parts(branch_path)
        if len(parts) < 2:
            continue

        if parts[0].lower() == "transformation":
            module = parts[1]
            fuel_group = ""
            fuel_name = ""
            for marker in ("Output Fuels", "Feedstock Fuels", "Auxiliary Fuels"):
                if marker in parts:
                    marker_index = parts.index(marker)
                    if marker_index + 1 < len(parts):
                        fuel_group = marker
                        fuel_name = parts[marker_index + 1]
                    break
            if fuel_name:
                rows.append(
                    {
                        "catalog_type": "transformation",
                        "source_workbook": path.name,
                        "scenario": scenario,
                        "module_or_root": module,
                        "fuel_group": fuel_group,
                        "fuel_name": fuel_name,
                        "branch_path": branch_path,
                        "variable": variable,
                        "catalog_source": "full_model_export",
                        "probe_status": "",
                    }
                )
            continue

        if parts[0].lower() == "resources" and len(parts) >= 3:
            root = parts[1]
            if root.lower() not in {"primary", "secondary"}:
                continue
            fuel_name = parts[2]
            rows.append(
                {
                    "catalog_type": "supply",
                    "source_workbook": path.name,
                    "scenario": scenario,
                    "module_or_root": root.title(),
                    "fuel_group": "",
                    "fuel_name": fuel_name,
                    "branch_path": branch_path,
                    "variable": variable,
                    "catalog_source": "full_model_export",
                    "probe_status": "",
                }
            )

    return rows


def _safe_leap_branch(app, path: str):
    """Return a LEAP branch object or None without raising."""
    branch_path = str(path or "").strip()
    if not branch_path:
        return None
    try:
        branches = app.Branches
        if not branches.Exists(branch_path):
            return None
        return branches.Item(branch_path)
    except Exception:
        return None


def _list_leap_child_branches(parent_branch) -> list[tuple[str, str]]:
    """List child branches as (name, full_path)."""
    rows: list[tuple[str, str]] = []
    if parent_branch is None:
        return rows
    try:
        children = parent_branch.Children
        count = int(children.Count)
    except Exception:
        return rows
    for idx in range(1, count + 1):
        try:
            child = children.Item(idx)
        except Exception:
            continue
        try:
            name = str(child.Name).strip()
        except Exception:
            name = ""
        try:
            full_name = str(child.FullName).strip()
        except Exception:
            full_name = ""
        if not name and full_name and "\\" in full_name:
            name = full_name.rsplit("\\", 1)[-1].strip()
        if name:
            rows.append((name, full_name or name))
    return rows


def _probe_branch_variable_expression(branch_obj, variable_candidates: Iterable[str]) -> tuple[str, str]:
    """Try candidate variables and read expression/value-like field to touch the branch."""
    for var_name in variable_candidates:
        candidate = str(var_name or "").strip()
        if not candidate:
            continue
        try:
            variable = branch_obj.Variable(candidate)
            if variable is None:
                continue
            # Touch one read path to validate branch-variable extraction.
            try:
                _ = str(variable.Expression)
            except Exception:
                _ = ""
            return candidate, "ok"
        except Exception:
            continue
    return "", "variable_not_found"


def refresh_fuel_branch_catalog_from_leap(
    output_path: Path | str = LEAP_FUEL_BRANCH_PROBE_OUTPUT_PATH,
) -> Path | None:
    """Touch transformation/supply fuel branches in LEAP and write a live probe CSV."""
    if get_analysis_input_write_mode() == "workbook":
        print(
            "[WORKBOOK MODE] Skipping live fuel-branch probe because it reads "
            "Analysis-view branches via LEAP API."
        )
        return None
    if not leap_api.is_available():
        print("[INFO] LEAP API unavailable; skipping live fuel-branch probe.")
        return None

    app = leap_api.connect()
    if app is None:
        print("[WARN] Failed to connect to LEAP for fuel-branch probe.")
        return None

    rows: list[dict[str, object]] = []
    try:
        active_scenario = str(getattr(app, "ActiveScenario", "") or "")
    except Exception:
        active_scenario = ""

    # Transformation module fuel branches.
    transformation_root = _safe_leap_branch(app, "Transformation")
    for module_name, module_full in _list_leap_child_branches(transformation_root):
        module_path = module_full or f"Transformation\\{module_name}"
        for fuel_group, probe_vars in (
            ("Output Fuels", ("Import Target", "Export Target", "Output Share", "Output")),
            ("Feedstock Fuels", ("Feedstock Fuel Share", "Inputs", "Output")),
            ("Auxiliary Fuels", ("Auxiliary Fuel Use", "Inputs", "Output")),
        ):
            group_path = f"{module_path}\\{fuel_group}"
            group_branch = _safe_leap_branch(app, group_path)
            if group_branch is None:
                continue
            for fuel_name, fuel_full in _list_leap_child_branches(group_branch):
                fuel_path = fuel_full or f"{group_path}\\{fuel_name}"
                fuel_branch = _safe_leap_branch(app, fuel_path)
                if fuel_branch is None:
                    continue
                variable_used, status = _probe_branch_variable_expression(fuel_branch, probe_vars)
                rows.append(
                    {
                        "catalog_type": "transformation",
                        "source_workbook": "__leap_probe__",
                        "scenario": active_scenario,
                        "module_or_root": module_name,
                        "fuel_group": fuel_group,
                        "fuel_name": fuel_name,
                        "branch_path": fuel_path,
                        "variable": variable_used,
                        "catalog_source": "leap_probe",
                        "probe_status": status,
                    }
                )

    # Supply fuel branches.
    for root_name in ("Primary", "Secondary"):
        root_path = f"Resources\\{root_name}"
        root_branch = _safe_leap_branch(app, root_path)
        if root_branch is None:
            continue
        for fuel_name, fuel_full in _list_leap_child_branches(root_branch):
            fuel_path = fuel_full or f"{root_path}\\{fuel_name}"
            fuel_branch = _safe_leap_branch(app, fuel_path)
            if fuel_branch is None:
                continue
            variable_used, status = _probe_branch_variable_expression(
                fuel_branch,
                ("Imports", "Exports", "Indigenous Production", "Unmet Requirements"),
            )
            rows.append(
                {
                    "catalog_type": "supply",
                    "source_workbook": "__leap_probe__",
                    "scenario": active_scenario,
                    "module_or_root": root_name,
                    "fuel_group": "",
                    "fuel_name": fuel_name,
                    "branch_path": fuel_path,
                    "variable": variable_used,
                    "catalog_source": "leap_probe",
                    "probe_status": status,
                }
            )

    out = _resolve(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    probe_df = pd.DataFrame(rows)
    if not probe_df.empty:
        probe_df = (
            probe_df.drop_duplicates(
                subset=[
                    "catalog_type",
                    "module_or_root",
                    "fuel_group",
                    "fuel_name",
                    "branch_path",
                ]
            )
            .sort_values(["catalog_type", "module_or_root", "fuel_group", "fuel_name"])
            .reset_index(drop=True)
        )
    probe_df.to_csv(out, index=False)
    print(f"[INFO] Wrote live LEAP fuel-branch probe catalog to {out}")
    return out


def _build_transformation_supply_fuel_catalog_df(
    *,
    transformation_export_paths: Iterable[Path],
    supply_export_paths: Iterable[Path],
    include_print_summary: bool = True,
) -> pd.DataFrame:
    """Build a transformation/supply fuel catalog dataframe."""
    rows: list[dict[str, object]] = []

    if USE_FULL_MODEL_EXPORT_CATALOG_SOURCE:
        full_model_rows = _extract_catalog_rows_from_full_model_export(
            source_path=FULL_MODEL_EXPORT_CATALOG_PATH,
            sheet_name=FULL_MODEL_EXPORT_CATALOG_SHEET,
        )
        if full_model_rows:
            rows.extend(full_model_rows)
            print(
                f"[INFO] Added {len(full_model_rows)} row(s) from full model export catalog source: "
                f"{_resolve(FULL_MODEL_EXPORT_CATALOG_PATH)}"
            )

    probe_path = _resolve(LEAP_FUEL_BRANCH_PROBE_OUTPUT_PATH)
    if probe_path.exists():
        try:
            probe_df = pd.read_csv(probe_path)
            if not probe_df.empty:
                for _, row in probe_df.iterrows():
                    rows.append(
                        {
                            "catalog_type": str(row.get("catalog_type") or ""),
                            "source_workbook": str(row.get("source_workbook") or "__leap_probe__"),
                            "scenario": str(row.get("scenario") or ""),
                            "module_or_root": str(row.get("module_or_root") or ""),
                            "fuel_group": str(row.get("fuel_group") or ""),
                            "fuel_name": str(row.get("fuel_name") or ""),
                            "branch_path": str(row.get("branch_path") or ""),
                            "variable": str(row.get("variable") or ""),
                            "catalog_source": str(row.get("catalog_source") or "leap_probe"),
                            "probe_status": str(row.get("probe_status") or ""),
                        }
                    )
        except Exception as exc:
            print(f"[WARN] Failed reading probe catalog {probe_path}: {exc}")

    def _parts(path_value: str) -> list[str]:
        return [part.strip() for part in str(path_value or "").split("\\") if str(part or "").strip()]

    for workbook in [Path(item) for item in transformation_export_paths]:
        if not workbook.exists():
            continue
        data = _read_leap_sheet_data_rows(workbook)
        if data.empty:
            continue
        for _, row in data.iterrows():
            branch_path = str(row.get("Branch Path") or "").strip()
            if not branch_path:
                continue
            parts = _parts(branch_path)
            if len(parts) < 4 or parts[0] != "Transformation":
                continue
            group_name = ""
            fuel_name = ""
            for marker in ("Output Fuels", "Feedstock Fuels", "Auxiliary Fuels"):
                if marker in parts:
                    marker_index = parts.index(marker)
                    if marker_index + 1 < len(parts):
                        group_name = marker
                        fuel_name = parts[marker_index + 1]
                    break
            if not fuel_name:
                continue
            rows.append(
                {
                    "catalog_type": "transformation",
                    "source_workbook": workbook.name,
                    "scenario": str(row.get("Scenario") or ""),
                    "module_or_root": parts[1],
                    "fuel_group": group_name,
                    "fuel_name": fuel_name,
                    "branch_path": branch_path,
                    "variable": str(row.get("Variable") or ""),
                    "catalog_source": "export",
                    "probe_status": "",
                }
            )

    for workbook in [Path(item) for item in supply_export_paths]:
        if not workbook.exists():
            continue
        data = _read_leap_sheet_data_rows(workbook)
        if data.empty:
            continue
        for _, row in data.iterrows():
            branch_path = str(row.get("Branch Path") or "").strip()
            if not branch_path:
                continue
            parts = _parts(branch_path)
            if len(parts) < 3 or parts[0] != "Resources":
                continue
            root_name = parts[1]
            if root_name not in {"Primary", "Secondary"}:
                continue
            rows.append(
                {
                    "catalog_type": "supply",
                    "source_workbook": workbook.name,
                    "scenario": str(row.get("Scenario") or ""),
                    "module_or_root": root_name,
                    "fuel_group": "",
                    "fuel_name": parts[2],
                    "branch_path": branch_path,
                    "variable": str(row.get("Variable") or ""),
                    "catalog_source": "export",
                    "probe_status": "",
                }
            )

    catalog_df = pd.DataFrame(rows)
    if catalog_df.empty:
        catalog_df = pd.DataFrame(
            columns=[
                "catalog_type",
                "source_workbook",
                "scenario",
                "module_or_root",
                "fuel_group",
                "fuel_name",
                "branch_path",
                "variable",
                "catalog_source",
                "probe_status",
            ]
        )
    else:
        catalog_df = (
            catalog_df.drop_duplicates(
                subset=[
                    "catalog_type",
                    "source_workbook",
                    "scenario",
                    "module_or_root",
                    "fuel_group",
                    "fuel_name",
                    "branch_path",
                    "variable",
                    "catalog_source",
                    "probe_status",
                ]
            )
            .sort_values(
                by=[
                    "catalog_type",
                    "catalog_source",
                    "module_or_root",
                    "fuel_group",
                    "fuel_name",
                    "branch_path",
                    "variable",
                ]
            )
            .reset_index(drop=True)
        )
    transformation_subset = catalog_df[catalog_df["catalog_type"] == "transformation"].copy()
    if include_print_summary and not transformation_subset.empty:
        print("\n=== Transformation Fuels By Module (catalog) ===")
        summary = (
            transformation_subset.groupby(["module_or_root", "fuel_group"], dropna=False)["fuel_name"]
            .nunique()
            .reset_index(name="unique_fuels")
        )
        for _, row in summary.sort_values(["module_or_root", "fuel_group"]).iterrows():
            print(
                f" - {row['module_or_root']} | {row['fuel_group']}: "
                f"{int(row['unique_fuels'])} fuel(s)"
            )

    supply_subset = catalog_df[catalog_df["catalog_type"] == "supply"].copy()
    if include_print_summary and not supply_subset.empty:
        print("\n=== Supply Fuels By Branch Root (catalog) ===")
        summary = (
            supply_subset.groupby(["module_or_root"], dropna=False)["fuel_name"]
            .nunique()
            .reset_index(name="unique_fuels")
        )
        for _, row in summary.sort_values(["module_or_root"]).iterrows():
            print(f" - {row['module_or_root']}: {int(row['unique_fuels'])} fuel(s)")

    return catalog_df


def _build_transformation_supply_fuel_catalog(
    *,
    transformation_export_paths: Iterable[Path],
    supply_export_paths: Iterable[Path],
    output_dir: Path | str = RESULTS_CHECKS_DIR,
) -> Path:
    """Build and save a CSV catalog of transformation/supply fuels by branch root."""
    output_path = _resolve(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    catalog_path = output_path / "transformation_supply_fuel_branch_catalog.csv"
    catalog_df = _build_transformation_supply_fuel_catalog_df(
        transformation_export_paths=transformation_export_paths,
        supply_export_paths=supply_export_paths,
        include_print_summary=True,
    )
    catalog_df.to_csv(catalog_path, index=False)
    print(f"[INFO] Wrote transformation/supply fuel catalog to {catalog_path}")
    return catalog_path


def build_year_balance_table(
    reconciliation_table: pd.DataFrame,
    year: int,
    economies: Iterable[str] | None = None,
    scenarios: Iterable[str] | None = None,
) -> pd.DataFrame:
    """Return a balance-table style long table for one year."""
    if reconciliation_table.empty:
        return pd.DataFrame(
            columns=[
                "economy",
                "scenario",
                "year",
                "esto_product",
                "balance_component",
                "value",
                "sign_convention",
            ]
        )

    working = reconciliation_table.copy()
    working["year"] = pd.to_numeric(working["year"], errors="coerce").astype("Int64")
    working = working[working["year"] == int(year)].copy()
    if economies:
        economy_list = {str(item) for item in economies}
        working = working[working["economy"].astype(str).isin(economy_list)].copy()
    if scenarios:
        scenario_list = {str(item) for item in scenarios}
        working = working[working["scenario"].astype(str).isin(scenario_list)].copy()
    if working.empty:
        return pd.DataFrame(
            columns=[
                "economy",
                "scenario",
                "year",
                "esto_product",
                "balance_component",
                "value",
                "sign_convention",
            ]
        )

    component_specs = [
        ("demand", "demand_value", "use_negative"),
        ("transformation_input", "transformation_input", "use_negative"),
        ("transformation_output", "constrained_transformation_output", "supply_positive"),
        ("transformation_losses", "transformation_losses", "use_negative"),
        ("production", "constrained_production", "supply_positive"),
        ("stock_changes", "stock_changes", "net_positive"),
        ("projected_imports", "projected_imports", "supply_positive"),
        ("projected_exports", "projected_exports", "use_negative"),
        ("adjusted_imports", "adjusted_imports", "supply_positive"),
        ("adjusted_exports", "adjusted_exports", "use_negative"),
        ("required_net_imports", "required_net_imports", "net_positive"),
        ("adjusted_net_imports", "adjusted_net_imports", "net_positive"),
        ("projected_net_imports", "projected_net_imports", "net_positive"),
        ("trade_adjustment", "trade_adjustment", "net_positive"),
        ("balance_residual", "adjusted_balance", "near_zero"),
    ]

    def _safe_number(value: object) -> float:
        numeric = pd.to_numeric(value, errors="coerce")
        if pd.isna(numeric):
            return 0.0
        return float(numeric)

    rows: list[dict[str, object]] = []
    for _, row in working.iterrows():
        base_record = {
            "economy": row["economy"],
            "scenario": row["scenario"],
            "year": int(row["year"]),
            "esto_product": row["esto_product"],
        }
        for component_name, column_name, sign_convention in component_specs:
            value = pd.to_numeric(row.get(column_name), errors="coerce")
            if pd.isna(value):
                continue
            signed_value = float(value)
            if sign_convention == "use_negative":
                signed_value = -abs(signed_value)
            elif sign_convention == "supply_positive":
                signed_value = abs(signed_value)
            rows.append(
                {
                    **base_record,
                    "balance_component": component_name,
                    "value": signed_value,
                    "raw_value": float(value),
                    "sign_convention": sign_convention,
                }
            )

        tpes_value = (
            _safe_number(row.get("constrained_production"))
            + _safe_number(row.get("stock_changes"))
            + _safe_number(row.get("adjusted_net_imports"))
            + _safe_number(row.get("constrained_transformation_output"))
            - abs(_safe_number(row.get("transformation_input")))
            - abs(_safe_number(row.get("transformation_losses")))
        )
        rows.append(
            {
                **base_record,
                "balance_component": "total_primary_energy_supply",
                "value": tpes_value,
                "raw_value": tpes_value,
                "sign_convention": "net_positive",
            }
        )
        final_balance = tpes_value - abs(_safe_number(row.get("demand_value")))
        rows.append(
            {
                **base_record,
                "balance_component": "final_balance_check",
                "value": final_balance,
                "raw_value": final_balance,
                "sign_convention": "near_zero",
            }
        )

    balance_table = pd.DataFrame(rows)
    if balance_table.empty:
        return balance_table

    totals = (
        balance_table.groupby(
            ["economy", "scenario", "year", "balance_component"],
            dropna=False,
            as_index=False,
        )
        .agg(value=("value", "sum"), raw_value=("raw_value", "sum"))
    )
    totals["esto_product"] = "Total"
    sign_lookup = {
        name: sign
        for name, _, sign in component_specs
    }
    sign_lookup["total_primary_energy_supply"] = "net_positive"
    sign_lookup["final_balance_check"] = "near_zero"
    totals["sign_convention"] = totals["balance_component"].map(sign_lookup).fillna("")

    balance_table = pd.concat([balance_table, totals], ignore_index=True, sort=False)
    balance_table = balance_table.sort_values(
        ["economy", "scenario", "esto_product", "balance_component"]
    ).reset_index(drop=True)
    return _zero_small_numeric_values(
        balance_table,
        label_columns=["economy", "scenario", "year", "esto_product", "balance_component", "sign_convention"],
        threshold=0.01,
    )


def save_year_balance_tables(
    reconciliation_table: pd.DataFrame,
    years: Iterable[int],
    output_dir: Path | str = YEARLY_BALANCE_DIR,
    economies: Iterable[str] | None = None,
    scenarios: Iterable[str] | None = None,
) -> list[Path]:
    """Write scenario/date-scoped CSV balance tables for the simple balance view."""
    output_path = _resolve(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    economy_list = [
        str(item).strip()
        for item in (
            economies
            if economies is not None
            else reconciliation_table.get("economy", pd.Series(dtype=str)).dropna().unique()
        )
        if str(item).strip()
    ]
    scenario_list = [
        str(item).strip()
        for item in (
            scenarios
            if scenarios is not None
            else reconciliation_table.get("scenario", pd.Series(dtype=str)).dropna().unique()
        )
        if str(item).strip()
    ]

    output_jobs: list[tuple[int, str, str, str, str, Path]] = []
    for year in years:
        year_int = int(year)
        for economy in economy_list:
            economy_token = _safe_filename_token(economy)
            for scenario in scenario_list:
                date_id, scenario_code = _balance_export_parts_for_scenario(scenario)
                stem = (
                    f"balance_table_{economy_token}_{_safe_filename_token(date_id)}_"
                    f"{_safe_filename_token(scenario_code)}_{year_int}"
                )
                output_jobs.append(
                    (
                        year_int,
                        economy,
                        scenario,
                        economy_token,
                        _safe_filename_token(scenario_code),
                        output_path / f"{stem}.csv",
                    )
                )

    saved_paths: list[Path] = []
    for year, economy, scenario, economy_token, scenario_code, csv_path in output_jobs:
        _archive_prior_year_balance_tables(
            output_path=output_path,
            economy_token=economy_token,
            scenario_code=scenario_code,
            year=int(year),
            current_csv_path=csv_path,
        )
        if csv_path.exists():
            saved_paths.append(csv_path)
            print(f"[INFO] Reusing existing yearly balance table CSV without overwrite: {csv_path}")
            continue
        table = build_year_balance_table(
            reconciliation_table,
            year=year,
            economies=[economy],
            scenarios=[scenario],
        )
        if table.empty:
            continue
        table.to_csv(csv_path, index=False)
        saved_paths.append(csv_path)
        print(f"Saved year balance table CSV to {csv_path}")
    return saved_paths


def _safe_filename_token(value: object) -> str:
    """Return a filesystem-safe token."""
    text = str(value or "").strip()
    if not text:
        return "item"
    safe = "".join(ch if ch.isalnum() or ch in {"_", "-"} else "_" for ch in text)
    return safe.strip("_") or "item"


def _archive_prior_year_balance_tables(
    *,
    output_path: Path,
    economy_token: str,
    scenario_code: str,
    year: int,
    current_csv_path: Path,
) -> list[Path]:
    """Move older matching yearly balance files out of the active output folder."""
    archive_dir = output_path / "archive"
    archived_paths: list[Path] = []
    current_stem = current_csv_path.stem
    patterns = [
        f"balance_table_{economy_token}_*_{scenario_code}_{int(year)}.csv",
        f"balance_table_{economy_token}_*_{scenario_code}_{int(year)}.xlsx",
    ]
    for pattern in patterns:
        for path in sorted(output_path.glob(pattern)):
            if path.name.startswith("~$"):
                continue
            if path.resolve() == current_csv_path.resolve():
                continue
            if path.suffix.lower() == ".xlsx" and path.stem == current_stem:
                pass
            elif path.stem == current_stem:
                continue
            archive_dir.mkdir(parents=True, exist_ok=True)
            target = archive_dir / path.name
            if target.exists():
                stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
                target = archive_dir / f"{path.stem}_{stamp}{path.suffix}"
            shutil.move(str(path), str(target))
            archived_paths.append(target)
            print(f"[INFO] Archived older yearly balance table to {target}")
    return archived_paths


def _balance_export_filename_parts(path: Path | str) -> tuple[str, str]:
    """Return (date_id, scenario_code) from a LEAP balance-export workbook name."""
    match = re.match(
        r"^full model output all years (?P<date_id>\d{5,8}) (?P<scenario>[A-Za-z]+)(?:\s[^.]*)?\.xlsx$",
        Path(path).name,
        flags=re.IGNORECASE,
    )
    if not match:
        return "unknown_date", "unknown_scenario"
    return match.group("date_id"), match.group("scenario").upper()


def _balance_export_parts_for_scenario(scenario: object) -> tuple[str, str]:
    """Return filename provenance for Reference/Target balance-demand source workbooks."""
    scenario_key = str(scenario or "").strip().lower()
    if scenario_key == "reference":
        return _balance_export_filename_parts(BALANCE_DEMAND_REF_WORKBOOK_PATH)
    if scenario_key == "target":
        return _balance_export_filename_parts(BALANCE_DEMAND_TGT_WORKBOOK_PATH)
    return "unknown_date", _safe_filename_token(scenario).upper()


def _zero_small_numeric_values(
    df: pd.DataFrame,
    *,
    label_columns: Iterable[str],
    threshold: float = 0.01,
) -> pd.DataFrame:
    """Set tiny numeric values to exactly zero for readability."""
    if df.empty:
        return df.copy()
    out = df.copy()
    label_set = {str(col) for col in label_columns}
    for column in out.columns:
        if str(column) in label_set:
            continue
        numeric = pd.to_numeric(out[column], errors="coerce")
        if numeric.notna().any():
            out[column] = numeric.where(numeric.abs() >= float(threshold), 0.0)
    return out


def _filter_balance_scenarios(scenarios: Iterable[str] | None) -> list[str]:
    """Return scenario labels excluding current-accounts style entries."""
    if scenarios is None:
        return []
    filtered: list[str] = []
    for value in scenarios:
        label = str(value or "").strip()
        if not label:
            continue
        if label.lower() in {"current accounts", "current account"}:
            continue
        filtered.append(label)
    return filtered


def _ensure_current_accounts_scenario(scenarios: Iterable[str] | None) -> list[str]:
    """Return scenarios with a canonical Current Accounts label appended if missing."""
    ordered: list[str] = []
    seen: set[str] = set()
    has_current_accounts = False
    for value in (scenarios or []):
        label = str(value or "").strip()
        if not label:
            continue
        key = label.lower()
        if key in seen:
            continue
        seen.add(key)
        ordered.append(label)
        if key in {"current accounts", "current account"}:
            has_current_accounts = True
    if not has_current_accounts:
        ordered.append("Current Accounts")
    return ordered


def _get_projection_value_for_flow_product(
    *,
    economy: str,
    flow: object,
    product: object,
    year: int,
) -> float:
    """Return the projected value for an ESTO flow/product pair in one year."""
    lookup = supply_data_pipeline.SUPPLY_PROJECTION_LOOKUP
    if lookup is None:
        return 0.0
    key = (
        supply_data_pipeline.normalize_economy_key(economy),
        str(flow or "").strip(),
        str(product or "").strip(),
    )
    if key not in lookup.index:
        return 0.0
    row = lookup.loc[key]
    if isinstance(row, pd.DataFrame):
        row = row.sum()
    value = pd.to_numeric(row.get(int(year), 0.0), errors="coerce")
    if pd.isna(value):
        return 0.0
    return float(value)


def _get_base_value_for_flow_product(
    *,
    base_df: pd.DataFrame,
    economy: str,
    flow: object,
    product: object,
    year: int,
) -> float:
    """Return the ESTO base-year value for an ESTO flow/product pair."""
    if base_df.empty:
        return 0.0
    if int(year) not in [col for col in base_df.columns if isinstance(col, int)]:
        return 0.0
    mask = (
        base_df.get("economy", pd.Series(index=base_df.index)).astype(str).eq(str(economy))
        & base_df.get("flows", pd.Series(index=base_df.index)).astype(str).eq(str(flow))
        & base_df.get("products", pd.Series(index=base_df.index)).astype(str).eq(str(product))
    )
    if not mask.any():
        return 0.0
    values = pd.to_numeric(base_df.loc[mask, int(year)], errors="coerce").fillna(0.0)
    return float(values.sum())


def _strip_esto_sector_prefix(label: object) -> str:
    """Remove leading numeric ESTO code prefixes for display rows."""
    text = str(label or "").strip()
    if not text:
        return ""
    parts = text.split(" ", 1)
    if len(parts) == 2 and all(part.isdigit() for part in parts[0].split(".")):
        return parts[1].strip()
    return text


def _build_conventional_row_backbone() -> list[str]:
    """Return the standard conventional balance row order."""
    preferred_rows = [
        "Production",
        "Imports",
        "Exports",
        "International marine bunkers",
        "International aviation bunkers",
        "Stock changes",
        "Total primary energy supply",
        "Transfers",
        "Upstream liquids transfers",
        "Refinery & blending transfers",
        "Transfers unallocated",
        "Recycled products",
        "Interproduct transfers",
        "Products transferred",
        "Gas separation",
        "Transformation nonspecified",
        "Total transformation sector",
        "Main activity producer",
        "Electricity plants",
        "CHP plants",
        "Heat plants",
        "Autoproducers",
        "Electricity plants (autoproducers)",
        "CHP plants (autoproducers)",
        "Heat plants (autoproducers)",
        "Heat pumps",
        "Electric boilers",
        "Chemical heat for electricity production",
        "Gas processing plants",
        "Gas works plants",
        "Liquefaction/regasification plants",
        "Natural gas blending plants",
        "Gas-to-liquids plants",
        "Oil refineries",
        "Coal transformation",
        "Coke ovens",
        "Blast furnaces",
        "Patent fuel plants",
        "BKB/PB plants",
        "Liquefaction (coal to oil)",
        "Petrochemical industry",
        "Biofuels processing",
        "Charcoal processing",
        "Hydrogen transformation",
        "Non-specified transformation",
        "Losses & own use",
        "Own Use",
        "Electricity, CHP and heat plants",
        "Gas works plants (own-use)",
        "Liquefaction/regasification plants (own-use)",
        "Gas-to-liquids plants (own-use)",
        "Coke ovens (own-use)",
        "Coal mines",
        "Blast furnaces (own-use)",
        "Patent fuel plants (own-use)",
        "BKB/PB plants (own-use)",
        "Liquefaction plants (Coal to Oil)",
        "Oil refineries (own-use)",
        "Oil and gas extraction",
        "Pump storage plants",
        "Nuclear industry",
        "Charcoal production plants",
        "Gasification plants for biogases",
        "Non-specified own uses",
        "Transmission and distribution losses",
        "Statistical discrepancy",
        "Total final consumption",
        "Total final energy consumption",
        "Industry sector",
        "Transport sector",
        "Other sector",
        "Non-energy use",
        "Total Final Energy Demand",
        "Unmet Requirements",
    ]

    dynamic_rows = [
        _strip_esto_sector_prefix(item)
        for item in ESTO_SECTORS
        if str(item or "").strip()
    ]

    normalized: list[str] = []
    seen: set[str] = set()
    for row in preferred_rows + dynamic_rows:
        name = str(row or "").strip()
        if not name or name in seen:
            continue
        seen.add(name)
        normalized.append(name)
    return normalized


def _normalize_conventional_sector_name(label: object) -> str:
    """Map internal labels to conventional balance row names."""
    text = _strip_esto_sector_prefix(label)
    replacements = {
        "NG Liquefaction": "Liquefaction/regasification plants",
        "NG Regasification": "Liquefaction/regasification plants",
        "09_13_hydrogen_transformation": "Hydrogen transformation",
        "Total Primary Supply": "Total primary energy supply",
        "Total Transformation": "Total transformation sector",
        "Transmission and Distribution": "Transmission and distribution losses",
        "Upstream & refinery transfers": "Transfers unallocated",
    }
    return replacements.get(text, text)


def _get_conventional_section_layout() -> list[tuple[str, list[str]]]:
    """Return ordered report sections and their canonical row labels."""
    return [
        (
            "Supply",
            [
                "Production",
                "Imports",
                "Exports",
                "International marine bunkers",
                "International aviation bunkers",
                "Stock changes",
                "Total primary energy supply",
            ],
        ),
        (
            "Transfers",
            [
                "Transfers",
                "Upstream liquids transfers",
                "Refinery & blending transfers",
                "Transfers unallocated",
                "Recycled products",
                "Interproduct transfers",
                "Products transferred",
                "Gas separation",
                "Transformation nonspecified",
            ],
        ),
        (
            "Transformation",
            [
                "Main activity producer",
                "Electricity plants",
                "CHP plants",
                "Heat plants",
                "Autoproducers",
                "Electricity plants (autoproducers)",
                "CHP plants (autoproducers)",
                "Heat plants (autoproducers)",
                "Heat pumps",
                "Electric boilers",
                "Chemical heat for electricity production",
                "Gas processing plants",
                "Gas works plants",
                "Liquefaction/regasification plants",
                "Natural gas blending plants",
                "Gas-to-liquids plants",
                "Oil refineries",
                "Coal transformation",
                "Coke ovens",
                "Blast furnaces",
                "Patent fuel plants",
                "BKB/PB plants",
                "Liquefaction (coal to oil)",
                "Petrochemical industry",
                "Biofuels processing",
                "Charcoal processing",
                "Hydrogen transformation",
                "Non-specified transformation",
                "Total transformation sector",
            ],
        ),
        (
            "Losses",
            [
                "Losses & own use",
                "Own Use",
                "Electricity, CHP and heat plants",
                "Gas works plants (own-use)",
                "Liquefaction/regasification plants (own-use)",
                "Gas-to-liquids plants (own-use)",
                "Coke ovens (own-use)",
                "Coal mines",
                "Blast furnaces (own-use)",
                "Patent fuel plants (own-use)",
                "BKB/PB plants (own-use)",
                "Liquefaction plants (Coal to Oil)",
                "Oil refineries (own-use)",
                "Oil and gas extraction",
                "Pump storage plants",
                "Nuclear industry",
                "Charcoal production plants",
                "Gasification plants for biogases",
                "Non-specified own uses",
                "Transmission and distribution losses",
                "Statistical discrepancy",
            ],
        ),
        (
            "Demand",
            [
                "Total final consumption",
                "Total final energy consumption",
                "Industry sector",
                "Transport sector",
                "Other sector",
                "Non-energy use",
                "Industry",
                "Transport",
                "Buildings",
                "Agriculture",
                "Other demand",
                "Total Final Energy Demand",
            ],
        ),
        (
            "Checks",
            [
                "Unmet Requirements",
            ],
        ),
    ]


def _build_conventional_section_lookup() -> tuple[dict[str, str], list[str]]:
    """Return row->section mapping and ordered row list from the report layout."""
    layout = _get_conventional_section_layout()
    row_to_section: dict[str, str] = {}
    ordered_rows: list[str] = []
    for section_name, rows in layout:
        for row in rows:
            row_to_section[row] = section_name
            ordered_rows.append(row)
    return row_to_section, ordered_rows


def _infer_top_level_demand_category(
    *,
    primary_sector_code: object,
    esto_flow: object,
) -> str:
    """Map a demand row to a broad category using mapped 9th/ESTO sector levels."""
    flow_label = str(esto_flow or "").strip()
    if not flow_label:
        flow_label = SECTOR_TO_ESTO_FLOW_LOOKUP.get(str(primary_sector_code or "").strip(), "")
    esto_text = flow_label.lower()
    seq = _sector_code_sequence(primary_sector_code)
    if esto_text.startswith("14.") or (seq and seq[0] == 14):
        return "Industry"
    if esto_text.startswith("15.") or (seq and seq[0] == 15):
        return "Transport"
    if esto_text.startswith("16.01") or esto_text.startswith("16.02"):
        return "Buildings"
    if esto_text.startswith("16.03") or esto_text.startswith("16.04"):
        return "Agriculture"
    if esto_text.startswith("17.") or (seq and seq[0] == 17):
        return "Non-energy use"
    if esto_text.startswith("16.05") or (seq and seq[0] == 16):
        return "Other demand"
    return "Other demand"


def _prepare_demand_rows_for_balance(
    demand: pd.DataFrame,
    *,
    drop_parent_rows: bool = DROP_PARENT_DEMAND_ROWS_WHEN_CHILDREN_PRESENT,
    include_top_level_categories: bool = INCLUDE_TOP_LEVEL_DEMAND_CATEGORY_ROWS,
    drop_disaggregated_rows: bool = DROP_DISAGGREGATED_DEMAND_SECTORS,
) -> pd.DataFrame:
    """Normalize demand rows using mapped hierarchy, drop parent rows, and add top-level aggregates."""
    if demand.empty:
        return pd.DataFrame(columns=["sector", "esto_product", "value", "is_top_level_aggregate"])

    working = (
        demand.groupby(
            ["sheet", "sector_code_9th", "esto_flow", "esto_product"],
            dropna=False,
            as_index=False,
        )["demand_value"]
        .sum(min_count=1)
    )
    working["sector"] = working["sheet"].map(_normalize_conventional_sector_name)
    working["primary_sector_code"] = working["sector_code_9th"].map(_select_primary_sector_code)
    working["sector_seq"] = working["primary_sector_code"].map(_sector_code_sequence)
    working["value"] = -working["demand_value"].abs()
    working["is_top_level_aggregate"] = False
    working = working[
        [
            "sector",
            "esto_product",
            "value",
            "primary_sector_code",
            "sector_seq",
            "esto_flow",
            "is_top_level_aggregate",
        ]
    ].copy()

    if drop_parent_rows:
        unique_codes = [
            seq
            for seq in {
                tuple(value) for value in working["sector_seq"].tolist()
                if isinstance(value, tuple) and value
            }
        ]
        parent_codes: set[tuple[int, ...]] = set()
        for seq in unique_codes:
            if any(
                len(other) > len(seq) and other[: len(seq)] == seq
                for other in unique_codes
            ):
                parent_codes.add(seq)
        if parent_codes:
            working = working[
                ~working["sector_seq"].map(
                    lambda seq: isinstance(seq, tuple) and tuple(seq) in parent_codes
                )
            ].copy()

    extra_rows: list[dict[str, object]] = []
    if include_top_level_categories and not working.empty:
        category_totals = (
            working.assign(
                top_level_category=working.apply(
                    lambda row: _infer_top_level_demand_category(
                        primary_sector_code=row.get("primary_sector_code"),
                        esto_flow=row.get("esto_flow"),
                    ),
                    axis=1,
                )
            )
            .groupby(["top_level_category", "esto_product"], dropna=False, as_index=False)["value"]
            .sum(min_count=1)
        )
        for _, row in category_totals.iterrows():
            extra_rows.append(
                {
                    "sector": str(row["top_level_category"]),
                    "esto_product": row["esto_product"],
                    "value": float(row["value"]),
                    "is_top_level_aggregate": True,
                }
            )
    if extra_rows:
        working = pd.concat([working, pd.DataFrame(extra_rows)], ignore_index=True, sort=False)

    if drop_disaggregated_rows:
        working = working[working["is_top_level_aggregate"].fillna(False)].copy()

    return (
        working.groupby(
            ["sector", "esto_product", "is_top_level_aggregate"],
            dropna=False,
            as_index=False,
        )["value"].sum(min_count=1)
    )


def build_reference_demand_rows_for_balance(
    sector_demand_table: pd.DataFrame,
    *,
    economy: str,
    scenario: str,
    year: int,
    base_df: pd.DataFrame,
) -> pd.DataFrame:
    """Build demand rows from ESTO base or projected values using dashboard mappings."""
    if sector_demand_table.empty:
        return pd.DataFrame(columns=["sector", "esto_product", "value", "is_top_level_aggregate"])

    working = sector_demand_table.copy()
    working["year"] = pd.to_numeric(working["year"], errors="coerce").astype("Int64")
    working = working[
        (working["year"] == int(year))
        & (working["economy"].astype(str) == str(economy))
        & (working["scenario"].astype(str) == str(scenario))
    ].copy()
    if working.empty:
        return pd.DataFrame(columns=["sector", "esto_product", "value", "is_top_level_aggregate"])

    def _resolve_source_value(row: pd.Series) -> float:
        if int(year) <= BASE_YEAR:
            return _get_base_value_for_flow_product(
                base_df=base_df,
                economy=str(economy),
                flow=row.get("esto_flow"),
                product=row.get("esto_product"),
                year=int(year),
            )
        return _get_projection_value_for_flow_product(
            economy=str(economy),
            flow=row.get("esto_flow"),
            product=row.get("esto_product"),
            year=int(year),
        )

    working["demand_value"] = working.apply(_resolve_source_value, axis=1)
    return _prepare_demand_rows_for_balance(
        working,
        drop_parent_rows=DROP_PARENT_DEMAND_ROWS_WHEN_CHILDREN_PRESENT,
        include_top_level_categories=INCLUDE_TOP_LEVEL_DEMAND_CATEGORY_ROWS,
        drop_disaggregated_rows=DROP_DISAGGREGATED_DEMAND_SECTORS,
    )


def build_reference_conventional_balance_matrix(
    *,
    reconciliation_table: pd.DataFrame,
    sector_demand_table: pd.DataFrame,
    transformation_sector_table: pd.DataFrame,
    supply_primary_table: pd.DataFrame,
    base_df: pd.DataFrame,
    year: int,
    economy: str,
    scenario: str,
) -> pd.DataFrame:
    """Build the same matrix shape using source ESTO/9th values before adjustments."""
    year_value = int(year)
    economy_value = str(economy)
    scenario_value = str(scenario)

    recon = reconciliation_table.copy()
    recon["year"] = pd.to_numeric(recon["year"], errors="coerce").astype("Int64")
    recon = recon[
        (recon["year"] == year_value)
        & (recon["economy"].astype(str) == economy_value)
        & (recon["scenario"].astype(str) == scenario_value)
    ].copy()

    trans = transformation_sector_table.copy()
    trans["year"] = pd.to_numeric(trans["year"], errors="coerce").astype("Int64")
    trans = trans[
        (trans["year"] == year_value)
        & (trans["economy"].astype(str) == economy_value)
    ].copy()

    supply = supply_primary_table.copy()
    supply["year"] = pd.to_numeric(supply["year"], errors="coerce").astype("Int64")
    supply = supply[
        (supply["year"] == year_value)
        & (supply["economy"].astype(str) == economy_value)
    ].copy()

    demand_grouped = build_reference_demand_rows_for_balance(
        sector_demand_table,
        economy=economy_value,
        scenario=scenario_value,
        year=year_value,
        base_df=base_df,
    )

    def _safe_number(value: object) -> float:
        numeric = pd.to_numeric(value, errors="coerce")
        if pd.isna(numeric):
            return 0.0
        return float(numeric)

    row_entries: list[dict[str, object]] = []

    if not supply.empty:
        for _, row in supply.iterrows():
            row_entries.append({"sector": "Production", "esto_product": row["esto_product"], "value": _safe_number(row.get("production"))})
            row_entries.append({"sector": "Stock changes", "esto_product": row["esto_product"], "value": _safe_number(row.get("stock_changes"))})

    if not recon.empty:
        for _, row in recon.iterrows():
            row_entries.append({"sector": "Imports", "esto_product": row["esto_product"], "value": abs(_safe_number(row.get("projected_imports")))})
            row_entries.append({"sector": "Exports", "esto_product": row["esto_product"], "value": -abs(_safe_number(row.get("projected_exports")))})

    if not supply.empty or not recon.empty:
        merged_primary = supply.merge(
            recon[["esto_product", "projected_imports", "projected_exports"]],
            on="esto_product",
            how="outer",
        )
        for _, row in merged_primary.iterrows():
            total_primary_supply = (
                _safe_number(row.get("production"))
                + _safe_number(row.get("projected_imports"))
                - _safe_number(row.get("projected_exports"))
                + _safe_number(row.get("stock_changes"))
            )
            row_entries.append({"sector": "Total primary energy supply", "esto_product": row["esto_product"], "value": total_primary_supply})

    trans_grouped = pd.DataFrame(columns=["sector", "esto_product", "value"])
    if not trans.empty:
        trans_grouped = trans.groupby(["sector", "esto_product"], dropna=False, as_index=False)["value"].sum(min_count=1)
    if trans_grouped.empty or not trans_grouped["sector"].astype(str).eq(REFINERY_SECTOR_NAME).any():
        refinery_fallback = _get_refinery_fallback_rows_for_balance(
            economy=economy_value,
            scenario=scenario_value,
            year=year_value,
        )
        if not refinery_fallback.empty:
            trans_grouped = pd.concat([trans_grouped, refinery_fallback], ignore_index=True, sort=False)
            trans_grouped = trans_grouped.groupby(["sector", "esto_product"], dropna=False, as_index=False)["value"].sum(min_count=1)
    if not trans_grouped.empty:
        for _, row in trans_grouped.iterrows():
            row_entries.append({"sector": row["sector"], "esto_product": row["esto_product"], "value": float(row["value"])})
        trans_totals = trans.groupby(["esto_product"], dropna=False, as_index=False)["value"].sum(min_count=1)
        if not trans_grouped.empty:
            trans_totals = trans_grouped.groupby(["esto_product"], dropna=False, as_index=False)["value"].sum(min_count=1)
        for _, row in trans_totals.iterrows():
            row_entries.append({"sector": "Total transformation sector", "esto_product": row["esto_product"], "value": float(row["value"])})

    if not demand_grouped.empty:
        for _, row in demand_grouped.iterrows():
            row_entries.append({"sector": _normalize_conventional_sector_name(row["sector"]), "esto_product": row["esto_product"], "value": float(row["value"])})
        demand_detail_rows = demand_grouped[~demand_grouped["is_top_level_aggregate"].fillna(False)].copy()
        demand_totals = demand_detail_rows.groupby(["esto_product"], dropna=False, as_index=False)["value"].sum(min_count=1)
        for _, row in demand_totals.iterrows():
            row_entries.append({"sector": "Total Final Energy Demand", "esto_product": row["esto_product"], "value": float(row["value"])})

    # Source datasets do not contain the post-adjustment unmet requirement concept; treat as zero baseline.
    if not row_entries:
        return pd.DataFrame()

    long_df = pd.DataFrame(row_entries)
    long_df["value"] = pd.to_numeric(long_df["value"], errors="coerce").fillna(0.0)
    long_df["fuel_group"] = long_df["esto_product"].map(lambda value: ESTO_PARENT_PRODUCT_LOOKUP.get(str(value), str(value)))
    long_df = long_df.groupby(["sector", "fuel_group"], dropna=False, as_index=False)["value"].sum(min_count=1).rename(columns={"fuel_group": "esto_product"})

    pivot = (
        long_df.pivot_table(index="sector", columns="esto_product", values="value", aggfunc="sum", fill_value=0.0)
        .reset_index()
    )
    pivot["sector"] = pivot["sector"].map(_normalize_conventional_sector_name)
    fuel_columns = [col for col in pivot.columns if col != "sector"]
    if fuel_columns:
        pivot["Total"] = pivot[fuel_columns].sum(axis=1)
    pivot = pivot.groupby("sector", as_index=False).sum(numeric_only=True)
    return _zero_small_numeric_values(pivot.rename(columns={"sector": "Sector"}), label_columns=["Sector"], threshold=0.01)


def build_conventional_balance_diff_matrix(
    shown_table: pd.DataFrame,
    reference_table: pd.DataFrame,
) -> pd.DataFrame:
    """Return shown minus reference with the same row/column layout as `shown_table`."""
    if shown_table.empty:
        return shown_table.copy()
    shown = shown_table.copy()
    if "Sector" not in shown.columns:
        return shown
    if reference_table.empty or "Sector" not in reference_table.columns:
        reference_aligned = pd.DataFrame()
    else:
        reference_aligned = reference_table.copy().set_index("Sector")

    value_columns = [col for col in shown.columns if col != "Sector"]
    diff = shown.copy()
    for column in value_columns:
        shown_values = pd.to_numeric(shown[column], errors="coerce").fillna(0.0)
        if not reference_aligned.empty and column in reference_aligned.columns:
            ref_values = pd.to_numeric(
                reference_aligned.reindex(shown["Sector"].astype(str).tolist())[column],
                errors="coerce",
            ).fillna(0.0)
        else:
            ref_values = pd.Series(0.0, index=shown.index)
        diff[column] = shown_values - ref_values.reset_index(drop=True)
    return _zero_small_numeric_values(diff, label_columns=["Sector"], threshold=0.01)


def build_conventional_balance_matrix(
    reconciliation_table: pd.DataFrame,
    sector_demand_table: pd.DataFrame,
    transformation_sector_table: pd.DataFrame,
    supply_primary_table: pd.DataFrame,
    year: int,
    economy: str,
    scenario: str,
) -> pd.DataFrame:
    """Return a conventional balance matrix: sectors on rows, fuels on columns."""
    year_value = int(year)
    economy_value = str(economy)
    scenario_value = str(scenario)

    recon = reconciliation_table.copy()
    recon["year"] = pd.to_numeric(recon["year"], errors="coerce").astype("Int64")
    recon = recon[
        (recon["year"] == year_value)
        & (recon["economy"].astype(str) == economy_value)
        & (recon["scenario"].astype(str) == scenario_value)
    ].copy()

    demand = sector_demand_table.copy()
    demand["year"] = pd.to_numeric(demand["year"], errors="coerce").astype("Int64")
    demand = demand[
        (demand["year"] == year_value)
        & (demand["economy"].astype(str) == economy_value)
        & (demand["scenario"].astype(str) == scenario_value)
    ].copy()

    trans = transformation_sector_table.copy()
    trans["year"] = pd.to_numeric(trans["year"], errors="coerce").astype("Int64")
    trans = trans[
        (trans["year"] == year_value)
        & (trans["economy"].astype(str) == economy_value)
    ].copy()

    supply = supply_primary_table.copy()
    supply["year"] = pd.to_numeric(supply["year"], errors="coerce").astype("Int64")
    supply = supply[
        (supply["year"] == year_value)
        & (supply["economy"].astype(str) == economy_value)
    ].copy()

    def _safe_number(value: object) -> float:
        numeric = pd.to_numeric(value, errors="coerce")
        if pd.isna(numeric):
            return 0.0
        return float(numeric)

    row_entries: list[dict[str, object]] = []

    if not supply.empty:
        production_override = None
        if not recon.empty and "constrained_production" in recon.columns:
            production_override = (
                recon[["esto_product", "constrained_production"]]
                .drop_duplicates(subset=["esto_product"], keep="last")
                .rename(columns={"constrained_production": "production_override"})
            )
            supply = supply.merge(production_override, on="esto_product", how="left")
        for _, row in supply.iterrows():
            production_value = row.get("production_override")
            if pd.isna(pd.to_numeric(production_value, errors="coerce")):
                production_value = row.get("production")
            row_entries.append(
                {
                    "sector": "Production",
                    "esto_product": row["esto_product"],
                    "value": _safe_number(production_value),
                }
            )
            row_entries.append(
                {
                    "sector": "Stock changes",
                    "esto_product": row["esto_product"],
                    "value": _safe_number(row.get("stock_changes")),
                }
            )

    if not recon.empty:
        for _, row in recon.iterrows():
            row_entries.append(
                {
                    "sector": "Imports",
                    "esto_product": row["esto_product"],
                    "value": abs(_safe_number(row.get("adjusted_imports"))),
                }
            )
            row_entries.append(
                {
                    "sector": "Exports",
                    "esto_product": row["esto_product"],
                    "value": -abs(_safe_number(row.get("adjusted_exports"))),
                }
            )

    if not supply.empty or not recon.empty:
        merged_primary = supply.merge(
            recon[["esto_product", "adjusted_imports", "adjusted_exports"]],
            on="esto_product",
            how="outer",
        )
        for _, row in merged_primary.iterrows():
            production = pd.to_numeric(row.get("production"), errors="coerce")
            adjusted_imports = pd.to_numeric(row.get("adjusted_imports"), errors="coerce")
            adjusted_exports = pd.to_numeric(row.get("adjusted_exports"), errors="coerce")
            stock_changes = pd.to_numeric(row.get("stock_changes"), errors="coerce")
            constrained_production = pd.to_numeric(row.get("constrained_production"), errors="coerce")
            if pd.notna(constrained_production):
                production_val = float(constrained_production)
            else:
                production_val = 0.0 if pd.isna(production) else float(production)
            imports_val = 0.0 if pd.isna(adjusted_imports) else float(adjusted_imports)
            exports_val = 0.0 if pd.isna(adjusted_exports) else float(adjusted_exports)
            stock_val = 0.0 if pd.isna(stock_changes) else float(stock_changes)
            total_primary_supply = production_val + imports_val - exports_val + stock_val
            row_entries.append(
                {
                    "sector": "Total primary energy supply",
                    "esto_product": row["esto_product"],
                    "value": total_primary_supply,
                }
            )

    trans_grouped = pd.DataFrame(columns=["sector", "esto_product", "value"])
    if not trans.empty:
        trans_grouped = (
            trans.groupby(["sector", "esto_product"], dropna=False, as_index=False)["value"]
            .sum(min_count=1)
        )
    if trans_grouped.empty or not trans_grouped["sector"].astype(str).eq(REFINERY_SECTOR_NAME).any():
        refinery_fallback = _get_refinery_fallback_rows_for_balance(
            economy=economy_value,
            scenario=scenario_value,
            year=year_value,
        )
        if not refinery_fallback.empty:
            trans_grouped = pd.concat([trans_grouped, refinery_fallback], ignore_index=True, sort=False)
            trans_grouped = (
                trans_grouped.groupby(["sector", "esto_product"], dropna=False, as_index=False)["value"]
                .sum(min_count=1)
            )
    if not trans_grouped.empty:
        for _, row in trans_grouped.iterrows():
            row_entries.append(
                {
                    "sector": row["sector"],
                    "esto_product": row["esto_product"],
                    "value": float(row["value"]),
                }
            )
        trans_totals = (
            trans_grouped.groupby(["esto_product"], dropna=False, as_index=False)["value"]
            .sum(min_count=1)
        )
        for _, row in trans_totals.iterrows():
            row_entries.append(
                {
                    "sector": "Total transformation sector",
                    "esto_product": row["esto_product"],
                    "value": float(row["value"]),
                }
            )

    if not demand.empty:
        demand_grouped = _prepare_demand_rows_for_balance(demand)
        for _, row in demand_grouped.iterrows():
            row_entries.append(
                {
                    "sector": _normalize_conventional_sector_name(row["sector"]),
                    "esto_product": row["esto_product"],
                    "value": float(row["value"]),
                }
            )
        demand_detail_rows = demand_grouped[~demand_grouped["is_top_level_aggregate"].fillna(False)].copy()
        demand_totals = (
            demand_detail_rows.groupby(["esto_product"], dropna=False, as_index=False)["value"]
            .sum(min_count=1)
        )
        for _, row in demand_totals.iterrows():
            row_entries.append(
                {
                    "sector": "Total Final Energy Demand",
                    "esto_product": row["esto_product"],
                    "value": float(row["value"]),
                }
            )

    if not recon.empty:
        residual_totals = (
            recon.groupby(["esto_product"], dropna=False, as_index=False)["adjusted_balance"]
            .sum(min_count=1)
        )
        for _, row in residual_totals.iterrows():
            row_entries.append(
                {
                    "sector": "Unmet Requirements",
                    "esto_product": row["esto_product"],
                    "value": -_safe_number(row["adjusted_balance"]),
                }
            )

    if not row_entries:
        return pd.DataFrame()

    long_df = pd.DataFrame(row_entries)
    long_df["value"] = pd.to_numeric(long_df["value"], errors="coerce").fillna(0.0)
    long_df["fuel_group"] = long_df["esto_product"].map(
        lambda value: ESTO_PARENT_PRODUCT_LOOKUP.get(str(value), str(value))
    )
    long_df = (
        long_df.groupby(["sector", "fuel_group"], dropna=False, as_index=False)["value"]
        .sum(min_count=1)
        .rename(columns={"fuel_group": "esto_product"})
    )

    pivot = (
        long_df.pivot_table(
            index="sector",
            columns="esto_product",
            values="value",
            aggfunc="sum",
            fill_value=0.0,
        )
        .reset_index()
    )
    pivot["sector"] = pivot["sector"].map(_normalize_conventional_sector_name)
    fuel_columns = [col for col in pivot.columns if col != "sector"]
    if fuel_columns:
        pivot["Total"] = pivot[fuel_columns].sum(axis=1)

    pivot = (
        pivot.groupby("sector", as_index=False)
        .sum(numeric_only=True)
    )
    row_to_section, ordered_backbone = _build_conventional_section_lookup()
    numeric_columns = [col for col in pivot.columns if col != "sector"]
    keep_always = {
        "Total primary energy supply",
        "Total transformation sector",
        "Total Final Energy Demand",
        "Unmet Requirements",
    }

    if numeric_columns:
        row_nonzero = (
            pivot[numeric_columns]
            .apply(pd.to_numeric, errors="coerce")
            .fillna(0.0)
            .abs()
            .sum(axis=1)
            > 0
        )
        pivot = pivot[row_nonzero | pivot["sector"].astype(str).isin(keep_always)].copy()

    fuel_columns = [col for col in pivot.columns if col not in {"sector", "Total"}]
    zero_fuel_columns = []
    for column in fuel_columns:
        column_values = pd.to_numeric(pivot[column], errors="coerce").fillna(0.0)
        if float(column_values.abs().sum()) == 0.0:
            zero_fuel_columns.append(column)
    if zero_fuel_columns:
        pivot = pivot.drop(columns=zero_fuel_columns)

    remaining_fuel_columns = [col for col in pivot.columns if col not in {"sector", "Total"}]
    if remaining_fuel_columns:
        pivot["Total"] = (
            pivot[remaining_fuel_columns]
            .apply(pd.to_numeric, errors="coerce")
            .fillna(0.0)
            .sum(axis=1)
        )
    elif "Total" in pivot.columns:
        pivot["Total"] = 0.0

    extra_rows = [
        name for name in pivot["sector"].astype(str).tolist()
        if name not in ordered_backbone
    ]
    demand_extra_rows = sorted(
        [name for name in extra_rows if row_to_section.get(name) is None]
    )
    ordered_rows = ordered_backbone + demand_extra_rows
    pivot = (
        pivot.set_index("sector")
        .reindex(ordered_rows, fill_value=0.0)
        .reset_index()
        .rename(columns={"sector": "Sector"})
    )
    if "Total" in pivot.columns:
        row_nonzero = (
            pivot[[col for col in pivot.columns if col != "Sector"]]
            .apply(pd.to_numeric, errors="coerce")
            .fillna(0.0)
            .abs()
            .sum(axis=1)
            > 0
        )
        pivot = pivot[row_nonzero | pivot["Sector"].astype(str).isin(keep_always)].copy()
    pivot = pivot.reset_index(drop=True)
    return _zero_small_numeric_values(
        pivot,
        label_columns=["Sector"],
        threshold=0.01,
    )


def _style_conventional_balance_worksheet(
    ws,
    table: pd.DataFrame,
    *,
    economy: str,
    scenario: str,
    year: int,
    is_diff: bool,
) -> None:
    """Apply report styling to a conventional balance worksheet."""
    subtotal_fill_by_row = {
        "Total primary energy supply": PatternFill(fill_type="solid", fgColor="FFFFFF"),
        "Total transformation sector": PatternFill(fill_type="solid", fgColor="E2F0D9"),
        "Total Final Energy Demand": PatternFill(fill_type="solid", fgColor="DDEBF7"),
        "Unmet Requirements": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    }
    subtotal_rows = set(subtotal_fill_by_row)
    strong_rows: set[str] = set()
    subtotal_font = Font(bold=True, color="1F1F1F")
    strong_font = Font(bold=True, color="9C0006")
    section_fills = {
        "Supply": PatternFill(fill_type="solid", fgColor="F2F2F2"),
        "Transfers": PatternFill(fill_type="solid", fgColor="FFF2CC"),
        "Transformation": PatternFill(fill_type="solid", fgColor="E2F0D9"),
        "Losses": PatternFill(fill_type="solid", fgColor="FCE4D6"),
        "Demand": PatternFill(fill_type="solid", fgColor="DDEBF7"),
        "Checks": PatternFill(fill_type="solid", fgColor="E4DFEC"),
    }
    row_to_section, _ = _build_conventional_section_lookup()
    ws["A1"] = f"Energy Balance for {economy}"
    subtitle = f"Scenario: {scenario}, Year: {year}, Units: Petajoule"
    if is_diff:
        subtitle += " | Values shown minus source dataset values"
    ws["A2"] = subtitle

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[column_letter].width = min(max_length + 2, 36)

    header_row = 3
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="EDEDED")

    for row_index in range(header_row + 1, header_row + 1 + len(table)):
        sector_value = ws.cell(row=row_index, column=1).value
        section_name = row_to_section.get(str(sector_value), "Demand")
        fill = section_fills.get(section_name)
        font = None
        if sector_value in subtotal_rows:
            fill = subtotal_fill_by_row.get(str(sector_value), fill)
            font = strong_font if sector_value in strong_rows else subtotal_font
        for column_index in range(1, len(table.columns) + 1):
            cell = ws.cell(row=row_index, column=column_index)
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font


def _write_formatted_conventional_balance_workbook(
    table: pd.DataFrame,
    path: Path,
    *,
    economy: str,
    scenario: str,
    year: int,
) -> None:
    """Write a single conventional balance worksheet to Excel."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        table.to_excel(writer, sheet_name="Balance", index=False, startrow=2)
        ws = writer.book["Balance"]
        _style_conventional_balance_worksheet(
            ws,
            table,
            economy=economy,
            scenario=scenario,
            year=year,
            is_diff=False,
        )


def save_conventional_balance_tables(
    reconciliation_table: pd.DataFrame,
    sector_demand_table: pd.DataFrame,
    transformation_sector_table: pd.DataFrame,
    supply_primary_table: pd.DataFrame,
    base_df: pd.DataFrame,
    years: Iterable[int],
    output_dir: Path | str = CONVENTIONAL_BALANCE_DIR,
    economies: Iterable[str] | None = None,
    scenarios: Iterable[str] | None = None,
) -> list[Path]:
    """Write one conventional balance workbook per economy/scenario with year and diff sheets."""
    output_path = _resolve(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    saved_paths: list[Path] = []

    economy_list = sorted(
        {
            str(value)
            for value in (economies or reconciliation_table.get("economy", pd.Series(dtype=str)).astype(str).unique())
            if str(value).strip()
        }
    )
    scenario_list = sorted(
        {
            str(value)
            for value in (scenarios or reconciliation_table.get("scenario", pd.Series(dtype=str)).astype(str).unique())
            if str(value).strip()
        }
    )

    for economy in economy_list:
        for scenario in scenario_list:
            workbook_path = output_path / (
                f"conventional_balance_{_safe_filename_token(economy)}_"
                f"{_safe_filename_token(scenario)}.xlsx"
            )
            sheets_written = 0
            with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
                for year in years:
                    year_int = int(year)
                    table = build_conventional_balance_matrix(
                        reconciliation_table=reconciliation_table,
                        sector_demand_table=sector_demand_table,
                        transformation_sector_table=transformation_sector_table,
                        supply_primary_table=supply_primary_table,
                        year=year_int,
                        economy=economy,
                        scenario=scenario,
                    )
                    if table.empty:
                        continue
                    reference_table = build_reference_conventional_balance_matrix(
                        reconciliation_table=reconciliation_table,
                        sector_demand_table=sector_demand_table,
                        transformation_sector_table=transformation_sector_table,
                        supply_primary_table=supply_primary_table,
                        base_df=base_df,
                        year=year_int,
                        economy=economy,
                        scenario=scenario,
                    )
                    diff_table = build_conventional_balance_diff_matrix(
                        shown_table=table,
                        reference_table=reference_table,
                    )

                    balance_sheet_name = str(year_int)
                    diff_sheet_name = f"{year_int}-diffs"
                    table.to_excel(writer, sheet_name=balance_sheet_name, index=False, startrow=2)
                    diff_table.to_excel(writer, sheet_name=diff_sheet_name, index=False, startrow=2)

                    _style_conventional_balance_worksheet(
                        writer.book[balance_sheet_name],
                        table,
                        economy=economy,
                        scenario=scenario,
                        year=year_int,
                        is_diff=False,
                    )
                    _style_conventional_balance_worksheet(
                        writer.book[diff_sheet_name],
                        diff_table,
                        economy=economy,
                        scenario=scenario,
                        year=year_int,
                        is_diff=True,
                    )
                    sheets_written += 2
            if sheets_written:
                saved_paths.append(workbook_path)
                print(f"Saved conventional balance workbook to {workbook_path}")
            elif workbook_path.exists():
                workbook_path.unlink()
    return saved_paths


def _archive_existing_results_file_if_needed(
    target_path: Path,
    *,
    archive_dir: Path | str,
    min_hours: int = 24,
) -> Path | None:
    """Archive an existing results file with timestamp unless archived recently."""
    if not target_path.exists():
        return None
    archive_root = _resolve(archive_dir)
    archive_root.mkdir(parents=True, exist_ok=True)
    pattern = f"{target_path.stem}_*{target_path.suffix}"
    prior_archives = sorted(archive_root.glob(pattern), key=lambda p: p.stat().st_mtime)
    now_utc = datetime.now(timezone.utc)
    if prior_archives:
        latest = prior_archives[-1]
        age_hours = (now_utc - datetime.fromtimestamp(latest.stat().st_mtime, tz=timezone.utc)).total_seconds() / 3600.0
        if age_hours < max(int(min_hours), 0):
            print(
                "[INFO] Skipping archive for current run workbook; latest archive is "
                f"{age_hours:.1f}h old (< {int(min_hours)}h window): {latest}"
            )
            return latest
    stamp = now_utc.strftime("%Y%m%d_%H%M%S")
    archived_path = archive_root / f"{target_path.stem}_{stamp}{target_path.suffix}"
    shutil.copy2(target_path, archived_path)
    print(f"[INFO] Archived previous run workbook to {archived_path}")
    return archived_path


def _archive_results_file_snapshot(
    target_path: Path,
    *,
    archive_dir: Path | str,
) -> Path:
    """Archive the current run workbook with a timestamped filename."""
    if not target_path.exists():
        raise FileNotFoundError(f"Cannot archive missing results workbook: {target_path}")
    archive_root = _resolve(archive_dir)
    archive_root.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    archived_path = archive_root / f"{target_path.stem}_{stamp}{target_path.suffix}"
    shutil.copy2(target_path, archived_path)
    print(f"[INFO] Archived current run workbook to {archived_path}")
    return archived_path


def _resolve_results_single_file_name(
    base_name: str,
    *,
    trade_mode: str,
    iteration_run_mode: str,
) -> str:
    """Return single-workbook filename with iterative mode suffix when applicable."""
    raw_name = str(base_name or "").strip() or "results_supply_link_run_test.xlsx"
    path = Path(raw_name)
    stem = path.stem
    suffix = path.suffix or ".xlsx"
    trade_token = str(trade_mode or "").strip().lower()
    mode_token = str(iteration_run_mode or "").strip().lower()
    if trade_token == "capacity_unmet_iterative_balanced" and mode_token:
        safe_mode = re.sub(r"[^a-z0-9_-]+", "_", mode_token).strip("_")
        if safe_mode and not stem.lower().endswith(f"_{safe_mode}".lower()):
            stem = f"{stem}_{safe_mode}"
    return f"{stem}{suffix}"


def save_results_linked_single_workbook(
    *,
    reconciliation_table: pd.DataFrame,
    sector_demand_table: pd.DataFrame,
    demand_table: pd.DataFrame,
    transformation_table: pd.DataFrame,
    transformation_sector_table: pd.DataFrame,
    supply_projection_table: pd.DataFrame,
    supply_primary_table: pd.DataFrame,
    transformation_target_rows: pd.DataFrame,
    fuel_branch_catalog_df: pd.DataFrame,
    base_df: pd.DataFrame,
    years: Iterable[int],
    economies: Iterable[str],
    scenarios: Iterable[str],
    export_paths: Iterable[Path],
    transformation_export_paths: Iterable[Path],
    transfer_export_paths: Iterable[Path],
    combined_export_path: Path | None,
    probe_catalog_path: Path | None,
    leap_import_result: dict[str, object],
    output_dir: Path | str = OUTPUT_DIR,
    file_name: str = RESULTS_SINGLE_FILE_NAME,
    archive_dir: Path | str = RESULTS_SINGLE_FILE_ARCHIVE_DIR,
    archive_min_hours: int | None = None,
    archive_every_run: bool = RESULTS_SINGLE_FILE_ARCHIVE_EVERY_RUN,
) -> Path:
    """
    Write one LEAP-style Export workbook matching the full-model template structure.

    Output format intentionally mirrors `data/full model export.xlsx`:
    - single sheet: `Export`
    - preamble row with `Area:` / `Ver:`
    - header columns:
      BranchID, VariableID, ScenarioID, RegionID, Branch Path, Variable, Scenario,
      Region, Scale, Units, Per..., Method, <year columns...>
    """
    _ = (
        reconciliation_table,
        sector_demand_table,
        demand_table,
        transformation_table,
        transformation_sector_table,
        supply_projection_table,
        supply_primary_table,
        transformation_target_rows,
        fuel_branch_catalog_df,
        base_df,
        years,
        economies,
        scenarios,
        export_paths,
        transformation_export_paths,
        transfer_export_paths,
        probe_catalog_path,
        leap_import_result,
    )
    if combined_export_path is None or not Path(combined_export_path).exists():
        raise FileNotFoundError(
            "Combined supply/transformation workbook is required for single-file output "
            "but was not found."
        )
    if archive_min_hours is None:
        archive_min_hours = int(RESULTS_SINGLE_FILE_ARCHIVE_MIN_HOURS)

    def _is_year_header(value: object) -> bool:
        text = str(value).strip()
        if not text:
            return False
        if re.fullmatch(r"\d{4}", text):
            return True
        if re.fullmatch(r"\d{4}\.0", text):
            return True
        return False

    def _infer_method_from_expression(expression: object) -> str:
        text = str(expression or "").strip()
        if not text:
            return "Interp"
        lowered = text.lower()
        for token in ("data", "interp", "step", "growth", "ramp"):
            if lowered.startswith(f"{token}("):
                if token == "data":
                    # LEAP import expects Method=Interp for imported Data(...) series.
                    return "Interp"
                return token.capitalize()
        if re.fullmatch(r"[-+]?\d+(\.\d+)?", text):
            return "Interp"
        return "Interp"

    def _normalize_merge_text(value: object) -> str:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return ""
        return str(value).strip().lower()

    def _normalize_metadata_text(value: object) -> str:
        """Normalize metadata cell values, treating NaN/None-like tokens as empty."""
        if value is None:
            return ""
        if pd.isna(value):
            return ""
        text = str(value).strip()
        if text.lower() in {"", "nan", "none", "null", "<na>", "na"}:
            return ""
        return text

    def _split_resource_branch_path(path_value: object) -> tuple[str, str]:
        """Return (`Resources\\<Root>`, leaf) for resource branch paths, else ('', '')."""
        parts = [part.strip() for part in str(path_value or "").split("\\") if part.strip()]
        if len(parts) < 3:
            return "", ""
        if parts[0].strip().lower() != "resources":
            return "", ""
        root = parts[1].strip().title()
        if root not in {"Primary", "Secondary"}:
            return "", ""
        return f"Resources\\{root}", parts[2].strip()

    def _branch_leaf_tokens(label: object) -> set[str]:
        """Tokenize branch leaf labels for safe fuzzy matching."""
        text = _normalize_merge_text(label)
        if not text:
            return set()
        tokens = re.findall(r"[a-z0-9]+", text)
        ignored = {
            "and",
            "of",
            "which",
            "the",
            "nonspecified",
            "non",
            "specified",
        }
        return {tok for tok in tokens if tok and tok not in ignored}

    def _remap_resource_branch_paths_from_reference(
        df: pd.DataFrame,
        source_data: pd.DataFrame,
    ) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """
        Remap resource branch paths to canonical reference paths when confidently resolvable.

        Matching order:
        1) exact key match (no remap)
        2) unique same-scope leaf exact match
        3) unique same-scope token-subset match
        """
        out = df.copy()
        key_cols = ["Branch Path", "Variable", "Scenario", "Region"]
        remap_cols = ["Branch Path", "Variable", "Scenario", "Region", "reference_branch_path", "match_type"]
        unresolved_cols = ["Branch Path", "Variable", "Scenario", "Region", "issue", "candidate_branch_paths"]
        empty_remap = pd.DataFrame(columns=remap_cols)
        empty_unresolved = pd.DataFrame(columns=unresolved_cols)
        if source_data is None or source_data.empty:
            return out, empty_remap, empty_unresolved
        if any(col not in source_data.columns for col in key_cols):
            return out, empty_remap, empty_unresolved
        if any(col not in out.columns for col in key_cols):
            return out, empty_remap, empty_unresolved

        source = source_data[key_cols].copy()
        for col in key_cols:
            source[f"__k_{col}"] = source[col].map(_normalize_merge_text)
            out[f"__k_{col}"] = out[col].map(_normalize_merge_text)

        source = source.drop_duplicates(
            subset=[f"__k_{col}" for col in key_cols],
            keep="first",
        ).copy()
        source["__root"], source["__leaf"] = zip(
            *source["Branch Path"].map(_split_resource_branch_path)
        )
        source = source[source["__root"] != ""].copy()
        if source.empty:
            out = out.drop(columns=[f"__k_{col}" for col in key_cols], errors="ignore")
            return out, empty_remap, empty_unresolved
        source["__k_root"] = source["__root"].map(_normalize_merge_text)
        source["__k_leaf"] = source["__leaf"].map(_normalize_merge_text)
        source["__leaf_tokens"] = source["__leaf"].map(_branch_leaf_tokens)

        source_exact_keys = {
            tuple(row[f"__k_{col}"] for col in key_cols)
            for _, row in source.iterrows()
        }
        source_scope_groups: dict[tuple[str, str, str, str], list[dict[str, object]]] = {}
        for _, row in source.iterrows():
            scope_key = (
                str(row["__k_Variable"]),
                str(row["__k_Scenario"]),
                str(row["__k_Region"]),
                str(row["__k_root"]),
            )
            source_scope_groups.setdefault(scope_key, []).append(
                {
                    "branch_path": str(row["Branch Path"]),
                    "k_leaf": str(row["__k_leaf"]),
                    "leaf_tokens": set(row["__leaf_tokens"]),
                }
            )

        remap_rows: list[dict[str, object]] = []
        unresolved_rows: list[dict[str, object]] = []
        for idx, row in out.iterrows():
            branch_path = str(row.get("Branch Path") or "")
            root, leaf = _split_resource_branch_path(branch_path)
            if not root:
                continue
            key_tuple = tuple(str(row.get(f"__k_{col}") or "") for col in key_cols)
            if key_tuple in source_exact_keys:
                continue
            scope_key = (
                str(row.get("__k_Variable") or ""),
                str(row.get("__k_Scenario") or ""),
                str(row.get("__k_Region") or ""),
                _normalize_merge_text(root),
            )
            candidates = source_scope_groups.get(scope_key, [])
            if not candidates:
                unresolved_rows.append(
                    {
                        "Branch Path": branch_path,
                        "Variable": row.get("Variable", ""),
                        "Scenario": row.get("Scenario", ""),
                        "Region": row.get("Region", ""),
                        "issue": "no_reference_candidates_in_scope",
                        "candidate_branch_paths": "",
                    }
                )
                continue
            leaf_norm = _normalize_merge_text(leaf)
            exact_leaf = [item for item in candidates if item.get("k_leaf") == leaf_norm and leaf_norm]
            if len(exact_leaf) == 1:
                new_path = str(exact_leaf[0]["branch_path"])
                out.at[idx, "Branch Path"] = new_path
                remap_rows.append(
                    {
                        "Branch Path": branch_path,
                        "Variable": row.get("Variable", ""),
                        "Scenario": row.get("Scenario", ""),
                        "Region": row.get("Region", ""),
                        "reference_branch_path": new_path,
                        "match_type": "leaf_exact_in_scope",
                    }
                )
                continue
            if len(exact_leaf) > 1:
                unresolved_rows.append(
                    {
                        "Branch Path": branch_path,
                        "Variable": row.get("Variable", ""),
                        "Scenario": row.get("Scenario", ""),
                        "Region": row.get("Region", ""),
                        "issue": "ambiguous_leaf_exact_candidates",
                        "candidate_branch_paths": " | ".join(
                            sorted(str(item["branch_path"]) for item in exact_leaf)
                        ),
                    }
                )
                continue

            leaf_tokens = _branch_leaf_tokens(leaf)
            fuzzy = [
                item
                for item in candidates
                if leaf_tokens
                and item.get("leaf_tokens")
                and (
                    leaf_tokens.issubset(set(item["leaf_tokens"]))
                    or set(item["leaf_tokens"]).issubset(leaf_tokens)
                )
            ]
            if len(fuzzy) == 1:
                new_path = str(fuzzy[0]["branch_path"])
                out.at[idx, "Branch Path"] = new_path
                remap_rows.append(
                    {
                        "Branch Path": branch_path,
                        "Variable": row.get("Variable", ""),
                        "Scenario": row.get("Scenario", ""),
                        "Region": row.get("Region", ""),
                        "reference_branch_path": new_path,
                        "match_type": "leaf_token_subset_in_scope",
                    }
                )
                continue

            issue = "no_confident_leaf_match"
            if len(fuzzy) > 1:
                issue = "ambiguous_leaf_token_subset_candidates"
            unresolved_rows.append(
                {
                    "Branch Path": branch_path,
                    "Variable": row.get("Variable", ""),
                    "Scenario": row.get("Scenario", ""),
                    "Region": row.get("Region", ""),
                    "issue": issue,
                    "candidate_branch_paths": " | ".join(
                        sorted(str(item["branch_path"]) for item in (fuzzy or candidates))
                    ),
                }
            )

        out = out.drop(columns=[f"__k_{col}" for col in key_cols], errors="ignore")
        remap_df = (
            pd.DataFrame(remap_rows, columns=remap_cols)
            .drop_duplicates()
            .reset_index(drop=True)
            if remap_rows
            else empty_remap
        )
        unresolved_df = (
            pd.DataFrame(unresolved_rows, columns=unresolved_cols)
            .drop_duplicates()
            .reset_index(drop=True)
            if unresolved_rows
            else empty_unresolved
        )
        return out, remap_df, unresolved_df

    def _merge_levels_from_reference_data(
        df: pd.DataFrame,
        reference_df: pd.DataFrame,
    ) -> pd.DataFrame:
        out = df.copy()
        if reference_df is None or reference_df.empty:
            reference_df = pd.DataFrame()
        key_cols = ["Branch Path", "Variable", "Scenario", "Region"]
        detected_level_cols = []
        for col in reference_df.columns:
            match = re.fullmatch(r"Level\s+(\d+)", str(col).strip(), flags=re.IGNORECASE)
            if match:
                detected_level_cols.append((int(match.group(1)), f"Level {int(match.group(1))}"))
        detected_level_cols = sorted({item for item in detected_level_cols}, key=lambda item: item[0])
        base_level_cols = [f"Level {idx}" for idx in range(1, 9)]
        merged_level_cols = []
        seen = set()
        for _, name in detected_level_cols:
            if name not in seen:
                seen.add(name)
                merged_level_cols.append(name)
        for name in base_level_cols:
            if name not in seen:
                seen.add(name)
                merged_level_cols.append(name)

        if all(col in reference_df.columns for col in key_cols):
            lookup_cols = key_cols + [col for col in merged_level_cols if col in reference_df.columns]
            level_lookup = reference_df[lookup_cols].copy()
            for col in key_cols:
                level_lookup[f"__k_{col}"] = level_lookup[col].map(_normalize_merge_text)
                out[f"__k_{col}"] = out[col].map(_normalize_merge_text)
            level_lookup = level_lookup.drop_duplicates(
                subset=[f"__k_{col}" for col in key_cols],
                keep="first",
            )
            merge_cols = [f"__k_{col}" for col in key_cols]
            value_cols = [col for col in merged_level_cols if col in level_lookup.columns]
            out = out.merge(
                level_lookup[merge_cols + value_cols],
                on=merge_cols,
                how="left",
            )
            out = out.drop(columns=merge_cols, errors="ignore")

        for col in merged_level_cols:
            if col not in out.columns:
                out[col] = ""

        # Fallback: derive levels from branch path when lookup rows were missing.
        parts_series = out["Branch Path"].fillna("").astype(str).map(
            lambda text: [part.strip() for part in text.split("\\") if part.strip()]
        )
        for idx, col in enumerate(merged_level_cols, start=1):
            existing = out[col].fillna("").astype(str)
            missing_mask = existing.str.strip().eq("")
            if missing_mask.any():
                fill_values = parts_series.map(
                    lambda parts: parts[idx - 1] if idx - 1 < len(parts) else ""
                )
                out.loc[missing_mask, col] = fill_values[missing_mask]
        return out

    def _filter_unmatched_zero_supply_rows_against_reference(
        df: pd.DataFrame,
        source_data: pd.DataFrame,
    ) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """
        Drop all-zero Resources rows not present in reference export keys.

        Returns:
        - filtered dataframe
        - dropped unmatched-zero rows report
        - unmatched-nonzero rows report (kept in output, requires review)
        """
        key_cols = ["Branch Path", "Variable", "Scenario", "Region"]
        empty_report = pd.DataFrame(columns=key_cols + ["year_abs_sum", "reason"])
        if source_data is None or source_data.empty:
            return df.copy(), empty_report, empty_report
        if any(col not in source_data.columns for col in key_cols):
            return df.copy(), empty_report, empty_report
        out = df.copy()
        if any(col not in out.columns for col in key_cols):
            return out, empty_report, empty_report

        source = source_data.copy()
        source_resource = source[
            source["Branch Path"].fillna("").astype(str).str.startswith("Resources\\")
        ].copy()
        if source_resource.empty:
            return out, empty_report, empty_report

        for col in key_cols:
            source_resource[f"__k_{col}"] = source_resource[col].map(_normalize_merge_text)
            out[f"__k_{col}"] = out[col].map(_normalize_merge_text)

        source_key_set = {
            tuple(row[f"__k_{col}"] for col in key_cols)
            for _, row in source_resource.drop_duplicates(
                subset=[f"__k_{col}" for col in key_cols],
                keep="first",
            ).iterrows()
        }

        resource_mask = out["Branch Path"].fillna("").astype(str).str.startswith("Resources\\")
        out["__resource_key"] = out.apply(
            lambda row: tuple(row[f"__k_{col}"] for col in key_cols),
            axis=1,
        )
        unmatched_resource_mask = resource_mask & ~out["__resource_key"].isin(source_key_set)
        if not unmatched_resource_mask.any():
            out = out.drop(
                columns=[*["__k_" + col for col in key_cols], "__resource_key"],
                errors="ignore",
            )
            return out, empty_report, empty_report

        year_cols = [col for col in out.columns if _is_year_header(col)]
        if year_cols:
            year_abs_sum = (
                out[year_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0).abs().sum(axis=1)
            )
        else:
            year_abs_sum = pd.Series(0.0, index=out.index)
        out["__year_abs_sum"] = year_abs_sum

        drop_mask = unmatched_resource_mask & (out["__year_abs_sum"] <= 0.0)
        keep_nonzero_mask = unmatched_resource_mask & (out["__year_abs_sum"] > 0.0)

        dropped_report = out.loc[drop_mask, key_cols + ["__year_abs_sum"]].copy()
        kept_nonzero_report = out.loc[keep_nonzero_mask, key_cols + ["__year_abs_sum"]].copy()
        if not dropped_report.empty:
            dropped_report = dropped_report.rename(columns={"__year_abs_sum": "year_abs_sum"})
            dropped_report["reason"] = "unmatched_resource_key_all_zero_row_dropped"
            dropped_report = dropped_report.drop_duplicates().reset_index(drop=True)
        if not kept_nonzero_report.empty:
            kept_nonzero_report = kept_nonzero_report.rename(columns={"__year_abs_sum": "year_abs_sum"})
            kept_nonzero_report["reason"] = "unmatched_resource_key_nonzero_row_kept"
            kept_nonzero_report = kept_nonzero_report.drop_duplicates().reset_index(drop=True)

        out = out.loc[~drop_mask].copy()
        out = out.drop(
            columns=[*["__k_" + col for col in key_cols], "__resource_key", "__year_abs_sum"],
            errors="ignore",
        )
        return out, dropped_report, kept_nonzero_report

    def _load_results_verification_data() -> tuple[pd.DataFrame, Path, str]:
        source_path = _resolve(RESULTS_VERIFICATION_EXPORT_PATH)
        source_sheet = RESULTS_VERIFICATION_EXPORT_SHEET
        if not USE_RESULTS_VERIFICATION_EXPORT_SOURCE:
            return pd.DataFrame(), source_path, source_sheet
        if not source_path.exists():
            print(f"[WARN] Verification export file not found: {source_path}")
            return pd.DataFrame(), source_path, source_sheet
        try:
            _, source_data, _ = _read_workbook_sheet_with_header_detection(
                source_path,
                source_sheet,
            )
            if source_data.empty:
                print(
                    f"[WARN] Verification export is empty: {source_path} (sheet={source_sheet})"
                )
            else:
                print(
                    "[INFO] Loaded verification export source from data/: "
                    f"{source_path} (sheet={source_sheet}, rows={len(source_data)})"
                )
            return source_data, source_path, source_sheet
        except Exception as exc:
            print(
                f"[WARN] Failed reading verification export source {source_path} "
                f"(sheet={source_sheet}): {exc}"
            )
            return pd.DataFrame(), source_path, source_sheet

    def _merge_ids_from_reference_export(
        df: pd.DataFrame,
        source_data: pd.DataFrame,
        source_path: Path,
    ) -> tuple[pd.DataFrame, pd.DataFrame]:
        if source_data is None or source_data.empty:
            out = df.copy()
            out["BranchID"] = -1
            out["VariableID"] = -1
            out["ScenarioID"] = -1
            out["RegionID"] = -1
            unmatched = out[["Branch Path", "Variable", "Scenario", "Region"]].copy()
            unmatched["reason"] = (
                "verification_export_missing"
                if not source_path.exists()
                else "verification_export_empty"
            )
            return out, unmatched

        required_source_cols = [
            "BranchID",
            "VariableID",
            "ScenarioID",
            "RegionID",
            "Branch Path",
            "Variable",
            "Scenario",
            "Region",
        ]
        missing_source = [col for col in required_source_cols if col not in source_data.columns]
        if missing_source:
            print(
                "[WARN] Verification export missing expected ID/key columns; using fallback -1 IDs: "
                f"{missing_source}"
            )
            out = df.copy()
            out["BranchID"] = -1
            out["VariableID"] = -1
            out["ScenarioID"] = -1
            out["RegionID"] = -1
            unmatched = out[["Branch Path", "Variable", "Scenario", "Region"]].copy()
            unmatched["reason"] = "verification_export_missing_required_columns"
            return out, unmatched

        source_subset = source_data[required_source_cols].copy()
        for col in ["BranchID", "VariableID", "ScenarioID", "RegionID"]:
            source_subset[col] = pd.to_numeric(source_subset[col], errors="coerce").astype("Int64")

        for col in ["Branch Path", "Variable", "Scenario", "Region"]:
            source_subset[f"__k_{col}"] = source_subset[col].map(_normalize_merge_text)

        key_cols = ["__k_Branch Path", "__k_Variable", "__k_Scenario", "__k_Region"]
        id_cols = ["BranchID", "VariableID", "ScenarioID", "RegionID"]

        # Detect conflicting duplicate keys in source IDs; keep first deterministic row.
        source_dedup = source_subset.copy()
        conflicting = (
            source_dedup.groupby(key_cols, dropna=False)[id_cols]
            .nunique(dropna=False)
            .reset_index()
        )
        conflicting = conflicting[
            (conflicting["BranchID"] > 1)
            | (conflicting["VariableID"] > 1)
            | (conflicting["ScenarioID"] > 1)
            | (conflicting["RegionID"] > 1)
        ]
        if not conflicting.empty:
            print(
                "[WARN] Verification export contains conflicting ID rows for "
                f"{len(conflicting)} key(s); using first match per key."
            )
        source_dedup = source_dedup.drop_duplicates(subset=key_cols, keep="first")
        source_dedup = source_dedup[key_cols + id_cols]

        out = df.copy()
        for col in ["Branch Path", "Variable", "Scenario", "Region"]:
            out[f"__k_{col}"] = out[col].map(_normalize_merge_text)
        out = out.merge(source_dedup, on=key_cols, how="left")
        matched = int(out["BranchID"].notna().sum())
        total = int(len(out))
        print(
            "[INFO] Merged IDs from verification export using "
            "Branch Path/Variable/Scenario/Region keys: "
            f"matched {matched}/{total}, unmatched {total - matched}."
        )
        unmatched = out[out["BranchID"].isna()][
            ["Branch Path", "Variable", "Scenario", "Region"]
        ].copy()
        if not unmatched.empty:
            unmatched["reason"] = "no_verification_export_id_match"
            unmatched = unmatched.drop_duplicates().reset_index(drop=True)
        for col in id_cols:
            out[col] = pd.to_numeric(out[col], errors="coerce").fillna(-1).astype(int)
        drop_cols = key_cols
        out = out.drop(columns=drop_cols, errors="ignore")
        return out, unmatched

    def _collect_metadata_mismatches_against_reference(
        df: pd.DataFrame,
        source_data: pd.DataFrame,
    ) -> pd.DataFrame:
        if source_data is None or source_data.empty:
            return pd.DataFrame(
                columns=[
                    "Branch Path",
                    "Variable",
                    "Scenario",
                    "Region",
                    "column",
                    "generated_value",
                    "reference_value",
                ]
            )
        key_cols = ["Branch Path", "Variable", "Scenario", "Region"]
        compare_cols = ["Scale", "Units", "Per..."]
        required = key_cols + compare_cols
        if any(col not in source_data.columns for col in required):
            return pd.DataFrame(
                columns=[
                    "Branch Path",
                    "Variable",
                    "Scenario",
                    "Region",
                    "column",
                    "generated_value",
                    "reference_value",
                ]
            )
        src = source_data[required].copy()
        for col in key_cols:
            src[f"__k_{col}"] = src[col].map(_normalize_merge_text)
        src = src.drop_duplicates(
            subset=[f"__k_{col}" for col in key_cols],
            keep="first",
        )
        out = df.copy()
        for col in key_cols:
            out[f"__k_{col}"] = out[col].map(_normalize_merge_text)
        merged = out.merge(
            src[
                [f"__k_{col}" for col in key_cols]
                + [f"{col}" for col in compare_cols]
            ].rename(columns={col: f"ref_{col}" for col in compare_cols}),
            on=[f"__k_{col}" for col in key_cols],
            how="left",
        )
        mismatches: list[dict[str, object]] = []
        for _, row in merged.iterrows():
            for col in compare_cols:
                left = _normalize_metadata_text(row.get(col))
                right = _normalize_metadata_text(row.get(f"ref_{col}"))
                if not right:
                    continue
                if left == right:
                    continue
                mismatches.append(
                    {
                        "Branch Path": row.get("Branch Path"),
                        "Variable": row.get("Variable"),
                        "Scenario": row.get("Scenario"),
                        "Region": row.get("Region"),
                        "column": col,
                        "generated_value": left,
                        "reference_value": right,
                    }
                )
        if not mismatches:
            return pd.DataFrame(
                columns=[
                    "Branch Path",
                    "Variable",
                    "Scenario",
                    "Region",
                    "column",
                    "generated_value",
                    "reference_value",
                ]
            )
        mismatch_df = pd.DataFrame(mismatches).drop_duplicates().reset_index(drop=True)
        return mismatch_df

    def _backfill_non_value_metadata_from_reference(
        df: pd.DataFrame,
        source_data: pd.DataFrame,
    ) -> tuple[pd.DataFrame, pd.DataFrame]:
        """Backfill empty non-year metadata fields from reference export values."""
        key_cols = ["Branch Path", "Variable", "Scenario", "Region"]
        if source_data is None or source_data.empty:
            return df.copy(), pd.DataFrame(
                columns=key_cols + ["column", "filled_value"]
            )
        reference_fields = ["Scale", "Units", "Per...", "Method"]
        available_ref_fields = [col for col in reference_fields if col in source_data.columns]
        if not available_ref_fields or any(col not in source_data.columns for col in key_cols):
            return df.copy(), pd.DataFrame(
                columns=key_cols + ["column", "filled_value"]
            )

        out = df.copy()
        src = source_data[key_cols + available_ref_fields].copy()
        for col in key_cols:
            src[f"__k_{col}"] = src[col].map(_normalize_merge_text)
            out[f"__k_{col}"] = out[col].map(_normalize_merge_text)
        key_join_cols = [f"__k_{col}" for col in key_cols]
        src = src.drop_duplicates(subset=key_join_cols, keep="first")
        merged = out.merge(
            src[key_join_cols + available_ref_fields].rename(
                columns={col: f"ref_{col}" for col in available_ref_fields}
            ),
            on=key_join_cols,
            how="left",
        )

        filled_rows: list[dict[str, object]] = []
        for col in available_ref_fields:
            if col not in merged.columns:
                merged[col] = ""
            current_values = merged[col].map(_normalize_metadata_text)
            reference_values = merged[f"ref_{col}"].map(_normalize_metadata_text)
            # Backfill only when generated metadata is missing; do not overwrite
            # explicit generated metadata (for example supply Petajoule units).
            apply_mask = current_values.eq("") & reference_values.ne("")
            if not apply_mask.any():
                continue
            merged.loc[apply_mask, col] = reference_values[apply_mask]
            for _, row in merged.loc[apply_mask, key_cols + [col]].iterrows():
                filled_rows.append(
                    {
                        "Branch Path": row.get("Branch Path"),
                        "Variable": row.get("Variable"),
                        "Scenario": row.get("Scenario"),
                        "Region": row.get("Region"),
                        "column": col,
                        "filled_value": row.get(col),
                    }
                )

        merged = merged.drop(columns=key_join_cols + [f"ref_{col}" for col in available_ref_fields], errors="ignore")
        fill_df = (
            pd.DataFrame(filled_rows).drop_duplicates().reset_index(drop=True)
            if filled_rows
            else pd.DataFrame(columns=key_cols + ["column", "filled_value"])
        )
        return merged, fill_df

    def _load_field_mapping_table_for_validation() -> pd.DataFrame:
        """Load configured analysis-input mapping workbook used for metadata checks."""
        path = Path(
            str(
                getattr(
                    workflow_cfg,
                    "ANALYSIS_INPUT_FIELD_MAPPING_PATH",
                    REPO_ROOT / "config" / "leap_export_workbook_mappings.xlsx",
                )
            ).replace("\\", "/")
        )
        if not path.is_absolute():
            path = REPO_ROOT / path
        sheet = str(
            getattr(
                workflow_cfg,
                "ANALYSIS_INPUT_FIELD_MAPPING_SHEET",
                "field_mappings",
            )
        ).strip() or "field_mappings"
        if not config_table_exists(path, sheet):
            return pd.DataFrame()
        try:
            table = read_config_table(path, sheet_name=sheet)
        except Exception as exc:
            print(
                "[WARN] Failed reading analysis-input field mapping workbook for validation: "
                f"{path} (sheet={sheet}) -> {exc}"
            )
            return pd.DataFrame()
        table.columns = [str(col).strip().lower() for col in table.columns]
        required = {
            "enabled",
            "match_scope",
            "branch_path",
            "variable",
            "units",
            "scale",
            "per",
            "confidence",
            "notes",
        }
        missing = sorted(required.difference(table.columns))
        if missing:
            print(
                "[WARN] Field mapping workbook missing required columns for validation: "
                f"{missing}"
            )
            return pd.DataFrame()
        return table

    def _collect_mapping_config_mismatches_against_reference(
        mapping_table: pd.DataFrame,
        source_data: pd.DataFrame,
    ) -> pd.DataFrame:
        """Compare enabled config mapping metadata values against reference export metadata."""
        if mapping_table is None or mapping_table.empty or source_data is None or source_data.empty:
            return pd.DataFrame(
                columns=[
                    "match_scope",
                    "branch_path",
                    "variable",
                    "field",
                    "config_value",
                    "reference_values",
                    "issue",
                ]
            )
        key_cols = ["Branch Path", "Variable"]
        compare_cols = {"units": "Units", "scale": "Scale", "per": "Per..."}
        required_source = ["Branch Path", "Variable", "Scale", "Units", "Per..."]
        if any(col not in source_data.columns for col in required_source):
            return pd.DataFrame(
                columns=[
                    "match_scope",
                    "branch_path",
                    "variable",
                    "field",
                    "config_value",
                    "reference_values",
                    "issue",
                ]
            )
        source = source_data[required_source].copy()
        source["__k_branch"] = source["Branch Path"].map(_normalize_merge_text)
        source["__k_variable"] = source["Variable"].map(_normalize_merge_text)

        def _is_enabled(value: object) -> bool:
            token = str(value or "").strip().lower()
            return token in {"1", "true", "yes", "y", "on"}

        mismatches: list[dict[str, object]] = []
        enabled_rows = mapping_table[mapping_table["enabled"].map(_is_enabled)].copy()
        for _, row in enabled_rows.iterrows():
            scope = str(row.get("match_scope") or "").strip().lower()
            branch = str(row.get("branch_path") or "").strip()
            variable = str(row.get("variable") or "").strip()
            if scope not in {"branch_variable", "variable", "branch"}:
                continue
            scoped = source
            if scope == "branch_variable":
                if not branch or not variable:
                    continue
                scoped = scoped[
                    (scoped["__k_branch"] == _normalize_merge_text(branch))
                    & (scoped["__k_variable"] == _normalize_merge_text(variable))
                ]
            elif scope == "variable":
                if not variable:
                    continue
                scoped = scoped[scoped["__k_variable"] == _normalize_merge_text(variable)]
            else:
                if not branch:
                    continue
                scoped = scoped[scoped["__k_branch"] == _normalize_merge_text(branch)]

            if scoped.empty:
                for cfg_field in compare_cols.keys():
                    cfg_value = str(row.get(cfg_field) or "").strip()
                    if cfg_value:
                        mismatches.append(
                            {
                                "match_scope": scope,
                                "branch_path": branch,
                                "variable": variable,
                                "field": cfg_field,
                                "config_value": cfg_value,
                                "reference_values": "",
                                "issue": "no_reference_match",
                            }
                        )
                continue

            for cfg_field, ref_col in compare_cols.items():
                cfg_value = _normalize_metadata_text(row.get(cfg_field))
                if not cfg_value:
                    continue
                ref_values = sorted(
                    {
                        _normalize_metadata_text(value)
                        for value in scoped[ref_col].tolist()
                        if _normalize_metadata_text(value)
                    }
                )
                if not ref_values:
                    mismatches.append(
                        {
                            "match_scope": scope,
                            "branch_path": branch,
                            "variable": variable,
                            "field": cfg_field,
                            "config_value": cfg_value,
                            "reference_values": "",
                            "issue": "reference_value_missing",
                        }
                    )
                    continue
                if cfg_value not in ref_values:
                    mismatches.append(
                        {
                            "match_scope": scope,
                            "branch_path": branch,
                            "variable": variable,
                            "field": cfg_field,
                            "config_value": cfg_value,
                            "reference_values": " | ".join(ref_values[:10]),
                            "issue": "config_reference_mismatch",
                        }
                    )
        if not mismatches:
            return pd.DataFrame(
                columns=[
                    "match_scope",
                    "branch_path",
                    "variable",
                    "field",
                    "config_value",
                    "reference_values",
                    "issue",
                ]
            )
        return pd.DataFrame(mismatches).drop_duplicates().reset_index(drop=True)

    def _extract_area_and_version(*preambles: pd.DataFrame) -> tuple[str, object]:
        default_area = "results_supply_link_run"
        default_version: object = 2
        resolved_area: str | None = None
        resolved_version: object | None = None
        for preamble in preambles:
            if preamble is None or preamble.empty:
                continue
            for row_idx in range(len(preamble.index)):
                row = [
                    _normalize_template_header_value(item)
                    for item in preamble.iloc[row_idx].tolist()
                ]
                for idx, value in enumerate(row):
                    token = str(value).strip().lower()
                    if token == "area:" and idx + 1 < len(row) and resolved_area is None:
                        candidate = _normalize_template_header_value(row[idx + 1])
                        if candidate:
                            resolved_area = candidate
                    if token == "ver:" and idx + 1 < len(row) and resolved_version is None:
                        candidate = row[idx + 1]
                        if candidate is not None and str(candidate).strip() != "":
                            resolved_version = candidate
        return (
            resolved_area if resolved_area is not None else default_area,
            resolved_version if resolved_version is not None else default_version,
        )

    combined_path = _resolve(combined_export_path)
    viewing_preamble, viewing_data, _ = _read_workbook_sheet_with_header_detection(
        combined_path,
        "FOR_VIEWING",
    )
    leap_preamble, leap_data, _ = _read_workbook_sheet_with_header_detection(
        combined_path,
        "LEAP",
    )
    if viewing_data.empty:
        raise ValueError(
            f"Combined workbook '{combined_path.name}' has no FOR_VIEWING data rows."
        )

    required = ["Branch Path", "Variable", "Scenario", "Region", "Scale", "Units", "Per..."]
    missing = [col for col in required if col not in viewing_data.columns]
    if missing:
        raise ValueError(
            f"Combined workbook '{combined_path.name}' is missing required columns for Export sheet: {missing}"
        )

    method_by_key: dict[tuple[str, str, str, str], str] = {}
    if not leap_data.empty and "Expression" in leap_data.columns:
        key_cols = ["Branch Path", "Variable", "Scenario", "Region"]
        if all(col in leap_data.columns for col in key_cols):
            for _, row in leap_data.iterrows():
                key = tuple(str(row.get(col) or "").strip() for col in key_cols)
                if not all(key):
                    continue
                method_by_key[key] = _infer_method_from_expression(row.get("Expression"))

    export_df = viewing_data.copy()
    key_cols = ["Branch Path", "Variable", "Scenario", "Region"]
    export_df["Method"] = [
        method_by_key.get(
            tuple(str(row.get(col) or "").strip() for col in key_cols),
            "Interp",
        )
        for _, row in export_df.iterrows()
    ]
    verification_data, verification_path, _verification_sheet = _load_results_verification_data()
    verification_preamble = pd.DataFrame()
    if verification_path.exists():
        try:
            verification_preamble, _, _ = _read_workbook_sheet_with_header_detection(
                verification_path,
                _verification_sheet,
            )
        except Exception as exc:
            print(
                "[WARN] Failed reading verification export preamble for Area/Ver: "
                f"{verification_path} (sheet={_verification_sheet}) -> {exc}"
            )
    (
        export_df,
        dropped_unmatched_zero_supply_rows,
        unmatched_nonzero_supply_rows,
    ) = _filter_unmatched_zero_supply_rows_against_reference(
        export_df,
        source_data=verification_data,
    )
    (
        export_df,
        remapped_resource_branch_rows,
        unresolved_resource_branch_rows,
    ) = _remap_resource_branch_paths_from_reference(
        export_df,
        source_data=verification_data,
    )
    level_source_df = verification_data if not verification_data.empty else leap_data
    export_df = _merge_levels_from_reference_data(export_df, level_source_df)
    export_df, metadata_backfill_rows = _backfill_non_value_metadata_from_reference(
        export_df,
        source_data=verification_data,
    )
    export_df, unmatched_id_rows = _merge_ids_from_reference_export(
        export_df,
        source_data=verification_data,
        source_path=verification_path,
    )
    metadata_mismatch_rows = _collect_metadata_mismatches_against_reference(
        export_df,
        source_data=verification_data,
    )
    mapping_table = _load_field_mapping_table_for_validation()
    mapping_config_mismatch_rows = _collect_mapping_config_mismatches_against_reference(
        mapping_table,
        source_data=verification_data,
    )

    year_columns = [col for col in export_df.columns if _is_year_header(col)]
    level_columns = [f"Level {idx}" for idx in range(1, 9)]
    level_spacer_column = ""
    non_year_order = [
        "BranchID",
        "VariableID",
        "ScenarioID",
        "RegionID",
        "Branch Path",
        "Variable",
        "Scenario",
        "Region",
        "Scale",
        "Units",
        "Per...",
        "Method",
    ]
    # Keep LEAP-import numeric year columns contiguous immediately after Method.
    # Keep one blank spacer column between the final year and Level 1.
    # Put Level hierarchy columns at the far right to avoid LEAP interpreting them
    # as year/value cells during import.
    export_columns = non_year_order + year_columns + [level_spacer_column] + level_columns
    export_df = export_df.reindex(columns=export_columns).copy()

    output_root = _resolve(output_dir)
    output_root.mkdir(parents=True, exist_ok=True)
    workbook_path = output_root / str(file_name).strip()
    _archive_existing_results_file_if_needed(
        workbook_path,
        archive_dir=archive_dir,
        min_hours=archive_min_hours,
    )
    area_name, version_value = _extract_area_and_version(
        verification_preamble,
        viewing_preamble,
        leap_preamble,
    )
    width = len(export_columns)
    row0 = [""] * width
    row1 = [""] * width
    if width >= 8:
        row0[4] = "Area:"
        row0[5] = area_name
        row0[6] = "Ver:"
        row0[7] = version_value
    preamble = pd.DataFrame([row0, row1])

    with pd.ExcelWriter(workbook_path, engine="openpyxl", mode="w") as writer:
        preamble.to_excel(writer, sheet_name="Export", index=False, header=False)
        pd.DataFrame([export_columns]).to_excel(
            writer,
            sheet_name="Export",
            index=False,
            header=False,
            startrow=len(preamble),
        )
        export_df.to_excel(
            writer,
            sheet_name="Export",
            index=False,
            header=False,
            startrow=len(preamble) + 1,
        )
    if bool(archive_every_run):
        _archive_results_file_snapshot(
            workbook_path,
            archive_dir=archive_dir,
        )

    if (
        isinstance(dropped_unmatched_zero_supply_rows, pd.DataFrame)
        and not dropped_unmatched_zero_supply_rows.empty
    ):
        dropped_supply_report_path = (
            _resolve(RESULTS_CHECKS_DIR) / RESULTS_DROPPED_UNMATCHED_ZERO_SUPPLY_ROWS_FILENAME
        )
        dropped_supply_report_path.parent.mkdir(parents=True, exist_ok=True)
        _sort_output_frame_for_csv(dropped_unmatched_zero_supply_rows).to_csv(
            dropped_supply_report_path,
            index=False,
        )
        print(
            "[INFO] Dropped unmatched all-zero Resources rows not present in "
            f"verification export: {len(dropped_unmatched_zero_supply_rows)} "
            f"(details saved to {dropped_supply_report_path})."
        )

    if (
        isinstance(unmatched_nonzero_supply_rows, pd.DataFrame)
        and not unmatched_nonzero_supply_rows.empty
    ):
        print(
            "\n[WARN] Found nonzero Resources rows not present in verification export; "
            "kept in output for review."
        )
        print(f"[WARN] Nonzero unmatched Resources rows: {len(unmatched_nonzero_supply_rows)}")
        for _, row in unmatched_nonzero_supply_rows.head(30).iterrows():
            print(
                "  - Branch Path='{bp}' | Variable='{var}' | Scenario='{sc}' | Region='{rg}' | "
                "year_abs_sum={ys}".format(
                    bp=str(row.get("Branch Path") or "").strip(),
                    var=str(row.get("Variable") or "").strip(),
                    sc=str(row.get("Scenario") or "").strip(),
                    rg=str(row.get("Region") or "").strip(),
                    ys=float(pd.to_numeric(row.get("year_abs_sum"), errors="coerce") or 0.0),
                )
            )
        if len(unmatched_nonzero_supply_rows) > 30:
            print(
                f"  ... plus {len(unmatched_nonzero_supply_rows) - 30} more nonzero unmatched Resources rows"
            )

    if (
        isinstance(remapped_resource_branch_rows, pd.DataFrame)
        and not remapped_resource_branch_rows.empty
    ):
        print(
            "[INFO] Remapped Resources branch paths to canonical verification-export paths for "
            f"{len(remapped_resource_branch_rows)} row(s)."
        )
        for _, row in remapped_resource_branch_rows.head(20).iterrows():
            print(
                "  - Branch Path='{old}' -> '{new}' | Variable='{var}' | Scenario='{sc}' | "
                "Region='{rg}' | match_type='{mt}'".format(
                    old=str(row.get("Branch Path") or "").strip(),
                    new=str(row.get("reference_branch_path") or "").strip(),
                    var=str(row.get("Variable") or "").strip(),
                    sc=str(row.get("Scenario") or "").strip(),
                    rg=str(row.get("Region") or "").strip(),
                    mt=str(row.get("match_type") or "").strip(),
                )
            )
        if len(remapped_resource_branch_rows) > 20:
            print(
                f"  ... plus {len(remapped_resource_branch_rows) - 20} more remapped Resources rows"
            )

    if (
        isinstance(unresolved_resource_branch_rows, pd.DataFrame)
        and not unresolved_resource_branch_rows.empty
    ):
        print(
            "\n[WARN] Unresolved Resources branch-path mappings against verification export; "
            "these rows may still receive -1 IDs and need explicit LEAP mapping confirmation."
        )
        print(
            f"[WARN] Unresolved resource branch mappings: {len(unresolved_resource_branch_rows)}"
        )
        for _, row in unresolved_resource_branch_rows.head(30).iterrows():
            print(
                "  - Branch Path='{bp}' | Variable='{var}' | Scenario='{sc}' | Region='{rg}' | "
                "issue='{issue}' | candidates='{cand}'".format(
                    bp=str(row.get("Branch Path") or "").strip(),
                    var=str(row.get("Variable") or "").strip(),
                    sc=str(row.get("Scenario") or "").strip(),
                    rg=str(row.get("Region") or "").strip(),
                    issue=str(row.get("issue") or "").strip(),
                    cand=str(row.get("candidate_branch_paths") or "").strip(),
                )
            )
        if len(unresolved_resource_branch_rows) > 30:
            print(
                f"  ... plus {len(unresolved_resource_branch_rows) - 30} more unresolved mapping rows"
            )

    unmatched_report_path = _resolve(RESULTS_CHECKS_DIR) / RESULTS_UNMATCHED_ID_REPORT_FILENAME
    unmatched_report_path.parent.mkdir(parents=True, exist_ok=True)
    if isinstance(unmatched_id_rows, pd.DataFrame) and not unmatched_id_rows.empty:
        _sort_output_frame_for_csv(unmatched_id_rows).to_csv(unmatched_report_path, index=False)
        print("\n[WARN] Unmatched verification-export IDs detected; these rows need LEAP alignment fixes.")
        print(
            f"[WARN] Unmatched rows: {len(unmatched_id_rows)} "
            f"(details saved to {unmatched_report_path})"
        )
        for _, row in unmatched_id_rows.head(30).iterrows():
            print(
                "  - Branch Path='{bp}' | Variable='{var}' | Scenario='{sc}' | Region='{rg}' | reason='{rsn}'".format(
                    bp=str(row.get("Branch Path") or "").strip(),
                    var=str(row.get("Variable") or "").strip(),
                    sc=str(row.get("Scenario") or "").strip(),
                    rg=str(row.get("Region") or "").strip(),
                    rsn=str(row.get("reason") or "").strip(),
                )
            )
        if len(unmatched_id_rows) > 30:
            print(f"  ... plus {len(unmatched_id_rows) - 30} more unmatched rows")
    else:
        if unmatched_report_path.exists():
            try:
                unmatched_report_path.unlink()
                print(
                    "[INFO] Cleared stale unmatched-ID report from previous run: "
                    f"{unmatched_report_path}"
                )
            except Exception as exc:
                print(
                    "[WARN] Could not remove stale unmatched-ID report "
                    f"{unmatched_report_path}: {exc}"
                )

    if isinstance(metadata_backfill_rows, pd.DataFrame) and not metadata_backfill_rows.empty:
        print(
            "[INFO] Backfilled non-year metadata from verification export for "
            f"{len(metadata_backfill_rows)} field(s)."
        )
        by_column = (
            metadata_backfill_rows.groupby("column", dropna=False).size().reset_index(name="count")
        )
        for _, row in by_column.iterrows():
            print(f"  - {row['column']}: {int(row['count'])} fill(s)")

    if isinstance(metadata_mismatch_rows, pd.DataFrame) and not metadata_mismatch_rows.empty:
        mismatch_report_path = _resolve(RESULTS_CHECKS_DIR) / RESULTS_METADATA_MISMATCH_REPORT_FILENAME
        mismatch_report_path.parent.mkdir(parents=True, exist_ok=True)
        _sort_output_frame_for_csv(metadata_mismatch_rows).to_csv(
            mismatch_report_path,
            index=False,
        )
        print(
            "\n[WARN] Verification-export metadata mismatches detected "
            "(Scale/Units/Per...)."
        )
        print(
            f"[WARN] Metadata mismatches: {len(metadata_mismatch_rows)} "
            f"(details saved to {mismatch_report_path})"
        )
        for _, row in metadata_mismatch_rows.head(30).iterrows():
            print(
                "  - Branch Path='{bp}' | Variable='{var}' | Scenario='{sc}' | Region='{rg}' | "
                "column='{col}' | generated='{gen}' | reference='{ref}'".format(
                    bp=str(row.get("Branch Path") or "").strip(),
                    var=str(row.get("Variable") or "").strip(),
                    sc=str(row.get("Scenario") or "").strip(),
                    rg=str(row.get("Region") or "").strip(),
                    col=str(row.get("column") or "").strip(),
                    gen=str(row.get("generated_value") or "").strip(),
                    ref=str(row.get("reference_value") or "").strip(),
                )
            )
        if len(metadata_mismatch_rows) > 30:
            print(
                f"  ... plus {len(metadata_mismatch_rows) - 30} more metadata mismatches"
            )

    if (
        isinstance(mapping_config_mismatch_rows, pd.DataFrame)
        and not mapping_config_mismatch_rows.empty
    ):
        mapping_mismatch_report_path = (
            _resolve(RESULTS_CHECKS_DIR) / RESULTS_CONFIG_MAPPING_MISMATCH_REPORT_FILENAME
        )
        mapping_mismatch_report_path.parent.mkdir(parents=True, exist_ok=True)
        _sort_output_frame_for_csv(mapping_config_mismatch_rows).to_csv(
            mapping_mismatch_report_path,
            index=False,
        )
        print(
            "\n[WARN] Analysis-input config mapping mismatches detected against "
            "full model export metadata."
        )
        print(
            f"[WARN] Mapping mismatches: {len(mapping_config_mismatch_rows)} "
            f"(details saved to {mapping_mismatch_report_path})"
        )
        for _, row in mapping_config_mismatch_rows.head(30).iterrows():
            print(
                "  - scope='{scope}' | branch='{branch}' | variable='{var}' | "
                "field='{field}' | config='{cfg}' | reference='{ref}' | issue='{issue}'".format(
                    scope=str(row.get("match_scope") or "").strip(),
                    branch=str(row.get("branch_path") or "").strip(),
                    var=str(row.get("variable") or "").strip(),
                    field=str(row.get("field") or "").strip(),
                    cfg=str(row.get("config_value") or "").strip(),
                    ref=str(row.get("reference_values") or "").strip(),
                    issue=str(row.get("issue") or "").strip(),
                )
            )
        if len(mapping_config_mismatch_rows) > 30:
            print(
                f"  ... plus {len(mapping_config_mismatch_rows) - 30} more mapping mismatches"
            )

    print(
        "[INFO] Saved single-file results workbook in full-model Export structure to "
        f"{workbook_path}"
    )
    return workbook_path


def run_results_linked_transformation_supply_workflow(
    economies: Iterable[str] | None = None,
    scenario_names: list[str] | None = None,
    export_dataset_key: str = EXPORT_DATASET_KEY,
    include_leap_import: bool | None = None,
    import_scenarios: Iterable[str] | str | None = LEAP_IMPORT_SCENARIOS,
    use_direct_leap_results_for_demand: bool | None = None,
    scrape_leap_results: bool | None = None,
) -> dict[str, object]:
    """Build reconciled transformation + supply exports driven by LEAP balance demand results."""
    timer = workflow_common.WorkflowTimer("results_supply_link", enabled=ENABLE_WORKFLOW_TIMING)
    timing_path = _resolve(RESULTS_RUNTIME_DIR) / WORKFLOW_TIMING_FILENAME
    global _CAPACITY_UNMET_RUNTIME_CAPACITY_ADDITIONS
    global _CAPACITY_UNMET_RUNTIME_PRIMARY_ADDITIONS
    global _CAPACITY_UNMET_RUNTIME_EXPORT_ADJUSTMENTS
    global _CAPACITY_UNMET_RUNTIME_PASS_SUMMARY
    _CAPACITY_UNMET_RUNTIME_CAPACITY_ADDITIONS = {}
    _CAPACITY_UNMET_RUNTIME_PRIMARY_ADDITIONS = {}
    _CAPACITY_UNMET_RUNTIME_EXPORT_ADJUSTMENTS = {}
    _CAPACITY_UNMET_RUNTIME_PASS_SUMMARY = None
    requested_include_leap_import = include_leap_import
    analysis_write_mode = get_analysis_input_write_mode()
    include_leap_import = analysis_write_mode == "api"
    if requested_include_leap_import is not None and bool(requested_include_leap_import) != include_leap_import:
        print(
            "[INFO] include_leap_import argument is ignored in this workflow. "
            "LEAP import execution is derived from ANALYSIS_INPUT_WRITE_MODE "
            f"('{analysis_write_mode}')."
        )
    if use_direct_leap_results_for_demand is not None and not bool(use_direct_leap_results_for_demand):
        print(
            "[INFO] use_direct_leap_results_for_demand=False is deprecated and ignored. "
            "Demand inputs are always loaded from LEAP balance exports."
        )
    # Balance-export demand sourcing is now always enabled in this workflow.
    use_direct_leap_results_for_demand = True
    if scrape_leap_results is None:
        scrape_leap_results = bool(SCRAPE_LEAP_RESULTS)
    if _use_capacity_unmet_iterative_any_mode() and get_analysis_input_write_mode() != "workbook":
        raise ValueError(
            "TRADE_TARGET_EXPORT_MODE iterative unmet modes require "
            "ANALYSIS_INPUT_WRITE_MODE='workbook' so Analysis-view writes stay manual-import only."
        )
    if _use_capacity_unmet_iterative_any_mode() and scrape_leap_results:
        print(
            "[INFO] capacity_unmet iterative mode will refresh LEAP results templates "
            "via LEAP Results API reads before downstream reconciliation steps."
        )
    should_pin_leap_session = bool(
        scrape_leap_results
        or REFRESH_TRANSFORMATION_MEASURES_FROM_LEAP_RESULTS
        or include_leap_import
    )
    if should_pin_leap_session and leap_api.is_available():
        try:
            pinned_app = leap_api.connect(force_rebuild=False)
            active_area = str(getattr(pinned_app, "ActiveArea", "") or "").strip()
            if active_area:
                print(f"[INFO] Pinned LEAP session for this run (Active area: {active_area}).")
            else:
                print("[INFO] Pinned LEAP session for this run.")
        except Exception as exc:
            print(f"[WARN] Failed to pin LEAP session at run start: {exc}")
    archive_config_dir_once_per_day()
    os.environ["LEAP_IMPORT_LOG_LEVEL"] = str(LEAP_IMPORT_LOG_LEVEL).strip()
    os.environ["LEAP_IMPORT_WARNING_PRINT_LIMIT"] = str(LEAP_IMPORT_WARNING_PRINT_LIMIT)
    scenario_list = workflow_common.normalize_workflow_scenarios(
        scenario_names,
        SCENARIOS,
    )
    export_scenario_list = list(scenario_list)
    if RUN_RESET_SUPPLY_AND_TRANSFORMATION_IMPORT_EXPORT:
        expanded = _ensure_current_accounts_scenario(export_scenario_list)
        if len(expanded) != len(export_scenario_list):
            print(
                "[INFO] Reset mode: appending 'Current Accounts' to export scenarios "
                "so zero-reset values are also written for Current Accounts."
            )
        export_scenario_list = expanded
    balance_scenario_list = _filter_balance_scenarios(scenario_list)
    economy_list = workflow_common.normalize_economies(economies or ECONOMIES)
    _print_reset_mode_reminder(
        run_economies=economy_list,
        run_scenarios=export_scenario_list,
    )
    timer.lap("setup")

    probe_catalog_path = None
    if RUN_LEAP_FUEL_BRANCH_PROBE_AT_START:
        probe_catalog_path = refresh_fuel_branch_catalog_from_leap(
            output_path=LEAP_FUEL_BRANCH_PROBE_OUTPUT_PATH
        )
        timer.lap("refresh fuel branch catalog from LEAP")

    if scrape_leap_results:
        _run_leap_results_template_scrape()
        timer.lap("scrape LEAP results templates")

    comparison_long_df, mapping_status_df, balance_demand_issues, balance_matching_diagnostics = load_balance_demand_inputs(
        economies=economy_list,
        scenarios=balance_scenario_list,
        workbook_dir=LEAP_RESULTS_TABLES_DIR,
    )
    balance_demand_issues = _annotate_balance_demand_issue_scope(balance_demand_issues)
    sector_demand_table = load_results_sector_demand_table(
        source_priority=DEMAND_SOURCE_PRIORITY,
        comparison_long_df=comparison_long_df,
        mapping_status_df=mapping_status_df,
    )
    demand_table = load_results_demand_table(
        source_priority=DEMAND_SOURCE_PRIORITY,
        comparison_long_df=comparison_long_df,
        mapping_status_df=mapping_status_df,
    )
    timer.lap("load balance demand inputs")
    if economy_list:
        sector_demand_table = sector_demand_table[
            sector_demand_table["economy"].isin(economy_list)
        ].copy()
        demand_table = demand_table[demand_table["economy"].isin(economy_list)].copy()
    transformation_table = build_transformation_balance_table(economies=economy_list)
    transformation_sector_table = build_transformation_sector_table(economies=economy_list)
    transformation_target_rows, transformation_process_records = build_transformation_trade_target_rows(
        economies=economy_list,
    )
    supply_projection_table, assets = prepare_projected_supply_table(
        economies=economy_list,
        dataset_key=export_dataset_key,
    )
    supply_primary_table = prepare_supply_primary_table(
        assets,
        economies=economy_list,
        dataset_key=export_dataset_key,
    )
    supply_constraints, transformation_constraints = load_leap_constraint_tables(
        template_paths=CONSTRAINT_TEMPLATE_PATHS,
        sheet_names=CONSTRAINT_TEMPLATE_SHEETS,
        economies=economy_list,
    )
    timer.lap("build transformation and supply inputs")
    reconciliation_table = build_reconciliation_table(
        demand_table,
        transformation_table,
        supply_projection_table,
        supply_primary_table=supply_primary_table,
        supply_constraints=supply_constraints,
        transformation_constraints=transformation_constraints,
    )
    reconciliation_table = apply_trade_split_between_transformation_and_supply(
        reconciliation_table,
        transformation_target_rows=(
            transformation_target_rows if _use_legacy_trade_split_mode() else None
        ),
    )
    if RUN_RESET_SUPPLY_AND_TRANSFORMATION_IMPORT_EXPORT:
        reset_economies = RESET_SCOPE_ECONOMIES if RESET_SCOPE_ECONOMIES is not None else economy_list
        reset_scenarios = RESET_SCOPE_SCENARIOS if RESET_SCOPE_SCENARIOS is not None else export_scenario_list
        reset_scenarios = _ensure_current_accounts_scenario(reset_scenarios)
        reconciliation_table, updated_process_records = (
            reset_supply_and_transformation_import_export_to_zero(
                reconciliation_table=reconciliation_table,
                transformation_process_records=transformation_process_records,
                economies=reset_economies,
                scenarios=reset_scenarios,
                sector_titles=RESET_SCOPE_SECTOR_TITLES,
                esto_products=RESET_SCOPE_ESTO_PRODUCTS,
                years=RESET_SCOPE_YEARS,
            )
        )
        if updated_process_records is not None:
            transformation_process_records = updated_process_records
    timer.lap("build reconciliation and apply trade rules")

    balance_paths = save_year_balance_tables(
        reconciliation_table,
        years=BALANCE_EXPORT_YEARS,
        economies=economy_list,
        scenarios=balance_scenario_list,
    )
    balance_csv_paths = [path for path in balance_paths if Path(path).suffix.lower() == ".csv"]
    timer.lap("write yearly balance tables")

    if _use_capacity_unmet_iterative_mode():
        _CAPACITY_UNMET_RUNTIME_PASS_SUMMARY = _run_capacity_unmet_iterative_pass(
            reconciliation_table=reconciliation_table,
            process_records=transformation_process_records,
            economies=economy_list,
            scenarios=export_scenario_list,
            results_dir=balance_csv_paths,
            state_path=CAPACITY_UNMET_STATE_PATH,
            allow_same_results_reuse=bool(CAPACITY_UNMET_ALLOW_SAME_RESULTS_REUSE),
        )
    elif _use_capacity_unmet_iterative_balanced_mode():
        if _is_capacity_unmet_first_clean_run_mode():
            seeded_state = _read_capacity_unmet_state(
                state_path=CAPACITY_UNMET_STATE_PATH,
                run_mode="first_clean",
            )
            seeded_state_path = _write_capacity_unmet_state(
                seeded_state, state_path=CAPACITY_UNMET_STATE_PATH
            )
            _CAPACITY_UNMET_RUNTIME_PASS_SUMMARY = {
                "timestamp_utc": datetime.now(timezone.utc).isoformat(),
                "mode": "capacity_unmet_iterative_balanced",
                "iteration_run_mode": "first_clean",
                "state_path": str(_resolve(CAPACITY_UNMET_STATE_PATH)),
                "state_seeded_path": str(seeded_state_path),
                "seed_action": (
                    "Baseline-only first pass: wrote imports=0 with baseline exports+capacity "
                    "with no residual allocation from existing LEAP results tables."
                ),
                "next_manual_step": (
                    "Import generated workbook into LEAP, recalculate, refresh results tables, "
                    "set CAPACITY_UNMET_ITERATION_RUN_MODE='consecutive', rerun."
                ),
            }
            print(
                "[CAPACITY_UNMET_ITERATIVE_BALANCED] first_clean baseline pass: "
                "skipping residual allocation and using imports=0 with baseline exports/capacity."
            )
        else:
            _CAPACITY_UNMET_RUNTIME_PASS_SUMMARY = _run_capacity_unmet_iterative_balanced_pass(
                reconciliation_table=reconciliation_table,
                process_records=transformation_process_records,
                economies=economy_list,
                scenarios=export_scenario_list,
                results_dir=balance_csv_paths,
                state_path=CAPACITY_UNMET_STATE_PATH,
                allow_same_results_reuse=bool(CAPACITY_UNMET_ALLOW_SAME_RESULTS_REUSE),
            )
    if _use_capacity_unmet_iterative_any_mode():
        timer.lap("capacity unmet handling")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    balance_demand_issue_path: Path | None = None
    reconciliation_path: Path | None = None
    conventional_balance_paths: list[Path] = []
    if RESULTS_WRITE_LEGACY_SIDECAR_FILES:
        reconciliation_path = OUTPUT_DIR / RECONCILIATION_FILENAME
        reconciliation_table.to_csv(reconciliation_path, index=False)
        print(f"Saved reconciliation table to {reconciliation_path}")
        conventional_balance_paths = save_conventional_balance_tables(
            reconciliation_table,
            sector_demand_table,
            transformation_sector_table,
            supply_primary_table,
            assets[4],
            years=BALANCE_EXPORT_YEARS,
            economies=economy_list,
            scenarios=balance_scenario_list,
        )
        timer.lap("write legacy sidecar outputs")

    overrides = build_supply_overrides(reconciliation_table)
    dataset_map, sector_config, code_to_name_mapping, _, _ = assets
    supply_measures = _build_supply_measures_for_trade_mode()
    export_paths = supply_data_pipeline.generate_supply_exports(
        dataset_map,
        sector_config,
        code_to_name_mapping,
        projection_lookup=supply_data_pipeline.SUPPLY_PROJECTION_LOOKUP,
        projection_years=supply_data_pipeline.PROJECTION_YEAR_RANGE,
        dataset_key=export_dataset_key,
        economies=economy_list,
        scenario_names=export_scenario_list,
        base_year=BASE_YEAR,
        final_year=FINAL_YEAR,
        export_output_dir=EXPORT_OUTPUT_DIR,
        filename_template=EXPORT_FILENAME_TEMPLATE,
        flow_value_overrides_by_economy=overrides,
        supply_measures=supply_measures,
        keep_all_zero_rows=bool(KEEP_ALL_ZERO_SUPPLY_ROWS),
    )
    # Build catalog from static sources (LEAP probe + full-model export) so aux-fuel
    # branches not covered by the current run can be explicitly zeroed in LEAP.
    pre_run_catalog_df = _build_transformation_supply_fuel_catalog_df(
        transformation_export_paths=[],
        supply_export_paths=[],
        include_print_summary=False,
    )
    transformation_export_paths = save_transformation_exports_with_split_targets(
        reconciliation_table,
        transformation_target_rows,
        transformation_process_records,
        scenarios=export_scenario_list,
        output_dir=TRANSFORMATION_EXPORT_OUTPUT_DIR,
        filename_template=TRANSFORMATION_EXPORT_FILENAME_TEMPLATE,
        full_branch_catalog_df=pre_run_catalog_df if not pre_run_catalog_df.empty else None,
    )
    transfer_export_paths = save_transfer_exports_with_supply_overrides(
        reconciliation_table,
        economies=economy_list,
        scenarios=export_scenario_list,
        output_dir=TRANSFORMATION_EXPORT_OUTPUT_DIR,
        filename_template=transfers_workflow.EXPORT_FILENAME_TEMPLATE,
    )
    combined_export_path = save_combined_supply_transformation_export(
        supply_export_paths=[path for _, path in export_paths],
        transformation_export_paths=transformation_export_paths,
        transfer_export_paths=transfer_export_paths,
        output_dir=EXPORT_OUTPUT_DIR,
        filename_template=COMBINED_EXPORT_FILENAME_TEMPLATE,
        economy_label=economy_list[0] if economy_list else "economy",
        scenarios=export_scenario_list,
    )
    fuel_branch_catalog_df = _build_transformation_supply_fuel_catalog_df(
        transformation_export_paths=transformation_export_paths,
        supply_export_paths=[path for _, path in export_paths],
        include_print_summary=True,
    )
    timer.lap("generate LEAP import workbooks")
    fuel_branch_catalog_path: Path | None = None
    if RESULTS_WRITE_LEGACY_SIDECAR_FILES:
        fuel_branch_catalog_path = _build_transformation_supply_fuel_catalog(
            transformation_export_paths=transformation_export_paths,
            supply_export_paths=[path for _, path in export_paths],
            output_dir=OUTPUT_DIR,
        )
    leap_import_result = {"supply_imported": [], "transformation_imported": [], "transfer_imported": []}
    if include_leap_import:
        leap_import_result = run_results_linked_leap_import(
            [path for _, path in export_paths],
            transformation_export_paths,
            transfer_export_paths=transfer_export_paths,
            scenarios=export_scenario_list,
            import_scenarios=import_scenarios,
            region=LEAP_IMPORT_REGION,
            create_branches=LEAP_IMPORT_CREATE_BRANCHES,
            fill_branches=LEAP_IMPORT_FILL_BRANCHES,
            include_current_accounts=(
                LEAP_IMPORT_INCLUDE_CURRENT_ACCOUNTS
                or RUN_RESET_SUPPLY_AND_TRANSFORMATION_IMPORT_EXPORT
            ),
            import_supply_to_leap=LEAP_IMPORT_SUPPLY_TO_LEAP,
            import_transformation_to_leap=LEAP_IMPORT_TRANSFORMATION_TO_LEAP,
            import_transfers_to_leap=LEAP_IMPORT_TRANSFERS_TO_LEAP,
        )
        timer.lap("run LEAP import")

    results_workbook_path: Path | None = None
    if RESULTS_SINGLE_FILE_OUTPUT:
        resolved_single_file_name = _resolve_results_single_file_name(
            RESULTS_SINGLE_FILE_NAME,
            trade_mode=TRADE_TARGET_EXPORT_MODE,
            iteration_run_mode=CAPACITY_UNMET_ITERATION_RUN_MODE,
        )
        results_workbook_path = save_results_linked_single_workbook(
            reconciliation_table=reconciliation_table,
            sector_demand_table=sector_demand_table,
            demand_table=demand_table,
            transformation_table=transformation_table,
            transformation_sector_table=transformation_sector_table,
            supply_projection_table=supply_projection_table,
            supply_primary_table=supply_primary_table,
            transformation_target_rows=transformation_target_rows,
            fuel_branch_catalog_df=fuel_branch_catalog_df,
            base_df=assets[4],
            years=BALANCE_EXPORT_YEARS,
            economies=economy_list,
            scenarios=balance_scenario_list,
            export_paths=[path for _, path in export_paths],
            transformation_export_paths=transformation_export_paths,
            transfer_export_paths=transfer_export_paths,
            combined_export_path=combined_export_path,
            probe_catalog_path=probe_catalog_path,
            leap_import_result=leap_import_result,
            output_dir=OUTPUT_DIR,
            file_name=resolved_single_file_name,
            archive_dir=RESULTS_SINGLE_FILE_ARCHIVE_DIR,
            archive_min_hours=RESULTS_SINGLE_FILE_ARCHIVE_MIN_HOURS,
            archive_every_run=RESULTS_SINGLE_FILE_ARCHIVE_EVERY_RUN,
        )
        timer.lap("write consolidated run workbook")

    balance_matching_diagnostics_path = _resolve(RESULTS_CHECKS_DIR) / RESULTS_BALANCE_MATCHING_DIAGNOSTICS_FILENAME
    balance_matching_diagnostics_path.parent.mkdir(parents=True, exist_ok=True)
    _sort_output_frame_for_csv(
        balance_matching_diagnostics,
        exclude_sort_columns=("source_workbook", "source_sheet"),
    ).to_csv(balance_matching_diagnostics_path, index=False)
    timer.lap("write balance matching diagnostics")

    if not balance_demand_issues.empty:
        balance_demand_issue_path = _resolve(RESULTS_CHECKS_DIR) / RESULTS_BALANCE_DEMAND_ISSUES_FILENAME
        balance_demand_issue_path.parent.mkdir(parents=True, exist_ok=True)
        _sort_output_frame_for_csv(
            balance_demand_issues,
            exclude_sort_columns=("source", "source_sheet"),
        ).to_csv(balance_demand_issue_path, index=False)
        actionable_balance_demand_issues = balance_demand_issues[
            balance_demand_issues.get("demand_relevant", False).fillna(False).astype(bool)
        ].copy()
        reason_counts = (
            actionable_balance_demand_issues.groupby("reason", dropna=False)
            .size()
            .reset_index(name="row_count")
            .sort_values(["row_count", "reason"], ascending=[False, True])
        )
        counts_text = ", ".join(
            f"{row.reason}: {int(row.row_count)}" for row in reason_counts.itertuples(index=False)
        )
        timer.lap("write balance-demand issue report")
        ignored_issue_count = int(len(balance_demand_issues) - len(actionable_balance_demand_issues))
        if (
            ignored_issue_count > 0
            and actionable_balance_demand_issues.empty
        ):
            print(
                "[INFO] Ignoring non-demand balance mapping issues that do not affect "
                f"results_supply_link demand inputs. See {balance_demand_issue_path}. "
                f"Ignored rows: {ignored_issue_count}"
            )
        elif ignored_issue_count > 0:
            print(
                "[INFO] Ignoring balance mapping issues outside demand-side inputs. "
                f"Actionable rows: {len(actionable_balance_demand_issues)}. "
                f"Ignored rows: {ignored_issue_count}. See {balance_demand_issue_path}."
            )
        if BALANCE_DEMAND_FAIL_ON_MAPPING_ISSUES and not actionable_balance_demand_issues.empty:
            timer.finish(status="failed")
            if WRITE_WORKFLOW_TIMING_CSV:
                timer.write_csv(timing_path)
            raise RuntimeError(
                "Demand-relevant balance-demand mapping issues remain unresolved after writing "
                "results_supply_link outputs. "
                f"See {balance_demand_issue_path}. Counts: {counts_text}"
            )
        if not actionable_balance_demand_issues.empty:
            print(
            "[WARN] Balance-demand mapping issues remain unresolved, but "
            "BALANCE_DEMAND_FAIL_ON_MAPPING_ISSUES=False so the workflow is continuing. "
            f"See {balance_demand_issue_path}. Counts: {counts_text}"
            )
    else:
        balance_demand_issue_path = None
    timer.finish()
    if WRITE_WORKFLOW_TIMING_CSV:
        timer.write_csv(timing_path)

    return {
        "results_workbook_path": results_workbook_path,
        "reconciliation_csv": reconciliation_path,
        "balance_table_paths": balance_paths,
        "conventional_balance_paths": conventional_balance_paths,
        "export_paths": [path for _, path in export_paths],
        "transformation_export_paths": transformation_export_paths,
        "transfer_export_paths": transfer_export_paths,
        "combined_export_path": combined_export_path,
        "fuel_branch_probe_path": probe_catalog_path,
        "fuel_branch_catalog_path": fuel_branch_catalog_path,
        "demand_mapping_issues_csv": balance_demand_issue_path,
        "direct_demand_mapping_gaps_csv": balance_demand_issue_path,
        "balance_matching_diagnostics_csv": balance_matching_diagnostics_path,
        "leap_import_result": leap_import_result,
        "capacity_unmet_iterative_summary": _CAPACITY_UNMET_RUNTIME_PASS_SUMMARY,
        "workflow_stage_timings_csv": str(timing_path),
        "row_count": int(len(reconciliation_table)),
    }


def run_results_linked_supply_workflow(
    economies: Iterable[str] | None = None,
    scenario_names: list[str] | None = None,
    export_dataset_key: str = EXPORT_DATASET_KEY,
    include_leap_import: bool | None = None,
    import_scenarios: Iterable[str] | str | None = LEAP_IMPORT_SCENARIOS,
    use_direct_leap_results_for_demand: bool | None = None,
    scrape_leap_results: bool | None = None,
) -> dict[str, object]:
    """Backward-compatible alias for the transformation+supply runner."""
    return run_results_linked_transformation_supply_workflow(
        economies=economies,
        scenario_names=scenario_names,
        export_dataset_key=export_dataset_key,
        include_leap_import=include_leap_import,
        import_scenarios=import_scenarios,
        use_direct_leap_results_for_demand=use_direct_leap_results_for_demand,
        scrape_leap_results=scrape_leap_results,
    )


# -----------------------------------------------------------------------------
# Notebook Runtime Variables (single editable block)
# -----------------------------------------------------------------------------
# Edit these values in notebooks before calling `run_with_config()`.
# ECONOMIES = ["20_USA"]
# SCENARIOS = list(workflow_cfg.SUPPLY_NOTEBOOK_SCENARIOS)
# TRADE_TARGET_EXPORT_MODE = "capacity_unmet_iterative_balanced"  # Options: "legacy_trade_split", "capacity_unmet_iterative", "capacity_unmet_iterative_balanced"
# SCRAPE_LEAP_RESULTS = False
# RUN_LEAP_FUEL_BRANCH_PROBE_AT_START = True
# RESULTS_SINGLE_FILE_OUTPUT = True
# RESULTS_WRITE_LEGACY_SIDECAR_FILES = False
# BALANCE_DEMAND_FAIL_ON_MAPPING_ISSUES = True
# ENABLE_WORKFLOW_TIMING = True
# WRITE_WORKFLOW_TIMING_CSV = True
# ENABLE_COMPLETION_BEEP = True
# RESULTS_SINGLE_FILE_ARCHIVE_MIN_HOURS = 24
# RESULTS_SINGLE_FILE_ARCHIVE_EVERY_RUN = True
# RUN_RESET_SUPPLY_AND_TRANSFORMATION_IMPORT_EXPORT = False
#%%
ECONOMIES = ["20_USA"]
SCENARIOS = ['Target', 'Current Accounts']#list(workflow_cfg.SUPPLY_NOTEBOOK_SCENARIOS)#'Target', 
TRADE_TARGET_EXPORT_MODE = "capacity_unmet_iterative_balanced"  # "legacy_split" | "capacity_unmet_iterative" | "capacity_unmet_iterative_balanced"

CAPACITY_UNMET_ITERATION_RUN_MODE = "consecutive"  # consecutive|first_clean
SCRAPE_LEAP_RESULTS = False
RUN_LEAP_FUEL_BRANCH_PROBE_AT_START = False
RESULTS_SINGLE_FILE_OUTPUT = True
RESULTS_WRITE_LEGACY_SIDECAR_FILES = False
BALANCE_DEMAND_FAIL_ON_MAPPING_ISSUES = True
ENABLE_WORKFLOW_TIMING = True
WRITE_WORKFLOW_TIMING_CSV = True
KEEP_ALL_ZERO_SUPPLY_ROWS = True
ENABLE_COMPLETION_BEEP = True
COMPLETION_BEEP_ON_ERROR = True
COMPLETION_BEEP_COUNT = 1
COMPLETION_BEEP_FREQUENCY_HZ = 880
COMPLETION_BEEP_DURATION_MS = 180
COMPLETION_BEEP_PAUSE_SECONDS = 0.12
RESULTS_SINGLE_FILE_ARCHIVE_MIN_HOURS = 24
RESULTS_SINGLE_FILE_ARCHIVE_EVERY_RUN = True
RUN_RESET_SUPPLY_AND_TRANSFORMATION_IMPORT_EXPORT = False

def run_with_config() -> dict[str, object]:
    """Run the notebook-configured results-linked transformation+supply workflow."""
    analysis_write_mode = get_analysis_input_write_mode()
    include_leap_import = analysis_write_mode == "api"
    resolved_single_file_name = _resolve_results_single_file_name(
        RESULTS_SINGLE_FILE_NAME,
        trade_mode=TRADE_TARGET_EXPORT_MODE,
        iteration_run_mode=CAPACITY_UNMET_ITERATION_RUN_MODE,
    )
    if int(supply_data_pipeline.EXPORT_FINAL_YEAR) > int(LEAP_IMPORT_MAX_YEAR):
        print(
            "[WARN] supply_data_pipeline.EXPORT_FINAL_YEAR="
            f"{supply_data_pipeline.EXPORT_FINAL_YEAR} exceeds LEAP max year {LEAP_IMPORT_MAX_YEAR}; "
            f"results_supply_link_workflow is clamping FINAL_YEAR to {FINAL_YEAR}."
        )
    print(
        "[INFO] run_with_config toggles: "
        f"TRADE_TARGET_EXPORT_MODE={TRADE_TARGET_EXPORT_MODE}, "
        f"CAPACITY_UNMET_ITERATION_RUN_MODE={CAPACITY_UNMET_ITERATION_RUN_MODE}, "
        "RUN_RESET_SUPPLY_AND_TRANSFORMATION_IMPORT_EXPORT="
        f"{RUN_RESET_SUPPLY_AND_TRANSFORMATION_IMPORT_EXPORT}, "
        f"ANALYSIS_INPUT_WRITE_MODE={analysis_write_mode}, "
        f"LEAP_IMPORT_LOG_LEVEL={LEAP_IMPORT_LOG_LEVEL}, "
        f"RUN_LEAP_FUEL_BRANCH_PROBE_AT_START={RUN_LEAP_FUEL_BRANCH_PROBE_AT_START}, "
        f"INCLUDE_LEAP_IMPORT={include_leap_import} (derived), "
        f"LEAP_IMPORT_SUPPLY_TO_LEAP={LEAP_IMPORT_SUPPLY_TO_LEAP}, "
        f"LEAP_IMPORT_TRANSFORMATION_TO_LEAP={LEAP_IMPORT_TRANSFORMATION_TO_LEAP}, "
        f"LEAP_IMPORT_TRANSFERS_TO_LEAP={LEAP_IMPORT_TRANSFERS_TO_LEAP}, "
        f"LEAP_IMPORT_INCLUDE_CURRENT_ACCOUNTS={LEAP_IMPORT_INCLUDE_CURRENT_ACCOUNTS}, "
        f"SCRAPE_LEAP_RESULTS={SCRAPE_LEAP_RESULTS}, "
        f"RESULTS_SINGLE_FILE_OUTPUT={RESULTS_SINGLE_FILE_OUTPUT}, "
        f"RESULTS_WRITE_LEGACY_SIDECAR_FILES={RESULTS_WRITE_LEGACY_SIDECAR_FILES}, "
        f"RESULTS_SINGLE_FILE_NAME={resolved_single_file_name}, "
        f"RESULTS_SINGLE_FILE_ARCHIVE_EVERY_RUN={RESULTS_SINGLE_FILE_ARCHIVE_EVERY_RUN}, "
        f"BALANCE_DEMAND_FAIL_ON_MAPPING_ISSUES={BALANCE_DEMAND_FAIL_ON_MAPPING_ISSUES}, "
        f"ENABLE_WORKFLOW_TIMING={ENABLE_WORKFLOW_TIMING}, "
        f"WRITE_WORKFLOW_TIMING_CSV={WRITE_WORKFLOW_TIMING_CSV}, "
        f"KEEP_ALL_ZERO_SUPPLY_ROWS={KEEP_ALL_ZERO_SUPPLY_ROWS}, "
        f"ENABLE_COMPLETION_BEEP={ENABLE_COMPLETION_BEEP}"
    )
    try:
        output = run_results_linked_transformation_supply_workflow(
            economies=ECONOMIES,
            scenario_names=SCENARIOS,
            export_dataset_key=EXPORT_DATASET_KEY,
            include_leap_import=include_leap_import,
            import_scenarios=LEAP_IMPORT_SCENARIOS,
            scrape_leap_results=SCRAPE_LEAP_RESULTS,
        )
    except Exception:
        if bool(COMPLETION_BEEP_ON_ERROR):
            _emit_completion_beep(success=False)
        raise
    _emit_completion_beep(success=True, style="chime")
    return output

#%%
if __name__ == "__main__":
    run_with_config()
#%%
